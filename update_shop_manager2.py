# coding:utf-8
import requests
import json
import time
import xlrd
from openpyxl import Workbook, load_workbook

#请求头部
authorization = 'Bearer f7YkroV6PTSqw5WDPw3FEQ'
hex_server_session = '5cc8acf0-328b-4a17-afa5-b9e46c5f9752'
test_url = 'http://test.dairyqueen.com.cn'
store_url = 'http://store.dairyqueen.com.cn'
store_ppj_url = 'http://store.papajohnshanghai.com'
test_be_url = 'http://hwstest.brutcakecafe.com'
store_be_url = 'http://store.brutcakecafe.com'

headers1 = {
    'authorization': '{}'.format(authorization),
    'Cookie': 'hex_server_session={}'.format(hex_server_session),
}

headers2 = {
        'authorization': '{}'.format(authorization),
        'Content-Type': 'charset=utf8',
        'Cookie': 'hex_server_session={}'.format(hex_server_session),
    }

branch_dict = {}
store_id_dict = {}


def get_product_id(id):
    """
    根据商品编码拿到商品id
    """
    url = test_url + '/api/v1/product?search={}&is_task=true&state=draft%2Cenabled&status=&include_request=true' \
                        '&is_new=true&include_state=true'.format(id)
    # print(url)
    response = requests.get(url, headers=headers1)
    # print(response.status_code)
    data = response.json()
    # print(type(data), data)
    product_id = None
    try:
        dic = data['payload']
        for i in dic:
            if i['code'] == str(id):
                product_id = i['id']
    except IndexError as e:
        print('物料编码{}不存在'.format(id))
        return
    # dic = data['payload'][0]
    # product_id = dic['id']
    # print('product_id', product_id)
    # data = json.dumps(data, sort_keys=True, indent=4, separators=(',', ':'))
    # print(type(data), data)
    # print(product_id)
    if not product_id:
        print('物料编码{}不存在'.format(id))
        return
    return product_id


def change_category(id, branch_id):
    url = test_url + '/api/v1/store/{}?stringified=true'.format(id)
    s = {"relation": {"branch": '{}'.format(branch_id)}}
    # print(s)
    data = json.dumps(s)
    response = requests.request("PUT", url, headers=headers1, data=data)
    data = response.json()
    return


def get_request_id2(id):
    """
    拿到所有变更计划的 request_id
    """
    url = test_url + '/api/v1/store/task?stringified=true&include_total=true&record_id={}&order=asc&offset=0&limit=1'.format(id)
    # print(url)
    response = requests.get(url, headers=headers1)
    # print(response.status_code)
    data = response.json()
    dic = data['payload']
    lis = dic['rows']
    # print('lis', lis)
    approve_id = lis[0]['id']
    return approve_id


def confirm_changes2(request_id):
    """
    审核变更计划
    """
    url = test_url + '/api/v1/store/task/{}/status/APPROVED?stringified=true'.format(request_id)
    # print(url)
    response = requests.request("PUT", url, headers=headers1)
    data = response.json()
    # print(data)
    # time.sleep(0.1)
    return


def get_request_id(code):
    url = test_url + '/api/v1/store/{}?is_task=true&relation=all&code=all&is_new=true&include_state=true&stringified=true'.format(code)
    print(url)
    response = requests.get(url, headers=headers1)
    data = response.json()
    # print(type(data), data)
    request_id = None
    try:
        dic = data['payload']
        request_id = dic['request_id']
        name = dic['name']
    except IndexError as e:
        print('门店{}不存在'.format(code))
        return
    if not request_id:
        print('门店{}修改出错'.format(code))
        return
    return request_id, name


def accept_action2(name, request_id):
    """
    接受更改操作,用map去拼接正确的名字 eg.DQ全国
    """

    url = test_url + '/api/v1/store/change/{}/to/task?stringified=true'.format(request_id)
    s = {"name": "{}".format(name), "immediate": False, "start_time": "2021-06-30 17:41:17"}
    # print(s)
    data = json.dumps(s, ensure_ascii=False).encode("utf-8")
    # print(data)
    response = requests.post(url, headers=headers2, data=data)
    data2 = response.json()
    # time.sleep(0.1)
    return


def one_trun(code, branch_id, branch_name):
    change_category(code, branch_id)
    request_id, name = get_request_id(code)

    accept_action2(name, request_id)
    request_id2 = get_request_id2(code)
    confirm_changes2(request_id2)
    print('门店{}修改区经理{}'.format(name, branch_name))
    return


def test():
    wb = Workbook()
    ws = wb.active
    file_path = '/Users/yjq/Desktop/test.xlsx'
    # 打开一个已有文件
    wb = load_workbook(file_path)
    sheet_list = wb.sheetnames
    sheet = wb[wb.sheetnames[0]]
    total_list = []
    for r in range(2, sheet.max_row + 1):
        row_list = []  # 每一行建立一个list
        for c in range(1, sheet.max_column + 1):
            v = str(sheet.cell(r, c).value)
            v = v.replace('\n', '')
            row_list.append(v)
        total_list.append(row_list)
    print(len(total_list), total_list)
    return total_list


def get_branch():
    url = test_url + '/api/v1/region/branch?stringified=true&include_total=true&include_state=true&include_parents=true&relation=all&search_fields=name,code&order=asc&offset=0&limit=300&state=draft%2Cenabled&status=&include_request=true&is_new=true&include_state=true&_=1625035086928'
    response = requests.get(url, headers=headers1)
    data = response.json()
    try:
        dic = data['payload']
        lis = dic['rows']
        total = dic['total']
        for i in lis:
            name = i['name']
            _id = i['id']
            if name in branch_dict:
                print('请检查{}是否唯一'.format(name))
            branch_dict[name] = _id
        # print(len(branch_dict), branch_dict)
    except IndexError as e:
        print('获取管理区域出错', e)
        return
    if total != len(branch_dict):
        print('管理区域数量不一致,请检查是否同名区经理')
    return branch_dict


def get_store_id():
    url = test_url + '/api/v1/store?code=all&include_state=true&include_total=true&relation=all&search_fields=extend_code.ex_code%2Cextend_code.us_id%2Cextend_code.ex_id%2Ccode%2Cname%2Caddress%2Crelation.geo_region.name%2Crelation.branch.name%2Crelation.distribution_region.name%2Crelation.attribute_region.name%2Crelation.formula_region.name%2Crelation.market_region.name%2Crelation.order_region.name&stringified=true&is_task=true&sort=extend_code.ex_code&order=asc&offset=0&limit=1500&state=draft%2Cenabled&status=&include_request=true&is_new=true&include_state=true&_=1625036536949'
    response = requests.get(url, headers=headers1)
    data = response.json()
    try:
        dic = data['payload']
        lis = dic['rows']
        total = dic['total']
        for i in lis:
            us_id = i['extend_code']['us_id']
            _id = i['id']
            if us_id in store_id_dict:
                print('请检查{}是否唯一'.format(us_id))
            store_id_dict[us_id] = _id
        # print(len(store_id_dict), store_id_dict)
    except IndexError as e:
        print('获取门店id出错', e)
        return
    if total != len(store_id_dict):
        print('门店数量不一致,请检查是否同名美编')
    return store_id_dict


if __name__ == '__main__':
    # lis = test()
    # dic_branch = get_branch()
    # dic_store = get_store_id()
    # for i in lis:
    #     one_trun(dic_store[i[0]], dic_branch[i[1]], i[1])
    #     print('----------------------------------')
    #     time.sleep(0.1)
    get_request_id(3935321603065405441)

    # 使用说明：打开网页获取此时的authorization; 全局修改url为想要执行的品牌; 转换模版格式
    # 批量执行时, 可能会因执行一直发请求等原因报错, 重新执行即可————这条不确定，建议从执行失败的重新执行
    # 文件格式如下：门店美编 ｜ 区经理名称
    #            42608   ｜ 陈超
    #            43460   ｜ 赵海迪







