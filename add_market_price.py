"""
# 新增价格区域
http://test.dairyqueen.com.cn/api/v1/product/3850201175919951873/region/market?stringified=true

allow_sale: true
currency: "CNY"
id: "4055613822868967425"
retail: 10
takeaway_dianping_mobile: false
tax_rate: 0
unit_id: "3932395201819324417"


# 生效价格
http://test.dairyqueen.com.cn/api/v1/product/region/market/4522654325176737793/state/enable?stringified=true

"""
import requests
import openpyxl
import pymongo
import time
from random import randint

# store_url = 'http://test.dairyqueen.com.cn'
store_url = 'http://store.dairyqueen.com.cn'

default_unit_id = '3932395201819324417'

headers = {
    'authorization': '{}'.format('Bearer eJpUMkgXOrWKFIvnm8WwBA'),  # 该env环境下的authorization
    'Cookie': 'hex_server_session={}'.format('c1d28cbd-8cec-4b5c-bac8-ca357567154c'),  # 该env环境下的Cookie
}

USER_AGENTS = [
	 "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; AcooBrowser; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
	 "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0; Acoo Browser; SLCC1; .NET CLR 2.0.50727; Media Center PC 5.0; .NET CLR 3.0.04506)",
	 "Mozilla/4.0 (compatible; MSIE 7.0; AOL 9.5; AOLBuild 4337.35; Windows NT 5.1; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
	 "Mozilla/5.0 (Windows; U; MSIE 9.0; Windows NT 9.0; en-US)",
	 "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Win64; x64; Trident/5.0; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 2.0.50727; Media Center PC 6.0)",
	 "Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 1.0.3705; .NET CLR 1.1.4322)",
	 "Mozilla/4.0 (compatible; MSIE 7.0b; Windows NT 5.2; .NET CLR 1.1.4322; .NET CLR 2.0.50727; InfoPath.2; .NET CLR 3.0.04506.30)",
	 "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN) AppleWebKit/523.15 (KHTML, like Gecko, Safari/419.3) Arora/0.3 (Change: 287 c9dfb30)",
	 "Mozilla/5.0 (X11; U; Linux; en-US) AppleWebKit/527+ (KHTML, like Gecko, Safari/419.3) Arora/0.6",
	 "Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.8.1.2pre) Gecko/20070215 K-Ninja/2.1.1",
	 "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9) Gecko/20080705 Firefox/3.0 Kapiko/3.0",
	 "Mozilla/5.0 (X11; Linux i686; U;) Gecko/20070322 Kazehakase/0.4.5",
	 "Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.9.0.8) Gecko Fedora/1.9.0.8-1.fc10 Kazehakase/0.5.6",
	 "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11",
	 "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_3) AppleWebKit/535.20 (KHTML, like Gecko) Chrome/19.0.1036.7 Safari/535.20",
	 "Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; fr) Presto/2.9.168 Version/11.52",
]


def get_product_id_by_mongo():
	mongo_client = pymongo.MongoClient(host='127.0.0.1', port=27072,)
	db_name = mongo_client.saas_dq

	query = {
		'status': 'ENABLED',
		'data_state': 'ENABLED'
	}

	result = db_name.product.find(query, {'_id': 1, 'name': 1, 'code': 1, 'relation.unit.id': 1})

	product_dic = {r['code']: r for r in result}
	product_ids = [i['_id'] for i in list(product_dic.values())]

	return product_dic, product_ids

def get_product_market_raltion(product_ids):
	mongo_client = pymongo.MongoClient(host='127.0.0.1', port=27072,)
	db_name = mongo_client.saas_dq

	# format_ids = ["NumberLong('{}'.format(i))" for i in product_ids]
	query = {
		'org_id': {'$in': product_ids},
		'to_org_type': 'market_region',
		'data_state': 'DRAFT'
	}

	result = db_name.product_relation.find(query, {'_id': 1, 'org_id': 1})
	product_market_relation_res = {i['org_id']: i['_id'] for i in result}
	return product_market_relation_res


path = '/Users/hws/Downloads/DQ商品售价-市场区域-20210810（拉萨门店）.xlsx'

wb = openpyxl.load_workbook(path)

sh = wb['Sheet2']
rows = sh.max_row
cols = sh.max_column

# print(rows, cols)

product_dic, product_ids = get_product_id_by_mongo()
# print(product_dic)

all_ids = []
for row in range(2, rows + 1):
	
	product_code = str(sh.cell(row, 2).value)
	
	market_code = str(sh.cell(row, 4).value)

	price = str(sh.cell(row, 5).value)


	product_id = product_dic.get(product_code, {}).get('_id')
	unit_id_res = product_dic.get(product_code, {}).get('relation', {}).get('unit', [])
	unit_id = unit_id_res and unit_id_res[0].get('id') or default_unit_id
	if not product_id:
		print('第{}行商品{}不存在'.format(row, product_code))
		continue
	all_ids.append(product_id)

	data = {
		'market': [{		
			'allow_sale': True,
			'currency': "CNY",
			'id': "4522620766904156160",  # 市场区域ID, 就改一个，先写成固定的
			'retail': price,
			'takeaway_dianping_mobile': False,
			'tax_rate': 0,
			'unit_id': unit_id	
		}]
	}
	# print(data)
	random_agent = USER_AGENTS[randint(0, len(USER_AGENTS)-1)]
	headers.update({
		'User-Agent':random_agent,
		})
	res = requests.post(store_url + '/api/v1/product/{}/region/market?stringified=true'.format(product_id), headers=headers, json=data)
	print(res.json())
	time.sleep(1)

print(len(all_ids))
product_market_relation_res = get_product_market_raltion(all_ids)

for i in all_ids:
	# if str(i) != '3850201109834498049':
	# 	continue
	relation_id = product_market_relation_res.get(i)
	print(i, relation_id)
	random_agent = USER_AGENTS[randint(0, len(USER_AGENTS)-1)]
	headers.update({
		'User-Agent':random_agent,
		})
	requests.put(store_url + '/api/v1/product/region/market/{}/state/enable?stringified=true'.format(relation_id), headers=headers)
	time.sleep(0.5)












