import requests
import json
import time
import xlrd
from openpyxl import Workbook, load_workbook

final_url = 'https://store.papajohnshanghai.com'


headers1 = {
    'authorization': 'Bearer Y-abXu0GPfCsg0foNf5IZQ',
    'Cookie': 'hex_server_session={}'.format('75034011-57fa-4415-b282-48f133fda13f'),
    'content-type': 'application/json'
}



def get_product_id(id):
    """
    根据商品编码拿到商品id及单位
    """
    url = final_url + '/api/v1/product?search={}&is_task=true&state=draft%2Cenabled&status=&include_request=true' \
                        '&is_new=true&relation=all&include_state=true'.format(id)
    # print(url)
    response = requests.get(url, headers=headers1)
    # print(response.status_code)
    data = response.json()
    # print(type(data), data)
    product_id = None
    try:
        dic = data['payload']
        for i in dic:
            if i['code'] == str(id):
                product_id = i['id']
                status = i['status']
                name = i['name']
    except IndexError as e:
        print('物料编码{}不存在IndexError'.format(id))
        return 0, 0
    if not product_id:
        print('物料编码{}不存在'.format(id))
        return 0, 0
    return product_id, status, name

def disable_status(pid):
    url = final_url + '/api/v1/product/{}?stringified=true'.format(pid)
    s = {"status": "DISABLED"}
    # print(s)
    data = json.dumps(s)
    response = requests.request("PUT", url, headers=headers1, data=data)
    data = response.json()
    time.sleep(3)
    return

def get_request_id(code):
    url = final_url + '/api/v1/product?search={}'.format(code)
    # print(url)
    response = requests.get(url, headers=headers1)
    data = response.json()
    # print(type(data), data)
    request_id = None
    try:
        dic = data['payload']
        for i in dic:
            if i['code'] == str(code):
                request_id = i['request_id']
                name = i['name']
    except IndexError as e:
        print('物料编码{}不存在11111'.format(id))
        return None, None
    if not request_id:
        print('物料编码{}不存在222'.format(id))
        return None, None
    return request_id, name

def accept_action2(name, request_id):
    """
    接受更改操作,用map去拼接正确的名字 eg.DQ全国
    """

    url = final_url + '/api/v1/product/change/{}/to/task?stringified=true'.format(request_id)
    s = {"name": "{}".format(name), "immediate": True, "start_time": ""}
    # print(s)
    data = json.dumps(s, ensure_ascii=False).encode("utf-8")
    # print(data)
    response = requests.post(url, headers=headers1, data=data)
    data2 = response.json()
    # time.sleep(0.1)
    return

def get_region_order_rel_id(product_id):
    url = final_url + '/api/v1/product/{}/region/order?stringified=true&include_total=true&is_task=true&order=asc' \
                     '&state=draft%2Cenabled&status=&include_request=true&is_new=true&include_state=true'.format(product_id)
    response = requests.get(url, headers=headers1)
    # print(response.status_code)
    data = response.json()
    dic = data['payload']
    lis = dic['rows']
    rel_id_new = [i.get('rel_id') for i in lis]
    # print('订货区域区域：%s' % rel_id_new)
    return rel_id_new

def accept_region_order(rel_id):
    url = final_url + '/api/v1/product/region/order/{}/state/disable?stringified=true'.format(rel_id)
    response = requests.request("PUT", url, headers=headers1)
    data = response.json()
    return

def get_distribution_order_rel_id(product_id):
    url = final_url + '/api/v1/product/{}/region/distribution?stringified=true&include_total=true&is_task=true&order=asc' \
                     '&state=draft%2Cenabled&status=&include_request=true&is_new=true&include_state=true'.format(product_id)
    response = requests.get(url, headers=headers1)
    # print(response.status_code)
    data = response.json()
    dic = data['payload']
    lis = dic['rows']
    rel_id_new = [i.get('rel_id') for i in lis]
    # print('配送区域：%s' % rel_id_new)
    return rel_id_new

def accept_distribution_order(rel_id):
    url = final_url + '/api/v1/product/region/distribution/{}/state/disable?stringified=true'.format(rel_id)
    response = requests.request("PUT", url, headers=headers1)
    data = response.json()
    return

def get_attribute_order_rel_id(product_id):
    url = final_url + '/api/v1/product/{}/region/attribute?stringified=true&include_total=true&is_task=true&order=asc' \
                     '&state=draft%2Cenabled&status=&include_request=true&is_new=true&include_state=true'.format(product_id)
    response = requests.get(url, headers=headers1)
    # print(response.status_code)
    data = response.json()
    dic = data['payload']
    lis = dic['rows']
    rel_id_new = [i.get('rel_id') for i in lis]
    # print('属性区域：%s' % rel_id_new)
    return rel_id_new

def accept_attribute_order(rel_id):
    url = final_url + '/api/v1/product/region/attribute/{}/state/disable?stringified=true'.format(rel_id)
    response = requests.request("PUT", url, headers=headers1)
    data = response.json()
    return

def get_task_id(code):
    url = final_url + '/api/v1/product/task?stringified=true&include_total=true&search_fields=name&order=asc&offset=0&limit=10'
    response = requests.get(url, headers=headers1)
    data = response.json()
    row = data['payload']['rows']
    print(row)
    now_id = None
    for i in row:
        if code == i.get('name') and i.get('process_status') == 'INITED':
            now_id = i.get('id')
            break
    return now_id

def put_task_id(task_id):
    url = final_url + 'api/v1/product/task/{}/status/APPROVED?stringified=true'.format(task_id)

    response = requests.request("PUT", url, headers=headers1)
    return 
def one_trun(code):
    # 禁用商品
    product_id, status,name = get_product_id(code)
    if not product_id:
        print('物料编码{}不存在'.format(code))
        return
    if status != 'DISABLED':
        disable_status(product_id)
        print('商品{}修改为禁用状态'.format(code))
        print(code, '*' * 10)
        request_id, name = get_request_id(code)
        print(request_id, name)
        accept_action2(name, request_id)
        print('商品{}接受生成变更计划'.format(code))
    else:
        print('商品{}已经是禁用状态'.format(code))
  
    # 禁用商品订货区域
    rel_ids = get_region_order_rel_id(product_id)
    for r in rel_ids:
        accept_region_order(r)
    print('商品订货区域：%s 完成' % code)
    # 禁用配送区域
    dis_ids = get_distribution_order_rel_id(product_id)
    for d in dis_ids:
        accept_distribution_order(d)
    print('商品配送区域：%s 完成' % code)
    # 禁用属性区域
    att_ids = get_attribute_order_rel_id(product_id)
    for a in att_ids:
        accept_attribute_order(a)
    print('商品属性区域：%s 完成' % code)
    print('商品：%s 完成' % code)
    

    
    time.sleep(3)
    return


def push_task(code):
    product_id, status,name = get_product_id(code)
    now_id = get_task_id(name)
    if now_id:
        put_task_id(now_id)
        print('code:%s 接受变更计划')


def test():
    wb = Workbook()
    ws = wb.active
    # file_path = '/Users/hws/Downloads/PPJHEX物料主档批量禁用-20221009 (1).xlsx'
    file_path = '/Users/hws/Downloads/工作簿1.xlsx'
    # 打开一个已有文件
    wb = load_workbook(file_path)
    sheet_list = wb.sheetnames
    sheet = wb[wb.sheetnames[0]]
    total_list = []
    for r in range(2, sheet.max_row + 1):
        row_list = []  # 每一行建立一个list
        for c in range(1, sheet.max_column + 1):
            v = str(sheet.cell(r, c).value)
            v = v.replace('\n', '')
            row_list.append(v)
        total_list.append(row_list)
    print(len(total_list), total_list)
    return total_list


lis = test()
for i in lis:
    print(i[1])
    one_trun(i[1])
# for i in lis:
# 	push_task(i[1])
