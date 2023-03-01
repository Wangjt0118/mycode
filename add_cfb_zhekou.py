###########################################################
# 1、获取现有已生效分类(用于判断是否已经存在)  GET
# /api/orgs/discount-category

# 2、新建折扣分类 (先建一级，后二级) POST
# /api/orgs/discount-category

# 3、分类生效
# /api/v1/discount-category/{category_id}/state/enable
###########################################################   

import requests
import openpyxl

host = {
	'dq_host': 'http://test.dairyqueen.com.cn',
	'ppj_host': 'http://test.dairyqueen.com.cn',
	'ym_host': 'http://teststore.meet-xiaomian.com',
	'be_host': 'http://hwstest.brutcakecafe.com',
}

# host = {
# 	'dq_host': 'http://store.dairyqueen.com.cn',
# 	'ppj_host': 'https://store.papajohnshanghai.com',
# 	'ym_host': 'http://store.meet-xiaomian.com',
# 	'be_host': 'http://store.brutcakecafe.com',
# }

headers = {
    'authorization': '{}'.format('Bearer oHAmikLuoc2m5hCMKYnGi0XIEApyEe'),  # 该env环境下的authorization
    'Cookie': 'hex_server_session={}'.format('be51beb2-6353-4bc4-ba9a-5176cc7c9704'),  # 该env环境下的Cookie
}

def get_env_host(env):

	host_info = {
		'category_url':'{}/api/orgs/discount-category'.format(host.get('{}_host'.format(env.lower()))),
		'enable_category_url': '%s/api/v1/discount-category/{}/state/enable' % host.get('{}_host'.format(env.lower())),
	}


	return host_info


# 根据Excel组合成层级关系
def read_excel_info_base(cfb_env, path):
	
	wb = openpyxl.load_workbook(path)

	sh = wb['%s折扣' % cfb_env]

	rows = sh.max_row
	cols = sh.max_column

	upload_dic = {}
	first_dic = {}
	second_dic = {}
	for row in range(2, rows + 1):

		first_name = sh.cell(row, 1).value
		first_code = str(sh.cell(row, 2).value)
		second_name = sh.cell(row, 3).value
		second_code = str(sh.cell(row, 4).value)
		if second_code and second_code != 'None':

			first_dic[first_code] = first_name
			second_dic[second_code] = second_name
			upload_dic.setdefault(first_code, set())
			upload_dic[first_code].add(second_code)
		
	print(upload_dic, '\n', first_dic, '\n', second_dic)
	return upload_dic, first_dic, second_dic


def read_sys_category(env):
	res = get_env_host(env)
	host = res['category_url']

	response = requests.get(host, headers=headers)

	category_res = response.json()['payload']

	return {r.get('code'): r.get('id') for r in category_res}

def create_zk_category(env, category_code, category_name, parent_id):
	# 根据现有名称、父级名称建折扣分类
	res = get_env_host(env)
	host = res['category_url']
	post_data = {
		'code': category_code,
		'name': category_name,
		'parent_id': parent_id
	}
	print(post_data, '添加折扣&&&&&&&&&&')
	response = requests.post(host, headers=headers, json=post_data)
	print(response.json(), '添加折扣======')
	category_res = response.json()['payload']

	return category_res.get('id')

def enable_zk_category(env, category_id):
	res = get_env_host(env)
	host = res['enable_category_url'].format(category_id)
	response = requests.put(host, headers=headers)
	
	return True

def main(env, path):
	upload_dic, first_dic, second_dic = read_excel_info_base(env, path)

	sys_category = read_sys_category(env)
	parent_code_id = sys_category.get('03')
	print(parent_code_id, '*' * 10, sys_category)

	if parent_code_id:

		for first_code, secode_code_list in upload_dic.items():
			first_id = sys_category.get(first_code)
			# 一级目录
			if not first_id:
				# 创建分类
				first_id = create_zk_category(env, first_code, first_dic[first_code], parent_code_id)
				sys_category[first_code] = first_id
				# 生效分类
				enable_zk_category(env, first_id)

			# 二级分类
			for sec_code  in list(secode_code_list):
				second_id = sys_category.get(sec_code)
				if not second_id:
					second_id = create_zk_category(env, sec_code, second_dic[sec_code], first_id)
					sys_category[sec_code] = second_id
					enable_zk_category(env, second_id)
	return True





if __name__ == "__main__":
	# path = '/Users/hws/Downloads/CFB新增商品类别数据.xlsx'
	# main('PPJ', path)
	env = 'BE'
	l = [('80016-1234', '80016-1234-test')]

	active_ids = []
	for i in l:
		ca_id = create_zk_category(env, i[0], i[1], "4163554550471278593")
		active_ids.append(ca_id)

	# ca_id = create_zk_category(env, '90013', 'OK卡', 4520115621821710336)
	# active_ids.append(ca_id)

	# print(active_ids)
