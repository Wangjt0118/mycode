import openpyxl

path = '/Users/hws/Downloads/product_market_price{}.xlsx'


result = []
for i in range(1, 6):
	wb = openpyxl.load_workbook(path.format(i))

	sh = wb['Sheet1']

	rows = sh.max_row
	cols = sh.max_column

	print(rows)
	result.append(())

	for c in range(2, rows + 1):
		if sh.cell(c, 1).value:
			result.append((sh.cell(c, 1).value, sh.cell(c, 2).value, sh.cell(c, 3).value, sh.cell(c, 4).value, sh.cell(c, 5).value))



outwb = openpyxl.Workbook()
outws = outwb.create_sheet(index=0)
title = ['商品名称', '商品编码', '市场区域名称', '市场区域编码', '价格']
for i in range(1, 6):
	outws.cell(1, i).value = title[i -1]

index = 2
for product in result:
	print(product)
	if not product:
		continue
	outws.cell(index, 1).value = product[0]
	outws.cell(index, 2).value = product[1]
	outws.cell(index, 3).value = product[2]
	outws.cell(index, 4).value = product[3]
	outws.cell(index, 5).value = product[4]
	index += 1


filename = '/Users/hws/Downloads/product_market_price_all.xlsx'
print(filename + '  down!!')
outwb.save(filename)	