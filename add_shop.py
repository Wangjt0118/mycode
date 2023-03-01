import openpyxl
import requests
import json

path = '/Users/hws/Downloads/测试门店清单.xlsx'
wb = openpyxl.load_workbook(path)

sh = wb['YM']

rows = sh.max_row
cols = sh.max_column
print(rows, cols)

shop_infos = []

for i in range(2, rows + 1):
	us_id = sh.cell(i, 1).value
	ex_cost_center_code = sh.cell(i, 5).value
	shop_name = sh.cell(i, 8).value
	name_en = sh.cell(i, 9).value
	address = sh.cell(i, 15).value
	store_type = "DRS" if '直营' in sh.cell(i, 20).value else 'FRS'

	shop_infos.append({
		"name": shop_name,
		"name_en": name_en,
		"store_type": store_type,
		"status": "OPENED",
		"code": "",
		"extend_code": {
		"us_id": us_id,
		"ex_cost_center_code": ex_cost_center_code
		},
		"currency": "CNY",
		"open_date": "2022-09-28",
		"close_date": "2099-09-28",
		"address": address,
		"relation": {}
		})
headers = {
    'authorization': 'Bearer wv1zjVbXP6OujYO7rwuGCw',
    'Cookie': 'hex_server_session={}'.format('4d24a349-00ea-4d63-8f2c-e6c4f80ee353'),
    'content-type': 'application/json'
}
get_shop_url = "http://teststore.meet-xiaomian.com/api/v1/store?code=all&include_state=true&include_total=true&relation=all&search_fields=extend_code.ex_code%2Cextend_code.us_id%2Cextend_code.ex_id%2Ccode%2Cname%2Caddress%2Crelation.geo_region.name%2Crelation.branch.name%2Crelation.distribution_region.name%2Crelation.attribute_region.name%2Crelation.formula_region.name%2Crelation.market_region.name%2Crelation.order_region.name&stringified=true&is_task=true&sort=extend_code.ex_code&order=asc&offset=0&limit=1000&state=draft%2Cenabled&status=&include_request=true&is_new=true&include_state=true&_=1664332102744"
res = requests.get(get_shop_url, headers=headers)
shop_us_id = [r.get('extend_code').get('us_id') for r in res.json().get('payload').get('rows')]
# print(shop_us_id)
url = "https://test.dairyqueen.com.cn/api/v1/store?stringified=true"
for shop in shop_infos:
	if shop.get('extend_code').get('us_id') not in shop_us_id:
		# print(shop)
		requests.post(url, json=shop, headers=headers)
		print(shop.get('name') + '创建成功！')
	else:
		print('已存在: %s' % shop)




# path = '/Users/hws/Downloads/测试门店清单.xlsx'
# wb = openpyxl.load_workbook(path)

# sh = wb['DQ']

# rows = sh.max_row
# cols = sh.max_column
# print(rows, cols)

# shop_infos = []

# for i in range(2, rows + 1):