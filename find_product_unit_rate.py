import requests
import openpyxl


######################参数修改部分################
env = 'be'

host = {
	'dq_host': 'https://store.dairyqueen.com.cn',
	'ppj_host':  'https://store.papajohnshanghai.com',
	'xm_host': 'http://store.meet-xiaomian.com',
	'be_host': 'http://store.brutcakecafe.com'
}
category_id = 4442541390362423297  # 该env环境下的分类ID

base_info = {
	'url': '{}/api/v1/product'.format(host.get('{}_host'.format(env))),
	'unit_url': '{}/api/v1/product/unit'.format(host.get('{}_host'.format(env))),
	'headers': {
	    'authorization': '{}'.format('Bearer 2fJE1BOPR1UKLmjCxH8ZhBvrTVtmIT'),  # 该env环境下的authorization
	    'Cookie': 'hex_server_session={}'.format('2ca26a03-7fd2-4670-864f-9ddff211cbfb'),  # 该env环境下的Cookie
	}
}

request_unit_url = '{}/api/v1/product/unit'.format(host.get('{}_host'.format(env))),

filename2 = '/Users/hws/Downloads/{}_product_unit_rate.xlsx'.format(env)
######################参数修改部分################


response = requests.get(request_unit_url, headers=base_info.get('headers'),)
unit_info = response.json()['payload']

unit_dic = {}
for unit in unit_info:
	unit_dic[str(unit['id'])] = unit['name']

response = requests.get(base_info.get('url'), headers=base_info.get('headers'), params={'category': category_id, 'relation': 'all'})
res_json = response.json()

product_list = res_json['payload']


outwb = openpyxl.Workbook()
outws = outwb.create_sheet(index=0)

title = ['商品编码', '商品名称', '名称', '换算比率', '核算单位', '销售单位', '配方单位', '订货单位', '盘点单位', '复核单位']
for i in range(1, 11):
	outws.cell(1, i).value = title[i -1]

index = 2
for product in product_list:
	
	p_code = product['code']
	p_name = product['name']

	
	unit_info = product.get('relation').get('unit') or []

	for col in range(len(unit_info)):
		outws.cell(index, 1).value = p_code
		outws.cell(index, 2).value = p_name

		outws.cell(index, 3).value = unit_dic[str(unit_info[col].get('id'))] or ''
		outws.cell(index, 4).value = unit_info[col].get('rate') or False
		outws.cell(index, 5).value = unit_info[col].get('default') or False
		outws.cell(index, 6).value = unit_info[col].get('sales') or False
		outws.cell(index, 7).value = unit_info[col].get('bom') or False
		outws.cell(index, 8).value = unit_info[col].get('order') or False
		outws.cell(index, 9).value = unit_info[col].get('stocktake') or False
		outws.cell(index, 10).value = unit_info[col].get('recheck') or False
		index += 1
	

print(filename2 + '  down!!')
outwb.save(filename2)