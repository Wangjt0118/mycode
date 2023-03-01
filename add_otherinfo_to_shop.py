# 先从MongoDB读取shop的信息，要手动格式化一下，或者有api读取shop信息
import openpyxl
import requests
shop_list = [{
    "_id": "3850146064724131841",
    "name": "嘉定罗宾森店",
    "extend_code": {
        "comm_shop_id": "3601d022edb545c5bf8aa2dab4bcd6a6",
        "ex_code": "20016",
        "alipay_id": "2015060900077000000000176457",
        "us_id": "42041",
        "comm_code": "301003400000115",
        "upcard_terminal": "02148860",
        "upcard_mer_id": "102210058120739",
        "ex_id": "223",
        "ex_cost_center_code": "1200042041",
        "dcore_store_appid": "s20170823000005197"
    }
},
{
    "_id": "3850146080528269313",
    "name": "上海仲盛店",
    "extend_code": {
        "comm_shop_id": "cddcc4bd4cf24ba4ade6d7606fa4b48d",
        "ex_code": "20043",
        "alipay_id": "2015060900077000000000178453",
        "us_id": "42256",
        "comm_code": "301003400000377",
        "upcard_terminal": "02194925",
        "upcard_mer_id": "102210058120718",
        "ex_id": "240",
        "ex_cost_center_code": "1200042256",
        "dcore_store_appid": "s20170823000005487"
    }
},
{
    "_id": "3850146089889955841",
    "name": "重庆洪崖洞店",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20053",
        "comm_shop_id": "b0d8987d48c043a1b2fdf1bec7f95fd4",
        "us_id": "42337",
        "alipay_id": "2015061200077000000000193010",
        "takeaway_eleme_id": '',
        "upcard_terminal": "02306269",
        "comm_code": "301003400000320",
        "upcard_mer_id": "102230058120110",
        "ex_id": "632",
        "ex_cost_center_code": "1200042337",
        "dcore_store_appid": "s20170823000005496"
    }
},
{
    "_id": "3850146093790658561",
    "name": "徐州彭城店",
    "extend_code": {
        "comm_shop_id": "71930119cd0d49fa8db1dbc99570087f",
        "ex_code": "20060",
        "alipay_id": "2015061100077000000000182834",
        "us_id": "42378",
        "comm_code": "301003400000225",
        "upcard_terminal": "51211252",
        "upcard_mer_id": "102512058120390",
        "ex_id": "991",
        "ex_cost_center_code": "1200042378",
        "dcore_store_appid": "s20170823000005562"
    }
},
{
    "_id": "3850146067915997185",
    "name": "南京水游城",
    "extend_code": {
        "comm_shop_id": "f23da3b0a54f4b4698a31a74df9c1bcc",
        "ex_code": "20022",
        "alipay_id": "2015061000077000000000191173",
        "us_id": "42134",
        "comm_code": "301003400000435",
        "upcard_terminal": "02516545",
        "upcard_mer_id": "102250058120833",
        "ex_id": "304",
        "ex_cost_center_code": "1200042134",
        "dcore_store_appid": "s20170823000005478"
    }
},
{
    "_id": "3850146075474132993",
    "name": "上海华诚店",
    "extend_code": {
        "comm_shop_id": "b64aa6c8e7be4a068064cff1e6d5fc6d",
        "ex_code": "20033",
        "alipay_id": "2015060900077000000000174612",
        "us_id": "42253",
        "comm_code": "301003400000334",
        "upcard_terminal": "02148897",
        "upcard_mer_id": "102210058120702",
        "ex_id": "241",
        "ex_cost_center_code": "1200042253",
        "dcore_store_appid": "s20170823000005486"
    }
},
{
    "_id": "3850146102749691905",
    "name": "大华店",
    "extend_code": {
        "comm_shop_id": "36327010dc4b4406ad080b6e88663884",
        "ex_code": "20078",
        "alipay_id": "2015060900077000000000174615",
        "us_id": "42427",
        "comm_code": "301003400000116",
        "upcard_terminal": "02194800",
        "upcard_mer_id": "102210058126900",
        "ex_id": "261",
        "ex_cost_center_code": "1200042427",
        "dcore_store_appid": "s20170823000005210"
    }
},
{
    "_id": "3850146106943995905",
    "name": "宁波银泰江东店",
    "extend_code": {
        "comm_shop_id": "49e6e190ceca4f1eb1de70d968aefdf4",
        "ex_code": "20086",
        "alipay_id": "2015060900077000000000178459",
        "us_id": "42467",
        "comm_code": "301003400000149",
        "upcard_terminal": "57401978",
        "upcard_mer_id": "102574058120109",
        "ex_id": "433",
        "ex_cost_center_code": "1200042467",
        "dcore_store_appid": "s20170823000005573"
    }
},
{
    "_id": "3850146109762568193",
    "name": "上海飞洲店",
    "extend_code": {
        "comm_shop_id": "4dc4941a29ed4c529a839cd9b9195922",
        "ex_code": "20092",
        "alipay_id": "2020011500077000000086880219",
        "us_id": "42503",
        "comm_code": "301003400000335",
        "upcard_terminal": "02194796",
        "upcard_mer_id": "102210058126896",
        "ex_id": "270",
        "ex_cost_center_code": "1200042503",
        "dianping_store_id": "4125018",
        "dcore_store_appid": "s20170823000005500"
    }
},
{
    "_id": "3850146111314460673",
    "name": "重庆西城天街店",
    "extend_code": {
        "comm_shop_id": "07a1ce65483744dd92f681d82438fe55",
        "ex_code": "20095",
        "alipay_id": "2015061200077000000000188807",
        "us_id": "42507",
        "comm_code": "301003400000033",
        "upcard_terminal": "02306262",
        "upcard_mer_id": "102230058120117",
        "ex_id": "638",
        "ex_cost_center_code": "1200042507",
        "dcore_store_appid": "s20170823000005213"
    }
},
{
    "_id": "3850146118616743937",
    "name": "太原铜锣湾店",
    "extend_code": {
        "comm_shop_id": "952bb1e3197045ae9b23d8b94cc969a9",
        "ex_code": "20109",
        "alipay_id": "2015061100077000000000191215",
        "us_id": "42572",
        "comm_code": "301003400000274",
        "upcard_terminal": "35103739",
        "upcard_mer_id": "102351058121620",
        "ex_id": "691",
        "ex_cost_center_code": "1200042572",
        "dcore_store_appid": "s20170823000005507"
    }
},
{
    "_id": "3850146062694088705",
    "name": "成都盐市口店",
    "extend_code": {
        "comm_shop_id": "c506ba759c104ee48adc51c1a3958f78",
        "ex_code": "20012",
        "alipay_id": "2015061200077000000000182874",
        "us_id": "41961",
        "comm_code": "301003400000355",
        "upcard_terminal": "02817522",
        "upcard_mer_id": "102280058121932",
        "ex_id": "209",
        "ex_cost_center_code": "1200041961",
        "dcore_store_appid": "s20170823000005471"
    }
},
{
    "_id": "3850146114422439937",
    "name": "淮安中央新亚店",
    "extend_code": {
        "comm_shop_id": "9e54150610e340fe8e677d68ded338b9",
        "ex_code": "20101",
        "alipay_id": "2015061100077000000000188753",
        "us_id": "42571",
        "comm_code": "301003400000293",
        "upcard_terminal": "51211882",
        "upcard_mer_id": "102512058120392",
        "ex_id": "611",
        "ex_cost_center_code": "1200042571",
        "dcore_store_appid": "s20170929000006494"
    }
},
{
    "_id": "3850146068947795969",
    "name": "常熟虞景文华店",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20024",
        "comm_shop_id": "f65c1f05e9e740e791f44c4e08a7d961",
        "us_id": "42144",
        "alipay_id": "2015061100077000000000191198",
        "takeaway_eleme_id": '',
        "upcard_terminal": "51215443",
        "comm_code": "301003400000443",
        "upcard_mer_id": "102512058123023",
        "ex_id": "554",
        "ex_cost_center_code": "1200042144",
        "dcore_store_appid": "s20170823000005199"
    }
},
{
    "_id": "3850146069987983361",
    "name": "福州万象城店",
    "extend_code": {
        "comm_shop_id": "169177d275b0427a82535c201bf57cba",
        "ex_code": "20026",
        "alipay_id": "2015061200077000000000192991",
        "us_id": "42207",
        "comm_code": "301003400000056",
        "upcard_terminal": "59105748",
        "upcard_mer_id": "102591058120151",
        "ex_id": "854",
        "ex_cost_center_code": "1200042207",
        "dcore_store_appid": "s20170823000005481"
    }
},
{
    "_id": "3850146123163369473",
    "name": "宜昌万达店",
    "extend_code": {
        "comm_shop_id": "ee762a825b4d4a1f9f31c99139ef16b8",
        "ex_code": "20117",
        "alipay_id": "2015061200077000000000194551",
        "us_id": "42614",
        "comm_code": "301003400000428",
        "upcard_terminal": "71700535",
        "upcard_mer_id": "102717058120032",
        "ex_id": "792",
        "ex_cost_center_code": "1200042614",
        "dcore_store_appid": "s20170823000005514"
    }
},
{
    "_id": "3850146071627956225",
    "name": "常州莱蒙店",
    "extend_code": {
        "comm_shop_id": "5963ee101ddd44ddaa4780a0541d8d6f",
        "ex_code": "20029",
        "alipay_id": "2015061000077000000000192926",
        "us_id": "42229",
        "comm_code": "301003400000181",
        "upcard_terminal": "51211275",
        "upcard_mer_id": "102512058120367",
        "ex_id": "556",
        "ex_cost_center_code": "1200042229",
        "dcore_store_appid": "s20170823000005484"
    }
},
{
    "_id": "3850146131107381249",
    "name": "上海我格广场店",
    "extend_code": {
        "comm_shop_id": "c0098057bb5445c5b0b182f4346c1b3c",
        "ex_code": "20132",
        "alipay_id": "2015060900077000000000176463",
        "us_id": "42582",
        "comm_code": "301003400000344",
        "upcard_terminal": "02148838",
        "upcard_mer_id": "102210058120844",
        "ex_id": "283",
        "ex_cost_center_code": "1200042582",
        "dcore_store_appid": "s20170823000005508"
    }
},
{
    "_id": "3850146072097718273",
    "name": "扬州珍园店",
    "extend_code": {
        "comm_shop_id": "a730f974c3e447d3be04e808ceecbdde",
        "ex_code": "20030",
        "alipay_id": "2015061100077000000000188755",
        "us_id": "42235",
        "comm_code": "301003400000312",
        "upcard_terminal": "51211881",
        "upcard_mer_id": "102512058120376",
        "ex_id": "702",
        "ex_cost_center_code": "1200042235",
        "dcore_store_appid": "s20170823000005201"
    }
},
{
    "_id": "3850146077982326785",
    "name": "上海调频壹店",
    "extend_code": {
        "comm_shop_id": "21e3a1578bd84f118e1861a1a997de63",
        "ex_code": "20038",
        "alipay_id": "2015060900077000000000166169",
        "us_id": "42264",
        "comm_code": "301003400000073",
        "upcard_terminal": "02194798",
        "upcard_mer_id": "102210058126898",
        "ex_id": "243",
        "ex_cost_center_code": "1200042264",
        "dcore_store_appid": "s20170823000005488"
    }
},
{
    "_id": "3850146129471602689",
    "name": "郑州嘉茂店",
    "extend_code": {
        "comm_shop_id": "58115443c255454c83954c9275a5aef1",
        "ex_code": "20129",
        "alipay_id": "2015061100077000000000194502",
        "us_id": "42637",
        "comm_code": "301003400000179",
        "upcard_terminal": "37108518",
        "upcard_mer_id": "102371058121212",
        "ex_id": "934",
        "ex_cost_center_code": "1200042637",
        "dcore_store_appid": "s20170823000005219"
    }
},
{
    "_id": "3850146095304802305",
    "name": "青岛延吉万达店",
    "extend_code": {
        "comm_shop_id": "90146645791e43caa835e5f52f0804d1",
        "ex_code": "20063",
        "alipay_id": "2015061100077000000000194507",
        "us_id": "42374",
        "comm_code": "301003400000269",
        "upcard_terminal": "53204092",
        "upcard_mer_id": "102532058120478",
        "ex_id": "181",
        "ex_cost_center_code": "1200042374",
        "dcore_store_appid": "s20170823000005206"
    }
},
{
    "_id": "3850146131585531905",
    "name": "苏州宫巷店",
    "extend_code": {
        "comm_shop_id": "f4bb237b5e88402ab574c34cb20ac94f",
        "ex_code": "20133",
        "alipay_id": "2015061000077000000000192925",
        "us_id": "42657",
        "comm_code": "301003400000442",
        "upcard_terminal": "51211247",
        "upcard_mer_id": "102512058120459",
        "ex_id": "509",
        "ex_cost_center_code": "1200042657",
        "dcore_store_appid": "s20170823000005521"
    }
},
{
    "_id": "3850146104419024897",
    "name": "郑州国贸店",
    "extend_code": {
        "comm_shop_id": "c5e8f8638cb5400791d06cf1ca821154",
        "ex_code": "20081",
        "alipay_id": "2015061100077000000000182837",
        "us_id": "42444",
        "comm_code": "301003400000356",
        "upcard_terminal": "37110808",
        "upcard_mer_id": "102371058121023",
        "ex_id": "933",
        "ex_cost_center_code": "1200042444",
        "dcore_store_appid": "s20170823000005570"
    }
},
{
    "_id": "3850146104876204033",
    "name": "南昌动壹店",
    "extend_code": {
        "comm_shop_id": "4c375abf3a914937b9c05233a58d3457",
        "ex_code": "20082",
        "alipay_id": "2015061000077000000000191159",
        "us_id": "42459",
        "comm_code": "301003400000151",
        "upcard_terminal": "79101366",
        "upcard_mer_id": "102791058120008",
        "ex_id": "981",
        "ex_cost_center_code": "1200042459",
        "dcore_store_appid": "s20170823000005572"
    }
},
{
    "_id": "3850146148555685889",
    "name": "莘庄龙之梦店",
    "extend_code": {
        "comm_shop_id": "4c5456e37d1f478a89cb2fbc4340f9a0",
        "ex_code": "20165",
        "alipay_id": "2015060900077000000000178454",
        "us_id": "42720",
        "comm_code": "301003400000152",
        "upcard_terminal": "02194926",
        "upcard_mer_id": "102210058120846",
        "ex_id": "288",
        "ex_cost_center_code": "1200042720",
        "dcore_store_appid": "s20170823000005223"
    }
},
{
    "_id": "3850146072571674625",
    "name": "成都锦里店",
    "extend_code": {
        "comm_shop_id": "532c7c9cf78d4e34a2bde16ee67bdc8d",
        "ex_code": "20031",
        "alipay_id": "2021090900077000000027663651",
        "us_id": "42236",
        "comm_code": "301003400000168",
        "upcard_terminal": "02817519",
        "upcard_mer_id": "102280058121935",
        "ex_id": "805",
        "ex_cost_center_code": "1200042236",
        "dcore_store_appid": "s20170823000005485"
    }
},
{
    "_id": "3850146094847623169",
    "name": "重庆沙坪坝店",
    "extend_code": {
        "comm_shop_id": "1e53dcf2eb4b401c91dc113085d8981d",
        "ex_code": "20062",
        "alipay_id": "2016112200077000000020268761",
        "us_id": "42371",
        "comm_code": "301003400000069",
        "upcard_terminal": "02306268",
        "upcard_mer_id": "102230058120111",
        "ex_id": "635",
        "ex_cost_center_code": "1200042371",
        "dcore_store_appid": "s20170823000005497"
    }
},
{
    "_id": "3850146121133326337",
    "name": "济南玉函银座店",
    "extend_code": {
        "comm_shop_id": "ccc22391582a4f70a384b259903d668c",
        "ex_code": "20114",
        "alipay_id": "2015061100077000000000194512",
        "us_id": "42567",
        "comm_code": "301003400000372",
        "upcard_terminal": "53101188",
        "upcard_mer_id": "102531058120052",
        "ex_id": "534",
        "ex_cost_center_code": "1200042567",
        "dcore_store_appid": "s20170823000005506"
    }
},
{
    "_id": "3850146122685218817",
    "name": "襄樊万达店",
    "extend_code": {
        "comm_shop_id": "c4f111fd30d04d789d46bd8d71ab0f73",
        "ex_code": "20116",
        "alipay_id": "2015061200077000000000194554",
        "us_id": "42619",
        "comm_code": "301003400000353",
        "upcard_terminal": "71003032",
        "upcard_mer_id": "102710058120531",
        "ex_id": "781",
        "ex_cost_center_code": "1200042619",
        "dcore_store_appid": "s20170823000005218"
    }
},
{
    "_id": "3850146125256327169",
    "name": "绍兴柯桥万达店",
    "extend_code": {
        "comm_shop_id": "181678e85a284c1582df74ee8d61367f",
        "ex_code": "20121",
        "alipay_id": "2016112200077000000020302520",
        "us_id": "42635",
        "comm_code": "301003400000059",
        "upcard_terminal": "57500298",
        "upcard_mer_id": "102575058120020",
        "ex_id": "362",
        "ex_cost_center_code": "1200042635",
        "dcore_store_appid": "s20170823000005516"
    }
},
{
    "_id": "3850146133250670593",
    "name": "淮安万达店",
    "extend_code": {
        "comm_shop_id": "ebd1a8688ac340ef8c322a55e7b29a0c",
        "ex_code": "20136",
        "alipay_id": "2015061100077000000000192953",
        "us_id": "42643",
        "comm_code": "301003400000426",
        "upcard_terminal": "51700220",
        "upcard_mer_id": "102517058120006",
        "ex_id": "612",
        "ex_cost_center_code": "1200042643",
        "dcore_store_appid": "s20170823000005519"
    }
},
{
    "_id": "3850146115601039361",
    "name": "宁波天一二店",
    "extend_code": {
        "comm_shop_id": "27d64f2e53764b8eba77416139d0d60c",
        "ex_code": "20103",
        "alipay_id": "2018041900077000000048257589",
        "us_id": "42563",
        "comm_code": "301003400000088",
        "upcard_terminal": "57401977",
        "upcard_mer_id": "102574058120110",
        "ex_id": "436",
        "ex_cost_center_code": "1200042563",
        "dcore_store_appid": "s20170823000005214"
    }
},
{
    "_id": "3850146137394642945",
    "name": "江阴新一城店",
    "extend_code": {
        "comm_shop_id": "576f2d79de764dceb9a3dec0fce66e09",
        "ex_code": "20144",
        "alipay_id": "2015061100077000000000188749",
        "us_id": "42699",
        "comm_code": "301003400000177",
        "upcard_terminal": "51000995",
        "upcard_mer_id": "102510058120023",
        "ex_id": "372",
        "ex_cost_center_code": "1200042699",
        "dcore_store_appid": "s20170823000005333"
    }
},
{
    "_id": "3850146125742866433",
    "name": "贵阳亨特国际店",
    "extend_code": {
        "comm_shop_id": "4f134bf6a9e241fa9f1f4a050bc203ad",
        "ex_code": "20122",
        "alipay_id": "2015093000077000000004452740",
        "us_id": "42149",
        "comm_code": "301003400000160",
        "upcard_terminal": "85101188",
        "upcard_mer_id": "102851058120125",
        "ex_id": "722",
        "ex_cost_center_code": "1200042149",
        "dcore_store_appid": "s20170823000005480"
    }
},
{
    "_id": "3850146200917377025",
    "name": "绵阳涪城万达",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20250",
        "comm_shop_id": "881caaafd1d645998425c1adb2b6cb08",
        "us_id": "43174",
        "alipay_id": "2015061200077000000000191239",
        "takeaway_eleme_id": '',
        "upcard_terminal": "81600335",
        "comm_code": "301003400000257",
        "upcard_mer_id": "102816058120025",
        "ex_id": "",
        "ex_cost_center_code": "1200043174",
        "dcore_store_appid": "s20170823000005360"
    }
},
{
    "_id": "3850146203522039809",
    "name": "苏州山塘街",
    "extend_code": {
        "comm_shop_id": "f240df6958f3403e98f026ff95ce0bda",
        "ex_code": "20253",
        "alipay_id": "2015061000077000000000188730",
        "us_id": "43342",
        "comm_code": "301003400000436",
        "upcard_terminal": "51211240",
        "upcard_mer_id": "102512058120798",
        "ex_id": "515",
        "ex_cost_center_code": "1200043342",
        "dcore_store_appid": "s20170823000005362"
    }
},
{
    "_id": "3850146224606806017",
    "name": "郑州建文店",
    "extend_code": {
        "comm_shop_id": "4c5bfea87a934b81bdca2c2d44859d87",
        "ex_code": "20280",
        "alipay_id": "2015061100077000000000194503",
        "us_id": "43553",
        "comm_code": "301003400000153",
        "upcard_terminal": "37110794",
        "upcard_mer_id": "102371058122198",
        "ex_id": "20023",
        "ex_cost_center_code": "1200043553",
        "dcore_store_appid": "s20170823000005263"
    }
},
{
    "_id": "3850146230332030977",
    "name": "余姚万达",
    "extend_code": {
        "comm_shop_id": "852f877c039e41a1a755165f4869a84a",
        "ex_code": "20291",
        "alipay_id": "2015061000077000000000188723",
        "us_id": "43516",
        "comm_code": "301003400000252",
        "upcard_terminal": "57402350",
        "upcard_mer_id": "102574058120259",
        "ex_id": "20042",
        "ex_cost_center_code": "1200043516",
        "dcore_store_appid": "s20170823000005253"
    }
},
{
    "_id": "3850146051738566657",
    "name": "诸暨雄风新天地",
    "extend_code": {
        "comm_shop_id": "f03c6490ddc04e9280581bd621fe5e2b",
        "ex_code": "10181",
        "alipay_id": "2015061000077000000000192913",
        "us_id": "43675",
        "comm_code": "301003400000432",
        "upcard_terminal": "57500814",
        "upcard_mer_id": "102575058120042",
        "ex_id": "20073",
        "ex_cost_center_code": "1200043675",
        "dcore_store_appid": "s20170823000005384"
    }
},
{
    "_id": "3850146139198193665",
    "name": "银川金花店",
    "extend_code": {
        "comm_shop_id": "b3431ff5dfd244878acfb4ea2488698c",
        "ex_code": "20147",
        "alipay_id": "2015061200077000000000182892",
        "us_id": "42705",
        "comm_code": "301003400000327",
        "upcard_terminal": "95100476",
        "upcard_mer_id": "102951058120211",
        "ex_id": "475",
        "ex_cost_center_code": "1200042705",
        "dcore_store_appid": "s20170823000005525"
    }
},
{
    "_id": "3850146155367235585",
    "name": "临沂九州商厦",
    "extend_code": {
        "comm_shop_id": "3d4efe338ea4468bb7645970616cea3a",
        "ex_code": "20178",
        "alipay_id": "2015061100077000000000182844",
        "us_id": "42819",
        "comm_code": "301003400000125",
        "upcard_terminal": "53900037",
        "upcard_mer_id": "102539058120001",
        "ex_id": "537",
        "ex_cost_center_code": "1200042819",
        "dcore_store_appid": "s20170823000005538"
    }
},
{
    "_id": "3850146066959695873",
    "name": "上海置地广场店",
    "extend_code": {
        "comm_shop_id": "018fc48bd88d4397a788f260171c80a6",
        "ex_code": "20020",
        "alipay_id": "2015060900077000000000178451",
        "us_id": "42107",
        "comm_code": "301003400000024",
        "upcard_terminal": "02148886",
        "upcard_mer_id": "102210058120713",
        "ex_id": "228",
        "ex_cost_center_code": "1200042107",
        "dcore_store_appid": "s20170823000005198"
    }
},
{
    "_id": "3850146150355042305",
    "name": "镇江万达",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20168",
        "comm_shop_id": "4dc63587b83347d9bc7c8036cc647387",
        "us_id": "42780",
        "alipay_id": "2015061100077000000000182831",
        "takeaway_eleme_id": '',
        "upcard_terminal": "51103377",
        "comm_code": "301003400000156",
        "upcard_mer_id": "102511058121212",
        "ex_id": "565",
        "ex_cost_center_code": "1200042780",
        "dcore_store_appid": "s20170823000005226"
    }
},
{
    "_id": "3850146107883520001",
    "name": "贵阳南国花锦店",
    "extend_code": {
        "comm_shop_id": "6b0f7697f0bf426daa74904b2128032a",
        "ex_code": "20088",
        "alipay_id": "2015061200077000000000182872",
        "us_id": "42478",
        "comm_code": "301003400000211",
        "upcard_terminal": "85101189",
        "upcard_mer_id": "102851058120028",
        "ex_id": "721",
        "ex_cost_center_code": "1200042478",
        "dcore_store_appid": "s20170823000005212"
    }
},
{
    "_id": "3850146086261882881",
    "name": "上海龙之梦店",
    "extend_code": {
        "comm_shop_id": "9a63cb3a782f4784aa8245ab79096ba9",
        "ex_code": "20046",
        "alipay_id": "2015060900077000000000166155",
        "us_id": "42307",
        "comm_code": "301003400000284",
        "upcard_terminal": "02193654",
        "upcard_mer_id": "102210058120721",
        "ex_id": "252",
        "ex_cost_center_code": "1200042307",
        "dcore_store_appid": "s20170823000005492"
    }
},
{
    "_id": "3850146158500380673",
    "name": "武汉东沙万达",
    "extend_code": {
        "comm_shop_id": "9c3023c5d62b4f6baffff4b9eb114dbd",
        "ex_code": "20184",
        "alipay_id": "2015061200077000000000188814",
        "us_id": "42778",
        "comm_code": "301003400000287",
        "upcard_terminal": "02713544",
        "upcard_mer_id": "102270058122423",
        "ex_id": "924",
        "ex_cost_center_code": "1200042778",
        "dcore_store_appid": "s20170823000005533"
    }
},
{
    "_id": "3850146116074995713",
    "name": "上海川沙东海岸店",
    "extend_code": {
        "comm_shop_id": "0ab7032bcc5c45d7a10e98aded3cc765",
        "ex_code": "20104",
        "alipay_id": "2015060900077000000000176454",
        "us_id": "42566",
        "comm_code": "301003400000039",
        "upcard_terminal": "02148849",
        "upcard_mer_id": "102210058120750",
        "ex_id": "275",
        "ex_cost_center_code": "1200042566",
        "dcore_store_appid": "s20170823000005505"
    }
},
{
    "_id": "3850146180763746305",
    "name": "深圳连城新天地",
    "extend_code": {
        "comm_shop_id": "2edcfab2ea6c49e4b8a3efd4a02ca941",
        "ex_code": "20219",
        "alipay_id": "2015061200077000000000182864",
        "us_id": "42775",
        "comm_code": "301003400000099",
        "upcard_terminal": "75516860",
        "upcard_mer_id": "102755058120866",
        "ex_id": "488",
        "ex_cost_center_code": "1200042775",
        "dcore_store_appid": "s20170823000005532"
    }
},
{
    "_id": "3850146088338063361",
    "name": "重庆北城店",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20050",
        "comm_shop_id": "fe542073e01749eeb6430ea415673b52",
        "us_id": "42343",
        "alipay_id": "2015061200077000000000191244",
        "takeaway_eleme_id": '',
        "upcard_terminal": "02306271",
        "comm_code": "301003400000458",
        "upcard_mer_id": "102230058120108",
        "ex_id": "634",
        "ex_cost_center_code": "1200042343",
        "dcore_store_appid": "s20170823000005205"
    }
},
{
    "_id": "3850146099742375937",
    "name": "成都红牌楼店",
    "extend_code": {
        "comm_shop_id": "b4191164e75e41d593455a557eac6275",
        "ex_code": "20072",
        "alipay_id": "2015061200077000000000188801",
        "us_id": "42393",
        "comm_code": "301003400000331",
        "upcard_terminal": "02817516",
        "upcard_mer_id": "102280058121938",
        "ex_id": "806",
        "ex_cost_center_code": "1200042393",
        "dcore_store_appid": "s20170823000005566"
    }
},
{
    "_id": "3850146159527985153",
    "name": "郑州中原万达",
    "extend_code": {
        "comm_shop_id": "e3ad385ac9af42e393c3a2bd02484a4c",
        "ex_code": "20186",
        "alipay_id": "2015061100077000000000188759",
        "us_id": "42866",
        "comm_code": "301003400000413",
        "upcard_terminal": "37108513",
        "upcard_mer_id": "102371058122078",
        "ex_id": "937",
        "ex_cost_center_code": "1200042866",
        "dcore_store_appid": "s20170823000005546"
    }
},
{
    "_id": "3850146182533742593",
    "name": "连云港中央商场",
    "extend_code": {
        "comm_shop_id": "3b850ebc59604315b38cd5f8cf71332e",
        "ex_code": "20222",
        "alipay_id": "2015061100077000000000191203",
        "us_id": "42944",
        "comm_code": "301003400000122",
        "upcard_terminal": "51800155",
        "upcard_mer_id": "102518058120008",
        "ex_id": "996",
        "ex_cost_center_code": "1200042944",
        "dcore_store_appid": "s20170823000005346"
    }
},
{
    "_id": "3850146294441967617",
    "name": "湛江鼎盛店",
    "extend_code": {
        "comm_shop_id": "7312336e40484ddc9c672f055157a0d9",
        "ex_code": "20413",
        "alipay_id": "2015061200077000000000192990",
        "us_id": "43954",
        "comm_code": "301003400000230",
        "upcard_terminal": "75900281",
        "upcard_mer_id": "102759058120013",
        "ex_id": "20202",
        "ex_cost_center_code": "1200043954",
        "dcore_store_appid": "s20170823000005585"
    }
},
{
    "_id": "3850146100228915201",
    "name": "上海宝山安信店",
    "extend_code": {
        "comm_shop_id": "a7b9d6ef4a2a4aae9deeec180423e725",
        "ex_code": "20073",
        "alipay_id": "2015060900077000000000169198",
        "us_id": "42384",
        "comm_code": "301003400000313",
        "upcard_terminal": "02194799",
        "upcard_mer_id": "102210058126899",
        "ex_id": "258",
        "ex_cost_center_code": "1200042384",
        "dcore_store_appid": "s20170823000005207"
    }
},
{
    "_id": "3850146301450649601",
    "name": "西安曲江金地店",
    "extend_code": {
        "comm_shop_id": "0a87c203472f4a8fa3ba258aaf042edb",
        "ex_code": "20422",
        "alipay_id": "2015093000077000000004460351",
        "us_id": "44013",
        "comm_code": "301003400000038",
        "upcard_terminal": "02990622",
        "upcard_mer_id": "102290058122792",
        "ex_id": "20211",
        "ex_cost_center_code": "1200044013",
        "dcore_store_appid": "s20170823000005297"
    }
},
{
    "_id": "3850146183477460993",
    "name": "驻马店新玛特",
    "extend_code": {
        "comm_shop_id": "912f8c4b0cf041af8b7c5c76130b35c6",
        "ex_code": "20224",
        "alipay_id": "2015061100077000000000194506",
        "us_id": "42943",
        "comm_code": "301003400004732",
        "upcard_terminal": "39601298",
        "upcard_mer_id": "102396058120019",
        "ex_id": "945",
        "ex_cost_center_code": "1200042943",
        "dcore_store_appid": "s20170823000005345"
    }
},
{
    "_id": "3850146101369765889",
    "name": "松江开元店",
    "extend_code": {
        "comm_shop_id": "b8e8d9a653cb4511980ccdebccd1c31d",
        "ex_code": "20075",
        "alipay_id": "2015060900077000000000176453",
        "us_id": "42413",
        "comm_code": "301003400000336",
        "upcard_terminal": "02148867",
        "upcard_mer_id": "102210058120732",
        "ex_id": "260",
        "ex_cost_center_code": "1200042413",
        "dcore_store_appid": "s20170823000005209"
    }
},
{
    "_id": "3850146302948016129",
    "name": "丽水万地店",
    "extend_code": {
        "us_id": "44016",
        "ex_id": "20220",
        "ex_cost_center_code": "1200044016",
        "ex_code": "20423"
    }
},
{
    "_id": "3850146066296995841",
    "name": "常熟印象城店",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20019",
        "comm_shop_id": "b109ace497354add9afaa0124aa44165",
        "us_id": "41997",
        "alipay_id": "2015061100077000000000182830",
        "takeaway_eleme_id": '',
        "upcard_terminal": "51215442",
        "comm_code": "301003400000322",
        "upcard_mer_id": "102512058123022",
        "ex_id": "552",
        "ex_cost_center_code": "1200041997",
        "dcore_store_appid": "s20170823000005364"
    }
},
{
    "_id": "3850146192285499393",
    "name": "晋江万达店",
    "extend_code": {
        "comm_shop_id": "5b3a28a0b9e047b5ae1300183c0b1591",
        "ex_code": "20233",
        "alipay_id": "2015061200077000000000194536",
        "us_id": "43062",
        "comm_code": "301003400000186",
        "upcard_terminal": "59500694",
        "upcard_mer_id": "102595058120137",
        "ex_id": "758",
        "ex_cost_center_code": "1200043062",
        "dcore_store_appid": "s20170823000005352"
    }
},
{
    "_id": "3850146173453074433",
    "name": "虹口龙之梦",
    "extend_code": {
        "comm_shop_id": "2dd7f92ca4ab4646a9214c242224634a",
        "ex_code": "20205",
        "alipay_id": "2015060900077000000000169215",
        "us_id": "42842",
        "comm_code": "301003400000095",
        "upcard_terminal": "02194859",
        "upcard_mer_id": "102210058120983",
        "ex_id": "2003",
        "ex_cost_center_code": "1200042842",
        "dcore_store_appid": "s20170823000005543"
    }
},
{
    "_id": "3850146319918170113",
    "name": "太原茂业店",
    "extend_code": {
        "comm_shop_id": "a713fb7c75004fe5989aac59f713ac73",
        "ex_code": "20450",
        "alipay_id": "2015093000077000000004438544",
        "us_id": "43922",
        "comm_code": "301003400000311",
        "upcard_terminal": "35103999",
        "upcard_mer_id": "102351058121525",
        "ex_id": "20222",
        "ex_cost_center_code": "1200043922",
        "dcore_store_appid": "s20170823000005577"
    }
},
{
    "_id": "3850146235088371713",
    "name": "徐州金鹰",
    "extend_code": {
        "comm_shop_id": "e2e28e217adf42d487a181066508a420",
        "ex_code": "20300",
        "alipay_id": "2019100900077000000083379295",
        "us_id": "43637",
        "comm_code": "301003400000411",
        "upcard_terminal": "51600590",
        "upcard_mer_id": "102516058120114",
        "ex_id": "20057",
        "ex_cost_center_code": "1200043637",
        "dcore_store_appid": "s20170823000005271"
    }
},
{
    "_id": "3850146073200820225",
    "name": "上海人民广场店",
    "extend_code": {
        "comm_shop_id": "6b09b5dfc5b147f581953f7feda99698",
        "ex_code": "20032",
        "alipay_id": "2015093000077000000004490205",
        "us_id": "42102",
        "comm_code": "301003400000210",
        "upcard_terminal": "02194795",
        "upcard_mer_id": "102210058126895",
        "ex_id": "237",
        "ex_cost_center_code": "1200042102",
        "dcore_store_appid": "s20170823000005476"
    }
},
{
    "_id": "3850146146143961089",
    "name": "江桥万达",
    "extend_code": {
        "comm_shop_id": "83f3050cb74849299edbdeab5f3b78b8",
        "ex_code": "20160",
        "alipay_id": "2015060900077000000000176465",
        "us_id": "42755",
        "comm_code": "301003400000250",
        "upcard_terminal": "02194856",
        "upcard_mer_id": "102210058120849",
        "ex_id": "292",
        "ex_cost_center_code": "1200042755",
        "dcore_store_appid": "s20170823000005530"
    }
},
{
    "_id": "3850146179299934209",
    "name": "济宁佳世客",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20216",
        "comm_shop_id": "1fe0462a4372433ead3e05f97ec00175",
        "us_id": "42945",
        "alipay_id": "2015061100077000000000194515",
        "takeaway_eleme_id": '',
        "upcard_terminal": "53700586",
        "comm_code": "301003400000070",
        "upcard_mer_id": "102537058120003",
        "ex_id": "545",
        "ex_cost_center_code": "1200042945",
        "dcore_store_appid": "s20170823000005236"
    }
},
{
    "_id": "3850146321524588545",
    "name": "武汉光谷世界城",
    "extend_code": {
        "comm_shop_id": "e0fcdeaf474b4eafa7f824956f93dfea",
        "ex_code": "20453",
        "alipay_id": "2015093000077000000004452737",
        "us_id": "43937",
        "comm_code": "301003400000409",
        "upcard_terminal": "02723116",
        "upcard_mer_id": "102270058125030",
        "ex_id": "20255",
        "ex_cost_center_code": "1200043937",
        "dcore_store_appid": "s20170823000005581"
    }
},
{
    "_id": "3850146108353282049",
    "name": "成山路店",
    "extend_code": {
        "comm_shop_id": "8b7dc996d2864d9facc64dcce6f3d27b",
        "ex_code": "20089",
        "alipay_id": "2015060900077000000000169200",
        "us_id": "42477",
        "comm_code": "301003400000263",
        "upcard_terminal": "02190200",
        "upcard_mer_id": "102210058120746",
        "ex_id": "267",
        "ex_cost_center_code": "1200042477",
        "dcore_store_appid": "s20170823000005575"
    }
},
{
    "_id": "3850146184463122433",
    "name": "长沙悦方",
    "extend_code": {
        "comm_shop_id": "843c10be51d542aa9d8fd70990891868",
        "ex_code": "20226",
        "alipay_id": "2015061200077000000000188816",
        "us_id": "42991",
        "comm_code": "301003400000251",
        "upcard_terminal": "73106061",
        "upcard_mer_id": "102731058120845",
        "ex_id": "738",
        "ex_cost_center_code": "1200042991",
        "dcore_store_appid": "s20170823000005349"
    }
},
{
    "_id": "3850146322732548097",
    "name": "郑州大卫城店",
    "extend_code": {
        "comm_shop_id": "d0135377e87d4314be52b09116fdc83c",
        "ex_code": "20455",
        "alipay_id": "2015093000077000000004451343",
        "us_id": "44077",
        "comm_code": "301003400000385",
        "upcard_terminal": "37110828",
        "upcard_mer_id": "102371058122403",
        "ex_id": "20245",
        "ex_cost_center_code": "1200044077",
        "dcore_store_appid": "s20170823000005611"
    }
},
{
    "_id": "3850146167618797569",
    "name": "舟山凯虹",
    "extend_code": {
        "comm_shop_id": "03981a70a5ff43e4927d260c04cc3125",
        "ex_code": "20194",
        "alipay_id": "2015061000077000000000188717",
        "us_id": "42870",
        "comm_code": "301003400000029",
        "upcard_terminal": "58000375",
        "upcard_mer_id": "102580058120040",
        "ex_id": "332",
        "ex_cost_center_code": "1200042870",
        "dcore_store_appid": "s20170823000005548"
    }
},
{
    "_id": "3850146324053753857",
    "name": "淄博万象汇",
    "extend_code": {
        "comm_shop_id": "6f654248590e45b49c81e24538b5ca2f",
        "ex_code": "20457",
        "alipay_id": "2015093000077000000004486033",
        "us_id": "44134",
        "comm_code": "301003400000221",
        "upcard_terminal": "53390075",
        "upcard_mer_id": "102533058120084",
        "ex_id": "",
        "ex_cost_center_code": "1200044134",
        "dcore_store_appid": "s20170823000005622"
    }
},
{
    "_id": "3850146096315629569",
    "name": "重庆万达店",
    "extend_code": {
        "comm_shop_id": "4e7bc86cf9c64520993ef0c6677c7548",
        "ex_code": "20065",
        "alipay_id": "2018041900077000000048354821",
        "us_id": "42390",
        "comm_code": "301003400000159",
        "upcard_terminal": "02306267",
        "upcard_mer_id": "102230058120112",
        "ex_id": "636",
        "ex_cost_center_code": "1200042390",
        "dcore_store_appid": "s20170823000005565"
    }
},
{
    "_id": "3850146154901667841",
    "name": "贵阳恒峰",
    "extend_code": {
        "comm_shop_id": "722cceb67dd247c0951bbf6226d142bb",
        "ex_code": "20177",
        "alipay_id": "2015061200077000000000192995",
        "us_id": "42828",
        "comm_code": "301003400000226",
        "upcard_terminal": "85101187",
        "upcard_mer_id": "102851058120161",
        "ex_id": "723",
        "ex_cost_center_code": "1200042828",
        "dcore_store_appid": "s20170823000005541"
    }
},
{
    "_id": "3850146327857987585",
    "name": "咸阳正兴店",
    "extend_code": {
        "comm_shop_id": "2e5d8ee430544a23b0fca6180072e338",
        "ex_code": "20464",
        "alipay_id": "2015093000077000000004460350",
        "us_id": "44163",
        "comm_code": "301003400000098",
        "upcard_terminal": "02990947",
        "upcard_mer_id": "102290058122902",
        "ex_id": "20265",
        "ex_cost_center_code": "1200044163",
        "dcore_store_appid": "s20170823000005625"
    }
},
{
    "_id": "3850146262212935681",
    "name": "银川西夏万达",
    "extend_code": {
        "comm_shop_id": "31489210ba974ccf858c004993385d2c",
        "ex_code": "20353",
        "alipay_id": "2015061200077000000000188822",
        "us_id": "43832",
        "comm_code": "301003400000106",
        "upcard_terminal": "95100525",
        "upcard_mer_id": "102951058120225",
        "ex_id": "20140",
        "ex_cost_center_code": "1200043832",
        "dcore_store_appid": "s20170823000005408"
    }
},
{
    "_id": "3850146172958146561",
    "name": "绵阳新世界",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20204",
        "comm_shop_id": "966b65c74f07418780864734084bad8c",
        "us_id": "42892",
        "alipay_id": "2015061200077000000000191236",
        "takeaway_eleme_id": '',
        "upcard_terminal": "81600336",
        "comm_code": "301003400000275",
        "upcard_mer_id": "102816058120005",
        "ex_id": "818",
        "ex_cost_center_code": "1200042892",
        "dcore_store_appid": "s20170823000005552"
    }
},
{
    "_id": "3850146227832225793",
    "name": "长沙开福万达店",
    "extend_code": {
        "comm_shop_id": "8ebedea25c694865b0532a75e9831157",
        "ex_code": "20286",
        "alipay_id": "2015093000077000000004511905",
        "us_id": "43542",
        "comm_code": "301003400000266",
        "upcard_terminal": "73109820",
        "upcard_mer_id": "102731058121596",
        "ex_id": "20021",
        "ex_cost_center_code": "1200043542",
        "dcore_store_appid": "s20170823000005256"
    }
},
{
    "_id": "3850146106012860417",
    "name": "绍兴世茂店",
    "extend_code": {
        "comm_shop_id": "2805cc9fe7a340b8a45b2e5a117fcb29",
        "ex_code": "20084",
        "alipay_id": "2015093000077000000004518859",
        "us_id": "42469",
        "comm_code": "301003400000089",
        "upcard_terminal": "57103822",
        "upcard_mer_id": "102571058120392",
        "ex_id": "361",
        "ex_cost_center_code": "1200042469",
        "dcore_store_appid": "s20170823000005574"
    }
},
{
    "_id": "3850146202389577729",
    "name": "宁波印象城店",
    "extend_code": {
        "comm_shop_id": "cab2672464a94b75883a8c07d7b4edc5",
        "ex_code": "20251",
        "alipay_id": "2015061000077000000000182803",
        "us_id": "43185",
        "comm_code": "301003400000368",
        "upcard_terminal": "57403308",
        "upcard_mer_id": "102574058120522",
        "ex_id": "441",
        "ex_cost_center_code": "1200043185",
        "dcore_store_appid": "s20170823000005361"
    }
},
{
    "_id": "3850146175042715649",
    "name": "武汉百联奥特莱斯",
    "extend_code": {
        "comm_shop_id": "9d363bae9f7c40eb8a18c65cb85a3b27",
        "ex_code": "20208",
        "alipay_id": "2015061200077000000000193011",
        "us_id": "42902",
        "comm_code": "301003400000290",
        "upcard_terminal": "02713541",
        "upcard_mer_id": "102270058122465",
        "ex_id": "928",
        "ex_cost_center_code": "1200042902",
        "dcore_store_appid": "s20170823000005234"
    }
},
{
    "_id": "3850146343452409857",
    "name": "宁波银泰环球城店",
    "extend_code": {
        "comm_shop_id": "da14515038c24458b9855bfbb29407aa",
        "ex_code": "20483",
        "alipay_id": "2015093000077000000004438545",
        "us_id": "44236",
        "comm_code": "301003400000401",
        "upcard_terminal": "57403154",
        "upcard_mer_id": "102574058120495",
        "ex_id": "20294",
        "ex_cost_center_code": "1200044236",
        "dcore_store_appid": "s20170823000005642"
    }
},
{
    "_id": "3850146235621048321",
    "name": "成都高新伊藤",
    "extend_code": {
        "comm_shop_id": "6f47a4c518da4ffd9e9c07e770a08e69",
        "ex_code": "20301",
        "alipay_id": "2015061200077000000000192999",
        "us_id": "43661",
        "comm_code": "301003400000220",
        "upcard_terminal": "02883679",
        "upcard_mer_id": "102280058123576",
        "ex_id": "99999",
        "ex_cost_center_code": "1200043661",
        "dcore_store_appid": "s20170823000005382"
    }
},
{
    "_id": "3850146227295354881",
    "name": "南京龙江新城店",
    "extend_code": {
        "comm_shop_id": "c3a352f0b33c418f8f0c7b52060a262e",
        "ex_code": "20285",
        "alipay_id": "2015061000077000000000191172",
        "us_id": "43549",
        "comm_code": "301003400000350",
        "upcard_terminal": "02592754",
        "upcard_mer_id": "102250058121909",
        "ex_id": "20025",
        "ex_cost_center_code": "1200043549",
        "dcore_store_appid": "s20170823000005260"
    }
},
{
    "_id": "3850146237164552193",
    "name": "深圳龙岗万科广场",
    "extend_code": {
        "comm_shop_id": "078573ba949c4ae1a1f7096e914ac079",
        "ex_code": "20304",
        "alipay_id": "2015061200077000000000192985",
        "us_id": "43672",
        "comm_code": "301003400000032",
        "upcard_terminal": "75512047",
        "upcard_mer_id": "102755058121836",
        "ex_id": "98765",
        "ex_cost_center_code": "1200043672",
        "dcore_store_appid": "s20170823000005383"
    }
},
{
    "_id": "3850146192801398785",
    "name": "宁德万达店",
    "extend_code": {
        "comm_shop_id": "e85573e530ff4bfc8d061be9d584eb7a",
        "ex_code": "20234",
        "alipay_id": "2015061200077000000000191230",
        "us_id": "43065",
        "comm_code": "301003400000417",
        "upcard_terminal": "59300038",
        "upcard_mer_id": "102593058120004",
        "ex_id": "865",
        "ex_cost_center_code": "1200043065",
        "dcore_store_appid": "s20170823000005696"
    }
},
{
    "_id": "3850146126778859521",
    "name": "武汉菱角湖万达店",
    "extend_code": {
        "comm_shop_id": "bfb337f6b3574719bae9ec05ad97d5ad",
        "ex_code": "20124",
        "alipay_id": "2015061200077000000000194558",
        "us_id": "42613",
        "comm_code": "301003400000342",
        "upcard_terminal": "02729439",
        "upcard_mer_id": "102270058122023",
        "ex_id": "916",
        "ex_cost_center_code": "1200042613",
        "dcore_store_appid": "s20170823000005513"
    }
},
{
    "_id": "3850146065198088193",
    "name": "西安万达店",
    "extend_code": {
        "comm_shop_id": "729976c8626848e19ea5217bfcceb7c8",
        "ex_code": "20017",
        "alipay_id": "2015061200077000000000191254",
        "us_id": "42062",
        "comm_code": "301003400000227",
        "upcard_terminal": "02903688",
        "upcard_mer_id": "102290058122437",
        "ex_id": "951",
        "ex_cost_center_code": "1200042062",
        "dcore_store_appid": "s20170823000005474"
    }
},
{
    "_id": "3850146065764319233",
    "name": "上海五莲店",
    "extend_code": {
        "comm_shop_id": "91cb11157d6b4267be46e3c18f4e4557",
        "ex_code": "20018",
        "alipay_id": "2015060900077000000000178440",
        "us_id": "42088",
        "comm_code": "301003400000273",
        "upcard_terminal": "02148910",
        "upcard_mer_id": "102210058120689",
        "ex_id": "502",
        "ex_cost_center_code": "1200042088",
        "dcore_store_appid": "s20170823000005475"
    }
},
{
    "_id": "3850146304504102913",
    "name": "青州泰华城",
    "extend_code": {
        "us_id": "44024",
        "ex_id": "20207",
        "ex_cost_center_code": "1200044024",
        "ex_code": "20425"
    }
},
{
    "_id": "3850146242680061953",
    "name": "上海浦江镇店",
    "extend_code": {
        "comm_shop_id": "5ad6bc7015c941e781f4d9e944dc99fa",
        "ex_code": "20315",
        "alipay_id": "2015060900077000000000169207",
        "us_id": "43659",
        "comm_code": "301003400000185",
        "upcard_terminal": "02149983",
        "upcard_mer_id": "102210058125753",
        "ex_id": "20071",
        "ex_cost_center_code": "1200043659",
        "dcore_store_appid": "s20170823000005381"
    }
},
{
    "_id": "3850146224061546497",
    "name": "南京商厦店",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20279",
        "comm_shop_id": "705e444cd279467391ad1162992b377b",
        "us_id": "43552",
        "alipay_id": "2015061000077000000000182814",
        "takeaway_eleme_id": '',
        "upcard_terminal": "02593529",
        "comm_code": "301003400000222",
        "upcard_mer_id": "102250058121864",
        "ex_id": "20022",
        "ex_cost_center_code": "1200043552",
        "dcore_store_appid": "s20170823000005262"
    }
},
{
    "_id": "3850146335319654401",
    "name": "黄石万达店",
    "extend_code": {
        "comm_shop_id": "d419e7488f1a4005bc7d1bd89ec87fe5",
        "ex_code": "20469",
        "alipay_id": "2015093000077000000004477748",
        "us_id": "44193",
        "comm_code": "301003400000390",
        "upcard_terminal": "71430057",
        "upcard_mer_id": "102714058120015",
        "ex_id": "20259",
        "ex_cost_center_code": "1200044193",
        "dcore_store_appid": "s20170823000005631"
    }
},
{
    "_id": "3850146244198400001",
    "name": "西安高新永辉",
    "extend_code": {
        "comm_shop_id": "e37909afe0d6408f83397cbc3f5db9b9",
        "ex_code": "20318",
        "alipay_id": "2015061200077000000000188818",
        "us_id": "43674",
        "comm_code": "301003400000412",
        "upcard_terminal": "02990372",
        "upcard_mer_id": "102290058122687",
        "ex_id": "20085",
        "ex_cost_center_code": "1200043674",
        "dcore_store_appid": "s20170823000005273"
    }
},
{
    "_id": "3850146225319837697",
    "name": "杭州城西银泰城",
    "extend_code": {
        "comm_shop_id": "896125a609bb476c924598cc60698cbc",
        "ex_code": "20281",
        "alipay_id": "2015060900077000000000178462",
        "us_id": "43548",
        "comm_code": "301003400000171",
        "upcard_terminal": "57104200",
        "upcard_mer_id": "102571058120812",
        "ex_id": "20019",
        "ex_cost_center_code": "1200043548",
        "dcore_store_appid": "s20170823000005259"
    }
},
{
    "_id": "3850146076438822913",
    "name": "武汉新佳丽店",
    "extend_code": {
        "comm_shop_id": "34f428487f1f4e2f998cee57eb6c37aa",
        "ex_code": "20035",
        "alipay_id": "2015061200077000000000193014",
        "us_id": "42257",
        "comm_code": "301003400000114",
        "upcard_terminal": "02713561",
        "upcard_mer_id": "102270058121832",
        "ex_id": "912",
        "ex_cost_center_code": "1200042257",
        "dcore_store_appid": "s20170823000005202"
    }
},
{
    "_id": "3850146256487710721",
    "name": "武汉汉口城市广场",
    "extend_code": {
        "comm_shop_id": "48e0a5e8597e4d638e01ab4f183cc79b",
        "ex_code": "20342",
        "alipay_id": "2015061200077000000000194552",
        "us_id": "43760",
        "comm_code": "301003400000146",
        "upcard_terminal": "02722083",
        "upcard_mer_id": "102270058124771",
        "ex_id": "20134",
        "ex_cost_center_code": "1200043760",
        "dcore_store_appid": "s20170823000005394"
    }
},
{
    "_id": "3850146376176369665",
    "name": "章丘唐人中心",
    "extend_code": {
        "comm_shop_id": "0f68fb88009644429a6edad37b931407",
        "ex_code": "20541",
        "alipay_id": "2016011100077000000014128752",
        "us_id": "44411",
        "comm_code": "301003400000043",
        "upcard_terminal": "53101201",
        "upcard_mer_id": "102531058120523",
        "ex_id": "20375",
        "ex_cost_center_code": "1200044411",
        "dcore_store_appid": "s20170823000005674"
    }
},
{
    "_id": "3850146245670600705",
    "name": "湛江万象金沙湾广场",
    "extend_code": {
        "comm_shop_id": "392d5fa2cbe642f89176364d8fcbb3c3",
        "ex_code": "20321",
        "alipay_id": "2015061200077000000000188792",
        "us_id": "43636",
        "comm_code": "301003400000119",
        "upcard_terminal": "75900241",
        "upcard_mer_id": "102759058120006",
        "ex_id": "99888",
        "ex_cost_center_code": "1200043636",
        "dcore_store_appid": "s20170823000005379"
    }
},
{
    "_id": "3850146361487917057",
    "name": "昆山九方店",
    "extend_code": {
        "comm_shop_id": "44e572d89e1b48a1b5bfc3036787ef54",
        "ex_code": "20513",
        "alipay_id": "2016020300077000000014719681",
        "us_id": "44333",
        "comm_code": "301003400000191",
        "upcard_terminal": "51215315",
        "upcard_mer_id": "102512058122983",
        "ex_id": "20326",
        "ex_cost_center_code": "1200044333",
        "dcore_store_appid": "s20170823000005320"
    }
},
{
    "_id": "3850146270186307585",
    "name": "顾村绿地",
    "extend_code": {
        "comm_shop_id": "6949c0c00bc347c997837b9c17c4b7ac",
        "ex_code": "20367",
        "alipay_id": "2015060900077000000000166172",
        "us_id": "43930",
        "comm_code": "301003400000207",
        "upcard_terminal": "02190306",
        "upcard_mer_id": "102210058125878",
        "ex_id": "20165",
        "ex_cost_center_code": "1200043930",
        "dcore_store_appid": "s20170823000005289"
    }
},
{
    "_id": "3850146077491593217",
    "name": "上海南站二店",
    "extend_code": {
        "comm_shop_id": "87b0a9b0abc242ecbd56af68f41f2d31",
        "ex_code": "20037",
        "alipay_id": "2016112200077000000020281356",
        "us_id": "42269",
        "comm_code": "301003400000256",
        "upcard_terminal": "02194793",
        "upcard_mer_id": "102210058126893",
        "ex_id": "244",
        "ex_cost_center_code": "1200042269",
        "dcore_store_appid": "s20170823000005490"
    }
},
{
    "_id": "3850146385483530241",
    "name": "青岛金狮广场店",
    "extend_code": {
        "comm_shop_id": "a5033490ef8941f1bf0cf05686a02807",
        "ex_code": "20558",
        "alipay_id": "2016081500077000000018020531",
        "us_id": "44367",
        "comm_code": "301003400000307",
        "upcard_terminal": "53205312",
        "upcard_mer_id": "102532058121382",
        "ex_id": "20411",
        "ex_cost_center_code": "1200044367",
        "dcore_store_appid": "s20170823000005323"
    }
},
{
    "_id": "3850146231749705729",
    "name": "南京马群花园城",
    "extend_code": {
        "comm_shop_id": "b13171ecb48c43fc9098daaa6c7f63b3",
        "ex_code": "20294",
        "alipay_id": "2015061000077000000000191169",
        "us_id": "43612",
        "comm_code": "301003400000323",
        "upcard_terminal": "02593531",
        "upcard_mer_id": "102250058121979",
        "ex_id": "20048",
        "ex_cost_center_code": "1200043612",
        "dcore_store_appid": "s20170823000005378"
    }
},
{
    "_id": "3850146277538922497",
    "name": "南充顺庆1227",
    "extend_code": {
        "comm_shop_id": "7675bd0e2ebd4d40b26f1252d00bd4de",
        "ex_code": "20381",
        "alipay_id": "2015061200077000000000182876",
        "us_id": "43929",
        "comm_code": "301003400000233",
        "upcard_terminal": "81700193",
        "upcard_mer_id": "102817058120011",
        "ex_id": "",
        "ex_cost_center_code": "1200043929",
        "dcore_store_appid": "s20170823000005579"
    }
},
{
    "_id": "3850146278096764929",
    "name": "濮阳万利",
    "extend_code": {
        "comm_shop_id": "8226ec0c686045758b7254f841bf5ea5",
        "ex_code": "20382",
        "alipay_id": "2019050500077000000077036041",
        "us_id": "43938",
        "comm_code": "301003400000247",
        "upcard_terminal": "39301031",
        "upcard_mer_id": "102393058120017",
        "ex_id": "20173",
        "ex_cost_center_code": "1200043938",
        "dcore_store_appid": "s20170823000005290"
    }
},
{
    "_id": "3850146232223662081",
    "name": "蚌埠万达",
    "extend_code": {
        "comm_shop_id": "3eeb9c24c93e474b89adfda137faced2",
        "ex_code": "20295",
        "alipay_id": "2015061200077000000000188785",
        "us_id": "43611",
        "comm_code": "301003400000131",
        "upcard_terminal": "55202300",
        "upcard_mer_id": "102552058120026",
        "ex_id": "20037",
        "ex_cost_center_code": "1200043611",
        "dcore_store_appid": "s20170823000005377"
    }
},
{
    "_id": "3850146407616872449",
    "name": "湘潭万达店",
    "extend_code": {
        "comm_shop_id": "8f74f89c807f4d48a7a52ad8f6bb1687",
        "ex_code": "20585",
        "alipay_id": "2016081600077000000018020831",
        "us_id": "44508",
        "comm_code": "301003400000267",
        "upcard_terminal": "73109946",
        "upcard_mer_id": "102731058121604",
        "ex_id": "20418",
        "ex_cost_center_code": "1200044508",
        "dcore_store_appid": "s20170823000005682"
    }
},
{
    "_id": "3850146240184451073",
    "name": "海宁银泰城",
    "extend_code": {
        "comm_shop_id": "fbfb044a0589479490d567fa6f5702da",
        "ex_code": "20310",
        "alipay_id": "2015061000077000000000188724",
        "us_id": "43529",
        "comm_code": "301003400000454",
        "upcard_terminal": "57302902",
        "upcard_mer_id": "102573058120062",
        "ex_id": "20033",
        "ex_cost_center_code": "1200043529",
        "dcore_store_appid": "s20170823000005254"
    }
},
{
    "_id": "3850146408111800321",
    "name": "资阳万达",
    "extend_code": {
        "comm_shop_id": "8dca11fdc19d4ca497befa62c28ddbed",
        "ex_code": "20612",
        "alipay_id": "2016111700077000000020010118",
        "us_id": "44504",
        "comm_code": "301003400000264",
        "upcard_terminal": "02894232",
        "upcard_mer_id": "102280058125188",
        "ex_id": "20414",
        "ex_cost_center_code": "1200044504",
        "dcore_store_appid": "s20170823000005328"
    }
},
{
    "_id": "3850146202985168897",
    "name": "衢州景文",
    "extend_code": {
        "comm_shop_id": "d6393bd30278481c930ac027a54772e7",
        "ex_code": "20252",
        "alipay_id": "2015061000077000000000188719",
        "us_id": "43184",
        "comm_code": "301003400000394",
        "upcard_terminal": "57000077",
        "upcard_mer_id": "102570058120002",
        "ex_id": "686",
        "ex_cost_center_code": "1200043184",
        "dcore_store_appid": "s20170823000005241"
    }
},
{
    "_id": "3850146255489466369",
    "name": "济宁贵和二店",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20340",
        "comm_shop_id": "236b918f099d4a39a000157659323e46",
        "us_id": "43784",
        "alipay_id": "2015061100077000000000188763",
        "takeaway_eleme_id": '',
        "upcard_terminal": "53700800",
        "comm_code": "301003400000079",
        "upcard_mer_id": "102537058120038",
        "ex_id": "20138",
        "ex_cost_center_code": "1200043784",
        "dcore_store_appid": "s20170823000005279"
    }
},
{
    "_id": "3850146243179184129",
    "name": "武汉凯德武胜店",
    "extend_code": {
        "comm_shop_id": "3fafb817aeaf4dd18ec426a09429f447",
        "ex_code": "20316",
        "alipay_id": "2015061200077000000000188820",
        "us_id": "43676",
        "comm_code": "301003400000132",
        "upcard_terminal": "02722339",
        "upcard_mer_id": "102270058124839",
        "ex_id": "20072",
        "ex_cost_center_code": "1200043676",
        "dcore_store_appid": "s20170823000005385"
    }
},
{
    "_id": "3850146293410168833",
    "name": "武汉永旺梦乐城店",
    "extend_code": {
        "comm_shop_id": "f9759cef87d9461cb67962818d4b4264",
        "ex_code": "20411",
        "alipay_id": "2015093000077000000004439948",
        "us_id": "44017",
        "comm_code": "301003400000448",
        "upcard_terminal": "02722412",
        "upcard_mer_id": "102270058124861",
        "ex_id": "20224",
        "ex_cost_center_code": "1200044017",
        "dcore_store_appid": "s20170823000005603"
    }
},
{
    "_id": "3850146097737498625",
    "name": "南京仙林金鹰店",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20069",
        "comm_shop_id": "62e6e0e7270746eca1650f075716d0d7",
        "us_id": "42415",
        "alipay_id": "2015061000077000000000191175",
        "takeaway_eleme_id": '',
        "upcard_terminal": "02516542",
        "comm_code": "301003400000201",
        "upcard_mer_id": "102250058120836",
        "ex_id": "309",
        "ex_cost_center_code": "1200042415",
        "dcore_store_appid": "s20170823000005568"
    }
},
{
    "_id": "3850146249198010369",
    "name": "临海银泰店",
    "extend_code": {
        "comm_shop_id": "88ee4def92004395a900bafcbf0f0f09",
        "ex_code": "20328",
        "alipay_id": "2015093000077000000004511912",
        "us_id": "43751",
        "comm_code": "301003400000258",
        "upcard_terminal": "57600854",
        "upcard_mer_id": "102576058120036",
        "ex_id": "20112",
        "ex_cost_center_code": "1200043751",
        "dcore_store_appid": "s20170823000005392"
    }
},
{
    "_id": "3850146405028986881",
    "name": "嘉定大融城店",
    "extend_code": {
        "comm_shop_id": "9fb052715006443c95c32036ec9f4def",
        "ex_code": "20603",
        "alipay_id": "2016112400077000000023066705",
        "us_id": "44502",
        "comm_code": "301003400004604",
        "upcard_terminal": "21098215",
        "upcard_mer_id": "102210058126921",
        "ex_id": "20415",
        "ex_cost_center_code": "1200844502",
        "dcore_store_appid": "s20170823000005173"
    }
},
{
    "_id": "3850146240679378945",
    "name": "淮北金鹰",
    "extend_code": {
        "comm_shop_id": "c381a4e2e91541e6909715acab49ef72",
        "ex_code": "20311",
        "alipay_id": "2015061200077000000000192986",
        "us_id": "43607",
        "comm_code": "301003400000349",
        "upcard_terminal": "56100669",
        "upcard_mer_id": "102561058120180",
        "ex_id": "20045",
        "ex_cost_center_code": "1200043607",
        "dcore_store_appid": "s20170823000005375"
    }
},
{
    "_id": "3850146117631082497",
    "name": "上海虹井店",
    "extend_code": {
        "comm_shop_id": "c1e7b2ce88b341dabcfcaa38f507a244",
        "ex_code": "20107",
        "alipay_id": "2015060900077000000000176452",
        "us_id": "42543",
        "comm_code": "301003400000348",
        "upcard_terminal": "02148858",
        "upcard_mer_id": "102210058120741",
        "ex_id": "277",
        "ex_cost_center_code": "1200042543",
        "dianping_store_id": "4288836",
        "dcore_store_appid": "s20170823000005502"
    }
},
{
    "_id": "3850146298950844417",
    "name": "深圳龙华九方购物中心店",
    "extend_code": {
        "us_id": "44012",
        "ex_id": "20226",
        "ex_cost_center_code": "1200044012",
        "ex_code": "20419"
    }
},
{
    "_id": "3850146183964000257",
    "name": "武汉摩尔城",
    "extend_code": {
        "comm_shop_id": "9f7bf893d82c43798e56e6951659911c",
        "ex_code": "20225",
        "alipay_id": "2017120400077000000046825993",
        "us_id": "42994",
        "comm_code": "301003400000297",
        "upcard_terminal": "02713539",
        "upcard_mer_id": "102270058122536",
        "ex_id": "929",
        "ex_cost_center_code": "1200042994",
        "dcore_store_appid": "s20170823000005237"
    }
},
{
    "_id": "3850146303988203521",
    "name": "青州泰华城",
    "extend_code": {
        "comm_shop_id": "ae1f7e7601e045d19213a8a6dacf2102",
        "ex_code": "20425",
        "alipay_id": "2015121500077000000013754527",
        "us_id": "44024",
        "comm_code": "301003400000318",
        "upcard_terminal": "53601843",
        "upcard_mer_id": "102536058120135",
        "ex_id": "20207",
        "ex_cost_center_code": "1200044024",
        "dcore_store_appid": "s20170823000005299"
    }
},
{
    "_id": "3850146191794765825",
    "name": "合肥天鹅湖万达",
    "extend_code": {
        "comm_shop_id": "eec51fb82c214364907cd90c1fb64e33",
        "ex_code": "20232",
        "alipay_id": "2015061100077000000000188771",
        "us_id": "42922",
        "comm_code": "301003400000429",
        "upcard_terminal": "55118202",
        "upcard_mer_id": "102551058122783",
        "ex_id": "158",
        "ex_cost_center_code": "1200042922",
        "dcore_store_appid": "s20170823000005559"
    }
},
{
    "_id": "3850146307691773953",
    "name": "重庆合川步步高",
    "extend_code": {
        "us_id": "43962",
        "ex_id": "20209",
        "ex_cost_center_code": "1200043962",
        "ex_code": "20428"
    }
},
{
    "_id": "3850146255996977153",
    "name": "宜宾洋洋百货",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20341",
        "comm_shop_id": "2e5ac6c1da224be38fd9063fa9511a30",
        "us_id": "43788",
        "alipay_id": "2015061200077000000000194547",
        "takeaway_eleme_id": '',
        "upcard_terminal": "83100169",
        "comm_code": "301003400000097",
        "upcard_mer_id": "102831058120008",
        "ex_id": "20136",
        "ex_cost_center_code": "1200043788",
        "dcore_store_appid": "s20170823000005280"
    }
},
{
    "_id": "3850146120185413633",
    "name": "东方浮庭店",
    "extend_code": {
        "comm_shop_id": "0fa2c075b15e44299263902ef1389474",
        "ex_code": "20112",
        "alipay_id": "2015060900077000000000169197",
        "us_id": "42606",
        "comm_code": "301003400000044",
        "upcard_terminal": "02148847",
        "upcard_mer_id": "102210058120752",
        "ex_id": "280",
        "ex_cost_center_code": "1200042606",
        "dcore_store_appid": "s20170823000005510"
    }
},
{
    "_id": "3850146267703279617",
    "name": "南京虹悦城店",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "43883",
        "comm_shop_id": "c90ad1b33a3f45188b0a7d4314707b50",
        "us_id": "43883",
        "alipay_id": "2021022600077000000017037541",
        "takeaway_eleme_id": '',
        "upcard_terminal": "02593483",
        "comm_code": "301003400004648",
        "upcard_mer_id": "102250058122298",
        "ex_id": "43883",
        "ex_cost_center_code": "1200043883",
        "dcore_store_appid": "s20170823000005156"
    }
},
{
    "_id": "3850146121645031425",
    "name": "济南魏家庄万达店",
    "extend_code": {
        "comm_shop_id": "6dc82249c3f44e2ba63f3fbeb7aaf75d",
        "ex_code": "20115",
        "alipay_id": "2015061100077000000000192963",
        "us_id": "42612",
        "comm_code": "301003400000216",
        "upcard_terminal": "53100856",
        "upcard_mer_id": "102531058120055",
        "ex_id": "535",
        "ex_cost_center_code": "1200042612",
        "dcore_store_appid": "s20170823000005217"
    }
},
{
    "_id": "3850146199424204801",
    "name": "重庆时代天街店",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20247",
        "comm_shop_id": "c7c9af9fe3a147918f3c66e2f7f0e874",
        "us_id": "43170",
        "alipay_id": "2015061200077000000000182879",
        "takeaway_eleme_id": '',
        "upcard_terminal": "02306258",
        "comm_code": "301003400000362",
        "upcard_mer_id": "102230058120635",
        "ex_id": "644",
        "ex_cost_center_code": "1200043170",
        "dcore_store_appid": "s20170823000005358"
    }
},
{
    "_id": "3850146288884514817",
    "name": "温州平阳万达",
    "extend_code": {
        "comm_shop_id": "f75825e314464a8a9a4d1e564cd6dd09",
        "ex_code": "20402",
        "alipay_id": "2015061100077000000000191194",
        "us_id": "43999",
        "comm_code": "301003400000446",
        "upcard_terminal": "57700529",
        "upcard_mer_id": "102577058120129",
        "ex_id": "",
        "ex_cost_center_code": "1200043999",
        "dcore_store_appid": "s20170823000005598"
    }
},
{
    "_id": "3850146271671091201",
    "name": "苏州繁花中心",
    "extend_code": {
        "comm_shop_id": "8085fbdb50424923a269655bb24bb1a7",
        "ex_code": "20370",
        "alipay_id": "2015061100077000000000194497",
        "us_id": "43876",
        "comm_code": "301003400000245",
        "upcard_terminal": "51213123",
        "upcard_mer_id": "102512058122265",
        "ex_id": "20162",
        "ex_cost_center_code": "1200043876",
        "dcore_store_appid": "s20170823000005413"
    }
},
{
    "_id": "3850146151290372097",
    "name": "武汉经开万达店",
    "extend_code": {
        "comm_shop_id": "ba397866d98d4d78aebc72739553078b",
        "ex_code": "20170",
        "alipay_id": "2015061200077000000000191251",
        "us_id": "42779",
        "comm_code": "301003400000338",
        "upcard_terminal": "02713547",
        "upcard_mer_id": "102270058122408",
        "ex_id": "921",
        "ex_cost_center_code": "1200042779",
        "dcore_store_appid": "s20170823000005534"
    }
},
{
    "_id": "3850146281057943553",
    "name": "嘉善恒利广场",
    "extend_code": {
        "comm_shop_id": "cadefa14887349eaaaec7caaa9738c62",
        "ex_code": "20387",
        "alipay_id": "2015060900077000000000169205",
        "us_id": "43907",
        "comm_code": "301003400000369",
        "upcard_terminal": "57302203",
        "upcard_mer_id": "102573058120071",
        "ex_id": "20184",
        "ex_cost_center_code": "1200043907",
        "dcore_store_appid": "s20170823000005418"
    }
},
{
    "_id": "3850146305741422593",
    "name": "南宁西关店",
    "extend_code": {
        "us_id": "43995",
        "ex_id": "20221",
        "ex_cost_center_code": "1200043995",
        "ex_code": "20426"
    }
},
{
    "_id": "3863171950377435137",
    "name": "高邮世贸广场店",
    "extend_code": {
        "comm_shop_id": "1d8b960d830c40ce908e27362daeb651",
        "ex_cost_center_code": "1200844541",
        "alipay_id": "2016111600077000000020025550",
        "us_id": "44541",
        "comm_code": "301003400004478",
        "upcard_terminal": "51400880",
        "upcard_mer_id": "102514058120266",
        "ex_id": "20431",
        "ex_code": "20608",
        "dcore_store_appid": "s20170823000005422"
    }
},
{
    "_id": "3850146157984481281",
    "name": "贵阳金阳",
    "extend_code": {
        "comm_shop_id": "c4fb752d9c684956bb1465500fd3d0c5",
        "ex_code": "20183",
        "alipay_id": "2015061200077000000000194540",
        "us_id": "42816",
        "comm_code": "301003400000354",
        "upcard_terminal": "85101186",
        "upcard_mer_id": "102851058120162",
        "ex_id": "724",
        "ex_cost_center_code": "1200042816",
        "dcore_store_appid": "s20170823000005537"
    }
},
{
    "_id": "3850146281582231553",
    "name": "盐城中南城店",
    "extend_code": {
        "comm_shop_id": "00f55c720f2d490abed0048123b2d42c",
        "ex_code": "20388",
        "alipay_id": "2015061100077000000000194500",
        "us_id": "43750",
        "comm_code": "301003400004442",
        "upcard_terminal": "51545352",
        "upcard_mer_id": "102515058120366",
        "ex_id": "20166",
        "ex_cost_center_code": "1200843750",
        "dcore_store_appid": "s20170823000005153"
    }
},
{
    "_id": "3850146317267369985",
    "name": "昆明嘉年华广场",
    "extend_code": {
        "comm_shop_id": "b3acca69ddcc4015bf9bc38ce2f4c545",
        "ex_code": "20445",
        "alipay_id": "2015061200077000000000191233",
        "us_id": "44079",
        "comm_code": "301003400000330",
        "upcard_terminal": "87109746",
        "upcard_mer_id": "102871058126161",
        "ex_id": "",
        "ex_cost_center_code": "1200044079",
        "dcore_store_appid": "s20170823000005613"
    }
},
{
    "_id": "3863178864440115201",
    "name": "杭州萧山银隆（加盟）",
    "extend_code": {
        "comm_shop_id": "40bece0ef221432ab200baa91f0ad110",
        "ex_cost_center_code": "1200844532",
        "alipay_id": "2016111500077000000020026700",
        "us_id": "44532",
        "comm_code": "301003400004511",
        "upcard_terminal": "57109657",
        "upcard_mer_id": "102571058122478",
        "ex_id": "20436",
        "ex_code": "20611",
        "dcore_store_appid": "s20170823000005175"
    }
},
{
    "_id": "3850146340617060353",
    "name": "东营万达店",
    "extend_code": {
        "comm_shop_id": "7d87f6e87d5f4d96aa2b1c2ca92675f8",
        "ex_code": "20478",
        "alipay_id": "2016031800077000000015089785",
        "us_id": "44210",
        "comm_code": "301003400000240",
        "upcard_terminal": "54600157",
        "upcard_mer_id": "102546058120067",
        "ex_id": "20283",
        "ex_cost_center_code": "1200044210",
        "dcore_store_appid": "s20170823000005633"
    }
},
{
    "_id": "3850146166582804481",
    "name": "潍坊泰华城",
    "extend_code": {
        "comm_shop_id": "fa355e5b759f42058b7fe22c895b48b4",
        "ex_code": "20192",
        "alipay_id": "2015061200077000000000182861",
        "us_id": "42883",
        "comm_code": "301003400000449",
        "upcard_terminal": "53600620",
        "upcard_mer_id": "102536058120051",
        "ex_id": "539",
        "ex_cost_center_code": "1200042883",
        "dcore_store_appid": "s20170823000005550"
    }
},
{
    "_id": "3850146262699474945",
    "name": "吉安天虹店",
    "extend_code": {
        "comm_shop_id": "c76cd16b7532447887c55e4184f70875",
        "ex_code": "20354",
        "alipay_id": "2015061000077000000000192921",
        "us_id": "43828",
        "comm_code": "301003400000361",
        "upcard_terminal": "79600227",
        "upcard_mer_id": "102796058120016",
        "ex_id": "20148",
        "ex_cost_center_code": "1200043828",
        "dcore_store_appid": "s20170823000005406"
    }
},
{
    "_id": "3850146063688138753",
    "name": "成都万达店",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20014",
        "comm_shop_id": "9bbeaff557334a53adb6580e5be1c432",
        "us_id": "42015",
        "alipay_id": "2015061200077000000000192996",
        "takeaway_eleme_id": '',
        "upcard_terminal": "02817520",
        "comm_code": "301003400000286",
        "upcard_mer_id": "102280058121934",
        "ex_id": "803",
        "ex_cost_center_code": "1200042015",
        "dcore_store_appid": "s20170823000005473"
    }
},
{
    "_id": "3850146178054225921",
    "name": "厦门SM广场",
    "extend_code": {
        "comm_shop_id": "0927485e3f7a48498cd5e1a379257c1e",
        "ex_code": "20214",
        "alipay_id": "2015061200077000000000194534",
        "us_id": "42947",
        "comm_code": "301003400000370",
        "upcard_terminal": "59204818",
        "upcard_mer_id": "102592058120359",
        "ex_id": "756",
        "ex_cost_center_code": "1200042947",
        "dcore_store_appid": "s20170823000005347"
    }
},
{
    "_id": "3850146349529956353",
    "name": "合肥百大鼓楼店",
    "extend_code": {
        "comm_shop_id": "5a21b27b1c674e5cb01b6af9ce80e841",
        "ex_code": "20494",
        "alipay_id": "2015093000077000000004515141",
        "us_id": "44261",
        "comm_code": "301003400000184",
        "upcard_terminal": "55126320",
        "upcard_mer_id": "102551058123558",
        "ex_id": "20297",
        "ex_cost_center_code": "1200044261",
        "dcore_store_appid": "s20170823000005316"
    }
},
{
    "_id": "3850146061125419009",
    "name": "上海大宁店",
    "extend_code": {
        "comm_shop_id": "8a6425413a3448c1985d4f329b791c22",
        "ex_code": "20009",
        "alipay_id": "2015060900077000000000176450",
        "us_id": "41765",
        "comm_code": "301003400000260",
        "upcard_terminal": "02194797",
        "upcard_mer_id": "102210058126897",
        "ex_id": "207",
        "ex_cost_center_code": "1200041765",
        "dianping_store_id": "1958392",
        "dcore_store_appid": "s20170823000005469"
    }
},
{
    "_id": "3850146372489576449",
    "name": "成都大悦城",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20534",
        "comm_shop_id": "70e1769c0ad14e54994cf200cd48d09d",
        "us_id": "44361",
        "alipay_id": "2016010800077000000014149218",
        "takeaway_eleme_id": '',
        "upcard_terminal": "02891111",
        "comm_code": "301003400000224",
        "upcard_mer_id": "102280058125064",
        "ex_id": "20349",
        "ex_cost_center_code": "1200044361",
        "dcore_store_appid": "s20170823000005322"
    }
},
{
    "_id": "3850146365829021697",
    "name": "柳州城中万达店",
    "extend_code": {
        "comm_shop_id": "262af0d9100348c6adb1435b32b3ac62",
        "ex_code": "20521",
        "alipay_id": "2016011100077000000014253180",
        "us_id": "44232",
        "comm_code": "301003400000084",
        "upcard_terminal": "77201445",
        "upcard_mer_id": "102772058120730",
        "ex_id": "20331",
        "ex_cost_center_code": "1200044232",
        "dcore_store_appid": "s20170823000005639"
    }
},
{
    "_id": "3850146061611958273",
    "name": "上海百联中环店",
    "extend_code": {
        "comm_shop_id": "15d293f2e4874550ac2b84065c7cb9d3",
        "ex_code": "20010",
        "alipay_id": "2015060900077000000000176447",
        "us_id": "41881",
        "comm_code": "301003400000053",
        "upcard_terminal": "02193652",
        "upcard_mer_id": "102210058120683",
        "ex_id": "217",
        "ex_cost_center_code": "1200041881",
        "dcore_store_appid": "s20170823000005470"
    }
},
{
    "_id": "3850146318303363073",
    "name": "聊城金鼎购物中心",
    "extend_code": {
        "comm_shop_id": "ca2efb3a185b4b6b88ec3435977d8b53",
        "ex_code": "20447",
        "alipay_id": "2015061200077000000000191221",
        "us_id": "44088",
        "comm_code": "301003400000367",
        "upcard_terminal": "63500056",
        "upcard_mer_id": "102635058120013",
        "ex_id": "",
        "ex_cost_center_code": "1200044088",
        "dcore_store_appid": "s20170823000005616"
    }
},
{
    "_id": "3850146295918362625",
    "name": "无锡荟聚中心",
    "extend_code": {
        "comm_shop_id": "29455f3e5090428fa7b9ae4f67f712dd",
        "ex_code": "20416",
        "alipay_id": "2015061000077000000000194480",
        "us_id": "43960",
        "comm_code": "301003400000090",
        "upcard_terminal": "51102342",
        "upcard_mer_id": "102510058121879",
        "ex_id": "20203",
        "ex_cost_center_code": "1200043960",
        "dcore_store_appid": "s20170823000005589"
    }
},
{
    "_id": "3850146175533449217",
    "name": "郑州CBD",
    "extend_code": {
        "comm_shop_id": "d8a11f853f2343e99f579c9b7a4a517c",
        "ex_code": "20209",
        "alipay_id": "2015061100077000000000191205",
        "us_id": "42886",
        "comm_code": "301003400000396",
        "upcard_terminal": "37111877",
        "upcard_mer_id": "102371058122434",
        "ex_id": "938",
        "ex_cost_center_code": "1200042886",
        "dcore_store_appid": "s20170823000005232"
    }
},
{
    "_id": "3850146371000598529",
    "name": "昆明海伦国际店",
    "extend_code": {
        "comm_shop_id": "fee78919a04f4b0eaf4c536c806500cf",
        "ex_code": "20531",
        "alipay_id": "2016011100077000000014087484",
        "us_id": "44364",
        "comm_code": "301003400000460",
        "upcard_terminal": "87111655",
        "upcard_mer_id": "102871058126571",
        "ex_id": "20373",
        "ex_cost_center_code": "1200044364",
        "dcore_store_appid": "s20170823000005667"
    }
},
{
    "_id": "3850146187067785217",
    "name": "宝山万达",
    "extend_code": {
        "comm_shop_id": "e0a7e8a852e0469a9dab10e26758a557",
        "ex_code": "20231",
        "alipay_id": "2015060900077000000000178444",
        "us_id": "42941",
        "comm_code": "301003400000407",
        "upcard_terminal": "02194801",
        "upcard_mer_id": "102210058126901",
        "ex_id": "2009",
        "ex_cost_center_code": "1200042941",
        "dcore_store_appid": "s20170823000005235"
    }
},
{
    "_id": "3850146320966746113",
    "name": "武汉宜家店",
    "extend_code": {
        "comm_shop_id": "89d2dd5793bb4d4fa59d9600ef9bb6a9",
        "ex_code": "20452",
        "alipay_id": "2015093000077000000004511902",
        "us_id": "44109",
        "comm_code": "301003400000259",
        "upcard_terminal": "02723063",
        "upcard_mer_id": "102270058125024",
        "ex_id": "20257",
        "ex_cost_center_code": "1200044109",
        "dcore_store_appid": "s20170823000005618"
    }
},
{
    "_id": "3850146177555103745",
    "name": "岳阳步步高",
    "extend_code": {
        "comm_shop_id": "a9d1b6813dcf45ab87e8fc9f4cee61de",
        "ex_code": "20213",
        "alipay_id": "2015061200077000000000182887",
        "us_id": "42898",
        "comm_code": "301003400000315",
        "upcard_terminal": "73000039",
        "upcard_mer_id": "102730058120004",
        "ex_id": "833",
        "ex_cost_center_code": "1200042898",
        "dcore_store_appid": "s20170823000005555"
    }
},
{
    "_id": "3850146197834563585",
    "name": "黄岛佳世客",
    "extend_code": {
        "comm_shop_id": "83bdca5abd1c49beb0fdfb94d5dd4c5a",
        "ex_code": "20244",
        "alipay_id": "2015061100077000000000182841",
        "us_id": "43101",
        "comm_code": "301003400000249",
        "upcard_terminal": "53204081",
        "upcard_mer_id": "102532058120851",
        "ex_id": "132",
        "ex_cost_center_code": "1200043101",
        "dcore_store_appid": "s20170823000005354"
    }
},
{
    "_id": "3850146374586728449",
    "name": "贵阳正德家邦店",
    "extend_code": {
        "comm_shop_id": "4ffe83f9925744ed8a809af10019c02e",
        "ex_code": "20538",
        "alipay_id": "2016011100077000000014076723",
        "us_id": "44246",
        "comm_code": "301003400000161",
        "upcard_terminal": "85101430",
        "upcard_mer_id": "102851058120424",
        "ex_id": "20341",
        "ex_cost_center_code": "1200044246",
        "dcore_store_appid": "s20170823000005647"
    }
},
{
    "_id": "3850146283670994945",
    "name": "龙岩万达店",
    "extend_code": {
        "comm_shop_id": "bfe5e4d9142b4644a3d061f0ffc0012c",
        "ex_code": "20392",
        "alipay_id": "2015061200077000000000188797",
        "us_id": "43958",
        "comm_code": "301003400000343",
        "upcard_terminal": "59700254",
        "upcard_mer_id": "102597058120113",
        "ex_id": "20169",
        "ex_cost_center_code": "1200043958",
        "dcore_store_appid": "s20170823000005587"
    }
},
{
    "_id": "3850146375090044929",
    "name": "都江堰百伦店",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20539",
        "comm_shop_id": "181c3bea399b41b18887e7ae206b21a0",
        "us_id": "44345",
        "alipay_id": "2016011100077000000014257689",
        "takeaway_eleme_id": '',
        "upcard_terminal": "02891194",
        "comm_code": "301003400000060",
        "upcard_mer_id": "102280058125088",
        "ex_id": "20347",
        "ex_cost_center_code": "1200044345",
        "dcore_store_appid": "s20170823000005662"
    }
},
{
    "_id": "3850146248648556545",
    "name": "合肥银泰店",
    "extend_code": {
        "comm_shop_id": "4dcefc0620c54fc6861ecf203be1c504",
        "ex_code": "20327",
        "alipay_id": "2015061100077000000000192966",
        "us_id": "43785",
        "comm_code": "301003400000157",
        "upcard_terminal": "55120663",
        "upcard_mer_id": "102551058123256",
        "ex_id": "20122",
        "ex_cost_center_code": "1200043785",
        "dcore_store_appid": "s20170823000005400"
    }
},
{
    "_id": "3850146284325306369",
    "name": "老闵行店",
    "extend_code": {
        "comm_shop_id": "ba413dc591fc47279e60f1a73bbfc73e",
        "ex_code": "20393",
        "alipay_id": "2015060900077000000000169204",
        "us_id": "43908",
        "comm_code": "301003400000339",
        "upcard_terminal": "02194924",
        "upcard_mer_id": "102210058125946",
        "ex_id": "20199",
        "ex_cost_center_code": "1200043908",
        "dcore_store_appid": "s20170823000005419"
    }
},
{
    "_id": "3850146057908387841",
    "name": "上海西郊百联店",
    "extend_code": {
        "comm_shop_id": "031efd37cf6b4c5eb5ea8d3e54b6457c",
        "ex_code": "20003",
        "alipay_id": "2015060900077000000000178445",
        "us_id": "41589",
        "comm_code": "301003400000027",
        "upcard_terminal": "02194921",
        "upcard_mer_id": "102210058120675",
        "ex_id": "205",
        "ex_cost_center_code": "1200041589",
        "dcore_store_appid": "s20170823000005467"
    }
},
{
    "_id": "3850146251794284545",
    "name": "连云港嘉瑞宝",
    "extend_code": {
        "comm_shop_id": "b9c604d14a1b4ad495b546b06efee524",
        "ex_code": "20333",
        "alipay_id": "2015061100077000000000182832",
        "us_id": "43715",
        "comm_code": "301003400000337",
        "upcard_terminal": "51800579",
        "upcard_mer_id": "102518058120014",
        "ex_id": "20125",
        "ex_cost_center_code": "1200043715",
        "dcore_store_appid": "s20170823000005275"
    }
},
{
    "_id": "3850146377275277313",
    "name": "平顶山鹰城世贸",
    "extend_code": {
        "comm_shop_id": "05c5a2b824a64e7196047cc79a9af122",
        "ex_code": "20543",
        "alipay_id": "2016011100077000000014135346",
        "us_id": "44366",
        "comm_code": "301003400000030",
        "upcard_terminal": "37501828",
        "upcard_mer_id": "102375058120285",
        "ex_id": "20370",
        "ex_cost_center_code": "1200044366",
        "dcore_store_appid": "s20170823000005668"
    }
},
{
    "_id": "3850146067429457921",
    "name": "南京1912店",
    "extend_code": {
        "comm_shop_id": "b38f50abe22b486895c39761634406ec",
        "ex_code": "20021",
        "alipay_id": "2015061000077000000000182809",
        "us_id": "42126",
        "comm_code": "301003400000328",
        "upcard_terminal": "02516550",
        "upcard_mer_id": "102250058120828",
        "ex_id": "305",
        "ex_cost_center_code": "1200042126",
        "dcore_store_appid": "s20170823000005477"
    }
},
{
    "_id": "3850146096793780225",
    "name": "西安解放万达店",
    "extend_code": {
        "comm_shop_id": "4631795460bb479599c0a17e7ecc752e",
        "ex_code": "20067",
        "alipay_id": "2015061200077000000000193015",
        "us_id": "42387",
        "comm_code": "301003400000141",
        "upcard_terminal": "02903681",
        "upcard_mer_id": "102290058122444",
        "ex_id": "958",
        "ex_cost_center_code": "1200042387",
        "dcore_store_appid": "s20170823000005564"
    }
},
{
    "_id": "3850146252788334593",
    "name": "中原城市广场",
    "extend_code": {
        "comm_shop_id": "0cba0c7b5f98485b90e4660d0636ef52",
        "ex_code": "20335",
        "alipay_id": "2015060900077000000000176456",
        "us_id": "43776",
        "comm_code": "301003400000042",
        "upcard_terminal": "02193285",
        "upcard_mer_id": "102210058125782",
        "ex_id": "20128",
        "ex_cost_center_code": "1200043776",
        "dcore_store_appid": "s20170823000005395"
    }
},
{
    "_id": "3850146328948506625",
    "name": "柳州步步高店",
    "extend_code": {
        "comm_shop_id": "5559d0ad6aee496b86a80e449ae4d98d",
        "ex_code": "20466",
        "alipay_id": "2015093000077000000004499871",
        "us_id": "44178",
        "comm_code": "301003400000172",
        "upcard_terminal": "77201405",
        "upcard_mer_id": "102772058120690",
        "ex_id": "20271",
        "ex_cost_center_code": "1200044178",
        "dcore_store_appid": "s20170823000005627"
    }
},
{
    "_id": "3850146292927823873",
    "name": "个旧丽水金湾",
    "extend_code": {
        "comm_shop_id": "33a9c52f61e347c492499be919f96225",
        "ex_code": "20410",
        "alipay_id": "2015061200077000000000193008",
        "us_id": "43993",
        "comm_code": "301003400000110",
        "upcard_terminal": "87300506",
        "upcard_mer_id": "102873058120363",
        "ex_id": "20206",
        "ex_cost_center_code": "1200043993",
        "dcore_store_appid": "s20170823000005295"
    }
},
{
    "_id": "3850146242206105601",
    "name": "三林金谊广场",
    "extend_code": {
        "comm_shop_id": "5f7225e33bbc49d4a436a744941fab01",
        "ex_code": "20314",
        "alipay_id": "2015060900077000000000166164",
        "us_id": "43678",
        "comm_code": "301003400000198",
        "upcard_terminal": "02193707",
        "upcard_mer_id": "102210058125752",
        "ex_id": "20080",
        "ex_cost_center_code": "1200043678",
        "dcore_store_appid": "s20170823000005386"
    }
},
{
    "_id": "3850146068461256705",
    "name": "南京金轮店",
    "extend_code": {
        "comm_shop_id": "8751a8393afe48cca5c78a42e9f2bf99",
        "ex_code": "20023",
        "alipay_id": "2015061000077000000000191171",
        "us_id": "42137",
        "comm_code": "301003400000255",
        "upcard_terminal": "02516544",
        "upcard_mer_id": "102250058120834",
        "ex_id": "306",
        "ex_cost_center_code": "1200042137",
        "dcore_store_appid": "s20170823000005479"
    }
},
{
    "_id": "3850146085720817665",
    "name": "上海惠南店",
    "extend_code": {
        "comm_shop_id": "9ea930cba93640a9bac3b9248a360449",
        "ex_code": "20045",
        "alipay_id": "2015060900077000000000174616",
        "us_id": "42302",
        "comm_code": "301003400000294",
        "upcard_terminal": "02148880",
        "upcard_mer_id": "102210058120719",
        "ex_id": "248",
        "ex_cost_center_code": "1200042302",
        "dcore_store_appid": "s20170823000005491"
    }
},
{
    "_id": "3850146250200449025",
    "name": "无锡清扬路",
    "extend_code": {
        "comm_shop_id": "3c5c9e21b5264dcd901369d2dcdb68a1",
        "ex_code": "20330",
        "alipay_id": "2020030200077000000092104834",
        "us_id": "43789",
        "comm_code": "301003400004506",
        "upcard_terminal": "51103375",
        "upcard_mer_id": "102510058122096",
        "ex_id": "20121",
        "ex_cost_center_code": "1200043789",
        "dcore_store_appid": "s20170823000005154"
    }
},
{
    "_id": "3850146097267736577",
    "name": "重庆百联店",
    "extend_code": {
        "comm_shop_id": "2cda330de808408e808ec1dd41a80d7d",
        "ex_code": "20068",
        "alipay_id": "2015061200077000000000188808",
        "us_id": "42386",
        "comm_code": "301003400000094",
        "upcard_terminal": "02306266",
        "upcard_mer_id": "102230058120113",
        "ex_id": "637",
        "ex_cost_center_code": "1200042386",
        "dcore_store_appid": "s20170823000005563"
    }
},
{
    "_id": "3850146200409866241",
    "name": "成都金牛万达店",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20249",
        "comm_shop_id": "cfa1636b843147e689c28d16d33e5501",
        "us_id": "43173",
        "alipay_id": "2015061200077000000000192998",
        "takeaway_eleme_id": '',
        "upcard_terminal": "02817503",
        "comm_code": "301003400000383",
        "upcard_mer_id": "102280058122947",
        "ex_id": "8002",
        "ex_cost_center_code": "1200043173",
        "dcore_store_appid": "s20170823000005359"
    }
},
{
    "_id": "3850146315124080641",
    "name": "武汉武商众圆广场",
    "extend_code": {
        "comm_shop_id": "077ad732b004427db9136f4ff395247d",
        "ex_code": "20441",
        "alipay_id": "2015061200077000000000188810",
        "us_id": "43991",
        "comm_code": "301003400000031",
        "upcard_terminal": "02722812",
        "upcard_mer_id": "102270058124931",
        "ex_id": "20204",
        "ex_cost_center_code": "1200043991",
        "dcore_store_appid": "s20170823000005593"
    }
},
{
    "_id": "3850146351216066561",
    "name": "太原龙湖万达",
    "extend_code": {
        "comm_shop_id": "fade296a0d29471a9f5dea7062299c13",
        "ex_code": "20497",
        "alipay_id": "2015113000077000000006669872",
        "us_id": "44237",
        "comm_code": "301003400000451",
        "upcard_terminal": "35103694",
        "upcard_mer_id": "102351058121580",
        "ex_id": "20291",
        "ex_cost_center_code": "1200044237",
        "dcore_store_appid": "s20170823000005313"
    }
},
{
    "_id": "3850146299441577985",
    "name": "盐城宝龙店",
    "extend_code": {
        "comm_shop_id": "c9d21b9803a64f0182469c5d664bb532",
        "ex_code": "20420",
        "alipay_id": "2015061100077000000000192954",
        "us_id": "43996",
        "comm_code": "301003400000365",
        "upcard_terminal": "51545365",
        "upcard_mer_id": "102515058120367",
        "ex_id": "20219",
        "ex_cost_center_code": "1200043996",
        "dcore_store_appid": "s20170823000005597"
    }
},
{
    "_id": "3850146284807651329",
    "name": "烟台万达店",
    "extend_code": {
        "comm_shop_id": "ff19931090de4fad8a464936a4aa8394",
        "ex_code": "20394",
        "alipay_id": "2015061100077000000000192965",
        "us_id": "44000",
        "comm_code": "301003400000463",
        "upcard_terminal": "53530348",
        "upcard_mer_id": "102535058120140",
        "ex_id": "20191",
        "ex_cost_center_code": "1200044000",
        "dcore_store_appid": "s20170823000005599"
    }
},
{
    "_id": "3850146124178391041",
    "name": "湖州爱山店",
    "extend_code": {
        "comm_shop_id": "9e3161b7f44d45f7b1fc015ba783ed36",
        "ex_code": "20119",
        "alipay_id": "2021092200077000000027981843",
        "us_id": "42608",
        "comm_code": "301003400000292",
        "upcard_terminal": "57200075",
        "upcard_mer_id": "102572058120006",
        "ex_id": "421",
        "ex_cost_center_code": "1200042608",
        "dcore_store_appid": "s20170823000005511"
    }
},
{
    "_id": "3850146368945389569",
    "name": "南通万达",
    "extend_code": {
        "comm_shop_id": "ef764caeffc04408af79c5023d4428c3",
        "ex_code": "20527",
        "alipay_id": "2016031800077000000014990612",
        "us_id": "44235",
        "comm_code": "301003400000431",
        "upcard_terminal": "51300830",
        "upcard_mer_id": "102513058120145",
        "ex_id": "20336",
        "ex_cost_center_code": "1200044235",
        "dcore_store_appid": "s20170823000005641"
    }
},
{
    "_id": "3850146222857781249",
    "name": "贵阳鸿通城",
    "extend_code": {
        "comm_shop_id": "2bd98fe0497f4a7596185c47a927810b",
        "ex_code": "20277",
        "alipay_id": "2015061200077000000000194542",
        "us_id": "43468",
        "comm_code": "301003400000092",
        "upcard_terminal": "85101432",
        "upcard_mer_id": "102851058120346",
        "ex_id": "",
        "ex_cost_center_code": "1200043468",
        "dcore_store_appid": "s20170823000005247"
    }
},
{
    "_id": "3850146384481091585",
    "name": "宁波高鑫广场",
    "extend_code": {
        "comm_shop_id": "6af19c6e0672487583d82c77e53af9d1",
        "ex_code": "20556",
        "alipay_id": "2016031800077000000015016816",
        "us_id": "44408",
        "comm_code": "301003400000209",
        "upcard_terminal": "57403294",
        "upcard_mer_id": "102574058120518",
        "ex_id": "20377",
        "ex_cost_center_code": "1200044408",
        "dcore_store_appid": "s20170823000005325"
    }
},
{
    "_id": "3850146272237322241",
    "name": "东莞东城万达店",
    "extend_code": {
        "comm_shop_id": "d8795b25a28448f29217ac8f4ab3de36",
        "ex_code": "20371",
        "alipay_id": "2015061200077000000000188790",
        "us_id": "43904",
        "comm_code": "301003400000395",
        "upcard_terminal": "75900242",
        "upcard_mer_id": "102759058120011",
        "ex_id": "20154",
        "ex_cost_center_code": "1200043904",
        "dcore_store_appid": "s20170823000005416"
    }
},
{
    "_id": "3850146316751470593",
    "name": "珠海富华里",
    "extend_code": {
        "comm_shop_id": "5bc1cefbec154c4e98ebfae9b1ca330f",
        "ex_code": "20444",
        "alipay_id": "2015061200077000000000182867",
        "us_id": "44086",
        "comm_code": "301003400000190",
        "upcard_terminal": "75601310",
        "upcard_mer_id": "102756058120158",
        "ex_id": "20234",
        "ex_cost_center_code": "1200044086",
        "dcore_store_appid": "s20170823000005614"
    }
},
{
    "_id": "3850146297424117761",
    "name": "郑州瀚海北金店",
    "extend_code": {
        "us_id": "43957",
        "ex_id": "20218",
        "ex_cost_center_code": "1200043957",
        "ex_code": "20417"
    }
},
{
    "_id": "3850146273227177985",
    "name": "荆州万达店",
    "extend_code": {
        "comm_shop_id": "68bd62f85a854849a7fa63016ebed89e",
        "ex_code": "20373",
        "alipay_id": "2015061200077000000000182880",
        "us_id": "43910",
        "comm_code": "301003400000206",
        "upcard_terminal": "71600159",
        "upcard_mer_id": "102716058120041",
        "ex_id": "20170",
        "ex_cost_center_code": "1200043910",
        "dcore_store_appid": "s20170823000005420"
    }
},
{
    "_id": "3850146234614415361",
    "name": "徐州云龙万达",
    "extend_code": {
        "comm_shop_id": "4e04f82fe6ec48f08a1350c017595e72",
        "ex_code": "20299",
        "alipay_id": "2015061100077000000000194499",
        "us_id": "43608",
        "comm_code": "301003400000158",
        "upcard_terminal": "51600591",
        "upcard_mer_id": "102516058120113",
        "ex_id": "20052",
        "ex_cost_center_code": "1200043608",
        "dcore_store_appid": "s20170823000005376"
    }
},
{
    "_id": "3850146274363834369",
    "name": "芜湖华强店",
    "extend_code": {
        "comm_shop_id": "c9e412d139af4525be057d518d751ea4",
        "ex_code": "20375",
        "alipay_id": "2015061100077000000000194520",
        "us_id": "43909",
        "comm_code": "301003400000366",
        "upcard_terminal": "55300828",
        "upcard_mer_id": "102553058120217",
        "ex_id": "20177",
        "ex_cost_center_code": "1200043909",
        "dcore_store_appid": "s20170823000005287"
    }
},
{
    "_id": "3850146403821027329",
    "name": "昆明广场",
    "extend_code": {
        "comm_shop_id": "afe42b24523d4e1c89288f5ebe700579",
        "ex_code": "20618",
        "alipay_id": "2016111600077000000019999004",
        "us_id": "44434",
        "comm_code": "301003400000319",
        "upcard_terminal": "87111711",
        "upcard_mer_id": "102871058126596",
        "ex_id": "20400",
        "ex_cost_center_code": "1200044434",
        "dcore_store_appid": "s20170823000005326"
    }
},
{
    "_id": "3850146327337893889",
    "name": "嘉兴万达",
    "extend_code": {
        "comm_shop_id": "e5d6f0c8b2aa4e8398afc628153b07d0",
        "ex_code": "20463",
        "alipay_id": "2015093000077000000004500946",
        "us_id": "44135",
        "comm_code": "301003400000415",
        "upcard_terminal": "57302985",
        "upcard_mer_id": "102573058120082",
        "ex_id": "",
        "ex_cost_center_code": "1200044135",
        "dcore_store_appid": "s20170823000005623"
    }
},
{
    "_id": "3850146277048188929",
    "name": "武汉群星城店",
    "extend_code": {
        "comm_shop_id": "cd8c0906bbd6473b89181d5f38734da0",
        "ex_code": "20380",
        "alipay_id": "2015061200077000000000194550",
        "us_id": "43955",
        "comm_code": "301003400000376",
        "upcard_terminal": "02722084",
        "upcard_mer_id": "102270058124770",
        "ex_id": "20186",
        "ex_cost_center_code": "1200043955",
        "dcore_store_appid": "s20170823000005586"
    }
},
{
    "_id": "3850146309465964545",
    "name": "宜昌水悦城",
    "extend_code": {
        "comm_shop_id": "51af7b725e0b45269ef88e5ea7bff8a6",
        "ex_code": "20430",
        "alipay_id": "2015093000077000000004515138",
        "us_id": "44078",
        "comm_code": "301003400000166",
        "upcard_terminal": "71701841",
        "upcard_mer_id": "102717058120173",
        "ex_id": "",
        "ex_cost_center_code": "1200044078",
        "dcore_store_appid": "s20170823000005612"
    }
},
{
    "_id": "3850146407067418625",
    "name": "湛江万达店",
    "extend_code": {
        "comm_shop_id": "da295504d8a9479590e1401b9c41774f",
        "ex_code": "20605",
        "alipay_id": "2016111700077000000020031207",
        "us_id": "44506",
        "comm_code": "301003400000402",
        "upcard_terminal": "75900414",
        "upcard_mer_id": "102759058120040",
        "ex_id": "20417",
        "ex_cost_center_code": "1200044506",
        "dcore_store_appid": "s20170823000005680"
    }
},
{
    "_id": "3850146283159289857",
    "name": "淮南新世界店",
    "extend_code": {
        "comm_shop_id": "64bc18a216c04d1789cf5c3744fb939f",
        "ex_code": "20391",
        "alipay_id": "2015061200077000000000192982",
        "us_id": "43923",
        "comm_code": "301003400000203",
        "upcard_terminal": "55400193",
        "upcard_mer_id": "102554058120071",
        "ex_id": "20190",
        "ex_cost_center_code": "1200043923",
        "dcore_store_appid": "s20170823000005578"
    }
},
{
    "_id": "3850146347965480961",
    "name": "无锡人民路百盛",
    "extend_code": {
        "comm_shop_id": "c3cac56713fa4ffab9e5e9d3b345713d",
        "ex_code": "20491",
        "alipay_id": "2015093000077000000004505788",
        "us_id": "44248",
        "comm_code": "301003400000351",
        "upcard_terminal": "51103159",
        "upcard_mer_id": "102510058122068",
        "ex_id": "20304",
        "ex_cost_center_code": "1200044248",
        "dcore_store_appid": "s20170823000005648"
    }
},
{
    "_id": "3850146404366286849",
    "name": "长沙梅溪湖步步高影院店",
    "extend_code": {
        "comm_shop_id": "73c4261111464b33917e0415c67673c7",
        "ex_code": "20561",
        "alipay_id": "2016081500077000000018032559",
        "us_id": "44479",
        "comm_code": "301003400000231",
        "upcard_terminal": "73110806",
        "upcard_mer_id": "102731058121608",
        "ex_id": "20420",
        "ex_cost_center_code": "1200044479",
        "dcore_store_appid": "s20170823000005677"
    }
},
{
    "_id": "3850146324682899457",
    "name": "南京中海环宇城店",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20458",
        "comm_shop_id": "08e63c0575264480b1a9ac2fbc9f241b",
        "us_id": "44137",
        "alipay_id": "2015121500077000000013820426",
        "takeaway_eleme_id": '',
        "upcard_terminal": "02595296",
        "comm_code": "301003400000037",
        "upcard_mer_id": "102250058122468",
        "ex_id": "20261",
        "ex_cost_center_code": "1200044137",
        "dcore_store_appid": "s20170823000005624"
    }
},
{
    "_id": "3850146412192858113",
    "name": "杭州东站加盟店",
    "extend_code": {
        "comm_shop_id": "1ec5943c41aa4014b0cbb78049020959",
        "ex_code": "20584",
        "alipay_id": "2018040900077000000048275897",
        "us_id": "44520",
        "comm_code": "301003400004523",
        "upcard_terminal": "57109726",
        "upcard_mer_id": "102571058122479",
        "ex_id": "20429",
        "ex_cost_center_code": "1200844520",
        "dcore_store_appid": "s20170929000006487"
    }
},
{
    "_id": "3850146126221017089",
    "name": "西藏路大悦城店",
    "extend_code": {
        "comm_shop_id": "3e878781ae1f49478421cb6404c9a970",
        "ex_code": "20123",
        "alipay_id": "2015060900077000000000176455",
        "us_id": "42636",
        "comm_code": "301003400000129",
        "upcard_terminal": "02148839",
        "upcard_mer_id": "102210058120781",
        "ex_id": "284",
        "ex_cost_center_code": "1200042636",
        "dcore_store_appid": "s20170823000005517"
    }
},
{
    "_id": "3850146297898074113",
    "name": "杭州水晶城购物中心",
    "extend_code": {
        "comm_shop_id": "8b43ed89f39044cf8e7191133ae1960a",
        "ex_code": "20418",
        "alipay_id": "2015060900077000000000166166",
        "us_id": "43989",
        "comm_code": "301003400000262",
        "upcard_terminal": "57104289",
        "upcard_mer_id": "102571058121029",
        "ex_id": "20223",
        "ex_cost_center_code": "1200043989",
        "dcore_store_appid": "s20170823000005592"
    }
},
{
    "_id": "3850146334573068289",
    "name": "缤谷广场一期",
    "extend_code": {
        "comm_shop_id": "cd5d30ded8d2492ab8cb31c31442f689",
        "ex_code": "20468",
        "alipay_id": "2015093000077000000004490206",
        "us_id": "44136",
        "comm_code": "301003400000374",
        "upcard_terminal": "02193914",
        "upcard_mer_id": "102210058126730",
        "ex_id": "20276",
        "ex_cost_center_code": "1200044136",
        "dcore_store_appid": "s20170823000005306"
    }
},
{
    "_id": "3850146338947727361",
    "name": "德阳洋洋百货南街店",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20475",
        "comm_shop_id": "64e5c303df4c4610890091e6f8ec955b",
        "us_id": "44121",
        "alipay_id": "2015061200077000000000188804",
        "takeaway_eleme_id": '',
        "upcard_terminal": "83890052",
        "comm_code": "301003400000204",
        "upcard_mer_id": "102838058120072",
        "ex_id": "20273",
        "ex_cost_center_code": "1200044121",
        "dcore_store_appid": "s20170823000005619"
    }
},
{
    "_id": "3850146301953966081",
    "name": "西安曲江金地店",
    "extend_code": {
        "us_id": "44013",
        "ex_id": "20211",
        "ex_cost_center_code": "1200044013",
        "ex_code": "20422"
    }
},
{
    "_id": "3850146339618816001",
    "name": "南通文峰城市广场店",
    "extend_code": {
        "comm_shop_id": "9a1d0639ee4643b68c79f0640361a88c",
        "ex_code": "20476",
        "alipay_id": "2018041900077000000048356376",
        "us_id": "44194",
        "comm_code": "301003400000283",
        "upcard_terminal": "51300623",
        "upcard_mer_id": "102513058120076",
        "ex_id": "20285",
        "ex_cost_center_code": "1200044194",
        "dcore_store_appid": "s20170823000005632"
    }
},
{
    "_id": "3850146366860820481",
    "name": "永城金博大",
    "extend_code": {
        "comm_shop_id": "1b4492f5d80f4086a0be3fbc6011b6d8",
        "ex_code": "20523",
        "alipay_id": "2016011100077000000014092239",
        "us_id": "44296",
        "comm_code": "301003400000064",
        "upcard_terminal": "37001237",
        "upcard_mer_id": "102370058120016",
        "ex_id": "20330",
        "ex_cost_center_code": "1200044296",
        "dcore_store_appid": "s20170823000005318"
    }
},
{
    "_id": "3850146340117938177",
    "name": "渭南万达店",
    "extend_code": {
        "comm_shop_id": "5ee859ae60dc4b03ad97034980f791f9",
        "ex_code": "20477",
        "alipay_id": "2015093000077000000004397091",
        "us_id": "44195",
        "comm_code": "301003400000197",
        "upcard_terminal": "91330019",
        "upcard_mer_id": "102913058120011",
        "ex_id": "20281",
        "ex_cost_center_code": "1200044195",
        "dcore_store_appid": "s20170823000005309"
    }
},
{
    "_id": "3850146370006548481",
    "name": "南沙万达店",
    "extend_code": {
        "comm_shop_id": "15ec81c87b5644569ed514f9c994c6ea",
        "ex_code": "20529",
        "alipay_id": "2016011100077000000014107389",
        "us_id": "44284",
        "comm_code": "301003400000054",
        "upcard_terminal": "02081978",
        "upcard_mer_id": "102200058120765",
        "ex_id": "20342",
        "ex_cost_center_code": "1200044284",
        "dcore_store_appid": "s20170823000005317"
    }
},
{
    "_id": "3850146145628061697",
    "name": "武汉徐东销品茂",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20159",
        "comm_shop_id": "eee38489bbb7468798e2b2259fa06535",
        "us_id": "42758",
        "alipay_id": "2015061200077000000000182884",
        "takeaway_eleme_id": '',
        "upcard_terminal": "02713548",
        "comm_code": "301003400000430",
        "upcard_mer_id": "102270058122304",
        "ex_id": "920",
        "ex_cost_center_code": "1200042758",
        "dcore_store_appid": "s20170823000005531"
    }
},
{
    "_id": "3850146382891450369",
    "name": "李沧乐客城",
    "extend_code": {
        "comm_shop_id": "31439f966dbf4bdd9cf13f146027d547",
        "ex_code": "20553",
        "alipay_id": "2016031800077000000015165667",
        "us_id": "44435",
        "comm_code": "301003400000105",
        "upcard_terminal": "53205649",
        "upcard_mer_id": "102532058121372",
        "ex_id": "20404",
        "ex_cost_center_code": "1200044435",
        "dcore_store_appid": "s20170823000005676"
    }
},
{
    "_id": "3850146146630500353",
    "name": "松江平高",
    "extend_code": {
        "comm_shop_id": "badf3f1d9d254d08bad834731023033f",
        "ex_code": "20161",
        "alipay_id": "2018051500077000000051645407",
        "us_id": "42724",
        "comm_code": "301003400000341",
        "upcard_terminal": "02148834",
        "upcard_mer_id": "102210058120848",
        "ex_id": "291",
        "ex_cost_center_code": "1200042724",
        "dcore_store_appid": "s20170823000005224"
    }
},
{
    "_id": "3850146341648859137",
    "name": "西安曲江龙湖星悦荟店",
    "extend_code": {
        "comm_shop_id": "bac2da8f84504027aba17f95decb0277",
        "ex_code": "20480",
        "alipay_id": "2015093000077000000004496093",
        "us_id": "44239",
        "comm_code": "301003400000340",
        "upcard_terminal": "02991003",
        "upcard_mer_id": "102290058122918",
        "ex_id": "20286",
        "ex_cost_center_code": "1200044239",
        "dcore_store_appid": "s20170823000005644"
    }
},
{
    "_id": "3850146108848209921",
    "name": "金华天地银泰店",
    "extend_code": {
        "comm_shop_id": "c6e191cb0402444294aa17cba0bc420d",
        "ex_code": "20090",
        "alipay_id": "2015061000077000000000188720",
        "us_id": "42493",
        "comm_code": "301003400000359",
        "upcard_terminal": "57103823",
        "upcard_mer_id": "102571058120390",
        "ex_id": "681",
        "ex_cost_center_code": "1200042493",
        "dcore_store_appid": "s20170823000005499"
    }
},
{
    "_id": "3850146345448898561",
    "name": "阜阳颍州万达",
    "extend_code": {
        "comm_shop_id": "cea7b0e32fd84cde968797e2fcdf897c",
        "ex_code": "20487",
        "alipay_id": "2015093000077000000004449376",
        "us_id": "44229",
        "comm_code": "301003400000382",
        "upcard_terminal": "55801250",
        "upcard_mer_id": "102558058120384",
        "ex_id": "20290",
        "ex_cost_center_code": "1200044229",
        "dcore_store_appid": "s20170823000005311"
    }
},
{
    "_id": "3850146405553274881",
    "name": "镇江苏宁（加盟）",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20593",
        "comm_shop_id": "447a8060d3bb47339d0f188d51ac4e71",
        "us_id": "44505",
        "alipay_id": "2016111600077000000019998970",
        "takeaway_eleme_id": '',
        "upcard_terminal": "51103399",
        "comm_code": "301003400004514",
        "upcard_mer_id": "102511058121216",
        "ex_id": "20422",
        "ex_cost_center_code": "1200844505",
        "dcore_store_appid": "s20170823000005338"
    }
},
{
    "_id": "3850146248157822977",
    "name": "郑州锦艺店",
    "extend_code": {
        "comm_shop_id": "6b63438f69da4c91a34d1ce641376ca5",
        "ex_code": "20326",
        "alipay_id": "2018040400077000000048185151",
        "us_id": "43714",
        "comm_code": "301003400000213",
        "upcard_terminal": "37110795",
        "upcard_mer_id": "102371058122279",
        "ex_id": "20110",
        "ex_cost_center_code": "1200043714",
        "dcore_store_appid": "s20170823000005388"
    }
},
{
    "_id": "3850146360992989185",
    "name": "西安印象城",
    "extend_code": {
        "comm_shop_id": "f3ca75a912a74aff8bd22d8fbf3e4d7b",
        "ex_code": "20512",
        "alipay_id": "2015120700077000000013343193",
        "us_id": "44238",
        "comm_code": "301003400000441",
        "upcard_terminal": "02991083",
        "upcard_mer_id": "102290058122930",
        "ex_id": "20318",
        "ex_cost_center_code": "1200044238",
        "dcore_store_appid": "s20170823000005643"
    }
},
{
    "_id": "3850146123696046081",
    "name": "青浦桥梓湾店",
    "extend_code": {
        "comm_shop_id": "9ce3d9fc145e46fa88881bfff3f6fb8a",
        "ex_code": "20118",
        "alipay_id": "2015060900077000000000169208",
        "us_id": "42609",
        "comm_code": "301003400000289",
        "upcard_terminal": "02194803",
        "upcard_mer_id": "102210058126903",
        "ex_id": "282",
        "ex_cost_center_code": "1200042609",
        "dcore_store_appid": "s20170823000005512"
    }
},
{
    "_id": "3850146352289808385",
    "name": "西双版纳万达",
    "extend_code": {
        "comm_shop_id": "029196710a444442b0bad9f54afffd58",
        "ex_code": "20499",
        "alipay_id": "2015093000077000000004511911",
        "us_id": "44240",
        "comm_code": "301003400000026",
        "upcard_terminal": "87110109",
        "upcard_mer_id": "102691058120118",
        "ex_id": "20302",
        "ex_cost_center_code": "1200044240",
        "dcore_store_appid": "s20170823000005314"
    }
},
{
    "_id": "3850146361974456321",
    "name": "绍兴汇金",
    "extend_code": {
        "comm_shop_id": "2de6c5c6a4584162b10e138cd2084cb4",
        "ex_code": "20514",
        "alipay_id": "2015121500077000000013667398",
        "us_id": "44293",
        "comm_code": "301003400000096",
        "upcard_terminal": "57500771",
        "upcard_mer_id": "102575058120106",
        "ex_id": "20327",
        "ex_cost_center_code": "1200044293",
        "dcore_store_appid": "s20170823000005653"
    }
},
{
    "_id": "3850146149830754305",
    "name": "西安立丰广场店",
    "extend_code": {
        "comm_shop_id": "2f8583111a08449094c519ce488ccc8b",
        "ex_code": "20167",
        "alipay_id": "2016112200077000000020268730",
        "us_id": "42793",
        "comm_code": "301003400000100",
        "upcard_terminal": "02903677",
        "upcard_mer_id": "102290058122453",
        "ex_id": "962",
        "ex_cost_center_code": "1200042793",
        "dcore_store_appid": "s20170823000005535"
    }
},
{
    "_id": "3850146355414564865",
    "name": "西塘古镇",
    "extend_code": {
        "comm_shop_id": "ab46f87884124d9a89636ee74c9a6051",
        "ex_code": "20505",
        "alipay_id": "2021102800077000000029382046",
        "us_id": "44241",
        "comm_code": "301003400000317",
        "upcard_terminal": "57303145",
        "upcard_mer_id": "102573058120136",
        "ex_id": "20303",
        "ex_cost_center_code": "1200044241",
        "dcore_store_appid": "s20170823000005645"
    }
},
{
    "_id": "3850146134831923201",
    "name": "重庆大渡口壹街店",
    "extend_code": {
        "comm_shop_id": "97ef30defe1d4c2a88d00e1a1fb965cf",
        "ex_code": "20139",
        "alipay_id": "2015061200077000000000188806",
        "us_id": "42685",
        "comm_code": "301003400000278",
        "upcard_terminal": "02306261",
        "upcard_mer_id": "102230058120198",
        "ex_id": "641",
        "ex_cost_center_code": "1200042685",
        "dcore_store_appid": "s20170823000005524"
    }
},
{
    "_id": "3850146063193210881",
    "name": "上虞百货店",
    "extend_code": {
        "comm_shop_id": "6b265f95477e4e23804e73ad79381eb0",
        "ex_code": "20013",
        "alipay_id": "2015060900077000000000174623",
        "us_id": "42006",
        "comm_code": "301003400000212",
        "upcard_terminal": "57103824",
        "upcard_mer_id": "102571058120389",
        "ex_id": "451",
        "ex_cost_center_code": "1200042006",
        "dcore_store_appid": "s20170823000005472"
    }
},
{
    "_id": "3850146410599022593",
    "name": "昆山昆城店",
    "extend_code": {
        "comm_shop_id": "9d8eb96312af441bb12786aeb41e8195",
        "ex_code": "20609",
        "alipay_id": "2016050900077000000015478520",
        "us_id": "44548",
        "comm_code": "301003400000291",
        "upcard_terminal": "51216860",
        "upcard_mer_id": "102512089993243",
        "ex_id": "",
        "ex_cost_center_code": "1200044548",
        "dcore_store_appid": "s20170823000005424"
    }
},
{
    "_id": "3850146160463314945",
    "name": "昌里",
    "extend_code": {
        "comm_shop_id": "a471961cbcda4ca6bac37175c42027a7",
        "ex_code": "20188",
        "alipay_id": "2015060900077000000000174619",
        "us_id": "42850",
        "comm_code": "301003400000304",
        "upcard_terminal": "02148816",
        "upcard_mer_id": "102210058120966",
        "ex_id": "2002",
        "ex_cost_center_code": "1200042850",
        "dcore_store_appid": "s20170823000005544"
    }
},
{
    "_id": "3850146342244450305",
    "name": "泰安万达广场店",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20481",
        "comm_shop_id": "870d4e946418483287522bd978b00744",
        "us_id": "44211",
        "alipay_id": "2015113000077000000006599324",
        "takeaway_eleme_id": '',
        "upcard_terminal": "53885523",
        "comm_code": "301003400000254",
        "upcard_mer_id": "102538058120492",
        "ex_id": "20274",
        "ex_cost_center_code": "1200044211",
        "dcore_store_appid": "s20170823000005634"
    }
},
{
    "_id": "3850146379577950209",
    "name": "嘉兴八佰伴",
    "extend_code": {
        "comm_shop_id": "f23d4145209245848d194ce38c5e6bee",
        "ex_code": "20547",
        "alipay_id": "2016031800077000000015026503",
        "us_id": "44406",
        "comm_code": "301003400000434",
        "upcard_terminal": "57303158",
        "upcard_mer_id": "102573058120140",
        "ex_id": "20394",
        "ex_cost_center_code": "1200044406",
        "dcore_store_appid": "s20170823000005673"
    }
},
{
    "_id": "3850146260568768513",
    "name": "南京澳林广场店",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20350",
        "comm_shop_id": "64af83e5988247fdba557fcbaf072b2f",
        "us_id": "43834",
        "alipay_id": "2015061000077000000000188729",
        "takeaway_eleme_id": '',
        "upcard_terminal": "02593482",
        "comm_code": "301003400000202",
        "upcard_mer_id": "102250058122251",
        "ex_id": "20146",
        "ex_cost_center_code": "1200043834",
        "dcore_store_appid": "s20170823000005409"
    }
},
{
    "_id": "3850146344484208641",
    "name": "青浦百联奥特莱斯",
    "extend_code": {
        "comm_shop_id": "07a62e94176c4a6e8153bcd53ea97cfa",
        "ex_code": "20485",
        "alipay_id": "2018041900077000000048365157",
        "us_id": "44256",
        "comm_code": "301003400004452",
        "upcard_terminal": "02194015",
        "upcard_mer_id": "102210058126747",
        "ex_id": "20301",
        "ex_cost_center_code": "1200844256",
        "dcore_store_appid": "s20170823000005164"
    }
},
{
    "_id": "3850146385965875201",
    "name": "蜀都万达店",
    "extend_code": {
        "comm_shop_id": "673e68ff23ff4925a97f867fc9c01d54",
        "ex_code": "20559",
        "alipay_id": "2016050900077000000015450642",
        "us_id": "44481",
        "comm_code": "301003400000205",
        "upcard_terminal": "02893985",
        "upcard_mer_id": "102280058125169",
        "ex_id": "20408",
        "ex_cost_center_code": "1200044481",
        "dcore_store_appid": "s20170823000005327"
    }
},
{
    "_id": "3850146170768719873",
    "name": "常州新北万达",
    "extend_code": {
        "comm_shop_id": "52ab400f3c5d4fd38b5e530d485e8b88",
        "ex_code": "20200",
        "alipay_id": "2015061000077000000000192928",
        "us_id": "42897",
        "comm_code": "301003400000167",
        "upcard_terminal": "51900953",
        "upcard_mer_id": "102519058120149",
        "ex_id": "558",
        "ex_cost_center_code": "1200042897",
        "dcore_store_appid": "s20170823000005233"
    }
},
{
    "_id": "3850146173973168129",
    "name": "成都王府井",
    "extend_code": {
        "comm_shop_id": "19bf63f85106443b8a4dd0eb12f03b43",
        "ex_code": "20206",
        "alipay_id": "2015061200077000000000191237",
        "us_id": "42896",
        "comm_code": "301003400000063",
        "upcard_terminal": "02817506",
        "upcard_mer_id": "102280058122415",
        "ex_id": "817",
        "ex_cost_center_code": "1200042896",
        "dcore_store_appid": "s20170823000005554"
    }
},
{
    "_id": "3850146180268818433",
    "name": "武汉光谷天地",
    "extend_code": {
        "comm_shop_id": "17596db283b64a3ea04591a9b71183db",
        "ex_code": "20218",
        "alipay_id": "2015061200077000000000193012",
        "us_id": "42885",
        "comm_code": "301003400000057",
        "upcard_terminal": "02713540",
        "upcard_mer_id": "102270058122469",
        "ex_id": "927",
        "ex_cost_center_code": "1200042885",
        "dcore_store_appid": "s20170823000005551"
    }
},
{
    "_id": "3850146360347066369",
    "name": "重庆巴南万达店",
    "extend_code": {
        "comm_shop_id": "7ec83f96f11a48788a747170aa7a19ce",
        "ex_code": "20511",
        "alipay_id": "2015113000077000000006713015",
        "us_id": "44227",
        "comm_code": "301003400000242",
        "upcard_terminal": "02385374",
        "upcard_mer_id": "102230058121781",
        "ex_id": "20296",
        "ex_cost_center_code": "1200044227",
        "dcore_store_appid": "s20170823000005635"
    }
},
{
    "_id": "3850146070461939713",
    "name": "上海七宝嘉茂店",
    "extend_code": {
        "comm_shop_id": "31520e8428c7478b80647bd8f35d44b2",
        "ex_code": "20027",
        "alipay_id": "2015060900077000000000174613",
        "us_id": "42210",
        "comm_code": "301003400000107",
        "upcard_terminal": "02194923",
        "upcard_mer_id": "102210058120699",
        "ex_id": "236",
        "ex_cost_center_code": "1200042210",
        "dcore_store_appid": "s20170823000005482"
    }
},
{
    "_id": "3850146273751465985",
    "name": "连云港苏宁广场店",
    "extend_code": {
        "comm_shop_id": "7b03e2bb670a47ff82197181c4ade072",
        "ex_code": "20374",
        "alipay_id": "2016011100077000000014099173",
        "us_id": "43939",
        "comm_code": "301003400004564",
        "upcard_terminal": "51800158",
        "upcard_mer_id": "102518058120020",
        "ex_id": "20182",
        "ex_cost_center_code": "1200843939",
        "dcore_store_appid": "s20170823000005159"
    }
},
{
    "_id": "3850146364088385537",
    "name": "西安阳光天地店",
    "extend_code": {
        "comm_shop_id": "145d13098a4f44ab8499dee8dd22179f",
        "ex_code": "20518",
        "alipay_id": "2015121500077000000013671574",
        "us_id": "44316",
        "comm_code": "301003400000050",
        "upcard_terminal": "02991139",
        "upcard_mer_id": "102290058122936",
        "ex_id": "20321",
        "ex_cost_center_code": "1200044316",
        "dcore_store_appid": "s20170823000005658"
    }
},
{
    "_id": "3850146280567209985",
    "name": "昆明西山万达店",
    "extend_code": {
        "comm_shop_id": "9cd0c2208d7a452e8800c0e374f4f30a",
        "ex_code": "20386",
        "alipay_id": "2015061200077000000000191231",
        "us_id": "43905",
        "comm_code": "301003400000288",
        "upcard_terminal": "87108973",
        "upcard_mer_id": "102871058125507",
        "ex_id": "20171",
        "ex_cost_center_code": "1200043905",
        "dcore_store_appid": "s20170823000005417"
    }
},
{
    "_id": "3850146075960672257",
    "name": "上海黄兴店",
    "extend_code": {
        "comm_shop_id": "5e45c205fcbf487890397415655c1615",
        "ex_code": "20034",
        "alipay_id": "2015060900077000000000178448",
        "us_id": "42265",
        "comm_code": "301003400000196",
        "upcard_terminal": "02148896",
        "upcard_mer_id": "102210058120703",
        "ex_id": "242",
        "ex_cost_center_code": "1200042265",
        "dcore_store_appid": "s20170823000005489"
    }
},
{
    "_id": "3850146193325686785",
    "name": "南昌红谷滩",
    "extend_code": {
        "comm_shop_id": "791c85e444d4421892852f5d424ed228",
        "ex_code": "20235",
        "alipay_id": "2015061000077000000000194476",
        "us_id": "43064",
        "comm_code": "301003400000235",
        "upcard_terminal": "79101363",
        "upcard_mer_id": "102791058120129",
        "ex_id": "984",
        "ex_cost_center_code": "1200043064",
        "dcore_store_appid": "s20170823000005353"
    }
},
{
    "_id": "3850146380056100865",
    "name": "扬州华懋购物中心加盟",
    "extend_code": {
        "comm_shop_id": "5d8f5d5b15094c0d87bac77eb7a5d7ff",
        "ex_code": "20548",
        "alipay_id": "2016011100077000000014227415",
        "us_id": "44372",
        "comm_code": "301003400004539",
        "upcard_terminal": "51400825",
        "upcard_mer_id": "102514058120231",
        "ex_id": "20390",
        "ex_cost_center_code": "1200844372",
        "dcore_store_appid": "s20170823000005334"
    }
},
{
    "_id": "3850146094247837697",
    "name": "上海金桥店",
    "extend_code": {
        "comm_shop_id": "a5062db5de3a47639bb9dd361dfce430",
        "ex_code": "20061",
        "alipay_id": "2015060900077000000000169201",
        "us_id": "42377",
        "comm_code": "301003400000308",
        "upcard_terminal": "02148871",
        "upcard_mer_id": "102210058120728",
        "ex_id": "220",
        "ex_cost_center_code": "1200042377",
        "dcore_store_appid": "s20170823000005561"
    }
},
{
    "_id": "3850146384980213761",
    "name": "常州文化宫延陵（加盟）",
    "extend_code": {
        "comm_shop_id": "04371ae82b284fffb16a1da0da21fb10",
        "ex_code": "20557",
        "alipay_id": "2016050900077000000015478521",
        "us_id": "44480",
        "comm_code": "301003400004446",
        "upcard_terminal": "51900952",
        "upcard_mer_id": "102519058120148",
        "ex_id": "20409",
        "ex_cost_center_code": "1200844480",
        "dcore_store_appid": "s20170823000005337"
    }
},
{
    "_id": "3850146095871033345",
    "name": "济南嘉华店",
    "extend_code": {
        "comm_shop_id": "eac4768deb544f6291e14b871b8916d0",
        "ex_code": "20064",
        "alipay_id": "2015061100077000000000192962",
        "us_id": "42373",
        "comm_code": "301003400000422",
        "upcard_terminal": "53100857",
        "upcard_mer_id": "102531058120053",
        "ex_id": "532",
        "ex_cost_center_code": "1200042373",
        "dcore_store_appid": "s20170823000005498"
    }
},
{
    "_id": "3850146194864996353",
    "name": "芜湖万达",
    "extend_code": {
        "comm_shop_id": "cd1c6196e27144c783d3006b1f116e1e",
        "ex_code": "20238",
        "alipay_id": "2015061100077000000000192970",
        "us_id": "42900",
        "comm_code": "301003400000373",
        "upcard_terminal": "55301396",
        "upcard_mer_id": "102553058120173",
        "ex_id": "464",
        "ex_cost_center_code": "1200042900",
        "dcore_store_appid": "s20170823000005556"
    }
},
{
    "_id": "3850146406060785665",
    "name": "荆门万达",
    "extend_code": {
        "comm_shop_id": "23969bddac5f4e648a965dbe1521b866",
        "ex_code": "20604",
        "alipay_id": "2016111700077000000019991158",
        "us_id": "44507",
        "comm_code": "301003400000080",
        "upcard_terminal": "72400282",
        "upcard_mer_id": "102724058120083",
        "ex_id": "20416",
        "ex_cost_center_code": "1200044507",
        "dcore_store_appid": "s20170823000005681"
    }
},
{
    "_id": "3850146209268236289",
    "name": "上海南翔",
    "extend_code": {
        "comm_shop_id": "448fd677478143cea99237d7d2bcd1b1",
        "ex_code": "20257",
        "alipay_id": "2015060900077000000000178446",
        "us_id": "43150",
        "comm_code": "301003400000139",
        "upcard_terminal": "02148808",
        "upcard_mer_id": "102210058121110",
        "ex_id": "2011",
        "ex_cost_center_code": "1200043150",
        "dcore_store_appid": "s20170823000005357"
    }
},
{
    "_id": "3850146198937665537",
    "name": "莆田万达店",
    "extend_code": {
        "comm_shop_id": "f042971d60cc4e04b6568057e05d51fb",
        "ex_code": "20246",
        "alipay_id": "2015061200077000000000182869",
        "us_id": "43171",
        "comm_code": "301003400000433",
        "upcard_terminal": "59400363",
        "upcard_mer_id": "102594058120022",
        "ex_id": "866",
        "ex_cost_center_code": "1200043171",
        "dcore_store_appid": "s20170929000006503"
    }
},
{
    "_id": "3850146214192349185",
    "name": "杭州湖滨名品",
    "extend_code": {
        "comm_shop_id": "2fee40a8f4354c0b90613f1587a43e9a",
        "ex_code": "20261",
        "alipay_id": "2015060900077000000000178456",
        "us_id": "43419",
        "comm_code": "301003400000101",
        "upcard_terminal": "57103814",
        "upcard_mer_id": "102571058120785",
        "ex_id": "412",
        "ex_cost_center_code": "1200043419",
        "dcore_store_appid": "s20170823000005368"
    }
},
{
    "_id": "3850146290969083905",
    "name": "淄博银泰城店",
    "extend_code": {
        "comm_shop_id": "5d8b87109753400fa78d584c30f303bc",
        "ex_code": "20406",
        "alipay_id": "2015061100077000000000188769",
        "us_id": "43992",
        "comm_code": "301003400000194",
        "upcard_terminal": "53390014",
        "upcard_mer_id": "102533058120066",
        "ex_id": "20216",
        "ex_cost_center_code": "1200043992",
        "dcore_store_appid": "s20170823000005594"
    }
},
{
    "_id": "3850146217124167681",
    "name": "湖州浙北",
    "extend_code": {
        "comm_shop_id": "567e5894194c417f8c53ae25c5d24973",
        "ex_code": "20266",
        "alipay_id": "2015061000077000000000192920",
        "us_id": "43457",
        "comm_code": "301003400000176",
        "upcard_terminal": "57200074",
        "upcard_mer_id": "102572058120008",
        "ex_id": "20004",
        "ex_cost_center_code": "1200043457",
        "dcore_store_appid": "s20170823000005371"
    }
},
{
    "_id": "3850146298405584897",
    "name": "深圳龙华九方购物中心店",
    "extend_code": {
        "comm_shop_id": "6ed4b5a05aba452facb800d0a70f645b",
        "ex_code": "20419",
        "alipay_id": "2015061200077000000000191227",
        "us_id": "44012",
        "comm_code": "301003400000218",
        "upcard_terminal": "75512777",
        "upcard_mer_id": "102755058122316",
        "ex_id": "20226",
        "ex_cost_center_code": "1200044012",
        "dcore_store_appid": "s20170823000005600"
    }
},
{
    "_id": "3850146214695665665",
    "name": "杭州西溪印象城",
    "extend_code": {
        "comm_shop_id": "e0f202315136422e9267cdf0ce74a496",
        "ex_code": "20262",
        "alipay_id": "2015060900077000000000176464",
        "us_id": "43420",
        "comm_code": "301003400000408",
        "upcard_terminal": "57103813",
        "upcard_mer_id": "102571058120786",
        "ex_id": "411",
        "ex_cost_center_code": "1200043420",
        "dcore_store_appid": "s20170823000005369"
    }
},
{
    "_id": "3850146099264225281",
    "name": "湘潭华隆步步高店",
    "extend_code": {
        "comm_shop_id": "270e43e469254701b10248a05f086ed0",
        "ex_code": "20071",
        "alipay_id": "2015061200077000000000182889",
        "us_id": "42409",
        "comm_code": "301003400000085",
        "upcard_terminal": "73203911",
        "upcard_mer_id": "102732058120152",
        "ex_id": "832",
        "ex_cost_center_code": "1200042409",
        "dcore_store_appid": "s20170823000005567"
    }
},
{
    "_id": "3850146413992214529",
    "name": "襄阳天元四季城",
    "extend_code": {
        "comm_shop_id": "22570bbfdfcb41d880fec7cc57d1d6fc",
        "ex_code": "20615",
        "alipay_id": "2016081500077000000018020536",
        "us_id": "44534",
        "comm_code": "301003400000075",
        "upcard_terminal": "71004522",
        "upcard_mer_id": "102710058121112",
        "ex_id": "",
        "ex_cost_center_code": "1200044534",
        "dcore_store_appid": "s20170823000005683"
    }
},
{
    "_id": "3850146219833688065",
    "name": "重庆万州万达",
    "extend_code": {
        "comm_shop_id": "2bc0457bde594535947d401c3b693e63",
        "ex_code": "20271",
        "alipay_id": "2015061200077000000000188809",
        "us_id": "43480",
        "comm_code": "301003400000091",
        "upcard_terminal": "02383942",
        "upcard_mer_id": "102230058120757",
        "ex_id": "20009",
        "ex_cost_center_code": "1200043480",
        "dcore_store_appid": "s20170823000005248"
    }
},
{
    "_id": "3850146303505858561",
    "name": "西安汉神",
    "extend_code": {
        "comm_shop_id": "f35e6bc7a9db46e3aec0d307b2488bdc",
        "ex_code": "20424",
        "alipay_id": "2015093000077000000004416507",
        "us_id": "44015",
        "comm_code": "301003400000440",
        "upcard_terminal": "02990620",
        "upcard_mer_id": "102290058122791",
        "ex_id": "",
        "ex_cost_center_code": "1200044015",
        "dcore_store_appid": "s20170823000005602"
    }
},
{
    "_id": "3850146112304316417",
    "name": "西安南大街店",
    "extend_code": {
        "comm_shop_id": "72d27066902e475a9d91e073c7a0c5e2",
        "ex_code": "20097",
        "alipay_id": "2015061200077000000000191252",
        "us_id": "42551",
        "comm_code": "301003400000228",
        "upcard_terminal": "02903680",
        "upcard_mer_id": "102290058122445",
        "ex_id": "959",
        "ex_cost_center_code": "1200042551",
        "dcore_store_appid": "s20170823000005503"
    }
},
{
    "_id": "3850146128976674817",
    "name": "合肥包河万达店",
    "extend_code": {
        "comm_shop_id": "b60b3b402c304a72ad279f7304ed7849",
        "ex_code": "20128",
        "alipay_id": "2015061100077000000000194518",
        "us_id": "42624",
        "comm_code": "301003400000333",
        "upcard_terminal": "55118205",
        "upcard_mer_id": "102551058121525",
        "ex_id": "155",
        "ex_cost_center_code": "1200042624",
        "dcore_store_appid": "s20170823000005515"
    }
},
{
    "_id": "3850146218231463937",
    "name": "信阳天润广场",
    "extend_code": {
        "comm_shop_id": "cd7ee407ef3149cfb1b21f1e67228a29",
        "ex_code": "20268",
        "alipay_id": "2015061100077000000000191209",
        "us_id": "43460",
        "comm_code": "301003400000375",
        "upcard_terminal": "37602307",
        "upcard_mer_id": "102376058120094",
        "ex_id": "20007",
        "ex_cost_center_code": "1200043460",
        "dcore_store_appid": "s20170823000005373"
    }
},
{
    "_id": "3850146154310270977",
    "name": "东莞星河城",
    "extend_code": {
        "comm_shop_id": "868e5a8067a548b4baea11643fa11a98",
        "ex_code": "20176",
        "alipay_id": "2015061200077000000000192989",
        "us_id": "42803",
        "comm_code": "301003400000253",
        "upcard_terminal": "76903426",
        "upcard_mer_id": "102769058120283",
        "ex_id": "496",
        "ex_cost_center_code": "1200042803",
        "dcore_store_appid": "s20170823000005536"
    }
},
{
    "_id": "3850146231267360769",
    "name": "昆明新西南",
    "extend_code": {
        "comm_shop_id": "c663e66af0024751ba3b930ea773a266",
        "ex_code": "43583",
        "alipay_id": "2015061200077000000000192993",
        "us_id": "43583",
        "comm_code": "301003400000358",
        "upcard_terminal": "87108951",
        "upcard_mer_id": "102871058125486",
        "ex_id": "43583",
        "ex_cost_center_code": "1200043583",
        "dcore_store_appid": "s20170823000005265"
    }
},
{
    "_id": "3850146139856699393",
    "name": "杭州星光大道店",
    "extend_code": {
        "comm_shop_id": "d453b5912b4a4c1cab468decfeeac74d",
        "ex_code": "20148",
        "alipay_id": "2015060900077000000000176462",
        "us_id": "42722",
        "comm_code": "301003400000391",
        "upcard_terminal": "57103821",
        "upcard_mer_id": "102571058120438",
        "ex_id": "405",
        "ex_cost_center_code": "1200042722",
        "dcore_store_appid": "s20170823000005527"
    }
},
{
    "_id": "3850146157518913537",
    "name": "武汉武商摩尔",
    "extend_code": {
        "comm_shop_id": "d2acbef3c7684e45892e8a01288a17d9",
        "ex_code": "20182",
        "alipay_id": "2015061200077000000000191249",
        "us_id": "42820",
        "comm_code": "301003400000387",
        "upcard_terminal": "02713545",
        "upcard_mer_id": "102270058122422",
        "ex_id": "923",
        "ex_cost_center_code": "1200042820",
        "dcore_store_appid": "s20170823000005539"
    }
},
{
    "_id": "3850146232823447553",
    "name": "西安大明宫",
    "extend_code": {
        "comm_shop_id": "2251bac0e03e4995962ae58691eda811",
        "ex_code": "20296",
        "alipay_id": "2015061200077000000000191257",
        "us_id": "43532",
        "comm_code": "301003400000074",
        "upcard_terminal": "02990374",
        "upcard_mer_id": "102290058122670",
        "ex_id": "20036",
        "ex_cost_center_code": "1200043532",
        "dcore_store_appid": "s20170823000005255"
    }
},
{
    "_id": "3850146086857474049",
    "name": "上海周浦万达店",
    "extend_code": {
        "comm_shop_id": "616b35dd0b0a42ed8d37f214f0b12491",
        "ex_code": "20047",
        "alipay_id": "2015060900077000000000169206",
        "us_id": "42309",
        "comm_code": "301003400000199",
        "upcard_terminal": "02190470",
        "upcard_mer_id": "102210058120706",
        "ex_id": "249",
        "ex_cost_center_code": "1200042309",
        "dcore_store_appid": "s20170823000005242"
    }
},
{
    "_id": "3850146245196644353",
    "name": "成都万象城店",
    "extend_code": {
        "comm_shop_id": "4cdaff88678e46e4b552ed4df69296b6",
        "ex_code": "20320",
        "alipay_id": "2015061200077000000000193002",
        "us_id": "43658",
        "comm_code": "301003400000155",
        "upcard_terminal": "02883677",
        "upcard_mer_id": "102280058123760",
        "ex_id": "20077",
        "ex_cost_center_code": "1200043658",
        "dcore_store_appid": "s20170823000005272"
    }
},
{
    "_id": "3850146171293007873",
    "name": "开封新玛特",
    "extend_code": {
        "comm_shop_id": "1112148ffae24cd7a4a010309d4da074",
        "ex_code": "20201",
        "alipay_id": "2015061100077000000000192959",
        "us_id": "42930",
        "comm_code": "301003400000046",
        "upcard_terminal": "37801837",
        "upcard_mer_id": "102378058120171",
        "ex_id": "939",
        "ex_cost_center_code": "1200042930",
        "dcore_store_appid": "s20170823000005344"
    }
},
{
    "_id": "3850146237667868673",
    "name": "南京江宁万达",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20305",
        "comm_shop_id": "5a02bed48ee54f0384def4339aec2b7b",
        "us_id": "43610",
        "alipay_id": "2015061000077000000000194481",
        "takeaway_eleme_id": '',
        "upcard_terminal": "02593530",
        "comm_code": "301003400000183",
        "upcard_mer_id": "102250058122030",
        "ex_id": "20047",
        "ex_cost_center_code": "1200043610",
        "dcore_store_appid": "s20170823000005270"
    }
},
{
    "_id": "3850146357935341569",
    "name": "成都金楠天街店",
    "extend_code": {
        "comm_shop_id": "ce1ad091047b409b95ac8cb750e46eea",
        "ex_code": "20509",
        "alipay_id": "2015113000077000000006583511",
        "us_id": "44294",
        "comm_code": "301003400000378",
        "upcard_terminal": "02890172",
        "upcard_mer_id": "102280058124947",
        "ex_id": "20322",
        "ex_cost_center_code": "1200044294",
        "dcore_store_appid": "s20170823000005654"
    }
},
{
    "_id": "3850146090930143233",
    "name": "昆明顺城店",
    "extend_code": {
        "comm_shop_id": "90b1e28d66c348d3b5f56123d29848a8",
        "ex_code": "20055",
        "alipay_id": "2015061200077000000000188796",
        "us_id": "42316",
        "comm_code": "301003400000270",
        "upcard_terminal": "87106454",
        "upcard_mer_id": "102871058120239",
        "ex_id": "773",
        "ex_cost_center_code": "1200042316",
        "dcore_store_appid": "s20170823000005494"
    }
},
{
    "_id": "3850146239706300417",
    "name": "舟山银泰",
    "extend_code": {
        "comm_shop_id": "f312543ceb78416b98ce0313ea57837c",
        "ex_code": "20309",
        "alipay_id": "2015061000077000000000192914",
        "us_id": "43588",
        "comm_code": "301003400000439",
        "upcard_terminal": "58000374",
        "upcard_mer_id": "102580058120019",
        "ex_id": "20053",
        "ex_cost_center_code": "1200043588",
        "dcore_store_appid": "s20170823000005269"
    }
},
{
    "_id": "3850146243653140481",
    "name": "启东店",
    "extend_code": {
        "comm_shop_id": "93572dc93fa54f5ea34d8ec59feb41f0",
        "ex_code": "20317",
        "alipay_id": "2015093000077000000004436861",
        "us_id": "43670",
        "comm_code": "301003400004593",
        "upcard_terminal": "51300824",
        "upcard_mer_id": "102513058120054",
        "ex_id": "20084",
        "ex_cost_center_code": "1200843670",
        "dcore_store_appid": "s20170823000005152"
    }
},
{
    "_id": "3850146247662895105",
    "name": "郑州万象城店",
    "extend_code": {
        "comm_shop_id": "a8a98a5deac145f487ae3c7e74c0726f",
        "ex_code": "20325",
        "alipay_id": "2015061100077000000000194501",
        "us_id": "43761",
        "comm_code": "301003400000314",
        "upcard_terminal": "37110793",
        "upcard_mer_id": "102371058122276",
        "ex_id": "20105",
        "ex_cost_center_code": "1200043761",
        "dcore_store_appid": "s20170823000005277"
    }
},
{
    "_id": "3850146238670307329",
    "name": "广州中华广场",
    "extend_code": {
        "comm_shop_id": "e6612273f9f14ba3a8f95337590453b4",
        "ex_code": "20307",
        "alipay_id": "2015061200077000000000191223",
        "us_id": "43638",
        "comm_code": "301003400000416",
        "upcard_terminal": "02081956",
        "upcard_mer_id": "102200058120371",
        "ex_id": "99777",
        "ex_cost_center_code": "1200043638",
        "dcore_store_appid": "s20170823000005380"
    }
},
{
    "_id": "3850146306223767553",
    "name": "厦门翔安汇景",
    "extend_code": {
        "comm_shop_id": "c8791ca9b1a74c2eba6f15d9365f87a1",
        "ex_code": "20427",
        "alipay_id": "2015061200077000000000192992",
        "us_id": "44014",
        "comm_code": "301003400000363",
        "upcard_terminal": "59204820",
        "upcard_mer_id": "102592058120542",
        "ex_id": "20214",
        "ex_cost_center_code": "1200044014",
        "dcore_store_appid": "s20170823000005601"
    }
},
{
    "_id": "3850146174447124481",
    "name": "西安曲江银泰",
    "extend_code": {
        "comm_shop_id": "fa6b27df58f14ceb85b2df548940aeef",
        "ex_code": "20207",
        "alipay_id": "2015061200077000000000194561",
        "us_id": "42901",
        "comm_code": "301003400000450",
        "upcard_terminal": "02903674",
        "upcard_mer_id": "102290058122463",
        "ex_id": "964",
        "ex_cost_center_code": "1200042901",
        "dcore_store_appid": "s20170823000005557"
    }
},
{
    "_id": "3850146287915630593",
    "name": "南京文鼎广场",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20400",
        "comm_shop_id": "81eec69d92a54c27a2124480bf56cc03",
        "us_id": "43880",
        "alipay_id": "2015061100077000000000188750",
        "takeaway_eleme_id": '',
        "upcard_terminal": "02594648",
        "comm_code": "301003400000246",
        "upcard_mer_id": "102250058122399",
        "ex_id": "20149",
        "ex_cost_center_code": "1200043880",
        "dcore_store_appid": "s20170823000005285"
    }
},
{
    "_id": "3850146127424782337",
    "name": "福州金融街万达店",
    "extend_code": {
        "comm_shop_id": "eb31692ddfb44e039bba30177c0b2601",
        "ex_code": "20125",
        "alipay_id": "2015061200077000000000188791",
        "us_id": "42638",
        "comm_code": "301003400000424",
        "upcard_terminal": "59106682",
        "upcard_mer_id": "102591058120180",
        "ex_id": "855",
        "ex_cost_center_code": "1200042638",
        "dcore_store_appid": "s20170823000005518"
    }
},
{
    "_id": "3850146147112845313",
    "name": "苏州绿宝广场",
    "extend_code": {
        "comm_shop_id": "226d608a44334afaa55fd8cfdd31192c",
        "ex_code": "20162",
        "alipay_id": "2015061000077000000000182815",
        "us_id": "42759",
        "comm_code": "301003400000076",
        "upcard_terminal": "51211245",
        "upcard_mer_id": "102512058120519",
        "ex_id": "511",
        "ex_cost_center_code": "1200042759",
        "dcore_store_appid": "s20170823000005225"
    }
},
{
    "_id": "3850146294953672705",
    "name": "九江九方店",
    "extend_code": {
        "comm_shop_id": "998a014478de481c807243be4bceaa4d",
        "ex_code": "20414",
        "alipay_id": "2015061000077000000000191165",
        "us_id": "44025",
        "comm_code": "301003400000282",
        "upcard_terminal": "79200395",
        "upcard_mer_id": "102792058120133",
        "ex_id": "20225",
        "ex_cost_center_code": "1200044025",
        "dcore_store_appid": "s20170823000005605"
    }
},
{
    "_id": "3850146318802485249",
    "name": "眉山万景国际",
    "extend_code": {
        "comm_shop_id": "279a374a3afe4c37bea14110245b2683",
        "ex_code": "20448",
        "alipay_id": "2015061200077000000000191238",
        "us_id": "44094",
        "comm_code": "301003400000087",
        "upcard_terminal": "83390005",
        "upcard_mer_id": "102833058120016",
        "ex_id": "20247",
        "ex_cost_center_code": "1200044094",
        "dcore_store_appid": "s20170823000005303"
    }
},
{
    "_id": "3850146155870552065",
    "name": "银川万达店",
    "extend_code": {
        "comm_shop_id": "fb16041eb666409e87aeab2e6395b2ff",
        "ex_code": "20179",
        "alipay_id": "2015061200077000000000182893",
        "us_id": "42829",
        "comm_code": "301003400000452",
        "upcard_terminal": "95100475",
        "upcard_mer_id": "102951058120212",
        "ex_id": "476",
        "ex_cost_center_code": "1200042829",
        "dcore_store_appid": "s20170823000005542"
    }
},
{
    "_id": "3850146323466551297",
    "name": "洛阳王府井达玛格利",
    "extend_code": {
        "comm_shop_id": "b31fea3c11744f5691e7a10532c6e421",
        "ex_code": "20456",
        "alipay_id": "2015093000077000000004529780",
        "us_id": "44087",
        "comm_code": "301003400000326",
        "upcard_terminal": "37922540",
        "upcard_mer_id": "102379058120420",
        "ex_id": "",
        "ex_cost_center_code": "1200044087",
        "dcore_store_appid": "s20170823000005615"
    }
},
{
    "_id": "3850146325832138753",
    "name": "兰州北京华联店",
    "extend_code": {
        "comm_shop_id": "e997cec00ccd4f2eaf824e9ecbcb7343",
        "ex_code": "20460",
        "alipay_id": "2015093000077000000004456914",
        "us_id": "44177",
        "comm_code": "301003400000418",
        "upcard_terminal": "93101009",
        "upcard_mer_id": "102931058120185",
        "ex_id": "20263",
        "ex_cost_center_code": "1200044177",
        "dcore_store_appid": "s20170823000005307"
    }
},
{
    "_id": "3850146168688345089",
    "name": "杭州南宋御街",
    "extend_code": {
        "comm_shop_id": "fd05a62bf76748e9b47c6bcae45fac5a",
        "ex_code": "20196",
        "alipay_id": "2015060900077000000000166168",
        "us_id": "42881",
        "comm_code": "301003400000456",
        "upcard_terminal": "57103817",
        "upcard_mer_id": "102571058120465",
        "ex_id": "408",
        "ex_cost_center_code": "1200042881",
        "dcore_store_appid": "s20170823000005231"
    }
},
{
    "_id": "3850146336464699393",
    "name": "安阳万达店",
    "extend_code": {
        "comm_shop_id": "36f0331f19194d4db3a8bfd2a591b671",
        "ex_code": "20471",
        "alipay_id": "2015093000077000000004454239",
        "us_id": "44190",
        "comm_code": "301003400000117",
        "upcard_terminal": "37202987",
        "upcard_mer_id": "102372058121256",
        "ex_id": "20269",
        "ex_cost_center_code": "1200044190",
        "dcore_store_appid": "s20170823000005629"
    }
},
{
    "_id": "3850146197335441409",
    "name": "南通文峰",
    "extend_code": {
        "comm_shop_id": "61da7f1f75ba49229fe304f5f03f82a0",
        "ex_code": "20243",
        "alipay_id": "2018041900077000000048255677",
        "us_id": "43128",
        "comm_code": "301003400000200",
        "upcard_terminal": "51300175",
        "upcard_mer_id": "102513058120041",
        "ex_id": "568",
        "ex_cost_center_code": "1200043128",
        "dcore_store_appid": "s20170823000005355"
    }
},
{
    "_id": "3850146341170708481",
    "name": "南京万谷慧",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20479",
        "comm_shop_id": "ff0a48ea50234da1821aa80c8d6d6d3f",
        "us_id": "44058",
        "alipay_id": "2015093000077000000004387800",
        "takeaway_eleme_id": '',
        "upcard_terminal": "02595909",
        "comm_code": "301003400000462",
        "upcard_mer_id": "102250058122640",
        "ex_id": "20227",
        "ex_cost_center_code": "1200044058",
        "dcore_store_appid": "s20170823000005301"
    }
},
{
    "_id": "3854097531449376769",
    "name": "苏州吴江万宝财富广场店",
    "extend_code": {
        "comm_shop_id": "b96aeb583e474406bbc759f0c3a4b826",
        "alipay_id": "2016111500077000000019999539",
        "us_id": "44539",
        "comm_code": "301003400004631",
        "upcard_terminal": "51215575",
        "upcard_mer_id": "102512058123108",
        "ex_cost_center_code": "1200844539",
        "dcore_store_appid": "s20170823000005421"
    }
},
{
    "_id": "3850146370501476353",
    "name": "临汾新安达圣店",
    "extend_code": {
        "comm_shop_id": "d9a0ff7c39554e67905f348e42d1ecb9",
        "ex_code": "20530",
        "alipay_id": "2016011100077000000014084490",
        "us_id": "44349",
        "comm_code": "301003400000399",
        "upcard_terminal": "35700738",
        "upcard_mer_id": "102357058120037",
        "ex_id": "20345",
        "ex_cost_center_code": "1200044349",
        "dcore_store_appid": "s20170823000005664"
    }
},
{
    "_id": "3850146215287062529",
    "name": "广州天河城店",
    "extend_code": {
        "comm_shop_id": "f26367822d654e91ac2ab3f3eb4135f2",
        "ex_code": "20263",
        "alipay_id": "2015061200077000000000194532",
        "us_id": "43461",
        "comm_code": "301003400000437",
        "upcard_terminal": "02003961",
        "upcard_mer_id": "102200058120301",
        "ex_id": "20003",
        "ex_cost_center_code": "1200043461",
        "dcore_store_appid": "s20170823000005246"
    }
},
{
    "_id": "3850146219288428545",
    "name": "无锡惠山万达店",
    "extend_code": {
        "comm_shop_id": "56262e121d96407495d814b39957a925",
        "ex_code": "20270",
        "alipay_id": "2015061000077000000000188725",
        "us_id": "43456",
        "comm_code": "301003400000174",
        "upcard_terminal": "51102375",
        "upcard_mer_id": "102510058120328",
        "ex_id": "660",
        "ex_cost_center_code": "1200043456",
        "dcore_store_appid": "s20170823000005370"
    }
},
{
    "_id": "3850146380655886337",
    "name": "广州丽影店",
    "extend_code": {
        "comm_shop_id": "993dcadbad1841b3a166dfa20b3b65da",
        "ex_code": "20549",
        "alipay_id": "2016042500077000000015407363",
        "us_id": "44245",
        "comm_code": "301003400000280",
        "upcard_terminal": "02081919",
        "upcard_mer_id": "102200058120750",
        "ex_id": "20310",
        "ex_cost_center_code": "1200044245",
        "dcore_store_appid": "s20170823000005646"
    }
},
{
    "_id": "3850146373567512577",
    "name": "许昌亨源通",
    "extend_code": {
        "comm_shop_id": "000c5c5e42374dcba16f37b8310336a0",
        "ex_code": "20536",
        "alipay_id": "2016011100077000000014214785",
        "us_id": "44374",
        "comm_code": "301003400000021",
        "upcard_terminal": "37401535",
        "upcard_mer_id": "102374058120057",
        "ex_id": "20371",
        "ex_cost_center_code": "1200044374",
        "dcore_store_appid": "s20170823000005672"
    }
},
{
    "_id": "3850146279938064385",
    "name": "兰州城关万达店",
    "extend_code": {
        "comm_shop_id": "0fff0f7da55f4f9b907ae73fd3689a09",
        "ex_code": "20385",
        "alipay_id": "2015061200077000000000188821",
        "us_id": "43961",
        "comm_code": "301003400000045",
        "upcard_terminal": "93100949",
        "upcard_mer_id": "102931058120161",
        "ex_id": "20167",
        "ex_cost_center_code": "1200043961",
        "dcore_store_appid": "s20170929000006492"
    }
},
{
    "_id": "3850146204251848705",
    "name": "深圳海雅缤纷城",
    "extend_code": {
        "comm_shop_id": "e9fc8fd3592e4b84baa2e45b986890e8",
        "ex_code": "20254",
        "alipay_id": "2015061200077000000000188787",
        "us_id": "43147",
        "comm_code": "301003400000420",
        "upcard_terminal": "75507872",
        "upcard_mer_id": "102755058121074",
        "ex_id": "490",
        "ex_cost_center_code": "1200043147",
        "dcore_store_appid": "s20170823000005356"
    }
},
{
    "_id": "3850146282081353729",
    "name": "贵阳中大国际店",
    "extend_code": {
        "comm_shop_id": "18e8dcfcc3f84acfad4d50280be5daa3",
        "ex_code": "20389",
        "alipay_id": "2015061200077000000000182871",
        "us_id": "43829",
        "comm_code": "301003400000062",
        "upcard_terminal": "85101270",
        "upcard_mer_id": "102851058120379",
        "ex_id": "20195",
        "ex_cost_center_code": "1200043829",
        "dcore_store_appid": "s20170823000005407"
    }
},
{
    "_id": "3850146229287649281",
    "name": "福州王府井",
    "extend_code": {
        "comm_shop_id": "4925b7280cd5452a969af748d0ff041a",
        "ex_code": "20289",
        "alipay_id": "2015061200077000000000191228",
        "us_id": "43127",
        "comm_code": "301003400000147",
        "upcard_terminal": "59106038",
        "upcard_mer_id": "102591058120294",
        "ex_id": "859",
        "ex_cost_center_code": "1200043127",
        "dcore_store_appid": "s20170929000006502"
    }
},
{
    "_id": "3850146233641336833",
    "name": "南昌红谷滩世茂",
    "extend_code": {
        "comm_shop_id": "304616b2753c47519b55acd119c6c9b3",
        "ex_code": "20297",
        "alipay_id": "2015061000077000000000194473",
        "us_id": "43596",
        "comm_code": "301003400000102",
        "upcard_terminal": "79190459",
        "upcard_mer_id": "102791058120462",
        "ex_id": "20035",
        "ex_cost_center_code": "1200043596",
        "dcore_store_appid": "s20170823000005374"
    }
},
{
    "_id": "3850146216599879681",
    "name": "宜兴万达",
    "extend_code": {
        "comm_shop_id": "5000003f62c046288aacf1cdcb2a70e9",
        "ex_code": "20265",
        "alipay_id": "2015061000077000000000194479",
        "us_id": "43422",
        "comm_code": "301003400000162",
        "upcard_terminal": "51000993",
        "upcard_mer_id": "102510058120318",
        "ex_id": "352",
        "ex_cost_center_code": "1200043422",
        "dcore_store_appid": "s20170823000005244"
    }
},
{
    "_id": "3850146378927833089",
    "name": "许昌时代广场店",
    "extend_code": {
        "comm_shop_id": "c6eb88e979cd4b419c97ca9efd8734d0",
        "ex_code": "20546",
        "alipay_id": "2016011100077000000014137251",
        "us_id": "44344",
        "comm_code": "301003400000360",
        "upcard_terminal": "37401536",
        "upcard_mer_id": "102374058120058",
        "ex_id": "20346",
        "ex_cost_center_code": "1200044344",
        "dcore_store_appid": "s20170823000005661"
    }
},
{
    "_id": "3850146296421679105",
    "name": "无锡宜家",
    "extend_code": {
        "us_id": "43960",
        "ex_id": "20203",
        "ex_cost_center_code": "1200043960",
        "ex_code": "20416"
    }
},
{
    "_id": "3850146382409105409",
    "name": "鹤壁爱之城",
    "extend_code": {
        "comm_shop_id": "69d44594d4934897bb55e9ecbdfc3537",
        "ex_code": "20552",
        "alipay_id": "2016050900077000000015488024",
        "us_id": "44332",
        "comm_code": "301003400000208",
        "upcard_terminal": "39201377",
        "upcard_mer_id": "102392058120291",
        "ex_id": "20406",
        "ex_cost_center_code": "1200044332",
        "dcore_store_appid": "s20170823000005659"
    }
},
{
    "_id": "3850146400234897409",
    "name": "西安奥特莱斯",
    "extend_code": {
        "comm_shop_id": "f86cb2e9b6fd4ec3af62ce7da6953279",
        "ex_code": "20597",
        "alipay_id": "2016111700077000000020165781",
        "us_id": "44283",
        "comm_code": "301003400000447",
        "upcard_terminal": "02991151",
        "upcard_mer_id": "102290058122944",
        "ex_id": "20344",
        "ex_cost_center_code": "1200044283",
        "dcore_store_appid": "s20170823000005652"
    }
},
{
    "_id": "3850146225802182657",
    "name": "济源信尧店",
    "extend_code": {
        "comm_shop_id": "3315c02333c3479995b65fe6841433bf",
        "ex_code": "20282",
        "alipay_id": "2015061100077000000000192960",
        "us_id": "43551",
        "comm_code": "301003400000109",
        "upcard_terminal": "39102437",
        "upcard_mer_id": "102391058120149",
        "ex_id": "20026",
        "ex_cost_center_code": "1200043551",
        "dcore_store_appid": "s20170823000005261"
    }
},
{
    "_id": "3850146259968983041",
    "name": "烟台大悦城",
    "extend_code": {
        "comm_shop_id": "184c5ca32a494bc184219ebfde2a35a6",
        "ex_code": "20349",
        "alipay_id": "2015061100077000000000188765",
        "us_id": "43825",
        "comm_code": "301003400000061",
        "upcard_terminal": "53530367",
        "upcard_mer_id": "102535058120116",
        "ex_id": "20145",
        "ex_cost_center_code": "1200043825",
        "dcore_store_appid": "s20170823000005404"
    }
},
{
    "_id": "3850146253920796673",
    "name": "潍坊万达店",
    "extend_code": {
        "comm_shop_id": "b101829a3a224d159abe5b59b2c0befb",
        "ex_code": "20337",
        "alipay_id": "2015061100077000000000191213",
        "us_id": "43797",
        "comm_code": "301003400000321",
        "upcard_terminal": "53601856",
        "upcard_mer_id": "102536058120095",
        "ex_id": "20109",
        "ex_cost_center_code": "1200043797",
        "dcore_store_appid": "s20170823000005402"
    }
},
{
    "_id": "3850146406572490753",
    "name": "济南高新万达",
    "extend_code": {
        "comm_shop_id": "08d405a788c94fd69524d97d0bcea960",
        "ex_code": "20592",
        "alipay_id": "2016080600077000000017911378",
        "us_id": "44503",
        "comm_code": "301003400000036",
        "upcard_terminal": "53101267",
        "upcard_mer_id": "102531058120531",
        "ex_id": "20419",
        "ex_cost_center_code": "1200044503",
        "dcore_store_appid": "s20170823000005679"
    }
},
{
    "_id": "3850146239190401025",
    "name": "武汉奥山世纪城",
    "extend_code": {
        "comm_shop_id": "4c80f0cad1ae408dbac95071c2fa384f",
        "ex_code": "20308",
        "alipay_id": "2015061200077000000000191256",
        "us_id": "43578",
        "comm_code": "301003400000154",
        "upcard_terminal": "02722801",
        "upcard_mer_id": "102270058124237",
        "ex_id": "20034",
        "ex_cost_center_code": "1200043578",
        "dcore_store_appid": "s20170823000005264"
    }
},
{
    "_id": "3850146263169236993",
    "name": "扬州时代广场店",
    "extend_code": {
        "comm_shop_id": "f2a9e15ebcca4366a9def358373af2c1",
        "ex_code": "20355",
        "alipay_id": "2015061100077000000000192956",
        "us_id": "43735",
        "comm_code": "301003400000438",
        "upcard_terminal": "51400315",
        "upcard_mer_id": "102514058120151",
        "ex_id": "20153",
        "ex_cost_center_code": "1200043735",
        "dcore_store_appid": "s20170823000005391"
    }
},
{
    "_id": "3850146244726882305",
    "name": "郑州西元",
    "extend_code": {
        "comm_shop_id": "3d5db88e96bc477cab8193246c5e7358",
        "ex_code": "20319",
        "alipay_id": "2015061100077000000000188760",
        "us_id": "43677",
        "comm_code": "301003400000126",
        "upcard_terminal": "37110792",
        "upcard_mer_id": "102371058122258",
        "ex_id": "20081",
        "ex_cost_center_code": "1200043677",
        "dcore_store_appid": "s20170823000005274"
    }
},
{
    "_id": "3850146254931623937",
    "name": "松江万达",
    "extend_code": {
        "comm_shop_id": "c4d7dd2ac8f7411b90cfbd91001cd3c6",
        "ex_code": "20339",
        "alipay_id": "2015060900077000000000178455",
        "us_id": "43781",
        "comm_code": "301003400000352",
        "upcard_terminal": "02190187",
        "upcard_mer_id": "102210058125837",
        "ex_id": "20103",
        "ex_cost_center_code": "1200043781",
        "dcore_store_appid": "s20170823000005398"
    }
},
{
    "_id": "3850146266717618177",
    "name": "常州武进万达",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20361",
        "comm_shop_id": "44b9427f30eb40d59999e0e9358bf3c3",
        "us_id": "43882",
        "alipay_id": "2015061100077000000000182829",
        "takeaway_eleme_id": '',
        "upcard_terminal": "51900658",
        "comm_code": "301003400000140",
        "upcard_mer_id": "102519058120079",
        "ex_id": "20151",
        "ex_cost_center_code": "1200043882",
        "dcore_store_appid": "s20170823000005415"
    }
},
{
    "_id": "3850146261747367937",
    "name": "济宁太白路万达店",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20352",
        "comm_shop_id": "802f1b3305b04eefbec71e95785bfeaa",
        "us_id": "43827",
        "alipay_id": "2015061100077000000000182843",
        "takeaway_eleme_id": '',
        "upcard_terminal": "53700799",
        "comm_code": "301003400000244",
        "upcard_mer_id": "102537058120039",
        "ex_id": "20129",
        "ex_cost_center_code": "1200043827",
        "dcore_store_appid": "s20170823000005405"
    }
},
{
    "_id": "3850146264356225025",
    "name": "金华万达店",
    "extend_code": {
        "comm_shop_id": "7f71b1e07f3f42cf843ef31f8274908a",
        "ex_code": "20357",
        "alipay_id": "2015061000077000000000182802",
        "us_id": "43835",
        "comm_code": "301003400000243",
        "upcard_terminal": "57990014",
        "upcard_mer_id": "102579058120128",
        "ex_id": "20127",
        "ex_cost_center_code": "1200043835",
        "dcore_store_appid": "s20170823000005283"
    }
},
{
    "_id": "3850146310455820289",
    "name": "赣州九方店",
    "extend_code": {
        "comm_shop_id": "422bb26511644408a57c8dbe9856ccc4",
        "ex_code": "20432",
        "alipay_id": "2015061000077000000000188722",
        "us_id": "44026",
        "comm_code": "301003400000137",
        "upcard_terminal": "79700565",
        "upcard_mer_id": "102797058120354",
        "ex_id": "20240",
        "ex_cost_center_code": "1200044026",
        "dcore_store_appid": "s20170823000005606"
    }
},
{
    "_id": "3850146286959329281",
    "name": "荆州人信汇店",
    "extend_code": {
        "comm_shop_id": "7c11f737e90b4b20b1564427113ec3bc",
        "ex_code": "20398",
        "alipay_id": "2015061200077000000000191247",
        "us_id": "43994",
        "comm_code": "301003400000236",
        "upcard_terminal": "71600161",
        "upcard_mer_id": "102716058120042",
        "ex_id": "20201",
        "ex_cost_center_code": "1200043994",
        "dcore_store_appid": "s20170823000005595"
    }
},
{
    "_id": "3850146296937578497",
    "name": "郑州瀚海北金店",
    "extend_code": {
        "comm_shop_id": "e4c661d5e3174ce3883615473389ad9d",
        "ex_code": "20417",
        "alipay_id": "2015061100077000000000191204",
        "us_id": "43957",
        "comm_code": "301003400000414",
        "upcard_terminal": "37110720",
        "upcard_mer_id": "102371058122384",
        "ex_id": "20218",
        "ex_cost_center_code": "1200043957",
        "dcore_store_appid": "s20170823000005292"
    }
},
{
    "_id": "3850146314616569857",
    "name": "张江长泰广场",
    "extend_code": {
        "comm_shop_id": "9768851747024e938fa5c78aee62c817",
        "ex_code": "20440",
        "alipay_id": "2015060900077000000000166156",
        "us_id": "43952",
        "comm_code": "301003400000277",
        "upcard_terminal": "02193305",
        "upcard_mer_id": "102210058126458",
        "ex_id": "20243",
        "ex_cost_center_code": "1200043952",
        "dcore_store_appid": "s20170823000005291"
    }
},
{
    "_id": "3850146299944894465",
    "name": "盐城宝龙店",
    "extend_code": {
        "us_id": "43996",
        "ex_id": "20219",
        "ex_cost_center_code": "1200043996",
        "ex_code": "20420"
    }
},
{
    "_id": "3850146315753226241",
    "name": "南宁会展航洋城",
    "extend_code": {
        "comm_shop_id": "467a67805b8c4dfba73d0eb13f665208",
        "ex_code": "20442",
        "alipay_id": "2016011100077000000014150644",
        "us_id": "44023",
        "comm_code": "301003400000142",
        "upcard_terminal": "77103788",
        "upcard_mer_id": "102771058121963",
        "ex_id": "20246",
        "ex_cost_center_code": "1200044023",
        "dcore_store_appid": "s20170823000005604"
    }
},
{
    "_id": "3850146292453867521",
    "name": "昆明红星爱琴海",
    "extend_code": {
        "comm_shop_id": "ebedf8527a0d4a04a5074f3c07c7e87b",
        "ex_code": "20409",
        "alipay_id": "2015061200077000000000194538",
        "us_id": "43990",
        "comm_code": "301003400000427",
        "upcard_terminal": "87111957",
        "upcard_mer_id": "102871058126647",
        "ex_id": "",
        "ex_cost_center_code": "1200043990",
        "dcore_store_appid": "s20170823000005294"
    }
},
{
    "_id": "3850146300951527425",
    "name": "杭州中大银泰店",
    "extend_code": {
        "us_id": "43063",
        "ex_id": "20200",
        "ex_cost_center_code": "1200043063",
        "ex_code": "20421"
    }
},
{
    "_id": "3850146308983619585",
    "name": "九江联盛店",
    "extend_code": {
        "us_id": "43998",
        "ex_id": "20230",
        "ex_cost_center_code": "1200043998",
        "ex_code": "20429"
    }
},
{
    "_id": "3850146302448893953",
    "name": "丽水万地店",
    "extend_code": {
        "comm_shop_id": "01c26919641a41b0972adac069f16f1a",
        "ex_code": "20423",
        "alipay_id": "2015061000077000000000192916",
        "us_id": "44016",
        "comm_code": "301003400000025",
        "upcard_terminal": "57800239",
        "upcard_mer_id": "102578058120009",
        "ex_id": "20220",
        "ex_cost_center_code": "1200044016",
        "dcore_store_appid": "s20170823000005298"
    }
},
{
    "_id": "3850146310954942465",
    "name": "南通圆融店",
    "extend_code": {
        "comm_shop_id": "65617c4cfca348d49bdec904b12ae9e7",
        "ex_code": "20433",
        "alipay_id": "2015093000077000000004515139",
        "us_id": "43956",
        "comm_code": "301003400004544",
        "upcard_terminal": "51300187",
        "upcard_mer_id": "102513058120065",
        "ex_id": "20193",
        "ex_cost_center_code": "1200843956",
        "dcore_store_appid": "s20170823000005160"
    }
},
{
    "_id": "3863190461975887873",
    "name": "亳州万达广场加盟店",
    "extend_code": {
        "comm_shop_id": "e313c53127b149c4be5582701b70db7a",
        "alipay_id": "2016111700077000000020149009",
        "us_id": "44568",
        "comm_code": "301003400004668",
        "upcard_terminal": "55801456",
        "upcard_mer_id": "102558058120429",
        "ex_id": "20437",
        "ex_cost_center_code": "1200844568",
        "dcore_store_appid": "s20170823000005425"
    }
},
{
    "_id": "3850146313572188161",
    "name": "济南领秀城",
    "extend_code": {
        "comm_shop_id": "014338012d9a4f8482115449b207bc9d",
        "ex_code": "20438",
        "alipay_id": "2015061100077000000000191212",
        "us_id": "44057",
        "comm_code": "301003400000023",
        "upcard_terminal": "53101187",
        "upcard_mer_id": "102531058120516",
        "ex_id": "",
        "ex_cost_center_code": "1200044057",
        "dcore_store_appid": "s20170823000005610"
    }
},
{
    "_id": "3850146342785515521",
    "name": "马鞍山金鹰天地",
    "extend_code": {
        "comm_shop_id": "fd8f622157b54bc3adef6a13bc2c4e6a",
        "ex_code": "20482",
        "alipay_id": "2015093000077000000004503888",
        "us_id": "44233",
        "comm_code": "301003400000457",
        "upcard_terminal": "55500393",
        "upcard_mer_id": "102555058120069",
        "ex_id": "20293",
        "ex_cost_center_code": "1200044233",
        "dcore_store_appid": "s20170823000005312"
    }
},
{
    "_id": "3850146311458258945",
    "name": "桂林万福广场",
    "extend_code": {
        "comm_shop_id": "5b60ec9e2b1f47879fa44151c80f8c3f",
        "ex_code": "20434",
        "alipay_id": "2015061200077000000000188805",
        "us_id": "44055",
        "comm_code": "301003400000187",
        "upcard_terminal": "77300711",
        "upcard_mer_id": "102773058120438",
        "ex_id": "",
        "ex_cost_center_code": "1200044055",
        "dcore_store_appid": "s20170823000005609"
    }
},
{
    "_id": "3850146317762297857",
    "name": "岳阳天虹店",
    "extend_code": {
        "comm_shop_id": "d2c1291e0a214aa88cb8c7b523fadb9e",
        "ex_code": "20446",
        "alipay_id": "2015093000077000000004488628",
        "us_id": "43824",
        "comm_code": "301003400000388",
        "upcard_terminal": "73000531",
        "upcard_mer_id": "102730058120035",
        "ex_id": "20250",
        "ex_cost_center_code": "1200043824",
        "dcore_store_appid": "s20170823000005403"
    }
},
{
    "_id": "3850146343993475073",
    "name": "杭州文一物美",
    "extend_code": {
        "comm_shop_id": "1c63d979184c493fa8a9067030cbbf1d",
        "ex_code": "20484",
        "alipay_id": "2015093000077000000004480416",
        "us_id": "44179",
        "comm_code": "301003400000066",
        "upcard_terminal": "57108821",
        "upcard_mer_id": "102571058122149",
        "ex_id": "20289",
        "ex_cost_center_code": "1200044179",
        "dcore_store_appid": "s20170823000005628"
    }
},
{
    "_id": "3850146320459235329",
    "name": "青岛万象城",
    "extend_code": {
        "comm_shop_id": "5631366691c146869010fc5bdfd18e24",
        "ex_code": "20451",
        "alipay_id": "2015093000077000000004464583",
        "us_id": "44107",
        "comm_code": "301003400000175",
        "upcard_terminal": "53205105",
        "upcard_mer_id": "102532058121351",
        "ex_id": "20254",
        "ex_cost_center_code": "1200044107",
        "dcore_store_appid": "s20170823000005617"
    }
},
{
    "_id": "3850146354420514817",
    "name": "苏州吴江吾悦加盟",
    "extend_code": {
        "comm_shop_id": "30f18deccc1c44dda3919ba6c5b19306",
        "ex_code": "20503",
        "alipay_id": "2015093000077000000004466436",
        "us_id": "44274",
        "comm_code": "301003400004496",
        "upcard_terminal": "51215129",
        "upcard_mer_id": "102512058122961",
        "ex_id": "20312",
        "ex_cost_center_code": "1200844274",
        "dcore_store_appid": "s20170823000005167"
    }
},
{
    "_id": "3850146365283762177",
    "name": "渭南信达",
    "extend_code": {
        "comm_shop_id": "2c2b7958d8d241c5b4767ff93656d0cd",
        "ex_code": "20520",
        "alipay_id": "2016012100077000000014458612",
        "us_id": "44350",
        "comm_code": "301003400000093",
        "upcard_terminal": "91330023",
        "upcard_mer_id": "102913058120013",
        "ex_id": "20343",
        "ex_cost_center_code": "1200044350",
        "dcore_store_appid": "s20170823000005665"
    }
},
{
    "_id": "3850146333826482177",
    "name": "长沙德思勤店",
    "extend_code": {
        "comm_shop_id": "a4d414ff82b044e3939c253e581a53ff",
        "ex_code": "20467",
        "alipay_id": "2015093000077000000004451342",
        "us_id": "44165",
        "comm_code": "301003400000305",
        "upcard_terminal": "73109129",
        "upcard_mer_id": "102731058121544",
        "ex_id": "20272",
        "ex_cost_center_code": "1200044165",
        "dcore_store_appid": "s20170823000005626"
    }
},
{
    "_id": "3850146381175980033",
    "name": "银川东方红广场店",
    "extend_code": {
        "comm_shop_id": "7e0e475195b741fe82b7ac43ba0c4e4b",
        "ex_code": "20550",
        "alipay_id": "2016111700077000000020171853",
        "us_id": "44433",
        "comm_code": "301003400000241",
        "upcard_terminal": "95100616",
        "upcard_mer_id": "102951058120244",
        "ex_id": "20402",
        "ex_cost_center_code": "1200044433",
        "dcore_store_appid": "s20170823000005675"
    }
},
{
    "_id": "3850146362481967105",
    "name": "德州万达店",
    "extend_code": {
        "comm_shop_id": "1472fd60b6424ca48d4e3fb57511b939",
        "ex_code": "20515",
        "alipay_id": "2016011100077000000014153677",
        "us_id": "44228",
        "comm_code": "301003400000051",
        "upcard_terminal": "53400045",
        "upcard_mer_id": "102534058120010",
        "ex_id": "20319",
        "ex_cost_center_code": "1200044228",
        "dcore_store_appid": "s20170823000005636"
    }
},
{
    "_id": "3850146328482938881",
    "name": "内江万达店",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20465",
        "comm_shop_id": "3700d853cc38457da21109e25bb8e980",
        "us_id": "44133",
        "alipay_id": "2015093000077000000004479169",
        "takeaway_eleme_id": '',
        "upcard_terminal": "83240047",
        "comm_code": "301003400000118",
        "upcard_mer_id": "102832058120004",
        "ex_id": "20266",
        "ex_cost_center_code": "1200044133",
        "dcore_store_appid": "s20170823000005621"
    }
},
{
    "_id": "3850146346455531521",
    "name": "长清大学城店",
    "extend_code": {
        "comm_shop_id": "72dd4f2e97854becbdecf9421a822851",
        "ex_code": "20489",
        "alipay_id": "2015093000077000000004411685",
        "us_id": "44231",
        "comm_code": "301003400000229",
        "upcard_terminal": "53101194",
        "upcard_mer_id": "102531058120518",
        "ex_id": "20288",
        "ex_cost_center_code": "1200044231",
        "dcore_store_appid": "s20170823000005638"
    }
},
{
    "_id": "3850146348980502529",
    "name": "桂林万达店",
    "extend_code": {
        "comm_shop_id": "17fc796a3b7947d6975d4d25d28db246",
        "ex_code": "20493",
        "alipay_id": "2015113000077000000006787409",
        "us_id": "44230",
        "comm_code": "301003400000058",
        "upcard_terminal": "77300780",
        "upcard_mer_id": "102773058120469",
        "ex_id": "20284",
        "ex_cost_center_code": "1200044230",
        "dcore_store_appid": "s20170823000005637"
    }
},
{
    "_id": "3850146356035321857",
    "name": "安康高新万达店",
    "extend_code": {
        "comm_shop_id": "ccae2caf49a341909ca68a0365e30fc0",
        "ex_code": "20506",
        "alipay_id": "2016012100077000000014371362",
        "us_id": "44244",
        "comm_code": "301003400000371",
        "upcard_terminal": "91500056",
        "upcard_mer_id": "102915058120004",
        "ex_id": "20316",
        "ex_cost_center_code": "1200044244",
        "dcore_store_appid": "s20170929000006491"
    }
},
{
    "_id": "3850146369440317441",
    "name": "丹阳吾悦",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20528",
        "comm_shop_id": "50d3079997d3468ab64bd4bb6af9cc7c",
        "us_id": "44362",
        "alipay_id": "2016011200077000000014111817",
        "takeaway_eleme_id": '',
        "upcard_terminal": "51103301",
        "comm_code": "301003400000164",
        "upcard_mer_id": "102511058121182",
        "ex_id": "20334",
        "ex_cost_center_code": "1200044362",
        "dcore_store_appid": "s20170823000005666"
    }
},
{
    "_id": "3850146383977775105",
    "name": "西宁力盟店",
    "extend_code": {
        "comm_shop_id": "eb8f67a541ee4ad78938efba440f0636",
        "ex_code": "20555",
        "alipay_id": "2016031800077000000015094619",
        "us_id": "44371",
        "comm_code": "301003400000425",
        "upcard_terminal": "97100382",
        "upcard_mer_id": "102971058120232",
        "ex_id": "20403",
        "ex_cost_center_code": "1200044371",
        "dcore_store_appid": "s20170823000005324"
    }
},
{
    "_id": "3869271134012702721",
    "name": "常德万达广场店",
    "extend_code": {
        "comm_shop_id": "dbdd13a644bf41eea38f3bf05f65cf9d",
        "ex_code": "20625",
        "alipay_id": "2016110200077000000019662896",
        "us_id": "44556",
        "comm_code": "301003400000404",
        "upcard_terminal": "73601475",
        "upcard_mer_id": "102736058120031",
        "ex_id": "20454",
        "ex_cost_center_code": "1200044556",
        "dcore_store_appid": "s20170823000005684"
    }
},
{
    "_id": "3850146363031420929",
    "name": "常州九洲新世界",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20516",
        "comm_shop_id": "beb5b08035814e66a859377e2c7618be",
        "us_id": "44271",
        "alipay_id": "2016011200077000000014263932",
        "takeaway_eleme_id": '',
        "upcard_terminal": "51900921",
        "comm_code": "301003400004636",
        "upcard_mer_id": "102519058120140",
        "ex_id": "20333",
        "ex_cost_center_code": "1200844271",
        "dcore_store_appid": "s20170823000005166"
    }
},
{
    "_id": "3850146367368331265",
    "name": "富阳玉长城",
    "extend_code": {
        "comm_shop_id": "7c590955193341439c1dc4736f465ddd",
        "ex_code": "20524",
        "alipay_id": "2016031800077000000015036324",
        "us_id": "44347",
        "comm_code": "301003400000238",
        "upcard_terminal": "57109443",
        "upcard_mer_id": "102571058122427",
        "ex_id": "20348",
        "ex_cost_center_code": "1200044347",
        "dcore_store_appid": "s20170823000005663"
    }
},
{
    "_id": "3850146411635015681",
    "name": "南昌万达茂店",
    "extend_code": {
        "comm_shop_id": "b15f4c0eff53445799966ca0edbce0a5",
        "ex_code": "20610",
        "alipay_id": "2016111600077000000019990441",
        "us_id": "44495",
        "comm_code": "301003400000325",
        "upcard_terminal": "79190465",
        "upcard_mer_id": "102791058120614",
        "ex_id": "",
        "ex_cost_center_code": "1200044495",
        "dcore_store_appid": "s20170823000005678"
    }
},
{
    "_id": "3873324792933253121",
    "name": "邳州新苏",
    "extend_code": {
        "comm_shop_id": "a7b8056287614209a7e7a37b31744274",
        "ex_code": "20620",
        "alipay_id": "2016111700077000000019992933",
        "us_id": "44583",
        "comm_code": "301003400004615",
        "upcard_terminal": "51600877",
        "upcard_mer_id": "102516058120230",
        "ex_id": "20459",
        "ex_cost_center_code": "1200844583",
        "dcore_store_appid": "s20170823000005429"
    }
},
{
    "_id": "3873587551059050497",
    "name": "杭州运河上街",
    "extend_code": {
        "comm_shop_id": "96f24e96ed4b4e708b0686bdf85dd6f1",
        "ex_code": "20627",
        "alipay_id": "2016111400077000000019961039",
        "us_id": "44584",
        "comm_code": "301003400004596",
        "upcard_terminal": "57109980",
        "upcard_mer_id": "102571058122507",
        "ex_id": "20460",
        "ex_cost_center_code": "1200844584",
        "dcore_store_appid": "s20170823000005180"
    }
},
{
    "_id": "3875522205441851393",
    "name": "合肥瑶海万达店",
    "extend_code": {
        "comm_shop_id": "75146db201ac4710bcc8a7c35ab6a73e",
        "ex_code": "20619",
        "alipay_id": "2016111500077000000020025227",
        "us_id": "44585",
        "comm_code": "301003400004555",
        "upcard_terminal": "55129620",
        "upcard_mer_id": "102551058123672",
        "ex_id": "20462",
        "ex_cost_center_code": "1200844585",
        "dcore_store_appid": "s20170823000005430"
    }
},
{
    "_id": "3876153153984397313",
    "name": "三门峡万达",
    "extend_code": {
        "comm_shop_id": "4002e8c31a2e4731bcedd1f8b0855b00",
        "ex_code": "20632",
        "alipay_id": "2016111600077000000020014296",
        "us_id": "44601",
        "comm_code": "301003400000134",
        "upcard_terminal": "39801114",
        "upcard_mer_id": "102398058120016",
        "ex_id": "20464",
        "ex_cost_center_code": "1200044601",
        "dcore_store_appid": "s20170823000005330"
    }
},
{
    "_id": "3876159232373948417",
    "name": "合肥万达茂",
    "extend_code": {
        "comm_shop_id": "e7f8488cf20841088d4e5402f7667c36",
        "ex_code": "20583",
        "alipay_id": "2016112100077000000020315788",
        "us_id": "44575",
        "comm_code": "301003400004673",
        "upcard_terminal": "55129615",
        "upcard_mer_id": "102551058123669",
        "ex_id": "20450",
        "ex_cost_center_code": "1200844575",
        "dcore_store_appid": "s20170823000005178"
    }
},
{
    "_id": "3878042887929200641",
    "name": "徐州铜山万达",
    "extend_code": {
        "comm_shop_id": "41024ba93e694f88aced25f862c4303a",
        "ex_code": "20629",
        "alipay_id": "2016111400077000000019938321",
        "us_id": "44607",
        "comm_code": "301003400004512",
        "upcard_terminal": "51600880",
        "upcard_mer_id": "102516058120232",
        "ex_id": "20468",
        "ex_cost_center_code": "1200844607",
        "dcore_store_appid": "s20170823000005433"
    }
},
{
    "_id": "3878414640677388289",
    "name": "宁海西子国际",
    "extend_code": {
        "comm_shop_id": "ed126009e0ce44c1b23ae9500ce728e8",
        "ex_code": "20622",
        "alipay_id": "2016111500077000000019990160",
        "us_id": "44579",
        "comm_code": "301003400004677",
        "upcard_terminal": "57403372",
        "upcard_mer_id": "102574058120527",
        "ex_id": "20451",
        "ex_cost_center_code": "1200844579",
        "dcore_store_appid": "s20170823000005427"
    }
},
{
    "_id": "3883396195845931009",
    "name": "蚌埠银泰城",
    "extend_code": {
        "comm_shop_id": "14ad34c86bd745b19bf8b6b9e7972c9c",
        "ex_code": "20638",
        "alipay_id": "2017120400077000000046754136",
        "us_id": "44605",
        "comm_code": "301003400004466",
        "upcard_terminal": "55202607",
        "upcard_mer_id": "102552058120097",
        "ex_cost_center_code": "1200844605",
        "dcore_store_appid": "s20170823000005181"
    }
},
{
    "_id": "3891058106678902785",
    "name": "连云港万达广场店",
    "extend_code": {
        "comm_shop_id": "c1769bacba45497ca84d7db0bb3752b1",
        "ex_code": "20635",
        "alipay_id": "2017120400077000000046744552",
        "us_id": "44593",
        "comm_code": "301003400004640",
        "upcard_terminal": "51800662",
        "upcard_mer_id": "102518058120066",
        "ex_cost_center_code": "1200844593",
        "dcore_store_appid": "s20170823000005439"
    }
},
{
    "_id": "3891708784841588737",
    "name": "郑州二七万达",
    "extend_code": {
        "comm_shop_id": "514a6cd3d22f436d8af6fbe29b24fc6c",
        "ex_code": "20637",
        "alipay_id": "2018040400077000000048193647",
        "us_id": "44628",
        "comm_code": "301003400000165",
        "upcard_terminal": "37112065",
        "upcard_mer_id": "102371058122461",
        "ex_cost_center_code": "1200044628",
        "dcore_store_appid": "s20170823000005688"
    }
},
{
    "_id": "3893230750010441729",
    "name": "成都青羊万达",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20639",
        "comm_shop_id": "fbe9f78d8cf4408b9cd28f2baf957a02",
        "us_id": "44632",
        "alipay_id": "2017120400077000000046745555",
        "takeaway_eleme_id": '',
        "upcard_terminal": "02895967",
        "comm_code": "301003400000453",
        "upcard_mer_id": "102280058125263",
        "ex_cost_center_code": "1200044632",
        "dcore_store_appid": "s20170823000005690"
    }
},
{
    "_id": "3893273023725174785",
    "name": "郑州惠济万达",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "ex_code": "20634",
        "comm_shop_id": "6f2d53b800ce4432bbb6fbb374f00e90",
        "us_id": "44602",
        "alipay_id": "2018040400077000000048193649",
        "takeaway_eleme_id": '',
        "upcard_terminal": "37112066",
        "comm_code": "301003400000219",
        "upcard_mer_id": "102371058122462",
        "ex_cost_center_code": "1200044602",
        "dcore_store_appid": "s20170823000005685"
    }
},
{
    "_id": "3896474707003703297",
    "name": "杭州滨江宝龙",
    "extend_code": {
        "comm_shop_id": "dbbb683dd83a4fe98a637d8a4fba8617",
        "ex_code": "20641",
        "alipay_id": "2017120400077000000046699854",
        "us_id": "44701",
        "comm_code": "301003400004662",
        "upcard_terminal": "57110445",
        "upcard_mer_id": "102571058122527",
        "ex_cost_center_code": "1200844701",
        "dcore_store_appid": "s20170823000005183"
    }
},
{
    "_id": "3896478014455676929",
    "name": "建德太平洋",
    "extend_code": {
        "comm_shop_id": "d87ad61122a84d208091d5fb8c9da435",
        "ex_code": "20640",
        "alipay_id": "2018040900077000000048270638",
        "us_id": "44699",
        "comm_code": "301003400004659",
        "upcard_terminal": "57110446",
        "upcard_mer_id": "102571058122528",
        "ex_cost_center_code": "1200844699",
        "dcore_store_appid": "s20170823000005436"
    }
},
{
    "_id": "3896481459271106561",
    "name": "烟台开发区万达店",
    "extend_code": {
        "comm_shop_id": "db271312c32742b4bade2db57e048c44",
        "ex_code": "20642",
        "alipay_id": "2018041200077000000048267150",
        "us_id": "44629",
        "comm_code": "301003400000403",
        "upcard_terminal": "53530406",
        "upcard_mer_id": "102535058120167",
        "ex_cost_center_code": "1200044629",
        "dcore_store_appid": "s20170823000005331"
    }
},
{
    "_id": "3898241725730127873",
    "name": "宿州万达",
    "extend_code": {
        "comm_shop_id": "546463e178434941978b5f889e359418",
        "ex_code": "20643",
        "alipay_id": "2017120400077000000046765144",
        "us_id": "44705",
        "comm_code": "301003400004528",
        "upcard_terminal": "55740026",
        "upcard_mer_id": "102557058120095",
        "ex_cost_center_code": "1200844705",
        "dcore_store_appid": "s20170823000005438"
    }
},
{
    "_id": "3901599326224252929",
    "name": "三明万达店",
    "extend_code": {
        "comm_shop_id": "55e5d6265c774f6bb4115facf615f0dd",
        "ex_code": "20646",
        "alipay_id": "2018040900077000000048249322",
        "us_id": "44630",
        "comm_code": "301003400000173",
        "upcard_terminal": "59800139",
        "upcard_mer_id": "102598058120041",
        "ex_cost_center_code": "1200044630",
        "dcore_store_appid": "s20170823000005689"
    }
},
{
    "_id": "3902216432686989313",
    "name": "滨州万达",
    "extend_code": {
        "comm_shop_id": "6c6dbc06acab49048bc20aa91aedfc05",
        "ex_code": "20647",
        "alipay_id": "2018040900077000000048281145",
        "us_id": "44633",
        "comm_code": "301003400000214",
        "upcard_terminal": "54300018",
        "upcard_mer_id": "102543058120085",
        "ex_cost_center_code": "1200044633",
        "dcore_store_appid": "s20170823000005332"
    }
},
{
    "_id": "3904032209438244865",
    "name": "象山博浪太平洋店",
    "extend_code": {
        "comm_shop_id": "208be415155b4c70927e7e02fa82c967",
        "ex_code": "20648",
        "alipay_id": "2017120400077000000046770747",
        "us_id": "44700",
        "comm_code": "301003400004482",
        "upcard_terminal": "57403414",
        "upcard_mer_id": "102574058120535",
        "ex_cost_center_code": "1200844700",
        "dcore_store_appid": "s20170823000005437"
    }
},
{
    "_id": "3904821037111443457",
    "name": "湖州万达",
    "extend_code": {
        "comm_shop_id": "b3647b4840a94e849e1b6f5b5193ba38",
        "ex_code": "20649",
        "alipay_id": "2017120400077000000046768872",
        "us_id": "44720",
        "comm_code": "301003400004626",
        "upcard_terminal": "57297181",
        "upcard_mer_id": "102572058120068",
        "ex_cost_center_code": "1200844720",
        "dcore_store_appid": "s20170823000005184"
    }
},
{
    "_id": "3906632910475624449",
    "name": "六安万达",
    "extend_code": {
        "comm_shop_id": "208fc01a65014d6dae44edee75b6e714",
        "ex_code": "20650",
        "alipay_id": "2018040900077000000048283968",
        "us_id": "44727",
        "comm_code": "301003400004483",
        "upcard_terminal": "56401490",
        "upcard_mer_id": "102564058120147",
        "ex_cost_center_code": "1200844727",
        "dcore_store_appid": "s20170823000005441"
    }
},
{
    "_id": "3907309584703815681",
    "name": "嘉善万联城",
    "extend_code": {
        "comm_shop_id": "f2691169f6224907bcfae967a5370c2f",
        "ex_code": "20651",
        "alipay_id": "2017120400077000000046728533",
        "us_id": "44726",
        "comm_code": "301003400004681",
        "upcard_terminal": "57303460",
        "upcard_mer_id": "102573058120167",
        "ex_cost_center_code": "1200844726",
        "dcore_store_appid": "s20170823000005440"
    }
},
{
    "_id": "3918904884010680321",
    "name": "盐城建军路金鹰店",
    "extend_code": {
        "comm_shop_id": "d31d4623d5de446fa2f673e6fd2ae1bd",
        "ex_code": "20653",
        "alipay_id": "2017120400077000000046768876",
        "us_id": "44770",
        "comm_code": "301003400004657",
        "upcard_terminal": "51500739",
        "upcard_mer_id": "102515058120461",
        "ex_cost_center_code": "1200844770",
        "dcore_store_appid": "s20170823000005443"
    }
},
{
    "_id": "3935323620915888129",
    "name": "上海百脑汇中金",
    "extend_code": {
        "comm_shop_id": "73373740e3544eb08ad97506c6020f33",
        "alipay_id": "2017120100077000000046764861",
        "us_id": "44799",
        "comm_code": "301003400004553",
        "upcard_terminal": "02100360",
        "upcard_mer_id": "102210058127338",
        "ex_cost_center_code": "1200844799",
        "dcore_store_appid": "s20170823000005186"
    }
},
{
    "_id": "3936702061001609217",
    "name": "合肥悦方",
    "extend_code": {
        "comm_shop_id": "111aaf6778874be6a6746d8344a1cfb5",
        "alipay_id": "2017120100077000000046705835",
        "us_id": "44800",
        "comm_code": "301003400004462",
        "upcard_terminal": "55108258",
        "upcard_mer_id": "102551058123725",
        "ex_cost_center_code": "1200844800",
        "dcore_store_appid": "s20170823000005445"
    }
},
{
    "_id": "3939969521955577857",
    "name": "武汉佰港城店",
    "extend_code": {
        "comm_shop_id": "08abf1638f6643e0a65881661fc55604",
        "ex_code": "",
        "alipay_id": "2017120100077000000046707781",
        "us_id": "44804",
        "comm_code": "301003400000035",
        "upcard_terminal": "02708725",
        "upcard_mer_id": "102270058125346",
        "ex_cost_center_code": "1200044804",
        "dcore_store_appid": "s20170823000005693"
    }
},
{
    "_id": "3949689889562189825",
    "name": "南通中南加盟店",
    "extend_code": {
        "comm_shop_id": "b72625ffd12b47f8b3ef693e0dac034f",
        "alipay_id": "2018040900077000000048286798",
        "us_id": "44818",
        "comm_code": "301003400004628",
        "upcard_terminal": "51300955",
        "upcard_mer_id": "102513058120211",
        "ex_cost_center_code": "1200844818",
        "dcore_store_appid": "s20170823000005448"
    }
},
{
    "_id": "3953061342424817665",
    "name": "义乌之心",
    "extend_code": {
        "comm_shop_id": "1e959784c6c0427bb06f3845786fd8fd",
        "alipay_id": "2018040900077000000048275914",
        "us_id": "44817",
        "comm_code": "301003400004480",
        "upcard_terminal": "57990577",
        "upcard_mer_id": "102579058120236",
        "ex_cost_center_code": "1200844817",
        "dcore_store_appid": "s20170823000005465"
    }
},
{
    "_id": "3957672566130307073",
    "name": "上海莲花国际",
    "extend_code": {
        "comm_shop_id": "bdcc7e0fba7c49d3ac110e8abc6dbe73",
        "alipay_id": "2017120100077000000046736390",
        "us_id": "44826",
        "comm_code": "301003400004635",
        "upcard_terminal": "02100302",
        "upcard_mer_id": "102210058127362",
        "ex_cost_center_code": "1200844826",
        "dcore_store_appid": "s20170823000005187"
    }
},
{
    "_id": "3958032643133870081",
    "name": "滁州苏宁广场",
    "extend_code": {
        "comm_shop_id": "8f6fa896f4e24c5a9edc8c0668bbeb8b",
        "alipay_id": "2018040900077000000048275898",
        "us_id": "44824",
        "comm_code": "301003400004590",
        "upcard_terminal": "55000331",
        "upcard_mer_id": "102550058120071",
        "ex_cost_center_code": "1200844824",
        "dcore_store_appid": "s20170823000005449"
    }
},
{
    "_id": "3966059407857238017",
    "name": "如皋文峰加盟店",
    "extend_code": {
        "comm_shop_id": "a5d7114096bb443ebf23cd73cf99d1c1",
        "alipay_id": "2017120400077000000046773987",
        "us_id": "44850",
        "comm_code": "301003400004612",
        "upcard_terminal": "51300962",
        "upcard_mer_id": "102513058120216",
        "ex_cost_center_code": "1200844850",
        "dcore_store_appid": "s20170823000005453"
    }
},
{
    "_id": "3966093447387537409",
    "name": "枣庄万达",
    "extend_code": {
        "comm_shop_id": "a33d2d5687b248f88ce661e498c6ff65",
        "alipay_id": "2017120400077000000046728539",
        "us_id": "44864",
        "comm_code": "301003400004607",
        "upcard_terminal": "63200098",
        "upcard_mer_id": "102632058120025",
        "ex_cost_center_code": "1200844864",
        "dcore_store_appid": "s20170823000005457"
    }
},
{
    "_id": "3966096594713075713",
    "name": "漯河大商新玛特",
    "extend_code": {
        "comm_shop_id": "7d1c25ae05d145d8bb4f4a572fa1d74e",
        "alipay_id": "2018040400077000000048200655",
        "us_id": "44862",
        "comm_code": "301003400004567",
        "upcard_terminal": "39501433",
        "upcard_mer_id": "102395089990085",
        "ex_cost_center_code": "1200844862",
        "dcore_store_appid": "s20170823000005455"
    }
},
{
    "_id": "3966100462889066497",
    "name": "焦作万达",
    "extend_code": {
        "comm_shop_id": "77dd986b49d4437c95a0e24b27249f7e",
        "alipay_id": "2017120100077000000046751943",
        "us_id": "44861",
        "comm_code": "301003400004558",
        "upcard_terminal": "39102509",
        "upcard_mer_id": "102391089990186",
        "ex_cost_center_code": "1200844861",
        "dcore_store_appid": "s20170823000005190"
    }
},
{
    "_id": "3968195667572068353",
    "name": "益阳万达广场店",
    "extend_code": {
        "comm_shop_id": "19da77983fca471bbc9cfec8b2c4d958",
        "alipay_id": "2017120400077000000046768863",
        "us_id": "44863",
        "comm_code": "301003400004472",
        "upcard_terminal": "73700870",
        "upcard_mer_id": "102737089990043",
        "ex_cost_center_code": "1200844863",
        "dcore_store_appid": "s20170823000005456"
    }
},
{
    "_id": "3970708804825374721",
    "name": "兴化东方商厦",
    "extend_code": {
        "comm_shop_id": "8412ccbf8ad84f4199950bf9ee838a4f",
        "alipay_id": "2017120400077000000046770756",
        "us_id": "44825",
        "comm_code": "301003400004576",
        "upcard_terminal": "52300736",
        "upcard_mer_id": "102523058120111",
        "ex_id": "",
        "ex_cost_center_code": "1200844825",
        "dcore_store_appid": "s20170823000005451"
    }
},
{
    "_id": "3970723154399453185",
    "name": "盐城万达广场",
    "extend_code": {
        "comm_shop_id": "5511192321874a8ca51fe3f0562dda01",
        "alipay_id": "2017120400077000000046745569",
        "us_id": "44856",
        "comm_code": "301003400004530",
        "upcard_terminal": "51500809",
        "upcard_mer_id": "102515058120472",
        "ex_cost_center_code": "1200844856",
        "dcore_store_appid": "s20170823000005454"
    }
},
{
    "_id": "3974341319478169601",
    "name": "昆山万达店",
    "extend_code": {
        "comm_shop_id": "79bf41c5af684b2aadbfb0057d22cce3",
        "alipay_id": "2017120100077000000046775297",
        "us_id": "44868",
        "comm_code": "301003400004561",
        "upcard_terminal": "51215951",
        "upcard_mer_id": "102512058123161",
        "ex_cost_center_code": "1200844868",
        "dcore_store_appid": "s20170823000005459"
    }
},
{
    "_id": "3977285180676460545",
    "name": "南京江宁金鹰（加盟）",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "comm_shop_id": "aaad2a3622a14910b21458e16461817b",
        "us_id": "44869",
        "alipay_id": "2017120100077000000046723049",
        "comm_code": "301003400004617",
        "takeaway_eleme_id": '',
        "upcard_mer_id": "102250058123272",
        "ex_cost_center_code": "1200844869",
        "dcore_store_appid": "s20170823000005460",
        "upcard_terminal": "02598814"
    }
},
{
    "_id": "3977287272995196929",
    "name": "常州新城吾悦（加盟店）",
    "extend_code": {
        "comm_shop_id": "2a0f3f99b9c049f2a3e8dd998f1c65ff",
        "alipay_id": "2018071000077000000059042277",
        "us_id": "44867",
        "comm_code": "301003400004488",
        "upcard_terminal": "51901161",
        "upcard_mer_id": "102519058120204",
        "ex_cost_center_code": "1200844867",
        "dcore_store_appid": "s20170823000005458"
    }
},
{
    "_id": "3979829786309996545",
    "name": "常熟万达",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "comm_shop_id": "6244df76e1cd4de9adf15c034baeb820",
        "us_id": "44875",
        "alipay_id": "2017120100077000000046723060",
        "comm_code": "301003400004541",
        "takeaway_eleme_id": '',
        "upcard_mer_id": "102520089990007",
        "ex_cost_center_code": "1200844875",
        "dcore_store_appid": "s20170823000005461",
        "upcard_terminal": "52000085"
    }
},
{
    "_id": "3984873648719810561",
    "name": "常熟江南印象",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "comm_shop_id": "3b73658c410846ae84e572e6f5c66aca",
        "us_id": "44886",
        "alipay_id": "2017120100077000000046721963",
        "comm_code": "301003400004505",
        "takeaway_eleme_id": '',
        "upcard_mer_id": "102512089993167",
        "ex_id": "",
        "ex_cost_center_code": "1200844886",
        "dcore_store_appid": "s20170929000006485",
        "upcard_terminal": "51215961"
    }
},
{
    "_id": "3985236110106099713",
    "name": "周口万顺达",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "comm_shop_id": "707acb78761f4ba08441fd9ce609c6b0",
        "us_id": "44896",
        "alipay_id": "2018040400077000000048176794",
        "comm_code": "301003400004550",
        "takeaway_eleme_id": '',
        "upcard_mer_id": "102394089990038",
        "ex_cost_center_code": "1200844896",
        "dcore_store_appid": "s20170823000005464",
        "upcard_terminal": "39401115"
    }
},
{
    "_id": "3994699034354524161",
    "name": "南东食品一店",
    "extend_code": {
        "comm_shop_id": "816f7969939d424089e8b63cd479070f",
        "alipay_id": "2017120400077000000046723324",
        "us_id": "44914",
        "comm_code": "301003400004573",
        "upcard_terminal": "02100411",
        "upcard_mer_id": "102210058227491",
        "ex_cost_center_code": "1200844914",
        "dcore_store_appid": "s20170929000006481"
    }
},
{
    "_id": "3994703758243516417",
    "name": "郑州乐尚生活广场",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "comm_shop_id": "b9516e09407143c5a1e44e39e6fe66e2",
        "us_id": "44938",
        "alipay_id": "2018040900077000000048288285",
        "comm_code": "301003400004630",
        "takeaway_eleme_id": '',
        "upcard_mer_id": "102371058222730",
        "ex_cost_center_code": "1200844938",
        "dcore_store_appid": "s20170929000006501",
        "upcard_terminal": "37112486"
    }
},
{
    "_id": "3995354263162851329",
    "name": "西安太白印象城",
    "extend_code": {
        "comm_shop_id": "5df937ed089b40678f8ce09b3ae0ebae",
        "alipay_id": "2017120400077000000046770741",
        "us_id": "44933",
        "comm_code": "301003400000195",
        "upcard_terminal": "02904311",
        "upcard_mer_id": "102290058223117",
        "ex_cost_center_code": "1200044933",
        "dcore_store_appid": "s20170929000006489"
    }
},
{
    "_id": "3995730255083913217",
    "name": "东阳银泰",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "comm_shop_id": "a2d59076155f486987639ce8f7be7247",
        "us_id": "44901",
        "alipay_id": "2017120400077000000046722217",
        "comm_code": "301003400004606",
        "takeaway_eleme_id": '',
        "upcard_mer_id": "102579089990249",
        "ex_cost_center_code": "1200844901",
        "dcore_store_appid": "s20170929000006488",
        "upcard_terminal": "57990603"
    }
},
{
    "_id": "3997174071516418049",
    "name": "济宁运河城",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "comm_shop_id": "e156237095a64288886b6152c823731f",
        "us_id": "44924",
        "alipay_id": "2017120600077000000046923728",
        "comm_code": "301003400004665",
        "takeaway_eleme_id": '',
        "upcard_mer_id": "102539089990304",
        "ex_cost_center_code": "1200844924",
        "dcore_store_appid": "s20170929000006496",
        "upcard_terminal": "53950776"
    }
},
{
    "_id": "3997176235735457793",
    "name": "连云港利群店",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "comm_shop_id": "851eaacfd20e4d8ba12744bb6184be6f",
        "us_id": "44935",
        "alipay_id": "2017120400077000000046750245",
        "comm_code": "301003400004577",
        "takeaway_eleme_id": '',
        "upcard_mer_id": "102518089990077",
        "ex_cost_center_code": "1200844935",
        "dcore_store_appid": "s20170929000006497",
        "upcard_terminal": "51800720"
    }
},
{
    "_id": "3997560956944306177",
    "name": "香港名都",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "comm_shop_id": "d9adc4bd46b04279bd7663741f72ca63",
        "us_id": "44926",
        "alipay_id": "2017120100077000000046777016",
        "comm_code": "301003400000400",
        "takeaway_eleme_id": '',
        "upcard_mer_id": "102210089997762",
        "ex_cost_center_code": "1200044926",
        "dcore_store_appid": "s20170929000006480",
        "upcard_terminal": "02105290"
    }
},
{
    "_id": "3999773029309284353",
    "name": "无锡梅村南服务区",
    "extend_code": {
        "comm_shop_id": "1cc61da075bc44e6ab1f82f167084709",
        "alipay_id": "2017120400077000000046759054",
        "us_id": "44945",
        "comm_code": "301003400004477",
        "upcard_terminal": "51001731",
        "upcard_mer_id": "102510089992602",
        "ex_cost_center_code": "1200844945",
        "dcore_store_appid": "s20170928000006468"
    }
},
{
    "_id": "3999774685956632577",
    "name": "无锡梅村北服务区",
    "extend_code": {
        "comm_shop_id": "a6f21e7b119a4f448cdba7106af51dac",
        "alipay_id": "2017120400077000000046759054",
        "us_id": "44947",
        "comm_code": "301003400004614",
        "upcard_terminal": "51001732",
        "upcard_mer_id": "102510089992603",
        "ex_cost_center_code": "1200844947",
        "dcore_store_appid": "s20170928000006470"
    }
},
{
    "_id": "4001376863112306689",
    "name": "国融测试门店02",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "comm_shop_id": "0905bf5cac37403695e02ed9197757ff",
        "pay100_terminal_id": "18097171",
        "us_id": "00002",
        "alipay_id": "2015060900077000000000176452",
        "takeaway_eleme_id": '',
        "pay100_merchant_no": "829005812003177",
        "comm_code": "301003400000348",
        "upcard_mer_id": "102210058126861",
        "pay100_access_token": "e26258cd27e44d9cb481ae20ae29891a",
        "dcore_store_appid": "s20160610000000406",
        "upcard_terminal": "02194635"
    }
},
{
    "_id": "4002249564373131265",
    "name": "杭州龙湖天街店",
    "extend_code": {
        "takeaway_honeymoon_id": '',
        "comm_shop_id": "94ab09c76e234165820c6e16c4300a97",
        "us_id": "44939",
        "alipay_id": "2017120100077000000046739063",
        "comm_code": "301003400004595",
        "takeaway_eleme_id": '',
        "upcard_mer_id": "102571089992670",
        "ex_cost_center_code": "1200844939",
        "dcore_store_appid": "s20171018000006595",
        "upcard_terminal": "57111371"
    }
},
{
    "_id": "4004830607795531777",
    "name": "衡阳万达",
    "extend_code": {
        "comm_shop_id": "5632ee0d5a6a4f07ba39cc3b2e88735b",
        "alipay_id": "2017120400077000000046738075",
        "us_id": "44975",
        "comm_code": "301003400004532",
        "upcard_terminal": "73401178",
        "upcard_mer_id": "102734089990075",
        "ex_cost_center_code": "1200844975",
        "dcore_store_appid": "s20171018000006603"
    }
},
{
    "_id": "4005915785624084481",
    "name": "新乡胖东来",
    "extend_code": {
        "comm_shop_id": "0a74b0275a38460883dd55a5f5e530a2",
        "alipay_id": "2018040400077000000048215754",
        "us_id": "44912",
        "comm_code": "301003400004456",
        "upcard_terminal": "37301800",
        "upcard_mer_id": "102373089990129",
        "ex_cost_center_code": "1200844912",
        "dcore_store_appid": "s20171018000006605"
    }
},
{
    "_id": "4007436278454431745",
    "name": "仪征北服务区",
    "extend_code": {
        "comm_shop_id": "926d0b5d2eff4f979393176ffe8cc447",
        "alipay_id": "2017120100077000000046775299",
        "us_id": "44957",
        "comm_code": "301003400004592",
        "upcard_terminal": "51401108",
        "upcard_mer_id": "102514089990290",
        "ex_cost_center_code": "1200844957",
        "dcore_store_appid": "s20170928000006467"
    }
},
{
    "_id": "4007436814637395969",
    "name": "仪征南服务区",
    "extend_code": {
        "comm_shop_id": "8e98876b901a47bc97c2c71c2c984b6b",
        "alipay_id": "2017120100077000000046775299",
        "us_id": "44958",
        "comm_code": "301003400004588",
        "upcard_terminal": "51401109",
        "upcard_mer_id": "102514089990291",
        "ex_cost_center_code": "1200844958",
        "dcore_store_appid": "s20170928000006466"
    }
},
{
    "_id": "4007676009175732225",
    "name": "重庆北站出发店",
    "extend_code": {
        "comm_shop_id": "bc6f6d06896a4f6b9554bbd58d1469b4",
        "alipay_id": "2018040900077000000048286799",
        "us_id": "44972",
        "comm_code": "301003400004632",
        "upcard_terminal": "02380235",
        "upcard_mer_id": "102230089992056",
        "ex_cost_center_code": "1200844972",
        "dcore_store_appid": "s20171020000006679"
    }
},
{
    "_id": "4008482139026063361",
    "name": "合肥商之都",
    "extend_code": {
        "comm_shop_id": "37890bde536c4adcae5a787fb1532a2f",
        "alipay_id": "2017120100077000000046777006",
        "us_id": "44955",
        "comm_code": "301003400004503",
        "upcard_terminal": "55700255",
        "upcard_mer_id": "102557089990115",
        "ex_cost_center_code": "1200844955",
        "dcore_store_appid": "s20171018000006608"
    }
},
{
    "_id": "4010288854749007873",
    "name": "合肥万象城",
    "extend_code": {
        "comm_shop_id": "4b4f2189f3224df4bf619cd811b9aec3",
        "alipay_id": "2017120300077000000046752172",
        "us_id": "44849",
        "comm_code": "301003400004519",
        "upcard_terminal": "55130567",
        "upcard_mer_id": "102551089993762",
        "ex_cost_center_code": "1200844849",
        "dcore_store_appid": "s20171020000006680"
    }
},
{
    "_id": "4015293472954216449",
    "name": "铜陵万达广场",
    "extend_code": {
        "comm_shop_id": "85deba940f49433692989ade6b19681d",
        "alipay_id": "2017120300077000000046766873",
        "us_id": "44986",
        "comm_code": "301003400004579",
        "upcard_terminal": "56210142",
        "upcard_mer_id": "102562089990106",
        "ex_cost_center_code": "1200844986",
        "dcore_store_appid": "s20171020000006682"
    }
},
{
    "_id": "4016018515533459457",
    "name": "西安高新万达",
    "extend_code": {
        "comm_shop_id": "15ca185291c9482baa1aa9310188ae3a",
        "ex_code": "",
        "alipay_id": "2017120300077000000046773913",
        "us_id": "44992",
        "comm_code": "301003400000052",
        "upcard_terminal": "02905015",
        "upcard_mer_id": "102290089993129",
        "ex_cost_center_code": "1200044992",
        "dcore_store_appid": "s20171020000006683"
    }
},
{
    "_id": "4016456551866130433",
    "name": "泸州佳乐世纪城店",
    "extend_code": {
        "comm_shop_id": "1bd11d7ce61d45c99dcc399163836f78",
        "alipay_id": "2018040900077000000048259004",
        "us_id": "45009",
        "comm_code": "301003400004476",
        "upcard_terminal": "83000976",
        "upcard_mer_id": "102830089990201",
        "ex_cost_center_code": "1200845009",
        "dcore_store_appid": "s20171020000006686"
    }
},
{
    "_id": "4022567895717613569",
    "name": "通州万达",
    "extend_code": {
        "comm_shop_id": "71c86122849a4674820c040cd0fc5c21",
        "ex_code": "",
        "alipay_id": "2018040900077000000048275899",
        "us_id": "45022",
        "comm_code": "301003400004551",
        "upcard_terminal": "51301218",
        "upcard_mer_id": "102513089990230",
        "ex_cost_center_code": "1200845022",
        "dcore_store_appid": "s20171102000006792"
    }
},
{
    "_id": "4022985230646005761",
    "name": "苏州中心店",
    "extend_code": {
        "comm_shop_id": "0822659ab7b241afac9b3a2998f48071",
        "alipay_id": "2017120300077000000046750182",
        "us_id": "45001",
        "comm_code": "301003400004455",
        "upcard_terminal": "51216565",
        "upcard_mer_id": "102512089993186",
        "ex_cost_center_code": "1200845001",
        "dcore_store_appid": "s20171102000006794"
    }
},
{
    "_id": "4024054442479214593",
    "name": "平顶山万达店",
    "extend_code": {
        "comm_shop_id": "78a5c52624534ed4a34c22f152496b1d",
        "alipay_id": "2018040900077000000048255945",
        "us_id": "45016",
        "comm_code": "301003400004560",
        "upcard_terminal": "37112760",
        "upcard_mer_id": "102371089992753",
        "ex_cost_center_code": "1200845016",
        "dcore_store_appid": "s20171114000006907"
    }
},
{
    "_id": "4024083696136933377",
    "name": "郑州朗悦公园茂店",
    "extend_code": {
        "comm_shop_id": "037971e53a864ae59b348dff8290a5f2",
        "alipay_id": "2018040400077000000048185150",
        "us_id": "44990",
        "comm_code": "301003400004445",
        "upcard_terminal": "37112762",
        "upcard_mer_id": "102371089992755",
        "ex_cost_center_code": "1200844990",
        "dcore_store_appid": "s20171114000006909"
    }
},
{
    "_id": "4027629077436919809",
    "name": "大丰美庐广场店",
    "extend_code": {
        "comm_shop_id": "c18178168c3a4076be27492ac6babf0f",
        "alipay_id": "2017120300077000000046740433",
        "us_id": "45007",
        "comm_code": "301003400004641",
        "upcard_terminal": "51501012",
        "upcard_mer_id": "102515089990476",
        "ex_cost_center_code": "1200845007",
        "dcore_store_appid": "s20171114000006910"
    }
},
{
    "_id": "4028778360471429121",
    "name": "漕河泾测试店",
    "extend_code": {
        "us_id": "00001",
        "dcore_store_appid": "s20171116000006958",
        "ex_cost_center_code": "1200000001",
        "alipay_id": "2017120400077000000046740477"
    }
},
{
    "_id": "4028800153556328449",
    "name": "上海肇嘉浜路店",
    "extend_code": {
        "comm_shop_id": "d544f97522c141b784b563bfd3218a33",
        "alipay_id": "2018040900077000000048255959",
        "us_id": "45036",
        "comm_code": "301003400000393",
        "upcard_terminal": "02110411",
        "upcard_mer_id": "102210089997915",
        "ex_cost_center_code": "1200045036",
        "dcore_store_appid": "s20171117000006961"
    }
},
{
    "_id": "4029145940016906241",
    "name": "上海颛桥万达店",
    "extend_code": {
        "comm_shop_id": "527550bdda0c48808d22e218bfb541d4",
        "alipay_id": "2018041200077000000048287424",
        "us_id": "45048",
        "comm_code": "301003400004527",
        "upcard_terminal": "02110415",
        "upcard_mer_id": "102210089997917",
        "ex_cost_center_code": "1200845048",
        "dcore_store_appid": "s20171120000006973"
    }
},
{
    "_id": "4030209818787983361",
    "name": "新乡万达店",
    "extend_code": {
        "comm_shop_id": "731eb48f3f6b4f28b282736a89af2d4b",
        "alipay_id": "2018041200077000000048287428",
        "us_id": "45044",
        "comm_code": "301003400004552",
        "upcard_terminal": "37301852",
        "upcard_mer_id": "102373089990132",
        "ex_cost_center_code": "1200845044",
        "dcore_store_appid": "s20171120000006974"
    }
},
{
    "_id": "4030526770225795073",
    "name": "衢州万达店",
    "extend_code": {
        "comm_shop_id": "12d5f5b2f56549af874dab400b0166ff",
        "alipay_id": "2018040900077000000048286782",
        "us_id": "45046",
        "comm_code": "301003400004465",
        "upcard_terminal": "57000331",
        "upcard_mer_id": "102570089990044",
        "ex_cost_center_code": "1200845046",
        "dcore_store_appid": "s20171218000007122"
    }
},
{
    "_id": "4034215829710839809",
    "name": "新乡宝龙店",
    "extend_code": {
        "comm_shop_id": "7a6ea4ade8764e2fb6afeb3b864c4362",
        "alipay_id": "2018041200077000000048276561",
        "us_id": "45058",
        "comm_code": "301003400004563",
        "upcard_terminal": "37301865",
        "upcard_mer_id": "102373089990133",
        "ex_cost_center_code": "1200845058",
        "dcore_store_appid": "s20171201000007071"
    }
},
{
    "_id": "4034221029533986817",
    "name": "南浔新世界店",
    "extend_code": {
        "comm_shop_id": "69fa93ceea5b409a91c93b0cbec45a06",
        "alipay_id": "2018040900077000000048281146",
        "us_id": "45027",
        "comm_code": "301003400004547",
        "upcard_terminal": "57297190",
        "upcard_mer_id": "102572089990073",
        "ex_cost_center_code": "1200845027",
        "dcore_store_appid": "s20171201000007070"
    }
},
{
    "_id": "4036409149526679553",
    "name": "岳阳步步高新天地店",
    "extend_code": {
        "comm_shop_id": "86cba843e9154fc0bc7dcfb3053b3046",
        "alipay_id": "2018041200077000000048267149",
        "us_id": "45059",
        "comm_code": "301003400004580",
        "upcard_terminal": "73001785",
        "upcard_mer_id": "102730089990073",
        "ex_cost_center_code": "1200845059",
        "dcore_store_appid": "s20171208000007093"
    }
},
{
    "_id": "4036411317603905537",
    "name": "长沙富兴中心店",
    "extend_code": {
        "comm_shop_id": "ac2e9a2a52234f54a9dca81664bc1252",
        "alipay_id": "2018040900077000000048270640",
        "us_id": "45068",
        "comm_code": "301003400004618",
        "upcard_terminal": "73117020",
        "upcard_mer_id": "102731089992001",
        "ex_cost_center_code": "1200045068",
        "dcore_store_appid": "s20171208000007092"
    }
},
{
    "_id": "4036748517250641921",
    "name": "开封星光天地店",
    "extend_code": {
        "comm_shop_id": "a65a628c040b4e3a8dcdc1609a2f1197",
        "alipay_id": "2018040400077000000048199601",
        "us_id": "45063",
        "comm_code": "301003400004613",
        "upcard_terminal": "37112988",
        "upcard_mer_id": "102371089992764",
        "ex_cost_center_code": "1200845063",
        "dcore_store_appid": "s20171213000007110"
    }
},
{
    "_id": "4037857286976069633",
    "name": "苏州盛泽购物公园店",
    "extend_code": {
        "comm_shop_id": "20578a2034a948e6a93c0c94a3d8a6c5",
        "alipay_id": "2018040900077000000048266504",
        "us_id": "45028",
        "comm_code": "301003400004481",
        "upcard_terminal": "51216596",
        "upcard_mer_id": "102512089993195",
        "ex_cost_center_code": "1200845028",
        "dcore_store_appid": "s20171213000007111"
    }
},
{
    "_id": "4038119643050315777",
    "name": "泗阳中央商场店",
    "extend_code": {
        "comm_shop_id": "7b6ba7f9ebe64e6da8230104239e5b32",
        "alipay_id": "2018040400077000000048183971",
        "us_id": "45049",
        "comm_code": "301003400004565",
        "upcard_terminal": "52701299",
        "upcard_mer_id": "102527089990301",
        "ex_cost_center_code": "1200845049",
        "dcore_store_appid": "s20171213000007113"
    }
},
{
    "_id": "4038122496011173889",
    "name": "合肥国购店",
    "extend_code": {
        "comm_shop_id": "f6a56d9aad1347d9b421057d457f1639",
        "alipay_id": "2018040900077000000048283970",
        "us_id": "45071",
        "comm_code": "301003400004685",
        "upcard_terminal": "55130910",
        "upcard_mer_id": "102551089993784",
        "ex_cost_center_code": "1200845071",
        "dcore_store_appid": "s20171213000007112"
    }
},
{
    "_id": "4038542360897060865",
    "name": "苏州相城天虹广场店",
    "extend_code": {
        "comm_shop_id": "3fc73d5480be442a97f2ab968f46ef66",
        "alipay_id": "2018040900077000000048281147",
        "us_id": "45056",
        "comm_code": "301003400004509",
        "upcard_terminal": "51216597",
        "upcard_mer_id": "102512089993196",
        "ex_cost_center_code": "1200845056",
        "dcore_store_appid": "s20171213000007114"
    }
},
{
    "_id": "4038952094685507585",
    "name": "邵阳友阿国际广场店",
    "extend_code": {
        "comm_shop_id": "bd8d89b2184243cda69a9288549e58f8",
        "alipay_id": "2018040900077000000048259005",
        "us_id": "45077",
        "comm_code": "301003400004634",
        "upcard_terminal": "73901012",
        "upcard_mer_id": "102739089990048",
        "ex_cost_center_code": "1200845077",
        "dcore_store_appid": "s20171215000007116"
    }
},
{
    "_id": "4041759819603394561",
    "name": "扬州江都金鹰店",
    "extend_code": {
        "comm_shop_id": "cd87eb76a8df4085bf6273098ab17922",
        "alipay_id": "2018040400077000000048176802",
        "us_id": "45100",
        "comm_code": "301003400004652",
        "upcard_terminal": "51401392",
        "upcard_mer_id": "102514089990300",
        "ex_cost_center_code": "1200845100",
        "dcore_store_appid": "s20171230000007137"
    }
},
{
    "_id": "4041828426640879617",
    "name": "延安万达店",
    "extend_code": {
        "comm_shop_id": "96810e0d6160419e817e6241983d8d56",
        "alipay_id": "2018040900077000000048249324",
        "us_id": "45081",
        "comm_code": "301003400000276",
        "upcard_terminal": "91100098",
        "upcard_mer_id": "102911089990023",
        "ex_cost_center_code": "1200045081",
        "dcore_store_appid": "s20171227000007131"
    }
},
{
    "_id": "4047933462918709249",
    "name": "湖州长兴东鱼坊店",
    "extend_code": {
        "comm_shop_id": "87ac10b59d964afba0b5798d07fe7275",
        "alipay_id": "2018040900077000000048281149",
        "us_id": "45029",
        "comm_code": "301003400004583",
        "upcard_terminal": "57297191",
        "upcard_mer_id": "102572089990074",
        "ex_cost_center_code": "1200845029",
        "dcore_store_appid": "s20180109000007153"
    }
},
{
    "_id": "4048346706482503681",
    "name": "安庆吾悦广场店",
    "extend_code": {
        "comm_shop_id": "e2a0e954e48847619dc35fbe5e89a61a",
        "ex_cost_center_code": "1200845101",
        "alipay_id": "2018040900077000000048281148",
        "us_id": "45101",
        "comm_code": "301003400004667",
        "upcard_terminal": "55600321",
        "upcard_mer_id": "102556089990117",
        "ex_id": "45101",
        "ex_code": "45101",
        "dcore_store_appid": "s20180112000007160"
    }
},
{
    "_id": "4048386376470007809",
    "name": "西安新乐汇店",
    "extend_code": {
        "comm_shop_id": "e23eacfb1bcc41f08f9731c8f47e9fe8",
        "alipay_id": "2018040400077000000048209499",
        "us_id": "45079",
        "comm_code": "301003400000410",
        "upcard_terminal": "02905370",
        "upcard_mer_id": "102290089993186",
        "ex_cost_center_code": "1200045079",
        "dcore_store_appid": "s20180112000007162"
    }
},
{
    "_id": "4049375426659905537",
    "name": "上海证大大拇指广场店",
    "extend_code": {
        "comm_shop_id": "76e8c9e3aa4d49e098091d60933a7eea",
        "alipay_id": "2018040900077000000048249338",
        "us_id": "45080",
        "comm_code": "301003400000234",
        "upcard_terminal": "02139377",
        "upcard_mer_id": "102210089998000",
        "ex_cost_center_code": "1200045080",
        "dcore_store_appid": "s20180116000007175"
    }
},
{
    "_id": "4051987334999449601",
    "name": "徐州东站店",
    "extend_code": {
        "comm_shop_id": "5015606cf0634f6891322bb56fa19faf",
        "us_id": "45110",
        "comm_code": "301003400004721",
        "upcard_terminal": "51601357",
        "upcard_mer_id": "102516089990265",
        "ex_cost_center_code": "1200845110",
        "dcore_store_appid": "s20180126000007200"
    }
},
{
    "_id": "4064314349316845569",
    "name": "苏州泉屋百货店",
    "extend_code": {
        "comm_shop_id": "93a35ea818ca4dd49188803d09c366e7",
        "alipay_id": "2018040900077000000048270642",
        "us_id": "45118",
        "comm_code": "301003400004594",
        "upcard_terminal": "51216691",
        "upcard_mer_id": "102512089993214",
        "ex_cost_center_code": "1200845118",
        "dcore_store_appid": "S2018031610007319"
    }
},
{
    "_id": "4076607790348185601",
    "name": "青岛影都万达茂店",
    "extend_code": {
        "comm_shop_id": "cdc84d4ba70449ae9ce7844490bf48ec",
        "alipay_id": "2018051500077000000051645406",
        "us_id": "45122",
        "comm_code": "301003400004653",
        "upcard_terminal": "53206369",
        "upcard_mer_id": "102532089991483",
        "ex_cost_center_code": "1200845122",
        "dcore_store_appid": "S2018032910007438"
    }
},
{
    "_id": "4079094118134743041",
    "name": "徐州苏宁店",
    "extend_code": {
        "comm_shop_id": "67dd4fe552904b49b3f28a4274f02dc8",
        "alipay_id": "2018051500077000000051727409",
        "us_id": "45154",
        "comm_code": "301003400004545",
        "upcard_terminal": "51601363",
        "upcard_mer_id": "102516089990267",
        "ex_cost_center_code": "1200845154",
        "dcore_store_appid": "S2018040910007462"
    }
},
{
    "_id": "4079127921389006849",
    "name": "郑州东站店",
    "extend_code": {
        "comm_shop_id": "174b38d2458c4bba95585db65189dd6d",
        "us_id": "45151",
        "comm_code": "301003400004468",
        "upcard_terminal": "39601449",
        "upcard_mer_id": "102396089990063",
        "ex_cost_center_code": "1200845151",
        "dcore_store_appid": "S2018040910007463"
    }
},
{
    "_id": "4079163645891452929",
    "name": "南昌西湖万达店",
    "extend_code": {
        "comm_shop_id": "7d8149a9fded4b6c99bfd9435cac43ee",
        "alipay_id": "2018051500077000000051631031",
        "us_id": "45153",
        "comm_code": "301003400004568",
        "upcard_terminal": "79191389",
        "upcard_mer_id": "102791089990672",
        "ex_cost_center_code": "1200845153",
        "dcore_store_appid": "S2018040910007465"
    }
},
{
    "_id": "4080945562105540609",
    "name": "洛阳建业凯旋店",
    "extend_code": {
        "comm_shop_id": "213d04e748054363b2834e0949df9750",
        "alipay_id": "2018051500077000000051629688",
        "us_id": "45155",
        "comm_code": "301003400004484",
        "upcard_terminal": "37901855",
        "upcard_mer_id": "102379089990507",
        "ex_cost_center_code": "1200845155",
        "dcore_store_appid": "S2018041110007497"
    }
},
{
    "_id": "4083475493077458945",
    "name": "大同百盛店",
    "extend_code": {
        "comm_shop_id": "8564eb9e9b024954aeb6667277b95f4d",
        "alipay_id": "2018070200077000000058090637",
        "us_id": "45161",
        "comm_code": "301003400004578",
        "upcard_terminal": "35200232",
        "upcard_mer_id": "102352089990049",
        "ex_cost_center_code": "1200845161",
        "dcore_store_appid": "S2018041610007624"
    }
},
{
    "_id": "4085973522101977089",
    "name": "瑞安吾悦店",
    "extend_code": {
        "comm_shop_id": "73b0916329eb4ed0acb693c21d9909a3",
        "alipay_id": "2018081500077000000061409413",
        "us_id": "45169",
        "comm_code": "301003400004554",
        "upcard_terminal": "57701270",
        "upcard_mer_id": "102577089990232",
        "ex_cost_center_code": "1200845169",
        "dcore_store_appid": "S2018042310007654"
    }
},
{
    "_id": "4085980156278366209",
    "name": "上海曹路家乐福店",
    "extend_code": {
        "comm_shop_id": "30b312af5cb14b21b17db5dff2f02e67",
        "alipay_id": "2018071000077000000058974770",
        "us_id": "45167",
        "comm_code": "301003400004495",
        "upcard_terminal": "02139643",
        "upcard_mer_id": "102210089998036",
        "ex_cost_center_code": "1200845167",
        "dcore_store_appid": "S2018042310007655"
    }
},
{
    "_id": "4086423399983869953",
    "name": "溧阳万达店",
    "extend_code": {
        "comm_shop_id": "1e17f000ab014d1aa2e214b4bfcd42ee",
        "alipay_id": "2018081500077000000061414820",
        "us_id": "45177",
        "comm_code": "301003400004479",
        "upcard_terminal": "51901534",
        "upcard_mer_id": "102519089990215",
        "ex_cost_center_code": "1200845177",
        "dcore_store_appid": "S2018042610007686"
    }
},
{
    "_id": "4087425920756776961",
    "name": "贵阳观山湖万达店",
    "extend_code": {
        "comm_shop_id": "2330ae456ac84fbf992bc62c18acf6d6",
        "alipay_id": "2018071000077000000058981444",
        "us_id": "45170",
        "comm_code": "301003400000077",
        "upcard_terminal": "85101969",
        "upcard_mer_id": "102851089990541",
        "ex_cost_center_code": "1200045170",
        "dcore_store_appid": "S2018042810007698"
    }
},
{
    "_id": "4087873879549825025",
    "name": "南京万达茂店",
    "extend_code": {
        "comm_shop_id": "87182bdb70b84952af932278c08fc82c",
        "alipay_id": "2018071000077000000058972962",
        "us_id": "45187",
        "comm_code": "301003400004582",
        "upcard_terminal": "02581210",
        "upcard_mer_id": "102250089993563",
        "ex_cost_center_code": "1200845187",
        "dcore_store_appid": "S2018050310007754"
    }
},
{
    "_id": "4089334362867511297",
    "name": "扬州京华城一店",
    "extend_code": {
        "comm_shop_id": "077c1349233a4d159679332642dd5ec3",
        "alipay_id": "2018072700077000000060272404",
        "us_id": "45176",
        "comm_code": "301003400004451",
        "upcard_terminal": "51401433",
        "upcard_mer_id": "102514089990318",
        "ex_cost_center_code": "1200845176",
        "dcore_store_appid": "S2018050310007755"
    }
},
{
    "_id": "4091471121304784897",
    "name": "黄冈万达店",
    "extend_code": {
        "comm_shop_id": "2a1910a92476469db17c384437eb36fb",
        "alipay_id": "2018071000077000000059037186",
        "us_id": "45189",
        "comm_code": "301003400004489",
        "upcard_terminal": "71300240",
        "upcard_mer_id": "102713089990039",
        "ex_cost_center_code": "1200845189",
        "dcore_store_appid": "S2018051610007893"
    }
},
{
    "_id": "4091478386262851585",
    "name": "晋城兰花城店",
    "extend_code": {
        "comm_shop_id": "1a605f2a9d5e46348fc765a1c6211f8e",
        "alipay_id": "2018071000077000000058981449",
        "us_id": "45191",
        "comm_code": "301003400004473",
        "upcard_terminal": "35600037",
        "upcard_mer_id": "102356089990022",
        "ex_cost_center_code": "1200845191",
        "dcore_store_appid": "S2018051610007894"
    }
},
{
    "_id": "4092169478572806145",
    "name": "长沙金茂店",
    "extend_code": {
        "comm_shop_id": "fc1b61839af84cf7a5084b0bc6965115",
        "alipay_id": "2018071000077000000058978334",
        "us_id": "45197",
        "comm_code": "301003400004691",
        "upcard_terminal": "73117588",
        "upcard_mer_id": "102731089992069",
        "ex_cost_center_code": "1200845197",
        "dcore_store_appid": "S2018051610007895"
    }
},
{
    "_id": "4092487770278039553",
    "name": "烟台莱山佳世客店",
    "extend_code": {
        "comm_shop_id": "cb5e67951f8648c4874bdd9a0fc5a7fe",
        "alipay_id": "2018071000077000000059001134",
        "us_id": "45196",
        "comm_code": "301003400004647",
        "upcard_terminal": "53500394",
        "upcard_mer_id": "102535089990205",
        "ex_cost_center_code": "1200845196",
        "dcore_store_appid": "S2018051610007898"
    }
},
{
    "_id": "4096921136813035521",
    "name": "上海静安大融城店",
    "extend_code": {
        "comm_shop_id": "34012477ac3a45b3b6efc968feae8838",
        "alipay_id": "2018070200077000000058067107",
        "us_id": "45205",
        "comm_code": "301003400004500",
        "upcard_terminal": "02139750",
        "upcard_mer_id": "102210089998091",
        "ex_cost_center_code": "1200845205",
        "dcore_store_appid": "S2018052410007967"
    }
},
{
    "_id": "4096922839413235713",
    "name": "威海威高店",
    "extend_code": {
        "comm_shop_id": "6975bdcb8e5d473fbb173fa2f0c3095c",
        "ex_code": "",
        "alipay_id": "2018081500077000000061409414",
        "us_id": "45204",
        "comm_code": "301003400004546",
        "upcard_terminal": "63103746",
        "upcard_mer_id": "102631089992119",
        "ex_cost_center_code": "1200845204",
        "dcore_store_appid": "S2018052410007968"
    }
},
{
    "_id": "4096924580198096897",
    "name": "日照万象汇店",
    "extend_code": {
        "comm_shop_id": "fee45ad2d2374a24863bfad6a25d5eb8",
        "alipay_id": "2018070200077000000058103829",
        "us_id": "45203",
        "comm_code": "301003400004693",
        "upcard_terminal": "63300006",
        "upcard_mer_id": "102633089990021",
        "ex_cost_center_code": "1200845203",
        "dcore_store_appid": "S2018052410007965"
    }
},
{
    "_id": "4100137793345470465",
    "name": "苏州市阳澄湖服务区北区店",
    "extend_code": {
        "comm_shop_id": "c0c9c578fbcd4218a960323b417cf298",
        "alipay_id": "2018081500077000000061413731",
        "us_id": "45211",
        "comm_code": "301003400004639",
        "upcard_terminal": "51216815",
        "upcard_mer_id": "102512089993230",
        "ex_cost_center_code": "1200845211",
        "dcore_store_appid": "S2018060110007985"
    }
},
{
    "_id": "4102278598728146945",
    "name": "晋中奥特莱斯店",
    "extend_code": {
        "comm_shop_id": "b28c34b273054aab96037616982d54c8",
        "alipay_id": "2018101200077000000063543383",
        "us_id": "45208",
        "comm_code": "301003400004625",
        "upcard_terminal": "35400080",
        "upcard_mer_id": "102354089990031",
        "ex_cost_center_code": "1200845208",
        "dcore_store_appid": "S2018060710008014"
    }
},
{
    "_id": "4103732800994172929",
    "name": "商丘万达店",
    "extend_code": {
        "comm_shop_id": "90e1e5a409b348faa3a89a994f53979c",
        "alipay_id": "2018081500077000000061405827",
        "us_id": "45209",
        "comm_code": "301003400004591",
        "upcard_terminal": "37001398",
        "upcard_mer_id": "102370089990039",
        "ex_cost_center_code": "1200845209",
        "dcore_store_appid": "S2018061210008019"
    }
},
{
    "_id": "4110238706563067905",
    "name": "青岛丽达店",
    "extend_code": {
        "comm_shop_id": "cd54d6397a3741a5ab68257bdbf75450",
        "alipay_id": "2018081500077000000061409415",
        "us_id": "45195",
        "comm_code": "301003400004651",
        "upcard_terminal": "53206386",
        "upcard_mer_id": "102532089991494",
        "ex_cost_center_code": "1200845195",
        "dcore_store_appid": "S2018070510008048"
    }
},
{
    "_id": "4113975942941487105",
    "name": "合肥港汇店",
    "extend_code": {
        "comm_shop_id": "0c19a00fd6ae470899467e90b0a90984",
        "alipay_id": "2018082300077000000061873604",
        "us_id": "45243",
        "comm_code": "301003400004460",
        "upcard_terminal": "55131087",
        "upcard_mer_id": "102551089993813",
        "ex_cost_center_code": "1200845243",
        "dcore_store_appid": "S2018071010008058"
    }
},
{
    "_id": "4114237766378766337",
    "name": "泰兴广陵东服务区店",
    "extend_code": {
        "comm_shop_id": "d8b2df697db64a7ea9f251c610956bca",
        "us_id": "45247",
        "comm_code": "301003400004660",
        "upcard_terminal": "52300873",
        "upcard_mer_id": "102523089990124",
        "ex_cost_center_code": "1200845247",
        "dcore_store_appid": "S2018071010008059"
    }
},
{
    "_id": "4114239187195158529",
    "name": "泰兴广陵西服务区店",
    "extend_code": {
        "comm_shop_id": "8b186bfcfb6644c19b7bb27cf6b38ac6",
        "us_id": "45246",
        "comm_code": "301003400004586",
        "upcard_terminal": "52300874",
        "upcard_mer_id": "102523089990125",
        "ex_cost_center_code": "1200845246",
        "dcore_store_appid": "S2018081310008105"
    }
},
{
    "_id": "4114581614087012353",
    "name": "南通永旺店",
    "extend_code": {
        "comm_shop_id": "8acc29019bbc4f9eaf757b7f6cb7b0df",
        "alipay_id": "2018080600077000000060849571",
        "us_id": "45249",
        "comm_code": "301003400004585",
        "upcard_terminal": "51301300",
        "upcard_mer_id": "102513089990242",
        "ex_cost_center_code": "1200845249",
        "dcore_store_appid": "S2018071110008062"
    }
},
{
    "_id": "4114641263291166721",
    "name": "上海五角场合生汇店",
    "extend_code": {
        "comm_shop_id": "f6c6c484fad641da955788722573fa5e",
        "alipay_id": "2019032700077000000073564862",
        "us_id": "45248",
        "comm_code": "301003400000445",
        "upcard_terminal": "02111253",
        "upcard_mer_id": "102210089998155",
        "ex_cost_center_code": "1200045248",
        "dcore_store_appid": "S2018071310008074"
    }
},
{
    "_id": "4115405479240224769",
    "name": "延安治平凤凰城店",
    "extend_code": {
        "comm_shop_id": "9aca77a1916f4c6eb677967d5c1cd478",
        "alipay_id": "2018081500077000000061413745",
        "us_id": "45233",
        "comm_code": "301003400000285",
        "upcard_terminal": "91100202",
        "upcard_mer_id": "102911089990027",
        "ex_cost_center_code": "1200045233",
        "dcore_store_appid": "S2018071910008081"
    }
},
{
    "_id": "4115407112690946049",
    "name": "杭州萧山机场店",
    "extend_code": {
        "comm_shop_id": "e500a5066dd5498299db6e4147b7752f",
        "alipay_id": "2019011400077000000069533575",
        "us_id": "45210",
        "comm_code": "301003400004671",
        "upcard_terminal": "57112492",
        "upcard_mer_id": "102571089992939",
        "ex_cost_center_code": "1200845210",
        "dcore_store_appid": "S2018071310008075"
    }
},
{
    "_id": "4116413591081840641",
    "name": "深圳宝安机场店",
    "extend_code": {
        "comm_shop_id": "adbdcd0f97804898b4e6449849b41b9e",
        "alipay_id": "2019010200077000000069210530",
        "us_id": "45234",
        "comm_code": "301003400004738",
        "upcard_terminal": "75519144",
        "upcard_mer_id": "102755089993786",
        "ex_cost_center_code": "1200045234",
        "dcore_store_appid": "S2018071810008077"
    }
},
{
    "_id": "4116830373533507585",
    "name": "太原万象城店",
    "extend_code": {
        "comm_shop_id": "40a0fa1a4af44132919522aa846db5ff",
        "alipay_id": "2018092700077000000063170418",
        "us_id": "45236",
        "comm_code": "301003400004510",
        "upcard_terminal": "35104472",
        "upcard_mer_id": "102351056991908",
        "ex_cost_center_code": "1200845236",
        "dcore_store_appid": "S2018071810008078"
    }
},
{
    "_id": "4116831979571179521",
    "name": "太原富力广场店",
    "extend_code": {
        "comm_shop_id": "f87159334248402faa2dde96106624c3",
        "alipay_id": "2018082300077000000061877007",
        "us_id": "45258",
        "comm_code": "301003400004687",
        "upcard_terminal": "35104473",
        "upcard_mer_id": "103351056991909",
        "ex_cost_center_code": "1200845258",
        "dcore_store_appid": "S2018071810008079"
    }
},
{
    "_id": "4117919376773586945",
    "name": "扬州仪征宝能店",
    "extend_code": {
        "comm_shop_id": "f52fd7d599f44f458f89081fb053b4d5",
        "alipay_id": "2018111500077000000065937133",
        "us_id": "45256",
        "comm_code": "301003400004684",
        "upcard_terminal": "51401512",
        "upcard_mer_id": "102514089990326",
        "ex_cost_center_code": "1200845256",
        "dcore_store_appid": "S2018072410008085"
    }
},
{
    "_id": "4118992973852532737",
    "name": "湖州德清银河城店",
    "extend_code": {
        "comm_shop_id": "c64ef9b5dd3f49f8893d12132588230f",
        "alipay_id": "2018110200077000000065441004",
        "us_id": "45242",
        "comm_code": "301003400004644",
        "upcard_terminal": "57297260",
        "upcard_mer_id": "102572089990081",
        "ex_cost_center_code": "1200845242",
        "dcore_store_appid": "S2018072410008087"
    }
},
{
    "_id": "4118997338543366145",
    "name": "九江万达店",
    "extend_code": {
        "comm_shop_id": "fc50007df0c2437faf20e5ccf16dfbb3",
        "alipay_id": "2018081500077000000061414821",
        "us_id": "45259",
        "comm_code": "301003400004692",
        "upcard_terminal": "79200762",
        "upcard_mer_id": "102792089990177",
        "ex_cost_center_code": "1200845259",
        "dcore_store_appid": "S2018072410008086"
    }
},
{
    "_id": "4121533209572458497",
    "name": "长治万达店",
    "extend_code": {
        "comm_shop_id": "ad9fae892de64ffe9cb3138079962634",
        "alipay_id": "2018092700077000000063133400",
        "us_id": "45273",
        "comm_code": "301003400004619",
        "upcard_terminal": "35500056",
        "upcard_mer_id": "102355089990028",
        "ex_cost_center_code": "1200845273",
        "dcore_store_appid": "S2018080210008096"
    }
},
{
    "_id": "4122677878704488449",
    "name": "郑州熙地港店",
    "extend_code": {
        "comm_shop_id": "2d99bb9722644d3880625a78b86a5da2",
        "alipay_id": "2019011400077000000069539833",
        "us_id": "45265",
        "comm_code": "301003400004492",
        "upcard_terminal": "37113912",
        "upcard_mer_id": "102371089992855",
        "ex_cost_center_code": "1200845265",
        "dcore_store_appid": "S2018080210008098"
    }
},
{
    "_id": "4122679299960795137",
    "name": "临汾生龙国际店",
    "extend_code": {
        "comm_shop_id": "12a9a3e6d5254ff5b9167507ed24504b",
        "alipay_id": "2018092700077000000063171913",
        "us_id": "45280",
        "comm_code": "301003400004464",
        "upcard_terminal": "35700764",
        "upcard_mer_id": "102357089990056",
        "ex_cost_center_code": "1200845280",
        "dcore_store_appid": "S2018080210008097"
    }
},
{
    "_id": "4124764025048219649",
    "name": "平湖吾悦店",
    "extend_code": {
        "comm_shop_id": "b71b2d25b8f04295ab2f8cd8332b0d12",
        "alipay_id": "2018110200077000000065441031",
        "us_id": "45281",
        "comm_code": "301003400004627",
        "upcard_terminal": "57304056",
        "upcard_mer_id": "102573089990195",
        "ex_cost_center_code": "1200845281",
        "dcore_store_appid": "S2018081310008106"
    }
},
{
    "_id": "4125152715792138241",
    "name": "郑州杉杉奥特莱斯店",
    "extend_code": {
        "comm_shop_id": "9f417f868b51479eb4139dfaaa653538",
        "us_id": "45261",
        "comm_code": "301003400004603",
        "upcard_terminal": "37113965",
        "upcard_mer_id": "102371089992858",
        "ex_cost_center_code": "1200845261",
        "dcore_store_appid": "S2018081310008108"
    }
},
{
    "_id": "4126601657147342849",
    "name": "郑州绿地新都会店",
    "extend_code": {
        "comm_shop_id": "c9c92223279f4197b5338eb557d38026",
        "alipay_id": "2018091300077000000062649175",
        "us_id": "45293",
        "comm_code": "301003400004646",
        "upcard_terminal": "37113966",
        "upcard_mer_id": "102371089992859",
        "ex_cost_center_code": "1200845293",
        "dcore_store_appid": "S2018081310008109"
    }
},
{
    "_id": "4126605730021203969",
    "name": "阜阳百太星马国际店",
    "extend_code": {
        "comm_shop_id": "6ebf5b72099844fcbbd7ee87b93e3042",
        "alipay_id": "2019012100077000000069848970",
        "us_id": "45266",
        "comm_code": "301003400004727",
        "upcard_terminal": "55801941",
        "upcard_mer_id": "102558089990688",
        "ex_cost_center_code": "1200845266",
        "dcore_store_appid": "S2018092910008214"
    }
},
{
    "_id": "4126611687771553793",
    "name": "湖州长兴九汇城店",
    "extend_code": {
        "comm_shop_id": "a50c33ad5ced4e929c33fcfe686810e0",
        "alipay_id": "2018110200077000000065439779",
        "us_id": "45285",
        "comm_code": "301003400004610",
        "upcard_terminal": "57297264",
        "upcard_mer_id": "102572089990083",
        "ex_cost_center_code": "1200845285",
        "dcore_store_appid": "S2018081310008110"
    }
},
{
    "_id": "4129103498775916545",
    "name": "郑州汇艺店",
    "extend_code": {
        "comm_shop_id": "af3e7313fb2a4986803d27891c11bcfb",
        "alipay_id": "2018092700077000000063168539",
        "us_id": "45299",
        "comm_code": "301003400004622",
        "upcard_terminal": "37113913",
        "upcard_mer_id": "102371089992856",
        "ex_cost_center_code": "1200845299",
        "dcore_store_appid": "S2018082010008150"
    }
},
{
    "_id": "4129104916787679233",
    "name": "淮北万达店",
    "extend_code": {
        "comm_shop_id": "d7cdede3d76e4f2a942a128659bd978f",
        "alipay_id": "2019041000077000000075866478",
        "us_id": "45300",
        "comm_code": "301003400004658",
        "upcard_terminal": "56100891",
        "upcard_mer_id": "102561089990247",
        "ex_cost_center_code": "1200845300",
        "dcore_store_appid": "S2018082010008149"
    }
},
{
    "_id": "4132033900400844801",
    "name": "南京溧水万达店",
    "extend_code": {
        "comm_shop_id": "517d7a869e87460480495f9e484362b2",
        "alipay_id": "2018101900077000000064357079",
        "us_id": "45308",
        "comm_code": "301003400004525",
        "upcard_terminal": "02582214",
        "upcard_mer_id": "102250089993824",
        "ex_cost_center_code": "1200845308",
        "dcore_store_appid": "S2018090310008173"
    }
},
{
    "_id": "4132035420274376705",
    "name": "南通万象城店",
    "extend_code": {
        "comm_shop_id": "f1e550f367b04354a9d59dcd2458a8c0",
        "alipay_id": "2018092700077000000063170419",
        "us_id": "45307",
        "comm_code": "301003400004680",
        "upcard_terminal": "51301420",
        "upcard_mer_id": "102513089990253",
        "ex_cost_center_code": "1200845307",
        "dcore_store_appid": "S2018090310008175"
    }
},
{
    "_id": "4133151762527371265",
    "name": "常州龙城天街店",
    "extend_code": {
        "comm_shop_id": "ee0d20a5199740f0a94126a4457433e5",
        "alipay_id": "2018122600077000000068543282",
        "us_id": "45312",
        "comm_code": "301003400004678",
        "upcard_terminal": "51902019",
        "upcard_mer_id": "102519089990226",
        "ex_cost_center_code": "1200845312",
        "dcore_store_appid": "S2018090310008176"
    }
},
{
    "_id": "4134204740237754369",
    "name": "宁波慈城东服务区店",
    "extend_code": {
        "comm_shop_id": "99db7ae4921a4f19887caf2fd2c52bfc",
        "us_id": "45322",
        "comm_code": "301003400004598",
        "upcard_terminal": "57404550",
        "upcard_mer_id": "102574089990675",
        "ex_cost_center_code": "1200845322",
        "dcore_store_appid": "S2018090610008186"
    }
},
{
    "_id": "4134210349346463745",
    "name": "宁波慈城西服务区店",
    "extend_code": {
        "comm_shop_id": "11818412aa594e09ad0dc90087b4c6b6",
        "us_id": "45321",
        "comm_code": "301003400004463",
        "upcard_terminal": "57404551",
        "upcard_mer_id": "102574089990676",
        "ex_cost_center_code": "1200845321",
        "dcore_store_appid": "S2018092910008211"
    }
},
{
    "_id": "4134215215138557953",
    "name": "上海奉贤宝龙店",
    "extend_code": {
        "comm_shop_id": "f7d3f546082a48c8a2e33ccccdeec5e4",
        "alipay_id": "2018092900077000000063208125",
        "us_id": "45316",
        "comm_code": "301003400004686",
        "upcard_terminal": "02111626",
        "upcard_mer_id": "102210089998307",
        "ex_cost_center_code": "1200845316",
        "dcore_store_appid": "S2018092910008210"
    }
},
{
    "_id": "4134635867201208321",
    "name": "扬州吾悦店",
    "extend_code": {
        "comm_shop_id": "07d2020a4fb243f98c15bc45dcabdeb5",
        "alipay_id": "2018111500077000000065942777",
        "us_id": "45298",
        "comm_code": "301003400004453",
        "upcard_terminal": "51401645",
        "upcard_mer_id": "102514089990335",
        "ex_cost_center_code": "1200845298",
        "dcore_store_appid": "S2018090610008188"
    }
},
{
    "_id": "4139286224734511105",
    "name": "达州升华广场店",
    "extend_code": {
        "comm_shop_id": "7fab9cbf57394feb9be0fddb7957861b",
        "alipay_id": "2018111500077000000065947582",
        "us_id": "45344",
        "comm_code": "301003400004570",
        "upcard_terminal": "81800321",
        "upcard_mer_id": "102818089990043",
        "ex_cost_center_code": "1200845344",
        "dcore_store_appid": "S2018092110008200"
    }
},
{
    "_id": "4139315531955630081",
    "name": "南昌酷加天虹店",
    "extend_code": {
        "comm_shop_id": "835fccebfd3847458da673a731c0aab3",
        "alipay_id": "2018101600077000000063658138",
        "us_id": "45332",
        "comm_code": "301003400004574",
        "upcard_terminal": "79191472",
        "upcard_mer_id": "102791089990719",
        "ex_cost_center_code": "1200845332",
        "dcore_store_appid": "S2018092110008199"
    }
},
{
    "_id": "4140016321382154241",
    "name": "驻马店爱家店",
    "extend_code": {
        "comm_shop_id": "afe1c69689d04b58b32016e01cd97a8b",
        "alipay_id": "2018110200077000000065438423",
        "us_id": "45343",
        "comm_code": "301003400004739",
        "upcard_terminal": "39601585",
        "upcard_mer_id": "102396089990065",
        "ex_cost_center_code": "1200845343",
        "dcore_store_appid": "S2018092110008201"
    }
},
{
    "_id": "4140421607509770241",
    "name": "忻州开来欣悦店",
    "extend_code": {
        "comm_shop_id": "8ebf87feaa2c41778eacd4a554095793",
        "alipay_id": "2018110200077000000065439778",
        "us_id": "45348",
        "comm_code": "301003400004589",
        "upcard_terminal": "35000073",
        "upcard_mer_id": "102350089990059",
        "ex_cost_center_code": "1200845348",
        "dcore_store_appid": "S2018092110008202"
    }
},
{
    "_id": "4140789178306994177",
    "name": "菏泽万达店",
    "extend_code": {
        "comm_shop_id": "33979b9dceff4ebcb54b25dc8931d176",
        "us_id": "45325",
        "comm_code": "301003400004499",
        "upcard_terminal": "53000575",
        "upcard_mer_id": "102530089990234",
        "ex_cost_center_code": "1200845325",
        "dcore_store_appid": "S2018092110008204"
    }
},
{
    "_id": "4140790602041192449",
    "name": "济南万虹银座店",
    "extend_code": {
        "comm_shop_id": "fa1186d8ca294f37841174105810590e",
        "alipay_id": "2018121900077000000068322823",
        "us_id": "45345",
        "comm_code": "301003400004690",
        "upcard_terminal": "53101612",
        "upcard_mer_id": "102531089990641",
        "ex_cost_center_code": "1200845345",
        "dcore_store_appid": "S2018092110008203"
    }
},
{
    "_id": "4144004231356600321",
    "name": "巢湖万达店",
    "extend_code": {
        "comm_shop_id": "d3146bb3e0474ecf8e25e84de9658cdd",
        "alipay_id": "2018110800077000000065641240",
        "us_id": "45365",
        "comm_code": "301003400004656",
        "upcard_terminal": "55131437",
        "upcard_mer_id": "102551089993838",
        "ex_cost_center_code": "1200845365",
        "dcore_store_appid": "S2018092910008217"
    }
},
{
    "_id": "4146845449628708865",
    "name": "郑州正弘城店",
    "extend_code": {
        "comm_shop_id": "345e0cfab09a4cb39487e9e6713ddb56",
        "alipay_id": "2019011000077000000069426365",
        "us_id": "45372",
        "comm_code": "301003400004501",
        "upcard_terminal": "37114056",
        "upcard_mer_id": "102371089992867",
        "ex_cost_center_code": "1200845372",
        "dcore_store_appid": "S2018100910008221"
    }
},
{
    "_id": "4150176012169773057",
    "name": "抚州硕果时代店",
    "extend_code": {
        "comm_shop_id": "28b9b4d62b1e43c6a3098836d544ddbc",
        "alipay_id": "2018122600077000000068537438",
        "us_id": "45284",
        "comm_code": "301003400004707",
        "upcard_terminal": "79400205",
        "upcard_mer_id": "102794089990050",
        "ex_cost_center_code": "1200845284",
        "dcore_store_appid": "S2018101710008256"
    }
},
{
    "_id": "4150196479186903041",
    "name": "金华永盛店",
    "extend_code": {
        "comm_shop_id": "3a73519256ac449cbb658445a987e526",
        "alipay_id": "2018110200077000000065430167",
        "us_id": "45364",
        "comm_code": "301003400004504",
        "upcard_terminal": "57990795",
        "upcard_mer_id": "102579089990277",
        "ex_cost_center_code": "1200845364",
        "dcore_store_appid": "S2018101710008258"
    }
},
{
    "_id": "4152649549586010113",
    "name": "合肥大洋百货店",
    "extend_code": {
        "comm_shop_id": "51cfa4f050f04be292900bba0300ae97",
        "alipay_id": "2018111500077000000066207550",
        "us_id": "45411",
        "comm_code": "301003400004526",
        "upcard_terminal": "55131445",
        "upcard_mer_id": "102551089993842",
        "ex_cost_center_code": "1200845411",
        "dcore_store_appid": "S2018102410008288"
    }
},
{
    "_id": "4153006374760136705",
    "name": "上海白玉兰广场店",
    "extend_code": {
        "comm_shop_id": "109104b27bc0448788f26770a5e30d58",
        "alipay_id": "2020060800077000000095144926",
        "us_id": "45405",
        "comm_code": "301003400004701",
        "upcard_terminal": "02111816",
        "upcard_mer_id": "102210089998402",
        "ex_cost_center_code": "1200045405"
    }
},
{
    "_id": "4153471261688762369",
    "name": "合肥滨湖世纪金源店",
    "extend_code": {
        "comm_shop_id": "ed28728be88f4024bc4f04a447ac010b",
        "alipay_id": "2019010200077000000069206207",
        "us_id": "45404",
        "comm_code": "301003400004421",
        "upcard_terminal": "55131446",
        "upcard_mer_id": "102551089993843",
        "ex_cost_center_code": "1200845404"
    }
},
{
    "_id": "4154464939744419841",
    "name": "合肥滨湖银泰店",
    "extend_code": {
        "comm_shop_id": "f060b01bb6df4c90a189409bc11ec957",
        "alipay_id": "2019010200077000000069209188",
        "us_id": "45410",
        "comm_code": "301003400004422",
        "upcard_terminal": "55131447",
        "upcard_mer_id": "102551089993844",
        "ex_cost_center_code": "1200845410"
    }
},
{
    "_id": "4154471226036903937",
    "name": "许昌万达店",
    "extend_code": {
        "comm_shop_id": "0a118489d8c34d3fa54f59fe0217f36f",
        "alipay_id": "2019041000077000000075864242",
        "us_id": "45413",
        "comm_code": "301003400004696",
        "upcard_terminal": "37401780",
        "upcard_mer_id": "102374089990077",
        "ex_cost_center_code": "1200845413"
    }
},
{
    "_id": "4154503143297040385",
    "name": "大同万达店",
    "extend_code": {
        "comm_shop_id": "765cc7c88ca548798ebfadd92b9dbcbb",
        "alipay_id": "2018112800077000000067691992",
        "us_id": "45418",
        "comm_code": "301003400004557",
        "upcard_terminal": "35200233",
        "upcard_mer_id": "102352089990050",
        "ex_cost_center_code": "1200845418"
    }
},
{
    "_id": "4154504983742676993",
    "name": "忻州开来欣悦影院店",
    "extend_code": {
        "comm_shop_id": "cf363c47c2f043d688a5ef2255775570",
        "us_id": "45373",
        "comm_code": "301003400004401",
        "upcard_terminal": "35000074",
        "upcard_mer_id": "102350089990060",
        "ex_cost_center_code": "1200845373"
    }
},
{
    "_id": "4159888449808629761",
    "name": "张家港万达店",
    "extend_code": {
        "comm_shop_id": "5387787ac56b4173b025bbb17436ca07",
        "alipay_id": "2019010200077000000069206208",
        "us_id": "45441",
        "comm_code": "301003400004722",
        "upcard_terminal": "51217436",
        "upcard_mer_id": "102512089993305",
        "ex_cost_center_code": "1200845441"
    }
},
{
    "_id": "4160976205647978497",
    "name": "上海金山万达店",
    "extend_code": {
        "comm_shop_id": "fb4b7c222e2e416eaf40de96a15c1251",
        "alipay_id": "2018121900077000000068328904",
        "us_id": "45445",
        "comm_code": "301003400004745",
        "upcard_terminal": "02111968",
        "upcard_mer_id": "102210089998439",
        "ex_cost_center_code": "1200845445"
    }
},
{
    "_id": "4160978965889052673",
    "name": "临沂泰盛店",
    "extend_code": {
        "comm_shop_id": "20937489ecca4b8188da24c234f3fffb",
        "alipay_id": "2018122600077000000068530306",
        "us_id": "45403",
        "comm_code": "301003400004704",
        "upcard_terminal": "53952341",
        "upcard_mer_id": "102539089990430",
        "ex_cost_center_code": "1200845403"
    }
},
{
    "_id": "4162431335909015553",
    "name": "青岛城阳万象汇店",
    "extend_code": {
        "comm_shop_id": "9b28cbe0f8004fb08be9bc4a23f70c65",
        "alipay_id": "2018122100077000000068395417",
        "us_id": "45444",
        "comm_code": "301003400004735",
        "upcard_terminal": "53206420",
        "upcard_mer_id": "102532089991518",
        "ex_cost_center_code": "1200845444"
    }
},
{
    "_id": "4165028107158962177",
    "name": "舟山定海凯虹店",
    "extend_code": {
        "comm_shop_id": "459f904ac2b64ec5905715c07290a4a1",
        "alipay_id": "2019011400077000000069537808",
        "us_id": "45420",
        "comm_code": "301003400004718",
        "upcard_terminal": "58000400",
        "upcard_mer_id": "102580089990051",
        "ex_cost_center_code": "1200045420"
    }
},
{
    "_id": "4165327831024205825",
    "name": "郑州局外太格茂店",
    "extend_code": {
        "comm_shop_id": "8193d54a29f7461999a5ca6f038386b5",
        "alipay_id": "2018122600077000000068530305",
        "us_id": "45425",
        "comm_code": "301003400004730",
        "upcard_terminal": "37114114",
        "upcard_mer_id": "102371089992872",
        "ex_cost_center_code": "1200845425"
    }
},
{
    "_id": "4165681062380634113",
    "name": "启东吾悦店",
    "extend_code": {
        "comm_shop_id": "989f5a492a45409b87fee6b72d793444",
        "alipay_id": "2018122600077000000068537439",
        "us_id": "45467",
        "comm_code": "301003400004734",
        "upcard_terminal": "51301431",
        "upcard_mer_id": "102513089990257",
        "ex_cost_center_code": "1200845467"
    }
},
{
    "_id": "4165729105024806913",
    "name": "运城万达店",
    "extend_code": {
        "comm_shop_id": "1634f8a81f1c47ca9abfee092de4c783",
        "alipay_id": "2019010200077000000069217892",
        "us_id": "45448",
        "comm_code": "301003400004400",
        "upcard_terminal": "35900133",
        "upcard_mer_id": "102359089990034",
        "ex_cost_center_code": "1200845448"
    }
},
{
    "_id": "4165730245892542465",
    "name": "岳阳步步高星都汇店",
    "extend_code": {
        "comm_shop_id": "24512b3854b14348876ffd4f2fd7110e",
        "alipay_id": "2019010200077000000069202176",
        "us_id": "45477",
        "comm_code": "301003400004705",
        "upcard_terminal": "73002209",
        "upcard_mer_id": "102730089990077",
        "ex_cost_center_code": "1200845477"
    }
},
{
    "_id": "4165742820760801281",
    "name": "厦门阿罗海店",
    "extend_code": {
        "comm_shop_id": "1036b5373d2343ca9f0077884c26c4a2",
        "alipay_id": "2018122600077000000068570705",
        "us_id": "45476",
        "comm_code": "301003400004700",
        "upcard_terminal": "59205666",
        "upcard_mer_id": "102592089990666",
        "ex_cost_center_code": "1200845476"
    }
},
{
    "_id": "4166039404921270273",
    "name": "扬州邗江万达店",
    "extend_code": {
        "comm_shop_id": "7a07dde0032e4f6bbe250d473ddcef52",
        "alipay_id": "2019010200077000000069214350",
        "us_id": "45472",
        "comm_code": "301003400004728",
        "upcard_terminal": "51401651",
        "upcard_mer_id": "102514089990338",
        "ex_cost_center_code": "1200845472"
    }
},
{
    "_id": "4167250700683948033",
    "name": "兰州中心店",
    "extend_code": {
        "comm_shop_id": "6bc8cc82b62941dfaefcc0b01ef7c78a",
        "alipay_id": "2019011400077000000069539785",
        "us_id": "45419",
        "comm_code": "301003400004726",
        "upcard_terminal": "93101216",
        "upcard_mer_id": "102931089990228",
        "ex_cost_center_code": "1200045419"
    }
},
{
    "_id": "4167588597452136449",
    "name": "济南和谐店",
    "extend_code": {
        "comm_shop_id": "3aba265663224256b72d75831d96dff3",
        "alipay_id": "2019042500077000000076477421",
        "us_id": "45483",
        "comm_code": "301003400004713",
        "upcard_terminal": "53101617",
        "upcard_mer_id": "102531089990644",
        "ex_cost_center_code": "1200845483"
    }
},
{
    "_id": "4168303011669086209",
    "name": "泰州茂业店",
    "extend_code": {
        "comm_shop_id": "f712ac32e3f74a88bdd3db0a559162ea",
        "alipay_id": "2019010200077000000069216289",
        "us_id": "45474",
        "comm_code": "301003400004744",
        "upcard_terminal": "52300991",
        "upcard_mer_id": "102523089990137",
        "ex_cost_center_code": "1200845474"
    }
},
{
    "_id": "4168318777077153793",
    "name": "三门峡梦之城店",
    "extend_code": {
        "comm_shop_id": "0e236304fe5040e79c7c2f0eeb6d42ea",
        "alipay_id": "2019041100077000000076000442",
        "us_id": "45490",
        "comm_code": "301003400004699",
        "upcard_terminal": "39801126",
        "upcard_mer_id": "102398089990029",
        "ex_cost_center_code": "1200845490"
    }
},
{
    "_id": "4168612141303578625",
    "name": "常州环球港店",
    "extend_code": {
        "comm_shop_id": "0d89f531a9cd457a94e6b520dd92cca4",
        "alipay_id": "2019010200077000000069214351",
        "us_id": "45447",
        "comm_code": "301003400004698",
        "upcard_terminal": "51902029",
        "upcard_mer_id": "102519089990232",
        "ex_cost_center_code": "1200845447"
    }
},
{
    "_id": "4168703113316777985",
    "name": "扬州五彩世界店",
    "extend_code": {
        "comm_shop_id": "29b59fd6a04c4daf8874b65ea1445126",
        "alipay_id": "2018122600077000000068530304",
        "us_id": "45443",
        "comm_code": "301003400004708",
        "upcard_terminal": "51401690",
        "upcard_mer_id": "102514089990341",
        "ex_cost_center_code": "1200845443"
    }
},
{
    "_id": "4171169440537362433",
    "name": "洛阳正大店",
    "extend_code": {
        "comm_shop_id": "e5f70d0277d24e5aa8065372b8f81c4c",
        "alipay_id": "2019051000077000000077229176",
        "us_id": "45501",
        "comm_code": "301003400004743",
        "upcard_terminal": "37902016",
        "upcard_mer_id": "102379089990512",
        "ex_cost_center_code": "1200845501"
    }
},
{
    "_id": "4178789021156511745",
    "name": "信阳罗山华鼎城店",
    "extend_code": {
        "comm_shop_id": "e30b00513a85448689385240af345840",
        "alipay_id": "2019021500077000000070980530",
        "us_id": "45515",
        "comm_code": "301003400004742",
        "upcard_terminal": "37602996",
        "upcard_mer_id": "102376089990155",
        "ex_cost_center_code": "1200845515"
    }
},
{
    "_id": "4179865584621322241",
    "name": "苏州园区永旺梦乐城店",
    "extend_code": {
        "comm_shop_id": "3b9256bd551744feada97e52ff8a5ac2",
        "alipay_id": "2019021500077000000070960588",
        "us_id": "45518",
        "comm_code": "301003400004714",
        "upcard_terminal": "51217451",
        "upcard_mer_id": "102512089993313",
        "ex_cost_center_code": "1200845518"
    }
},
{
    "_id": "4180622686182809601",
    "name": "广东大槐东服务区店",
    "extend_code": {
        "comm_shop_id": "25e27720a5db4ee899fb1951f4aa621e",
        "us_id": "45517",
        "comm_code": "301003400004706",
        "upcard_terminal": "75090152",
        "upcard_mer_id": "102750089990064",
        "ex_cost_center_code": "1200045517"
    }
},
{
    "_id": "4180623942064128001",
    "name": "广东大槐西服务区店",
    "extend_code": {
        "comm_shop_id": "a0b6d6829fff4c51a2947f697e9478d2",
        "us_id": "45516",
        "comm_code": "301003400004831",
        "upcard_terminal": "75090153",
        "upcard_mer_id": "102750089990065",
        "ex_cost_center_code": "1200045516",
        "dianping_store_id": ""
    }
},
{
    "_id": "4198046181180911617",
    "name": "厦门湖里万达店",
    "extend_code": {
        "comm_shop_id": "a2481a3a22494022a2425896fb4002f3",
        "alipay_id": "2019122000077000000085656944",
        "us_id": "45549",
        "comm_code": "301003400004806",
        "upcard_terminal": "59204819",
        "upcard_mer_id": "102592058120541",
        "ex_cost_center_code": "1200845549"
    }
},
{
    "_id": "4202675140379578369",
    "name": "鄂州新亚太国际店",
    "extend_code": {
        "comm_shop_id": "f8622fec00224cedab85596e1d55c3eb",
        "us_id": "45529",
        "comm_code": "301003400004809",
        "upcard_terminal": "71150020",
        "upcard_mer_id": "102711089990013",
        "ex_cost_center_code": "1200845529"
    }
},
{
    "_id": "4207713452165398529",
    "name": "合肥万科店",
    "extend_code": {
        "comm_shop_id": "432b5dedfadb4a4da14d8fb4649c5ef5",
        "alipay_id": "2019051500077000000077394467",
        "us_id": "45556",
        "comm_code": "301003400004821",
        "upcard_terminal": "55131871",
        "upcard_mer_id": "102551089993874",
        "ex_cost_center_code": "1200845556"
    }
},
{
    "_id": "4208132735067004929",
    "name": "合肥肥东吾悦店",
    "extend_code": {
        "comm_shop_id": "eb09f8b90bea41f498f17307e00c52ab",
        "alipay_id": "2019101800077000000083653837",
        "us_id": "45563",
        "upcard_terminal": "55131872",
        "upcard_mer_id": "102551089993875",
        "ex_cost_center_code": "1200845563"
    }
},
{
    "_id": "4208926120304644097",
    "name": "新昌世贸店",
    "extend_code": {
        "comm_shop_id": "24c60da1e67f44a89c4bbd86d136d9d0",
        "alipay_id": "2019122500077000000085760325",
        "us_id": "45574",
        "comm_code": "301003400004816",
        "upcard_terminal": "57501409",
        "upcard_mer_id": "102575089990178",
        "ex_cost_center_code": "1200845574"
    }
},
{
    "_id": "4208931230917259265",
    "name": "苏州龙湖狮山天街店",
    "extend_code": {
        "comm_shop_id": "19cc8b7ad28741b1b9f5c97fb6872435",
        "alipay_id": "2019062700077000000079582295",
        "us_id": "45570",
        "comm_code": "301003400004813",
        "upcard_terminal": "51217487",
        "upcard_mer_id": "102512089993326",
        "ex_cost_center_code": "1200845570"
    }
},
{
    "_id": "4211050697250721793",
    "name": "青浦万达茂店",
    "extend_code": {
        "comm_shop_id": "88ce89b4670442a69be9694cdaea5f5f",
        "ex_code": "45569",
        "alipay_id": "2019101400077000000083562898",
        "us_id": "45569",
        "comm_code": "301003400004804",
        "upcard_terminal": "02199659",
        "upcard_mer_id": "102210089998554",
        "ex_cost_center_code": "1200045569"
    }
},
{
    "_id": "4213219671288963073",
    "name": "南京华采天地店",
    "extend_code": {
        "comm_shop_id": "49786b9de04c4609a1250c2b5983f9f8",
        "alipay_id": "2019111900077000000084756998",
        "us_id": "45581",
        "comm_code": "301003400004823",
        "upcard_terminal": "02583214",
        "upcard_mer_id": "102250089993932",
        "ex_cost_center_code": "1200845581"
    }
},
{
    "_id": "4214312227799261185",
    "name": "郑州木色店",
    "extend_code": {
        "comm_shop_id": "011b00773e9f48b28c9dc30e04552c00",
        "alipay_id": "2019061000077000000078946814",
        "us_id": "45588",
        "comm_code": "301003400004810",
        "upcard_terminal": "37114329",
        "upcard_mer_id": "102371089992891",
        "ex_cost_center_code": "1200845588"
    }
},
{
    "_id": "4216038488840843265",
    "name": "苏州吴江华润万象汇店",
    "extend_code": {
        "comm_shop_id": "0680c7a030a94c8b8b143b8c7bb0b27d",
        "alipay_id": "2019092600077000000082997803",
        "us_id": "45571",
        "comm_code": "301003400004811",
        "upcard_terminal": "51217498",
        "upcard_mer_id": "102512089993329",
        "ex_cost_center_code": "1200845571"
    }
},
{
    "_id": "4217975891788771329",
    "name": "杭州金沙印象城店",
    "extend_code": {
        "comm_shop_id": "7149d86affc24cb992070e0924c13972",
        "alipay_id": "2019122500077000000085765655",
        "us_id": "45591",
        "comm_code": "301003400004827",
        "upcard_terminal": "57113808",
        "upcard_mer_id": "102571089993135",
        "ex_cost_center_code": "1200845591"
    }
},
{
    "_id": "4220455818191876097",
    "name": "南昌王府井购物中心店",
    "extend_code": {
        "comm_shop_id": "f28cff788b534c4d94338b32c1df3685",
        "alipay_id": "2019070800077000000080192207",
        "us_id": "45593",
        "comm_code": "301003400004842",
        "upcard_terminal": "07950002",
        "upcard_mer_id": "102791058220742",
        "ex_cost_center_code": "1200845593"
    }
},
{
    "_id": "4223316429929222145",
    "name": "济南印象城店",
    "extend_code": {
        "comm_shop_id": "b4edd73222b949cdbfe6f7dd0f5a757b",
        "alipay_id": "2019062500077000000079456854",
        "us_id": "45600",
        "comm_code": "301003400004833",
        "upcard_terminal": "53101650",
        "upcard_mer_id": "102531089990654",
        "ex_cost_center_code": "1200845600"
    }
},
{
    "_id": "4224486993269325825",
    "name": "七宝万科店",
    "extend_code": {
        "comm_shop_id": "f03f26abd5bf4150a9255d10db43a709",
        "alipay_id": "2019101400077000000083561582",
        "us_id": "45604",
        "comm_code": "301003400004808",
        "upcard_terminal": "02108095",
        "upcard_mer_id": "102210058228754",
        "ex_cost_center_code": "1200045604"
    }
},
{
    "_id": "4224494527853682689",
    "name": "贵州凯里国贸店",
    "extend_code": {
        "comm_shop_id": "ab1a856f45d742d7939f891296627105",
        "alipay_id": "2019121900077000000085649238",
        "us_id": "45606",
        "upcard_terminal": "85590017",
        "upcard_mer_id": "102855089990017",
        "ex_cost_center_code": "1200045606"
    }
},
{
    "_id": "4225931375332503553",
    "name": "广州云门new park店",
    "extend_code": {
        "comm_shop_id": "6ae7a0d27c014bac985cb4f2beb36989",
        "alipay_id": "2019080800077000000081078204",
        "us_id": "45603",
        "comm_code": "301003400004826",
        "upcard_terminal": "02082894",
        "upcard_mer_id": "102200089991106",
        "ex_cost_center_code": "1200045603"
    }
},
{
    "_id": "4225933325107351553",
    "name": "海门龙信广场店",
    "extend_code": {
        "comm_shop_id": "15ea2aa2a3934a878753fe5e9bd26266",
        "alipay_id": "2019061700077000000079152148",
        "us_id": "45599",
        "comm_code": "301003400004812",
        "upcard_terminal": "51301450",
        "upcard_mer_id": "102513089990264",
        "ex_cost_center_code": "1200845599"
    }
},
{
    "_id": "4225934705314746369",
    "name": "洛阳泉舜店",
    "extend_code": {
        "comm_shop_id": "3a4fcdc2d3ac435da5fd05f1940ae088",
        "alipay_id": "2019061000077000000078945191",
        "us_id": "45611",
        "comm_code": "301003400004819",
        "upcard_terminal": "37902017",
        "upcard_mer_id": "102379089990513",
        "ex_cost_center_code": "1200845611"
    }
},
{
    "_id": "4228822079782264832",
    "name": "杭州余杭万达店",
    "extend_code": {
        "comm_shop_id": "c1c0427fde174a27bb5f243d5162cc8c",
        "alipay_id": "2019091600077000000082640203",
        "us_id": "45615",
        "upcard_terminal": "57114077",
        "upcard_mer_id": "102571089993146",
        "ex_cost_center_code": "1200845615"
    }
},
{
    "_id": "4230971390993371136",
    "name": "成都温江新尚天地店",
    "extend_code": {
        "comm_shop_id": "1eaa4bd0ceb94d819700e70a384daa14",
        "alipay_id": "2019071500077000000080410984",
        "us_id": "45629",
        "comm_code": "301003400004815",
        "upcard_terminal": "02829701",
        "upcard_mer_id": "102280089996659",
        "ex_cost_center_code": "1200845629"
    }
},
{
    "_id": "4230973845076447232",
    "name": "武汉人信汇店",
    "extend_code": {
        "comm_shop_id": "6a6516e63fa34459b5e6464e21a99d88",
        "alipay_id": "2019072200077000000080601687",
        "us_id": "45627",
        "comm_code": "301003400004824",
        "upcard_terminal": "02731111",
        "upcard_mer_id": "102270089995588",
        "ex_cost_center_code": "1200845627"
    }
},
{
    "_id": "4230976043718021120",
    "name": "滨海县海悦城店",
    "extend_code": {
        "comm_shop_id": "41d35065507444f081fd508b4e24509c",
        "us_id": "45614",
        "comm_code": "301003400004820",
        "upcard_terminal": "51501044",
        "upcard_mer_id": "102515089990487",
        "ex_cost_center_code": "1200845614"
    }
},
{
    "_id": "4233186903303655424",
    "name": "重庆来福士店",
    "extend_code": {
        "comm_shop_id": "d1e5ef83b0fe4b36b9f35a941f348391",
        "alipay_id": "2019091200077000000082537399",
        "us_id": "45605",
        "upcard_terminal": "02310097",
        "upcard_mer_id": "102230089992578",
        "ex_cost_center_code": "1200045605"
    }
},
{
    "_id": "4233433928653869056",
    "name": "常熟永旺梦乐城店",
    "extend_code": {
        "comm_shop_id": "1d64812f8616417bbe0311eb443eea53",
        "alipay_id": "2019070200077000000079858871",
        "us_id": "45613",
        "comm_code": "301003400004814",
        "upcard_terminal": "51217547",
        "upcard_mer_id": "102512089993340",
        "ex_cost_center_code": "1200845613"
    }
},
{
    "_id": "4233443346170777600",
    "name": "无锡融创茂店",
    "extend_code": {
        "comm_shop_id": "48f2ad7d5c464129b887d35516d518fd",
        "alipay_id": "2019062700077000000079578463",
        "us_id": "45630",
        "comm_code": "301003400004822",
        "upcard_terminal": "51004304",
        "upcard_mer_id": "102510089993828",
        "ex_cost_center_code": "1200845630"
    }
},
{
    "_id": "4235704935947177984",
    "name": "深圳金光华店",
    "extend_code": {
        "us_id": "45633",
        "upcard_mer_id": "102755089994224",
        "ex_cost_center_code": "1200045633",
        "comm_shop_id": "df117370e97247708e30491a795beac2",
        "upcard_terminal": "75520555"
    }
},
{
    "_id": "4238199221091569664",
    "name": "京东",
    "extend_code": {
        "us_id": "45406",
        "ex_cost_center_code": "1200045406",
        "comm_shop_id": "42"
    }
},
{
    "_id": "4238265115767476224",
    "name": "淮安吾悦广场店",
    "extend_code": {
        "comm_shop_id": "9feb5c06645c44bcb886ccb4b909dc20",
        "alipay_id": "2019091200077000000082536132",
        "us_id": "45635",
        "upcard_terminal": "51701460",
        "upcard_mer_id": "102517089990103",
        "ex_cost_center_code": "1200845635"
    }
},
{
    "_id": "4241179079203622912",
    "name": "宁波杭州湾世纪金源店",
    "extend_code": {
        "comm_shop_id": "1b9ca6ce5c5842a48ea16c7287ffad47",
        "alipay_id": "2019080200077000000080895579",
        "us_id": "45655",
        "upcard_terminal": "57404985",
        "upcard_mer_id": "102574089990707",
        "ex_cost_center_code": "1200845655"
    }
},
{
    "_id": "4241185397750632448",
    "name": "巩义德丰香榭里店",
    "extend_code": {
        "comm_shop_id": "867881346d864613b2befbdbc0f4b0d5",
        "alipay_id": "2020032400077000000092589564",
        "us_id": "45634",
        "upcard_terminal": "37114418",
        "upcard_mer_id": "102371089992921",
        "ex_cost_center_code": "1200845634"
    }
},
{
    "_id": "4243303888272236544",
    "name": "苏州新区永旺店",
    "extend_code": {
        "comm_shop_id": "8ec3fe99dd15435e94af6327a067c221",
        "alipay_id": "2019101400077000000083561589",
        "us_id": "45659",
        "upcard_terminal": "51217552",
        "upcard_mer_id": "102512089993346",
        "ex_cost_center_code": "1200845659"
    }
},
{
    "_id": "4244779874385936384",
    "name": "南通如东欧尚店",
    "extend_code": {
        "comm_shop_id": "d70a399743894c7698a84052efd5103b",
        "alipay_id": "2019103000077000000084080875",
        "us_id": "45645",
        "upcard_terminal": "51301455",
        "upcard_mer_id": "102513089990267",
        "ex_cost_center_code": "1200845645"
    }
},
{
    "_id": "4244786433493057536",
    "name": "徐州三胞国际广场店",
    "extend_code": {
        "comm_shop_id": "78fb94bb9ac94377b010652caa87f775",
        "alipay_id": "2019082900077000000081987572",
        "us_id": "45661",
        "upcard_terminal": "51601718",
        "upcard_mer_id": "102516089990292",
        "ex_cost_center_code": "1200845661"
    }
},
{
    "_id": "4246917558256549888",
    "name": "太原公园时代店",
    "extend_code": {
        "us_id": "45669",
        "upcard_mer_id": "102351089992037",
        "ex_cost_center_code": "1200845669",
        "comm_shop_id": "0d9ae84e5e50498da294231655cfc363",
        "upcard_terminal": "35104673"
    }
},
{
    "_id": "4248412012570230784",
    "name": "平湖南河头店",
    "extend_code": {
        "comm_shop_id": "8ebd653b9e4e4f54850af35cddffe842",
        "alipay_id": "2019123100077000000085897658",
        "us_id": "45668",
        "upcard_terminal": "57304200",
        "upcard_mer_id": "102573058220211",
        "ex_cost_center_code": "1200845668"
    }
},
{
    "_id": "4253546321522200576",
    "name": "蛋糕仓配",
    "extend_code": {
        "us_id": "40001",
        "ex_id": "40001",
        "ex_code": "40001",
        "ex_cost_center_code": "1200040001"
    }
},
{
    "_id": "4253847668826050560",
    "name": "临沂万象汇店",
    "extend_code": {
        "comm_shop_id": "55f49e0165b54ee1ad02f3491505c9d0",
        "ex_code": "45676",
        "alipay_id": "2020042200077000000093741372",
        "us_id": "45676",
        "upcard_terminal": "53952418",
        "upcard_mer_id": "102539089990459",
        "ex_id": "45676",
        "ex_cost_center_code": "1200845676"
    }
},
{
    "_id": "4253850283970400257",
    "name": "临沂万达店",
    "extend_code": {
        "comm_shop_id": "0db049f6ad1b4b1d9d59308c6a2b6a78",
        "ex_code": "45675",
        "alipay_id": "2020042200077000000093739953",
        "us_id": "45675",
        "upcard_terminal": "53952417",
        "upcard_mer_id": "102539089990458",
        "ex_id": "45675",
        "ex_cost_center_code": "1200845675"
    }
},
{
    "_id": "4253851784402976768",
    "name": "连云港赣榆吾悦广场店",
    "extend_code": {
        "comm_shop_id": "fc1f7551d37d4fa3a8956d73dcfafff7",
        "ex_code": "45677",
        "alipay_id": "2019121600077000000085554033",
        "us_id": "45677",
        "upcard_terminal": "51800798",
        "upcard_mer_id": "102518089990095",
        "ex_id": "45677",
        "ex_cost_center_code": "1200845677"
    }
},
{
    "_id": "4256991223148974080",
    "name": "合肥万象汇店",
    "extend_code": {
        "comm_shop_id": "6b5bc066f2784877a4a4abd7fece87c2",
        "ex_code": "",
        "alipay_id": "2020042300077000000093779511",
        "us_id": "45681",
        "upcard_terminal": "55132218",
        "upcard_mer_id": "102551089993923",
        "ex_cost_center_code": "1200845681"
    }
},
{
    "_id": "4256996464334475264",
    "name": "西安正荣彩虹谷店",
    "extend_code": {
        "comm_shop_id": "e3ea958951034c0d88c720afbe743beb",
        "alipay_id": "2019092700077000000083077258",
        "us_id": "45674",
        "upcard_terminal": "02907294",
        "upcard_mer_id": "102290089993361",
        "ex_cost_center_code": "1200845674"
    }
},
{
    "_id": "4258552441118658560",
    "name": "濮阳恒丰店",
    "extend_code": {
        "comm_shop_id": "ba472111360848c29749af9af50cb2b8",
        "alipay_id": "2019111800077000000084701157",
        "us_id": "45685",
        "upcard_terminal": "39301050",
        "upcard_mer_id": " 102393089990032",
        "ex_cost_center_code": "1200845685"
    }
},
{
    "_id": "4258555715762786304",
    "name": "武汉凯德西城店",
    "extend_code": {
        "comm_shop_id": "b3643764e7d54040af481ad92be8308d",
        "alipay_id": "2019091000077000000082470265",
        "us_id": "45693",
        "upcard_terminal": "02731326",
        "upcard_mer_id": "102270089995776",
        "ex_cost_center_code": "1200845693"
    }
},
{
    "_id": "4258560512280760320",
    "name": "洛阳万达店",
    "extend_code": {
        "comm_shop_id": "50cb236ab8764882ac54d0e4c46e83a6",
        "alipay_id": "2019091000077000000082473368",
        "us_id": "45698",
        "upcard_terminal": "37902018",
        "upcard_mer_id": "102379089990514",
        "ex_cost_center_code": "1200845698"
    }
},
{
    "_id": "4258812682901131264",
    "name": "广州百脑汇店",
    "extend_code": {
        "comm_shop_id": "a3d504ee159442ce946cea5af7dd7e51",
        "alipay_id": "2019100900077000000083379296",
        "us_id": "45699",
        "upcard_terminal": "02083949",
        "upcard_mer_id": "102200089991152",
        "ex_cost_center_code": "1200845699"
    }
},
{
    "_id": "4258875212885397504",
    "name": "西双版纳王府广场店",
    "extend_code": {
        "comm_shop_id": "897d356ef46e4f16ac7581658b09328e",
        "alipay_id": "2019121900077000000085649236",
        "us_id": "45701",
        "upcard_terminal": "69100030",
        "upcard_mer_id": "102691089990126",
        "ex_cost_center_code": "1200845701"
    }
},
{
    "_id": "4261110271038197760",
    "name": "广汉百伦店",
    "extend_code": {
        "us_id": "45700",
        "upcard_mer_id": "102838089990117",
        "ex_cost_center_code": "1200845700",
        "comm_shop_id": "0594619f605b4cdfafb079991e4f5575",
        "upcard_terminal": "83890645"
    }
},
{
    "_id": "4261341418020147200",
    "name": "西安西咸吾悦店",
    "extend_code": {
        "comm_shop_id": "c79de7a4b76742598ea4e93251d8dd9f",
        "alipay_id": "2019120200077000000085147972",
        "us_id": "45687",
        "upcard_terminal": "02907304",
        "upcard_mer_id": "102290089993366",
        "ex_cost_center_code": "1200845687"
    }
},
{
    "_id": "4262192023844425728",
    "name": "昆明瑞鼎城店",
    "extend_code": {
        "comm_shop_id": "16d770df66ac41d8afa2f6a7c1b201a8",
        "alipay_id": "2019121900077000000085603174",
        "us_id": "45712",
        "upcard_terminal": "87113853",
        "upcard_mer_id": "102871089996988",
        "ex_cost_center_code": "1200845712"
    }
},
{
    "_id": "4264698043607027712",
    "name": "淮安茂业天地店",
    "extend_code": {
        "comm_shop_id": "d8cc5affb2b049c780b803b36751b789",
        "alipay_id": "2020062400077000000095670713",
        "us_id": "45717",
        "upcard_terminal": "51701461",
        "upcard_mer_id": "102517089990104",
        "ex_cost_center_code": "1200845717"
    }
},
{
    "_id": "4264699648242880512",
    "name": "常州金坛新天地店",
    "extend_code": {
        "comm_shop_id": "f0a46195e4d649b79526d58a75e170a4",
        "alipay_id": "2019110600077000000084412677",
        "us_id": "45686",
        "upcard_terminal": "51902140",
        "upcard_mer_id": "102519089990239",
        "ex_cost_center_code": "1200845686"
    }
},
{
    "_id": "4264701090664026112",
    "name": "沭阳雨润中央商场店",
    "extend_code": {
        "comm_shop_id": "288f3180114f4288a0137e8e7bdc62ec",
        "alipay_id": "2019111800077000000084702855",
        "us_id": "45716",
        "upcard_terminal": "52701928",
        "upcard_mer_id": "102527089990314",
        "ex_cost_center_code": "1200845716"
    }
},
{
    "_id": "4266445588981092352",
    "name": "南京黄栗墅南服务区店",
    "extend_code": {
        "us_id": "45720",
        "upcard_mer_id": "102250089993999",
        "ex_cost_center_code": "1200845720",
        "comm_shop_id": "6fa657b1abe84859a6b50a91313a2777",
        "upcard_terminal": "02584500"
    }
},
{
    "_id": "4267164412483211264",
    "name": "扬州宝应吾悦店",
    "extend_code": {
        "comm_shop_id": "9a6633812bff45ddad0ce12c3b4809ee",
        "alipay_id": "2019103000077000000084077898",
        "us_id": "45715",
        "upcard_terminal": "51401746",
        "upcard_mer_id": "102514089990353",
        "ex_cost_center_code": "1200845715"
    }
},
{
    "_id": "4267167954568744960",
    "name": "聊城万达店",
    "extend_code": {
        "comm_shop_id": "f4b4363a629c416c94cd7da21ab5a60d",
        "alipay_id": "2020041300077000000093511665",
        "us_id": "45722",
        "upcard_terminal": "63500060",
        "upcard_mer_id": "102635089990030",
        "ex_cost_center_code": "1200845722"
    }
},
{
    "_id": "4267169396348162048",
    "name": "济南万象城店",
    "extend_code": {
        "comm_shop_id": "e22611ab85014076b2c8bc3c941b5017",
        "alipay_id": "2020033100077000000092776087",
        "us_id": "45708",
        "upcard_terminal": "53101688",
        "upcard_mer_id": "102531089990659",
        "ex_cost_center_code": "1200845708"
    }
},
{
    "_id": "4268712347750633472",
    "name": "济南东北服务区店",
    "extend_code": {
        "us_id": "45728",
        "upcard_mer_id": "102531089990660",
        "ex_cost_center_code": "1200845728",
        "comm_shop_id": "04f01b8a735d4df19974ba2d39505eab",
        "upcard_terminal": "53101689"
    }
},
{
    "_id": "4268713828314779648",
    "name": "济南东南服务区店",
    "extend_code": {
        "us_id": "45729",
        "upcard_mer_id": "102531089990661",
        "ex_cost_center_code": "1200845729",
        "comm_shop_id": "440827dff5b74627b4dc3b2f9e42cb6d",
        "upcard_terminal": "53101690"
    }
},
{
    "_id": "4271978999430778880",
    "name": "泰州姜堰时代店",
    "extend_code": {
        "comm_shop_id": "97971c6742d345778cb658cbcd408987",
        "alipay_id": "2019103000077000000084077899",
        "us_id": "45727",
        "upcard_terminal": "52301029",
        "upcard_mer_id": "102523089990144",
        "ex_cost_center_code": "1200845727"
    }
},
{
    "_id": "4273761756603486208",
    "name": "常州天宁吾悦店",
    "extend_code": {
        "comm_shop_id": "1e95bc36c52b42d189fe3a88d9ad5d96",
        "alipay_id": "2019110600077000000084405309",
        "us_id": "45731",
        "upcard_terminal": "51902145",
        "upcard_mer_id": "102519089990242",
        "ex_cost_center_code": "1200845731"
    }
},
{
    "_id": "4274857481202569216",
    "name": "青岛凯德茂店",
    "extend_code": {
        "comm_shop_id": "118c3f45848149be8ec043d7f9af7ebc",
        "alipay_id": "2020010700077000000086341666",
        "us_id": "45732",
        "upcard_terminal": "53206492",
        "upcard_mer_id": "102532089991555",
        "ex_cost_center_code": "1200845732"
    }
},
{
    "_id": "4275957033708032000",
    "name": "随州万达店",
    "extend_code": {
        "us_id": "45750",
        "upcard_mer_id": "102722089990020",
        "ex_cost_center_code": "1200845750",
        "comm_shop_id": "8074dd7b59b64bfbb9874412fcef14f5",
        "upcard_terminal": "72200125"
    }
},
{
    "_id": "4275958556798226432",
    "name": "福州长乐万星店",
    "extend_code": {
        "comm_shop_id": "0f97dba5df384d2a96bc1c1b805cb2dc",
        "alipay_id": "2019111100077000000084532014",
        "us_id": "45754",
        "upcard_terminal": "59113831",
        "upcard_mer_id": "102591089990674",
        "ex_cost_center_code": "1200845754"
    }
},
{
    "_id": "4275959968726781952",
    "name": "郑州航海路丹尼斯店",
    "extend_code": {
        "comm_shop_id": "8f0ed18592774651adc1af126b638fce",
        "alipay_id": "2020032400077000000092599667",
        "us_id": "45746",
        "upcard_terminal": "37114444",
        "upcard_mer_id": "102371089992999",
        "ex_cost_center_code": "1200845746"
    }
},
{
    "_id": "4275961306562957312",
    "name": "吉安天虹二店",
    "extend_code": {
        "comm_shop_id": "28a08582c9d9497ea7cb91d0d1c7e81a",
        "alipay_id": "2020032400077000000092589565",
        "us_id": "45748",
        "upcard_terminal": "79600974",
        "upcard_mer_id": "102796089990046",
        "ex_cost_center_code": "1200845748"
    }
},
{
    "_id": "4276282522205163520",
    "name": "佛山金沙洲金铂天地店",
    "extend_code": {
        "comm_shop_id": "c1fc75ebb8f04c2685dce5aca95a90c8",
        "alipay_id": "2019112700077000000085023649",
        "us_id": "45747",
        "upcard_terminal": "75703882",
        "upcard_mer_id": "102757089990665",
        "ex_cost_center_code": "1200045747"
    }
},
{
    "_id": "4281387215881244672",
    "name": "杭州西溪龙湖天街店",
    "extend_code": {
        "comm_shop_id": "0e850b6dcc6340aba832e0c29eca5ca6",
        "alipay_id": "2019121900077000000085649237",
        "us_id": "45760",
        "upcard_terminal": "57114976",
        "upcard_mer_id": "102571089993201",
        "ex_cost_center_code": "1200845760"
    }
},
{
    "_id": "4284187678389309440",
    "name": "庆阳东方丽晶MALL店",
    "extend_code": {
        "comm_shop_id": "eb68266041f447b9a4b203c6c5932421",
        "alipay_id": "2020042100077000000093732010",
        "us_id": "45776",
        "upcard_terminal": "93400018",
        "upcard_mer_id": "102934089990015",
        "ex_cost_center_code": "1200045776"
    }
},
{
    "_id": "4285358875844022272",
    "name": "滕州万达店",
    "extend_code": {
        "comm_shop_id": "f222647e4075496db2916c0f48f28d2a",
        "alipay_id": "2019123000077000000085880428",
        "us_id": "45752",
        "upcard_terminal": "63200099",
        "upcard_mer_id": "102632089990034",
        "ex_cost_center_code": "1200845752"
    }
},
{
    "_id": "4285360871560974336",
    "name": "郑州机场南中心店",
    "extend_code": {
        "comm_shop_id": "13a3f9e5a578462d9b4f5ce51740b9f4",
        "company_code": "",
        "us_id": "45694",
        "upcard_terminal": "97100417",
        "upcard_mer_id": "102971089990249",
        "ex_cost_center_code": "1200845694"
    }
},
{
    "_id": "4285362435658878976",
    "name": "杭州博雅城店",
    "extend_code": {
        "comm_shop_id": "572f84cf71044ff0a437a2a9500afae1",
        "alipay_id": "2020010700077000000086343469",
        "us_id": "45778",
        "upcard_terminal": "57114977",
        "upcard_mer_id": "102571089993202",
        "ex_cost_center_code": "1200845778"
    }
},
{
    "_id": "4285363821876350976",
    "name": "汉中吾悦店",
    "extend_code": {
        "comm_shop_id": "91d85d3d4d4a44218bd43a84ee61886f",
        "alipay_id": "2019120200077000000085123092",
        "us_id": "45777",
        "upcard_terminal": "91600102",
        "upcard_mer_id": "102916089990019",
        "ex_cost_center_code": "1200845777"
    }
},
{
    "_id": "4287529483126181888",
    "name": "徐州丰县欢乐城店",
    "extend_code": {
        "comm_shop_id": "90fc24179f6c463d9c4dd26b22e87b3c",
        "alipay_id": "2019121900077000000085651239",
        "us_id": "45781",
        "upcard_terminal": "51601817",
        "upcard_mer_id": "102516089990307",
        "ex_cost_center_code": "1200845781"
    }
},
{
    "_id": "4287531202274922496",
    "name": "青岛黄岛永旺店",
    "extend_code": {
        "comm_shop_id": "e9bf736f060f48bfbd080d5cc8b0ce74",
        "alipay_id": "2019121900077000000085651238",
        "us_id": "45765",
        "upcard_terminal": "53206502",
        "upcard_mer_id": "102532089991646",
        "ex_cost_center_code": "1200845765"
    }
},
{
    "_id": "4287532561766289408",
    "name": "上海浦江万达店",
    "extend_code": {
        "comm_shop_id": "6e183f3df42441c4aa161a95de6e6923",
        "alipay_id": "2020033000077000000092775681",
        "us_id": "45780",
        "upcard_terminal": "02109461",
        "upcard_mer_id": "102210089999228",
        "ex_cost_center_code": "1200845780"
    }
},
{
    "_id": "4287533562703384576",
    "name": "银川大阅城店",
    "extend_code": {
        "comm_shop_id": "e370315435bf40beac855ea8ef2689de",
        "alipay_id": "2020010700077000000086341665",
        "us_id": "45785",
        "upcard_terminal": "95102588",
        "upcard_mer_id": "102951089991004",
        "ex_cost_center_code": "1200045785"
    }
},
{
    "_id": "4289350741081985024",
    "name": "重庆江北机场T2A店",
    "extend_code": {
        "us_id": "45788",
        "upcard_mer_id": "102230089992761",
        "ex_cost_center_code": "1200845788",
        "comm_shop_id": "b803f6fa06fc4def9fce6e9644c7a68b",
        "upcard_terminal": "02311529"
    }
},
{
    "_id": "4289352218714312704",
    "name": "重庆江北机场T2B店",
    "extend_code": {
        "us_id": "45789",
        "upcard_mer_id": "102230089992762",
        "ex_cost_center_code": "1200845789",
        "comm_shop_id": "8923a69f2d64470c85d6781dc0728a90",
        "upcard_terminal": "02311530"
    }
},
{
    "_id": "4289354905203773440",
    "name": "桂林东西巷店",
    "extend_code": {
        "comm_shop_id": "0e3abddfbe5d452281c2caeba1c3956a",
        "alipay_id": "2020010700077000000086344667",
        "us_id": "45758",
        "upcard_terminal": "77301673",
        "upcard_mer_id": "102773089990524",
        "ex_cost_center_code": "1200845758"
    }
},
{
    "_id": "4289357921629769728",
    "name": "连云港海州吾悦店",
    "extend_code": {
        "comm_shop_id": "e41d3d9e0bc04874a54b65aa6c8a7485",
        "alipay_id": "2019121900077000000085652439",
        "us_id": "45753",
        "upcard_terminal": "51800803",
        "upcard_mer_id": "102518089990100",
        "ex_cost_center_code": "1200845753"
    }
},
{
    "_id": "4290405516382896128",
    "name": "上海复地活力城店",
    "extend_code": {
        "comm_shop_id": "8ef9300090fb4776afd887b4f8821da2",
        "alipay_id": "2020042100077000000093730699",
        "us_id": "45787",
        "upcard_terminal": "02109687",
        "upcard_mer_id": "102210089999245",
        "ex_cost_center_code": "1200845787"
    }
},
{
    "_id": "4290406980945739776",
    "name": "上海剑川路龙湖店",
    "extend_code": {
        "comm_shop_id": "cbf4b55dfb1b4606907c5351ccf3003d",
        "alipay_id": "2019123100077000000086162353",
        "us_id": "45793",
        "upcard_terminal": "02109688",
        "upcard_mer_id": "102210089999246",
        "ex_cost_center_code": "1200845793"
    }
},
{
    "_id": "4290408081153622016",
    "name": "南京南站店",
    "extend_code": {
        "us_id": "45795",
        "upcard_mer_id": "102250089994037",
        "ex_cost_center_code": "1200845795",
        "comm_shop_id": "c730c8c6676049ab9e89e62e529c70ee",
        "upcard_terminal": "02585099"
    }
},
{
    "_id": "4292610335877169152",
    "name": "泉州浦西万达店",
    "extend_code": {
        "comm_shop_id": "99ab76ad555840cca70b76abf721c22d",
        "alipay_id": "2019122500077000000085760324",
        "us_id": "45799",
        "upcard_terminal": "59501753",
        "upcard_mer_id": "102595089990342",
        "ex_cost_center_code": "1200045799"
    }
},
{
    "_id": "4292612759085940736",
    "name": "杭州大悦城店",
    "extend_code": {
        "comm_shop_id": "38a0c10de8304574ab449f5ff314ca20",
        "alipay_id": "2020010700077000000086343470",
        "us_id": "45798",
        "upcard_terminal": "57115185",
        "upcard_mer_id": "102571089993213",
        "ex_cost_center_code": "1200845798"
    }
},
{
    "_id": "4292613943221878784",
    "name": "郑州中牟天泽城店",
    "extend_code": {
        "comm_shop_id": "9d8bd2e5981142c38a48db0df2853d01",
        "alipay_id": "2020051800077000000094408977",
        "us_id": "45797",
        "upcard_terminal": "37114494",
        "upcard_mer_id": "102371089993007",
        "ex_cost_center_code": "1200845797"
    }
},
{
    "_id": "4292615657425502208",
    "name": "淮安楚州万达店",
    "extend_code": {
        "comm_shop_id": "b1169e3cb5654743b5a5a0075f1a4868",
        "alipay_id": "2020062400077000000095661413",
        "us_id": "45805",
        "upcard_terminal": "51701473",
        "upcard_mer_id": "102517089990112",
        "ex_cost_center_code": "1200845805"
    }
},
{
    "_id": "4295431863522066432",
    "name": "苏州盛泽碧桂园店",
    "extend_code": {
        "comm_shop_id": "c8c1039e9f8f4958a7b78e1775b410f3",
        "alipay_id": "2020101900077000000007345883",
        "us_id": "45796",
        "upcard_terminal": "51217652",
        "upcard_mer_id": "102512089993393",
        "ex_cost_center_code": "1200845796"
    }
},
{
    "_id": "4295433169955487744",
    "name": "怀化万达店",
    "extend_code": {
        "us_id": "45804",
        "upcard_mer_id": "102745089990056",
        "ex_cost_center_code": "1200845804",
        "comm_shop_id": "6c12ae4990174ab8889211ccd7890dac",
        "upcard_terminal": "74501336"
    }
},
{
    "_id": "4295434461952114688",
    "name": "如皋吾悦店",
    "extend_code": {
        "comm_shop_id": "509ec681fe124c0a926c292947ef9931",
        "alipay_id": "2020040300077000000092869011",
        "us_id": "45792",
        "upcard_terminal": "51301460",
        "upcard_mer_id": "102513089990276",
        "ex_cost_center_code": "1200845792"
    }
},
{
    "_id": "4296602450353225728",
    "name": "郑州万锦城店",
    "extend_code": {
        "comm_shop_id": "923f6533e6d448b3a4e4ccc2bf385995",
        "alipay_id": "2020011500077000000086882010",
        "us_id": "45810",
        "upcard_terminal": "37114496",
        "upcard_mer_id": "102371089993009",
        "ex_cost_center_code": "1200845810"
    }
},
{
    "_id": "4296604174614822913",
    "name": "扬州东关街店",
    "extend_code": {
        "us_id": "45803",
        "upcard_mer_id": "102514089990365",
        "ex_cost_center_code": "1200845803",
        "comm_shop_id": "8e965260c44e4c94b07c87ef1ffff891",
        "upcard_terminal": "51402002"
    }
},
{
    "_id": "4296966128328015872",
    "name": "泰安爱琴海店",
    "extend_code": {
        "us_id": "45809",
        "upcard_mer_id": "102538089990572",
        "ex_cost_center_code": "1200845809",
        "comm_shop_id": "44ef870247524135bf01862a8188e385",
        "upcard_terminal": "53800667"
    }
},
{
    "_id": "4296967508887371776",
    "name": "合肥北城万达店",
    "extend_code": {
        "comm_shop_id": "b19295d6e3e741babe05665ee1206e2b",
        "alipay_id": "2020030900077000000092243223",
        "us_id": "45812",
        "upcard_terminal": "55132404",
        "upcard_mer_id": "102551089993954",
        "ex_cost_center_code": "1200845812"
    }
},
{
    "_id": "4297230402816344064",
    "name": "濮阳万达店",
    "extend_code": {
        "comm_shop_id": "6b7c3c1d42b84d0ba9f9be2bd2cb8297",
        "alipay_id": "2020011500077000000086880218",
        "us_id": "45816",
        "upcard_terminal": "39301051",
        "upcard_mer_id": "102393089990039",
        "ex_cost_center_code": "1200845816"
    }
},
{
    "_id": "4297607682621243392",
    "name": "高邮吾悦店",
    "extend_code": {
        "comm_shop_id": "361f6bb508eb445cae9d86b0317690df",
        "alipay_id": "2020011600077000000087239442",
        "us_id": "45811",
        "upcard_terminal": "51402003",
        "upcard_mer_id": "102514089990366",
        "ex_cost_center_code": "1200845811"
    }
},
{
    "_id": "4300231256288854016",
    "name": "郑州瀚海海尚店",
    "extend_code": {
        "comm_shop_id": "b298064c907c4529961650046054dc4f",
        "alipay_id": "2020032400077000000092589566",
        "us_id": "45817",
        "upcard_terminal": "37114577",
        "upcard_mer_id": "102371089993013",
        "ex_cost_center_code": "1200845817"
    }
},
{
    "_id": "4300232608515653632",
    "name": "南京六合龙湖天街店",
    "extend_code": {
        "comm_shop_id": "81bb605227f643e39d4dc59522c3133a",
        "alipay_id": "2020030200077000000092103141",
        "us_id": "45800",
        "upcard_terminal": "02585220",
        "upcard_mer_id": "102250089994048",
        "ex_cost_center_code": "1200845800"
    }
},
{
    "_id": "4300234219631738880",
    "name": "孝感服务区北区店",
    "extend_code": {
        "us_id": "45825",
        "upcard_mer_id": "102712089990426",
        "ex_cost_center_code": "1200845825",
        "comm_shop_id": "d81aa83ab0a041a1a2810daccdb4c1a0",
        "upcard_terminal": "71200082"
    }
},
{
    "_id": "4300492797798023168",
    "name": "西安大雁塔东街店",
    "extend_code": {
        "us_id": "45826",
        "upcard_mer_id": "102290089993400",
        "ex_cost_center_code": "1200845826",
        "comm_shop_id": "333b14c53ecb4cf0a72dd31afe41c34c",
        "upcard_terminal": "02910257"
    }
},
{
    "_id": "4302053469753212928",
    "name": "南浔服务区北区店",
    "extend_code": {
        "comm_shop_id": "b9d0a67a4667459fa1f2545a7e7934c2",
        "us_id": "45801",
        "upcard_terminal": "57297325",
        "upcard_mer_id": "102572089990094",
        "ex_id": "45801",
        "ex_cost_center_code": "1200845801"
    }
},
{
    "_id": "4302055864797265920",
    "name": "阜阳临泉丰泽悦城店",
    "extend_code": {
        "comm_shop_id": "e48cd774aeaa4387afcdf106d95e9e79",
        "alipay_id": "2020031200077000000092343792",
        "us_id": "45832",
        "upcard_terminal": "55802185",
        "upcard_mer_id": "102558089990754",
        "ex_cost_center_code": "1200845832"
    }
},
{
    "_id": "4302799970209136640",
    "name": "盐城新弄里店",
    "extend_code": {
        "comm_shop_id": "cb462629eb3b40c1a6431271373285a8",
        "alipay_id": "2020030200077000000092101862",
        "us_id": "45828",
        "upcard_terminal": "51501055",
        "upcard_mer_id": "102515089990495",
        "ex_cost_center_code": "1200845828"
    }
},
{
    "_id": "4303117755711782912",
    "name": "太仓沙溪北服务区店",
    "extend_code": {
        "us_id": "45840",
        "upcard_mer_id": "102512089993395",
        "ex_cost_center_code": "1200845840",
        "comm_shop_id": "ee5e8556f2e14a87a15dd7d00aab1e18",
        "upcard_terminal": "51217661"
    }
},
{
    "_id": "4303124525293305856",
    "name": "西安立丰城市生活广场店",
    "extend_code": {
        "us_id": "45844",
        "upcard_mer_id": "102290089993402",
        "ex_cost_center_code": "1200845844",
        "comm_shop_id": "6eb32203fe3e4ea9923645a017035926",
        "upcard_terminal": "02910259"
    }
},
{
    "_id": "4303125910155984896",
    "name": "西安万和城店",
    "extend_code": {
        "comm_shop_id": "610e1c15d9e84e4b9f668599fd83ec4d",
        "alipay_id": "2020010700077000000086338363",
        "us_id": "45845",
        "upcard_terminal": "02910258",
        "upcard_mer_id": "102290089993401",
        "ex_cost_center_code": "1200845845"
    }
},
{
    "_id": "4303129458742689792",
    "name": "无锡万象城店",
    "extend_code": {
        "comm_shop_id": "481c232ac84d42debbf7f14b6e413955",
        "alipay_id": "2019123100077000000085900209",
        "us_id": "45847",
        "upcard_terminal": "51004730",
        "upcard_mer_id": "102510089994210",
        "ex_cost_center_code": "1200845847"
    }
},
{
    "_id": "4303131624928772096",
    "name": "桐乡吾悦店",
    "extend_code": {
        "comm_shop_id": "71070c319aa24e89a89805b97b22a70f",
        "alipay_id": "2020061600077000000095297669",
        "us_id": "45846",
        "upcard_terminal": "57304208",
        "upcard_mer_id": "102573089990217",
        "ex_cost_center_code": "1200845846"
    }
},
{
    "_id": "4304888050655199232",
    "name": "济南恒隆店",
    "extend_code": {
        "comm_shop_id": "242d0a78e1c84f1684e8b258865c108d",
        "alipay_id": "2020040300077000000092873412",
        "us_id": "45851",
        "upcard_terminal": "53101700",
        "upcard_mer_id": "102531089990758",
        "ex_cost_center_code": "1200845851"
    }
},
{
    "_id": "4307010121820569600",
    "name": "荆州监利宏泰店",
    "extend_code": {
        "us_id": "45866",
        "upcard_mer_id": "102716089990081",
        "ex_cost_center_code": "1200845866",
        "comm_shop_id": "00e75bba309f43d88325ee3ac4155284",
        "upcard_terminal": "71600283"
    }
},
{
    "_id": "4307011575247568896",
    "name": "盐城吾悦店",
    "extend_code": {
        "comm_shop_id": "20fe63801986477e8dd4f8a100dd8909",
        "alipay_id": "2020033000077000000092775680",
        "us_id": "45870",
        "upcard_terminal": "51501062",
        "upcard_mer_id": "102515089990496",
        "ex_cost_center_code": "1200845870"
    }
},
{
    "_id": "4307012786214436864",
    "name": "银川中海环宇天地店",
    "extend_code": {
        "us_id": "45867",
        "upcard_mer_id": "102951089991006",
        "ex_cost_center_code": "1200845867",
        "comm_shop_id": "3efd2abe38b9476c93d5f03287e4e5a9",
        "upcard_terminal": "95102619"
    }
},
{
    "_id": "4307014746283999232",
    "name": "马鞍山含山县玉龙湖店",
    "extend_code": {
        "comm_shop_id": "1c2089204847429288d405f38dd63c14",
        "alipay_id": "2020041600077000000093600871",
        "us_id": "45865",
        "upcard_terminal": "55500588",
        "upcard_mer_id": "102555089990096",
        "ex_cost_center_code": "1200845865"
    }
},
{
    "_id": "4307015917891551232",
    "name": "砀山万达店",
    "extend_code": {
        "comm_shop_id": "4bdaa1dd2ed14bcd9e6ae39998f0966c",
        "alipay_id": "2020031200077000000092343791",
        "us_id": "45873",
        "upcard_terminal": "55700481",
        "upcard_mer_id": "102557089990125",
        "ex_cost_center_code": "1200845873"
    }
},
{
    "_id": "4307740354441936896",
    "name": "登封万佳中心城店",
    "extend_code": {
        "comm_shop_id": "e59ff61364ab403d8cb3ff53fe449cc7",
        "alipay_id": "2020123000077000000013302500",
        "us_id": "45872",
        "upcard_terminal": "37114623",
        "upcard_mer_id": "102371089993015",
        "ex_cost_center_code": "1200845872"
    }
},
{
    "_id": "4348048881001463808",
    "name": "中山大信新都会店",
    "extend_code": {
        "comm_shop_id": "20d9415206cc47e8921e3ff2c288657b",
        "alipay_id": "2020071400077000000098797734",
        "us_id": "45903",
        "upcard_terminal": "76001565",
        "upcard_mer_id": "102760089990335",
        "ex_id": "45903",
        "ex_cost_center_code": "1200845903"
    }
},
{
    "_id": "4348051545059819520",
    "name": "昆明世纪金源店",
    "extend_code": {
        "comm_shop_id": "a60c818ab1e24ce09a04d3755845045c",
        "ex_code": "45904",
        "alipay_id": "2020051800077000000094411123",
        "us_id": "45904",
        "upcard_terminal": "87113933",
        "upcard_mer_id": "102871089996997",
        "ex_cost_center_code": "1200845904"
    }
},
{
    "_id": "4356411764739440640",
    "name": "重庆金沙龙湖天街店",
    "extend_code": {
        "comm_shop_id": "838b89a1888f436792e4a7c032143a66",
        "ex_code": "45909",
        "alipay_id": "2021010800077000000013636950",
        "us_id": "45909",
        "upcard_terminal": "02312136",
        "upcard_mer_id": "102230089992840",
        "ex_id": "45909",
        "ex_cost_center_code": "1200045909"
    }
},
{
    "_id": "4357806338053570560",
    "name": "深圳深业上城店",
    "extend_code": {
        "comm_shop_id": "e2a8d2caf7b5421e886b85d56d4dc266",
        "ex_code": "45914",
        "alipay_id": "2020061600077000000095315239",
        "us_id": "45914",
        "upcard_terminal": "75524777",
        "upcard_mer_id": "102755089995666",
        "ex_id": "45914",
        "ex_cost_center_code": "1200845914"
    }
},
{
    "_id": "4361446207728418816",
    "name": "兴义欢乐橙店",
    "extend_code": {
        "comm_shop_id": "7cab2eebcf0f47379ac6ae0a02622b00",
        "ex_code": "45916",
        "alipay_id": "2020061600077000000095297668",
        "us_id": "45916",
        "upcard_terminal": "85900106",
        "upcard_mer_id": "102859089990014",
        "ex_id": "45916",
        "ex_cost_center_code": "1200845916"
    }
},
{
    "_id": "4361449197763887104",
    "name": "蒙自南湖荟店",
    "extend_code": {
        "comm_shop_id": "514c527a8b3a4a51a486f543993c29d1",
        "ex_code": "45918",
        "us_id": "45918",
        "upcard_terminal": "87300727",
        "upcard_mer_id": "102873089990422",
        "ex_id": "45918",
        "ex_cost_center_code": "1200845918"
    }
},
{
    "_id": "4361451488617201664",
    "name": "贵阳花溪万科店",
    "extend_code": {
        "comm_shop_id": "b06e0299fbeb4a2088bc59c86081b330",
        "ex_code": "45919",
        "alipay_id": "2020062400077000000095669110",
        "us_id": "45919",
        "upcard_terminal": "85102849",
        "upcard_mer_id": "102851089990592",
        "ex_id": "45919",
        "ex_cost_center_code": "1200845919"
    }
},
{
    "_id": "4361471575621435392",
    "name": "南通海安万达店",
    "extend_code": {
        "comm_shop_id": "ee8a6f4fa409420aae91d9679d3d109a",
        "ex_code": "45917",
        "alipay_id": "2020071400077000000098792965",
        "us_id": "45917",
        "upcard_terminal": "51301529",
        "upcard_mer_id": "102513089990282",
        "ex_id": "45917",
        "ex_cost_center_code": "1200845917"
    }
},
{
    "_id": "4364025192370995200",
    "name": "兰州国芳杉杉奥特莱斯店",
    "extend_code": {
        "comm_shop_id": "f5318a0e90e84886a6b24681fa2ee3c2",
        "ex_code": "45910",
        "alipay_id": "2021110500077000000029761090",
        "us_id": "45910",
        "upcard_terminal": "93101318",
        "upcard_mer_id": "102931089990241",
        "ex_id": "45910",
        "ex_cost_center_code": "1200045910"
    }
},
{
    "_id": "4365069702291062784",
    "name": "无锡锡东大润发店",
    "extend_code": {
        "comm_shop_id": "9b28b230f2f0414c8f569ef811e039cf",
        "ex_code": "45920",
        "alipay_id": "2020080600077000000099554197",
        "us_id": "45920",
        "upcard_terminal": "51005020",
        "upcard_mer_id": "102510089994349",
        "ex_id": "45920",
        "ex_cost_center_code": "1200845920"
    }
},
{
    "_id": "4365071990074146816",
    "name": "深圳卓悦汇店",
    "extend_code": {
        "comm_shop_id": "f52e732532b24dac844f3108cdbe2d23",
        "ex_code": "45921",
        "alipay_id": "2020072700077000000099267864",
        "us_id": "45921",
        "upcard_terminal": "75524877",
        "upcard_mer_id": "102755089995750",
        "ex_id": "45921",
        "ex_cost_center_code": "1200045921"
    }
},
{
    "_id": "4366496911824322560",
    "name": "西安高新大都荟店",
    "extend_code": {
        "comm_shop_id": "f9863c68d06543eb8bf9401fa5acd348",
        "ex_code": "45922",
        "alipay_id": "2020072400077000000099094919",
        "us_id": "45922",
        "upcard_terminal": "02910294",
        "upcard_mer_id": "102290089993420",
        "ex_id": "45922",
        "ex_cost_center_code": "1200845922"
    }
},
{
    "_id": "4372965043023708160",
    "name": "南丰城店",
    "extend_code": {
        "comm_shop_id": "b815027097a0460fa12ce38b3f8d29e9",
        "ex_code": "45934",
        "alipay_id": "2020081300077000000001063214",
        "us_id": "45934",
        "upcard_terminal": "02112841",
        "upcard_mer_id": "102210089999379",
        "ex_id": "45934",
        "ex_cost_center_code": "1200045934"
    }
},
{
    "_id": "4372967935851921408",
    "name": "渭南吾悦广场店",
    "extend_code": {
        "comm_shop_id": "e8d57395802c456a9ea2eee1e5069e66",
        "ex_code": "45925",
        "alipay_id": "2021062800077000000023581561",
        "us_id": "45925",
        "upcard_terminal": "91300372",
        "upcard_mer_id": "102913089990036",
        "ex_id": "45925",
        "ex_cost_center_code": "1200845925"
    }
},
{
    "_id": "4372971656501592064",
    "name": "寿光万达店",
    "extend_code": {
        "comm_shop_id": "7db20711f5ea435ba4e5f14eb77d8183",
        "ex_code": "45931",
        "alipay_id": "2020081300077000000001063001",
        "us_id": "45931",
        "upcard_terminal": "53606374",
        "upcard_mer_id": "102536089990739",
        "ex_id": "45931",
        "ex_cost_center_code": "1200845931"
    }
},
{
    "_id": "4372973901964148736",
    "name": "海口友谊阳光城店",
    "extend_code": {
        "comm_shop_id": "58022baff2704fe8b2f6ff714ee9b583",
        "ex_code": "45933",
        "alipay_id": "2020072900077000000099334097",
        "us_id": "45933",
        "upcard_terminal": "89801976",
        "upcard_mer_id": "102898089990638",
        "ex_id": "45933",
        "ex_cost_center_code": "1200845933"
    }
},
{
    "_id": "4372975937875116032",
    "name": "三亚滨海世界店",
    "extend_code": {
        "comm_shop_id": "27705a81f7ed49ff93a9e4991e078000",
        "ex_code": "45911",
        "alipay_id": "2020081300077000000001063002",
        "us_id": "45911",
        "upcard_terminal": "89801977",
        "upcard_mer_id": "102898089990639",
        "ex_id": "45911",
        "ex_cost_center_code": "1200045911"
    }
},
{
    "_id": "4376300270651506688",
    "name": "孝感服务区南区店",
    "extend_code": {
        "comm_shop_id": "7c63c7419147476aaa7e284aefca1c64",
        "ex_code": "45936",
        "us_id": "45936",
        "upcard_terminal": "71200083",
        "upcard_mer_id": "102712089990427",
        "ex_id": "45936",
        "ex_cost_center_code": "1200845936"
    }
},
{
    "_id": "4376319295544688640",
    "name": "西安西咸万象城店",
    "extend_code": {
        "comm_shop_id": "5c655320468546fa8fe298bca1a93432",
        "ex_code": "45942",
        "alipay_id": "2020072700077000000099267865",
        "us_id": "45942",
        "upcard_terminal": "02910373",
        "upcard_mer_id": "102290089993425",
        "ex_id": "45942",
        "ex_cost_center_code": "1200845942"
    }
},
{
    "_id": "4378520938667343872",
    "name": "太原国金中心店",
    "extend_code": {
        "comm_shop_id": "2d2f7a91468541778d73e2539a18625f",
        "ex_code": "45940",
        "alipay_id": "2020092900077000000003519020",
        "us_id": "45940",
        "upcard_terminal": "35104887",
        "upcard_mer_id": "102351089992164",
        "ex_id": "45940",
        "ex_cost_center_code": "1200845940"
    }
},
{
    "_id": "4383485710831779840",
    "name": "阜阳吾悦店",
    "extend_code": {
        "comm_shop_id": "19c068af11174993b900e840214c6ec0",
        "ex_code": "45943",
        "alipay_id": "2021102600077000000029271048",
        "us_id": "45943",
        "upcard_terminal": "55802415",
        "upcard_mer_id": "102558089990764",
        "ex_id": "45943",
        "ex_cost_center_code": "1200845943"
    }
},
{
    "_id": "4386130885617942528",
    "name": "兴化吾悦店",
    "extend_code": {
        "comm_shop_id": "76b0faea91f247478dd97daa984744e0",
        "ex_code": "45937",
        "alipay_id": "2021102800077000000029382054",
        "us_id": "45937",
        "upcard_terminal": "52301063",
        "upcard_mer_id": "102523089990152",
        "ex_id": "45937",
        "ex_cost_center_code": "1200845937"
    }
},
{
    "_id": "4387941392733732864",
    "name": "毕节花园城店",
    "extend_code": {
        "comm_shop_id": "11f6ced34a394a43ae95f8713d6e0a67",
        "ex_code": "45954",
        "us_id": "45954",
        "upcard_terminal": "85700062",
        "upcard_mer_id": "102857089990021",
        "ex_id": "45954",
        "ex_cost_center_code": "1200845954"
    }
},
{
    "_id": "4387947305460891648",
    "name": "福州泰禾广场店",
    "extend_code": {
        "comm_shop_id": "0f5836388b074813a27224f548ee3a65",
        "ex_code": "45953",
        "alipay_id": "2020090900077000000002553407",
        "us_id": "45953",
        "upcard_terminal": "59114146",
        "upcard_mer_id": "102591089990694",
        "ex_id": "45953",
        "ex_cost_center_code": "1200845953"
    }
},
{
    "_id": "4388255581494902784",
    "name": "珠海优特汇店",
    "extend_code": {
        "comm_shop_id": "4e85c3aff0c342239eb0528bc5d230ec",
        "ex_code": "45950",
        "alipay_id": "2020101900077000000007345882",
        "us_id": "45950",
        "upcard_terminal": "75602231",
        "upcard_mer_id": "102756089990222",
        "ex_id": "45950",
        "ex_cost_center_code": "1200845950"
    }
},
{
    "_id": "4388256959403458560",
    "name": "三亚青春颂店",
    "extend_code": {
        "comm_shop_id": "29a9b4d1035e4001899c0ff3e78331e6",
        "ex_code": "45957",
        "alipay_id": "2020092800077000000003401922",
        "us_id": "45957",
        "upcard_terminal": "89801983",
        "upcard_mer_id": "102898089990641",
        "ex_id": "45957",
        "ex_cost_center_code": "1200845957"
    }
},
{
    "_id": "4389379616115064832",
    "name": "重庆西站店",
    "extend_code": {
        "comm_shop_id": "645ebc7639cc427c8f4a49fc3e045007",
        "ex_code": "45956",
        "alipay_id": "2021102600077000000029272278",
        "us_id": "45956",
        "upcard_terminal": "02312240",
        "upcard_mer_id": "102230089992885",
        "ex_id": "45956",
        "ex_cost_center_code": "1200845956"
    }
},
{
    "_id": "4390804055180476416",
    "name": "青岛合肥路佳世客店",
    "extend_code": {
        "comm_shop_id": "36f64239a2744b17b8844d9557dba4b2",
        "ex_code": "45949",
        "alipay_id": "2020092200077000000003024303",
        "us_id": "45949",
        "upcard_terminal": "53206610",
        "upcard_mer_id": "102532089991679",
        "ex_id": "45949",
        "ex_cost_center_code": "1200845949"
    }
},
{
    "_id": "4391895151423848448",
    "name": "广东悦汇城店",
    "extend_code": {
        "comm_shop_id": "1437b0ec9eab4aba957694b4721eaf88",
        "ex_code": "45962",
        "alipay_id": "2020092800077000000003403578",
        "us_id": "45962",
        "upcard_terminal": "02005455",
        "upcard_mer_id": "102200089991261",
        "ex_id": "45962",
        "ex_cost_center_code": "1200045962"
    }
},
{
    "_id": "4391897109522087936",
    "name": "西安UPlaza店",
    "extend_code": {
        "comm_shop_id": "5091acc79cfb401e89ebfba8b9c63243",
        "ex_code": "45963",
        "alipay_id": "2020092200077000000003024304",
        "us_id": "45963",
        "upcard_terminal": "02910476",
        "upcard_mer_id": "102290089993441",
        "ex_id": "45963",
        "ex_cost_center_code": "1200845963"
    }
},
{
    "_id": "4391899144959066112",
    "name": "新南宁万象城店",
    "extend_code": {
        "comm_shop_id": "d607cd14a0c14b578581ef4a42a8aed1",
        "ex_code": "45965",
        "alipay_id": "2020101200077000000007013526",
        "us_id": "45965",
        "upcard_terminal": "77106041",
        "upcard_mer_id": "102771089992196",
        "ex_id": "45965",
        "ex_cost_center_code": "1200045965"
    }
},
{
    "_id": "4398052426228236288",
    "name": "衡阳杉杉奥特莱斯店",
    "extend_code": {
        "comm_shop_id": "7e8526296a8e47be8ce3a9ae1ed00b97",
        "ex_code": "45968",
        "us_id": "45968",
        "upcard_terminal": "83890691",
        "upcard_mer_id": "102838089990120",
        "ex_id": "45968",
        "ex_cost_center_code": "1200845968"
    }
},
{
    "_id": "4398054206588616704",
    "name": "常州文化广场店",
    "extend_code": {
        "comm_shop_id": "d8670ac5bc8149a38e09ee09c912ec40",
        "ex_code": "45971",
        "alipay_id": "2020120900077000000010445232",
        "us_id": "45971",
        "upcard_terminal": "51902170",
        "upcard_mer_id": "102519089990250",
        "ex_id": "45971",
        "ex_cost_center_code": "1200845971"
    }
},
{
    "_id": "4398055515450572800",
    "name": "海盐吾悦店",
    "extend_code": {
        "comm_shop_id": "a325a947394f4340bbfd454e92aef873",
        "ex_code": "45973",
        "alipay_id": "2020110500077000000008032337",
        "us_id": "45973",
        "upcard_terminal": "57304366",
        "upcard_mer_id": "102573089990224",
        "ex_id": "45973",
        "ex_cost_center_code": "1200845973"
    }
},
{
    "_id": "4398808738801156096",
    "name": "西安昆明池店",
    "extend_code": {
        "comm_shop_id": "dd5bfda3055e4d438629cacd65356ee0",
        "ex_code": "45932",
        "us_id": "45932",
        "upcard_terminal": "02910493",
        "upcard_mer_id": "102290089993444",
        "ex_id": "4593245932",
        "ex_cost_center_code": "1200845932"
    }
},
{
    "_id": "4400596504904007680",
    "name": "南京雨花吾悦店",
    "extend_code": {
        "comm_shop_id": "dbf6c5e4a3f7456f92871fcfad8f7bb1",
        "ex_code": "45976",
        "alipay_id": "2020120900077000000010443361",
        "us_id": "45976",
        "upcard_terminal": "02588792",
        "upcard_mer_id": "102250089994109",
        "ex_id": "45976",
        "ex_cost_center_code": "1200845976"
    }
},
{
    "_id": "4400604720048177152",
    "name": "金湖苏宁店",
    "extend_code": {
        "comm_shop_id": "e6c4dec499c647c1a705076e8b6e8ecb",
        "ex_code": "45951",
        "alipay_id": "2020120900077000000010443363",
        "us_id": "45951",
        "upcard_terminal": "51701585",
        "upcard_mer_id": "102517089990120",
        "ex_id": "4595145951",
        "ex_cost_center_code": "1200845951"
    }
},
{
    "_id": "4400606041526566912",
    "name": "郑州高新万达店",
    "extend_code": {
        "comm_shop_id": "c678b5ecd6234c9f934f9d43040d3558",
        "ex_code": "45977",
        "us_id": "45977",
        "upcard_terminal": "37115009",
        "upcard_mer_id": "102371089993040",
        "ex_id": "45977",
        "ex_cost_center_code": "1200845977"
    }
},
{
    "_id": "4401656886586277888",
    "name": "安顺国贸店",
    "extend_code": {
        "comm_shop_id": "0b280e6e86484484ad135a60e1c8d6ff",
        "ex_code": "45994",
        "alipay_id": "2020101200077000000007013527",
        "us_id": "45994",
        "upcard_terminal": "85102854",
        "upcard_mer_id": "102851089990595",
        "ex_id": "45994",
        "ex_cost_center_code": "1200845994"
    }
},
{
    "_id": "4401659878639173632",
    "name": "上海港城新天地店",
    "extend_code": {
        "comm_shop_id": "795aa53fe2b344f3b520721585ce07ff",
        "ex_code": "45984",
        "us_id": "45984",
        "upcard_terminal": "02113303",
        "upcard_mer_id": "102210089999463",
        "ex_id": "45984",
        "ex_cost_center_code": "1200845984"
    }
},
{
    "_id": "4401661382624313344",
    "name": "成都东航中心店",
    "extend_code": {
        "comm_shop_id": "f1becacad17047f19635c0f028253b3c",
        "ex_code": "45986",
        "us_id": "45986",
        "upcard_terminal": "02833443",
        "upcard_mer_id": "102280089997757",
        "ex_id": "45986",
        "ex_cost_center_code": "1200845986"
    }
},
{
    "_id": "4402072802507685888",
    "name": "西安香醍龙湖店",
    "extend_code": {
        "comm_shop_id": "647d4a9e6e334f2aa7ff9a68affdc0c2",
        "ex_code": "45985",
        "alipay_id": "2020122200077000000012946463",
        "us_id": "45985",
        "upcard_terminal": "02907484",
        "upcard_mer_id": "102290089993445",
        "ex_id": "45985",
        "ex_cost_center_code": "1200845985"
    }
},
{
    "_id": "4402074319969124352",
    "name": "合肥新桥机场店",
    "extend_code": {
        "comm_shop_id": "6b257c07f6e44d1fa9142ba9f9cdabb4",
        "ex_code": "45982",
        "us_id": "45982",
        "upcard_terminal": "55132629",
        "upcard_mer_id": "102551089993972",
        "ex_id": "45982",
        "ex_cost_center_code": "1200845982"
    }
},
{
    "_id": "4403513639892451328",
    "name": "徐州沛县城投广场店",
    "extend_code": {
        "comm_shop_id": "efd70c9b5a174b6eaf310439a0f750eb",
        "ex_code": "45988",
        "alipay_id": "2021070700077000000023828581",
        "us_id": "45988",
        "upcard_terminal": "51601883",
        "upcard_mer_id": "102516089990316",
        "ex_id": "45988",
        "ex_cost_center_code": "1200845988"
    }
},
{
    "_id": "4403786764840402944",
    "name": "合肥罍街店",
    "extend_code": {
        "comm_shop_id": "04c6e542ccae463f9cd9cebdee554381",
        "ex_code": "45983",
        "alipay_id": "2020120900077000000010444092",
        "us_id": "45983",
        "upcard_terminal": "55132628",
        "upcard_mer_id": "102551089993971",
        "ex_id": "45983",
        "ex_cost_center_code": "1200845983"
    }
},
{
    "_id": "4403788646925271040",
    "name": "淮南吾悦店",
    "extend_code": {
        "comm_shop_id": "6dbf1788ce874413a13c0baf17edd2ba",
        "ex_code": "45997",
        "alipay_id": "2020110500077000000008038966",
        "us_id": "45997",
        "upcard_terminal": "55401101",
        "upcard_mer_id": "102554089990141",
        "ex_id": "45997",
        "ex_cost_center_code": "1200845997"
    }
},
{
    "_id": "4404589051254996992",
    "name": "西宁万达店",
    "extend_code": {
        "comm_shop_id": "7cf1d4e7e0074733adeb4b6b1401ec5d",
        "ex_code": "45998",
        "alipay_id": "2021110500077000000029760139",
        "us_id": "45998",
        "upcard_terminal": "97100419",
        "upcard_mer_id": "102971089990251",
        "ex_id": "45998",
        "ex_cost_center_code": "1200845998"
    }
},
{
    "_id": "4404591784699658240",
    "name": "东台黄海森林公园店",
    "extend_code": {
        "comm_shop_id": "7ecfcb5027344440a0a57384e3efb15d",
        "ex_code": "45996",
        "us_id": "45996",
        "upcard_terminal": "51501134",
        "upcard_mer_id": "102515089990501",
        "ex_id": "45996",
        "ex_cost_center_code": "1200845996"
    }
},
{
    "_id": "4405687701762801664",
    "name": "射阳吾悦店",
    "extend_code": {
        "comm_shop_id": "2fb53e98e018475488fef0494193ff9b",
        "ex_code": "45993",
        "alipay_id": "2021101900077000000029063759",
        "us_id": "45993",
        "upcard_terminal": "51501135",
        "upcard_mer_id": "102515089990502",
        "ex_id": "45993",
        "ex_cost_center_code": "1200845993"
    }
},
{
    "_id": "4405936798340022272",
    "name": "信阳万达店",
    "extend_code": {
        "comm_shop_id": "cd7d6ed3592449429d415249446bb320",
        "ex_code": "45990",
        "alipay_id": "2021111100077000000030078261",
        "us_id": "45990",
        "upcard_terminal": "37603181",
        "upcard_mer_id": "102376089990175",
        "ex_id": "45990",
        "ex_cost_center_code": "1200845990"
    }
},
{
    "_id": "4407843470004256768",
    "name": "西安Momopark店",
    "extend_code": {
        "comm_shop_id": "17d5397a67cd4522a0671c813b4bfc83",
        "ex_code": "45991",
        "alipay_id": "2020110500077000000008032354",
        "us_id": "45991",
        "upcard_terminal": "02907537",
        "upcard_mer_id": "102290089993456",
        "ex_id": "45991",
        "ex_cost_center_code": "1200845991"
    }
},
{
    "_id": "4407845446486097920",
    "name": "杭州永旺梦乐城店",
    "extend_code": {
        "comm_shop_id": "4fa93966c9fc4c97a2b551c7da76b0c0",
        "ex_code": "46005",
        "alipay_id": "2020122200077000000012948913",
        "us_id": "46005",
        "upcard_terminal": "57115507",
        "upcard_mer_id": "102571089993277",
        "ex_id": "46005",
        "ex_cost_center_code": "1200846005"
    }
},
{
    "_id": "4408217690907377664",
    "name": "贵阳益田假日里店",
    "extend_code": {
        "comm_shop_id": "4a7676c354c444c5bd66495b3113e981",
        "ex_code": "46011",
        "alipay_id": "2021102600077000000029273728",
        "us_id": "46011",
        "upcard_terminal": "85102861",
        "upcard_mer_id": "102851089990600",
        "ex_id": "46011",
        "ex_cost_center_code": "1200846011"
    }
},
{
    "_id": "4408219408118349824",
    "name": "崇明万达店",
    "extend_code": {
        "comm_shop_id": "3b9d9a7811154c0980243332dc8e62c7",
        "ex_code": "46009",
        "us_id": "46009",
        "upcard_terminal": "02113355",
        "upcard_mer_id": "102210089999481",
        "ex_id": "46009",
        "ex_cost_center_code": "1200846009"
    }
},
{
    "_id": "4408220817438048256",
    "name": "湖州吾悦店",
    "extend_code": {
        "comm_shop_id": "2371be9663154f528453766644d8f33a",
        "ex_code": "46010",
        "alipay_id": "2020110600077000000008084336",
        "us_id": "46010",
        "upcard_terminal": "57297333",
        "upcard_mer_id": "102572089990100",
        "ex_id": "46010",
        "ex_cost_center_code": "1200846010"
    }
},
{
    "_id": "4408222122382491649",
    "name": "徐州贾汪吾悦店",
    "extend_code": {
        "comm_shop_id": "10e53a4fb4bb4e959af6581564f9c074",
        "ex_code": "46006",
        "us_id": "46006",
        "upcard_terminal": "51601884",
        "upcard_mer_id": "102516089990317",
        "ex_id": "46006",
        "ex_cost_center_code": "1200846006"
    }
},
{
    "_id": "4408849400971984896",
    "name": "赣州万象城店",
    "extend_code": {
        "comm_shop_id": "3dffd5a0308744f1b14c3f73d9dd50eb",
        "ex_code": "46013",
        "us_id": "46013",
        "upcard_terminal": "79701556",
        "upcard_mer_id": "102797089990406",
        "ex_id": "46013",
        "ex_cost_center_code": "1200846013"
    }
},
{
    "_id": "4414360281961005056",
    "name": "宝鸡天下汇店",
    "extend_code": {
        "comm_shop_id": "c84bb139ee234319900fefa00ad78d79",
        "ex_code": "46016",
        "alipay_id": "2021102600077000000029272258",
        "us_id": "46016",
        "upcard_terminal": "91700305",
        "upcard_mer_id": "102917089990037",
        "ex_id": "46016",
        "ex_cost_center_code": "1200846016"
    }
},
{
    "_id": "4414362017975042048",
    "name": "如皋金雅店",
    "extend_code": {
        "comm_shop_id": "f2d8a20a7a594dd29bd0e574588840de",
        "ex_code": "46008",
        "us_id": "46008",
        "upcard_terminal": "51301541",
        "upcard_mer_id": "102513089990289 ",
        "ex_id": "46008",
        "ex_cost_center_code": "1200846008"
    }
},
{
    "_id": "4415803368423882752",
    "name": "海门狮山店",
    "extend_code": {
        "comm_shop_id": "b42c3b8f0e9d47b19c101b7ba045c277",
        "ex_code": "46037",
        "us_id": "46037",
        "upcard_terminal": "51301542",
        "upcard_mer_id": "102513089990290",
        "ex_id": "46037",
        "ex_cost_center_code": "1200846037"
    }
},
{
    "_id": "4415822172902162432",
    "name": "三亚蓝海广场店",
    "extend_code": {
        "comm_shop_id": "c299839c3b954b4f8500aa1e9725aa7b",
        "ex_code": "46030",
        "alipay_id": "2020120900077000000010443362",
        "us_id": "46030",
        "upcard_terminal": "89801987",
        "upcard_mer_id": "102898089990645",
        "ex_id": "46030",
        "ex_cost_center_code": "1200846030"
    }
},
{
    "_id": "4416157027632447488",
    "name": "东台吾悦店",
    "extend_code": {
        "comm_shop_id": "b2e0f12b41714b47b549b7ca2ed00ef5",
        "ex_code": "46032",
        "us_id": "46032",
        "upcard_terminal": "51501196",
        "upcard_mer_id": "102515089990505",
        "ex_id": "46032",
        "ex_cost_center_code": "1200846032"
    }
},
{
    "_id": "4416842214259326976",
    "name": "苏州星湖龙湖天街店",
    "extend_code": {
        "comm_shop_id": "f9c229bf2e4445109fd8c5055321eafd",
        "ex_code": "46039",
        "alipay_id": "2020123000077000000013302468",
        "us_id": "46039",
        "upcard_terminal": "51218007",
        "upcard_mer_id": "102512089993429",
        "ex_id": "46039",
        "ex_cost_center_code": "1200846039"
    }
},
{
    "_id": "4416875245665845248",
    "name": "银川Ccmall店",
    "extend_code": {
        "comm_shop_id": "cabcea1ee47f43cfbe09818df2226a17",
        "ex_code": "46038",
        "alipay_id": "2020122200077000000012948914",
        "us_id": "46038",
        "upcard_terminal": "95102632",
        "upcard_mer_id": "102951089991008",
        "ex_id": "46038",
        "ex_cost_center_code": "1200846038"
    }
},
{
    "_id": "4416880132327800832",
    "name": "海南儋州恒大海花岛沙滩吧店",
    "extend_code": {
        "comm_shop_id": "41f48e2c38b64e6db4bd2154b5911df4",
        "ex_code": "46031",
        "alipay_id": "2021020900077000000016669999",
        "us_id": "46031",
        "upcard_terminal": "89801988",
        "upcard_mer_id": "102898089990646",
        "ex_id": "46031",
        "ex_cost_center_code": "1200846031"
    }
},
{
    "_id": "4416885632851279872",
    "name": "济南龙湖奥体天街店",
    "extend_code": {
        "comm_shop_id": "8c6f63c98b474f50b758097fae3450c9",
        "ex_code": "46015",
        "alipay_id": "2021010800077000000013636949",
        "us_id": "46015",
        "upcard_terminal": "53101824",
        "upcard_mer_id": "102531089990776",
        "ex_id": "46015",
        "ex_cost_center_code": "1200846015"
    }
},
{
    "_id": "4416887417410224128",
    "name": "东台德润店",
    "extend_code": {
        "comm_shop_id": "c301336e623945689c01cc3b24d6a862",
        "ex_code": "46033",
        "us_id": "46033",
        "upcard_terminal": "51501197",
        "upcard_mer_id": "102515089990506",
        "ex_id": "46033",
        "ex_cost_center_code": "1200846033"
    }
},
{
    "_id": "4417276235204329472",
    "name": "盐城悦达889店",
    "extend_code": {
        "comm_shop_id": "d221c1fe18b4436790eaded140f2a74d",
        "ex_code": "46043",
        "alipay_id": "2021101900077000000029058685",
        "us_id": "46043",
        "upcard_terminal": "51501198",
        "upcard_mer_id": "102515089990507",
        "ex_id": "46043",
        "ex_cost_center_code": "1200846043"
    }
},
{
    "_id": "4419470449266688000",
    "name": "自贡万达店",
    "extend_code": {
        "comm_shop_id": "2f3ca04ff9b9499ba53e24a8aee7fa93",
        "ex_code": "46051",
        "alipay_id": "2021102600077000000029273855",
        "us_id": "46051",
        "upcard_terminal": "81301079",
        "upcard_mer_id": "102813089990063",
        "ex_id": "46051",
        "ex_cost_center_code": "1200846051"
    }
},
{
    "_id": "4420799278480424960",
    "name": "海口日月广场店",
    "extend_code": {
        "comm_shop_id": "6f271266cc5d45a0a48375e80e022658",
        "ex_code": "46047",
        "alipay_id": "2020120900077000000010443364",
        "us_id": "46047",
        "upcard_terminal": "89801989",
        "upcard_mer_id": "102898089990647",
        "ex_id": "46047",
        "ex_cost_center_code": "1200846047"
    }
},
{
    "_id": "4420800342109749248",
    "name": "济宁兖州贵和广场店",
    "extend_code": {
        "comm_shop_id": "e3390a49a14947a4a121ae6a4cf0d6d0",
        "ex_code": "46046",
        "alipay_id": "2020123000077000000013302471",
        "us_id": "46046",
        "upcard_terminal": "53701206",
        "upcard_mer_id": "102537089990184",
        "ex_id": "46046",
        "ex_cost_center_code": "1200846046"
    }
},
{
    "_id": "4420802479942303744",
    "name": "海门大有镜店",
    "extend_code": {
        "comm_shop_id": "aaed0dca070a45acb3329062b3e04c50",
        "ex_code": "46036",
        "alipay_id": "2020123000077000000013302472",
        "us_id": "46036",
        "upcard_terminal": "51301593",
        "upcard_mer_id": "102513089990292",
        "ex_id": "46036",
        "ex_cost_center_code": "1200846036"
    }
},
{
    "_id": "4420917981876355072",
    "name": "南昌铜锣湾店",
    "extend_code": {
        "comm_shop_id": "d464acd58d584618b5218182ac98d21f",
        "ex_code": "46041",
        "alipay_id": "2020122200077000000012946462",
        "us_id": "46041",
        "upcard_terminal": "79192352",
        "upcard_mer_id": "102791089990771",
        "ex_id": "46041",
        "ex_cost_center_code": "1200046041"
    }
},
{
    "_id": "4421619118006927360",
    "name": "贵阳玖福城店",
    "extend_code": {
        "comm_shop_id": "27e3a766da6c4ffd89cf99dce248954d",
        "ex_code": "46042",
        "alipay_id": "2021010800077000000013636948",
        "us_id": "46042",
        "upcard_terminal": "85102967",
        "upcard_mer_id": "102851089990603",
        "ex_id": "46042",
        "ex_cost_center_code": "1200046042"
    }
},
{
    "_id": "4421622254952022016",
    "name": "象山万达店",
    "extend_code": {
        "comm_shop_id": "fe4dbad191424c068d2554461669d60b",
        "ex_code": "46053",
        "alipay_id": "2020122200077000000012945558",
        "us_id": "46053",
        "upcard_terminal": "57405350",
        "upcard_mer_id": "102574089990747",
        "ex_id": "46053",
        "ex_cost_center_code": "1200846053"
    }
},
{
    "_id": "4421623610530103296",
    "name": "天长吾悦店",
    "extend_code": {
        "comm_shop_id": "3e62f0ffe039473b8bcea7204b872cc1",
        "ex_code": "46055",
        "alipay_id": "2021102600077000000029271066",
        "us_id": "46055",
        "upcard_terminal": "55000701",
        "upcard_mer_id": "102550089990093",
        "ex_id": "46055",
        "ex_cost_center_code": "1200846055"
    }
},
{
    "_id": "4421624774831144960",
    "name": "荆州吾悦店",
    "extend_code": {
        "comm_shop_id": "e0fdf674cdad42cf9d7c4017f5ddca70",
        "ex_code": "46054",
        "alipay_id": "2021092300077000000027995340",
        "us_id": "46054",
        "upcard_terminal": "71600314",
        "upcard_mer_id": "102716089990083",
        "ex_id": "46054",
        "ex_cost_center_code": "1200846054"
    }
},
{
    "_id": "4421631425105690624",
    "name": "商丘港汇万达店",
    "extend_code": {
        "comm_shop_id": "1f298050992f451abb5a932d361ff679",
        "ex_code": "46049",
        "us_id": "46049",
        "upcard_terminal": "37001399",
        "upcard_mer_id": "102370089990045",
        "ex_id": "46049",
        "ex_cost_center_code": "1200846049"
    }
},
{
    "_id": "4422356409788727296",
    "name": "上海LCM置汇旭辉店",
    "extend_code": {
        "comm_shop_id": "273e93cd3d7b438f91e773c638610ed7",
        "ex_code": "46040",
        "alipay_id": "2020123000077000000013302467",
        "us_id": "46040",
        "upcard_terminal": "02113504",
        "upcard_mer_id": "102210089999500",
        "ex_id": "46040",
        "ex_cost_center_code": "1200046040"
    }
},
{
    "_id": "4422362271861211136",
    "name": "温州万象城店",
    "extend_code": {
        "comm_shop_id": "ca8989ff40714b67a39d4ea9221c0daa",
        "ex_code": "46048",
        "alipay_id": "2020120900077000000010444090",
        "us_id": "46048",
        "upcard_terminal": "57701378",
        "upcard_mer_id": "102577089990258",
        "ex_id": "46048",
        "ex_cost_center_code": "1200846048"
    }
},
{
    "_id": "4423789834601005056",
    "name": "扬州蜀冈万达店",
    "extend_code": {
        "comm_shop_id": "4ba859d9d27649daa74a2a479f7a0ec6",
        "ex_code": "46050",
        "alipay_id": "2021010800077000000013701394",
        "us_id": "46050",
        "upcard_terminal": "51402125",
        "upcard_mer_id": "102514089990371",
        "ex_id": "46050",
        "ex_cost_center_code": "1200846050"
    }
},
{
    "_id": "4423791452918677504",
    "name": "宜昌夷陵万达店",
    "extend_code": {
        "comm_shop_id": "46ec6eeface54f69a2dd10d3ed190dc5",
        "ex_code": "46056",
        "alipay_id": "2020123000077000000013301149",
        "us_id": "46056",
        "upcard_terminal": "71703761",
        "upcard_mer_id": "102717089990305",
        "ex_id": "46056",
        "ex_cost_center_code": "1200846056"
    }
},
{
    "_id": "4423794107384659968",
    "name": "周口开元万达店",
    "extend_code": {
        "comm_shop_id": "d4ffc5f60c0a4f8fbdc4e9a02ce98f48",
        "ex_code": "46063",
        "us_id": "46063",
        "upcard_terminal": "39401207",
        "upcard_mer_id": "102394089990056",
        "ex_id": "46063",
        "ex_cost_center_code": "1200846063"
    }
},
{
    "_id": "4423798981933465600",
    "name": "重庆悦荟店",
    "extend_code": {
        "comm_shop_id": "7bbed0fbd9a14bd0adcbcdbc40fec165",
        "ex_code": "46062",
        "us_id": "46062",
        "upcard_terminal": "02312354",
        "upcard_mer_id": "102230089992937",
        "ex_id": "46062",
        "ex_cost_center_code": "1200046062"
    }
},
{
    "_id": "4426307962639810560",
    "name": "张家港锦丰东服务区店",
    "extend_code": {
        "comm_shop_id": "b743d456c22649adaee9465495c82aa1",
        "ex_code": "45938",
        "us_id": "45938",
        "upcard_terminal": "51218046",
        "upcard_mer_id": "102512089993434",
        "ex_id": "45938",
        "ex_cost_center_code": "1200845938"
    }
},
{
    "_id": "4426309182842503168",
    "name": "张家港锦丰西服务区店",
    "extend_code": {
        "comm_shop_id": "7c62e06f0b064ddfa41d59c84161471f",
        "ex_code": "45939",
        "us_id": "45939",
        "upcard_terminal": "51218011",
        "upcard_mer_id": "102512089993430 ",
        "ex_id": "45939",
        "ex_cost_center_code": "1200845939"
    }
},
{
    "_id": "4426312133959024640",
    "name": "西安砂之船奥特莱斯店",
    "extend_code": {
        "comm_shop_id": "9a9c6b66e4774090b9db3ff09e678e9c",
        "ex_code": "46069",
        "alipay_id": "2021010800077000000013701404",
        "us_id": "46069",
        "upcard_terminal": "02907561",
        "upcard_mer_id": "102290089993459",
        "ex_id": "46069",
        "ex_cost_center_code": "1200846069"
    }
},
{
    "_id": "4427401656923095040",
    "name": "西安盛安广场店",
    "extend_code": {
        "comm_shop_id": "3a1df4ce53844084b47049ef34e6a6be",
        "ex_code": "46070",
        "alipay_id": "2021061000077000000021912673",
        "us_id": "46070",
        "upcard_terminal": "02907562",
        "upcard_mer_id": "102290089993460",
        "ex_id": "46070",
        "ex_cost_center_code": "1200846070"
    }
},
{
    "_id": "4427409902698266624",
    "name": "银川新华联店",
    "extend_code": {
        "comm_shop_id": "1602ea988fc74fbca57c24b7b2a11cd6",
        "ex_code": "46071",
        "alipay_id": "2021010800077000000013636947",
        "us_id": "46071",
        "upcard_terminal": "95102633",
        "upcard_mer_id": "102951089991009",
        "ex_id": "46071",
        "ex_cost_center_code": "1200846071"
    }
},
{
    "_id": "4429875275636965376",
    "name": "深圳龙华壹方天地店",
    "extend_code": {
        "comm_shop_id": "30574b52d0ff4000a1c9f25178ab1741",
        "ex_code": "46072",
        "alipay_id": "2020123000077000000013302469",
        "us_id": "46072",
        "upcard_terminal": "75525697",
        "upcard_mer_id": "102755089996010",
        "ex_id": "46072",
        "ex_cost_center_code": "1200846072"
    }
},
{
    "_id": "4429880672435372032",
    "name": "西安回民街店",
    "extend_code": {
        "comm_shop_id": "32cf750f67cc4ab3ace3d1938e60504b",
        "ex_code": "46057",
        "alipay_id": "2020123000077000000013306039",
        "us_id": "46057",
        "upcard_terminal": "02907619",
        "upcard_mer_id": "102290089993517",
        "ex_id": "46057",
        "ex_cost_center_code": "1200846057"
    }
},
{
    "_id": "4429882886415482880",
    "name": "嘉兴旭辉广场店",
    "extend_code": {
        "comm_shop_id": "dbe95e8f84d94a1c952ec45e9d328713",
        "ex_code": "46076",
        "alipay_id": "2021010800077000000013701392",
        "us_id": "46076",
        "upcard_terminal": "57304467",
        "upcard_mer_id": "102573089990227",
        "ex_id": "46076",
        "ex_cost_center_code": "1200846076"
    }
},
{
    "_id": "4432440369235427328",
    "name": "大同百盛2店",
    "extend_code": {
        "comm_shop_id": "59e63db5f2fe46109926798599291611",
        "ex_code": "46079",
        "alipay_id": "2021070600077000000023799114",
        "us_id": "46079",
        "upcard_terminal": "35200304",
        "upcard_mer_id": "102352089990053",
        "ex_id": "46079",
        "ex_cost_center_code": "1200846079"
    }
},
{
    "_id": "4432443063366877184",
    "name": "昆明1903店",
    "extend_code": {
        "comm_shop_id": "b7e6dda0f1ef4456a331a65f7313ab52",
        "ex_code": "46080",
        "alipay_id": "2021102600077000000029271052",
        "us_id": "46080",
        "upcard_terminal": "87114112",
        "upcard_mer_id": "102871089997006",
        "ex_id": "46080",
        "ex_cost_center_code": "1200846080"
    }
},
{
    "_id": "4432465507276357632",
    "name": "涟水吾悦店",
    "extend_code": {
        "comm_shop_id": "7c93cb75b32545ac89e2c5f58467d1a6",
        "ex_code": "46014",
        "alipay_id": "2020123000077000000013301150",
        "us_id": "46014",
        "upcard_terminal": "51701986",
        "upcard_mer_id": "102517089990122",
        "ex_id": "46014",
        "ex_cost_center_code": "1200846014"
    }
},
{
    "_id": "4432466407449493504",
    "name": "银川宁阳广场店",
    "extend_code": {
        "comm_shop_id": "c77b80353ca44103ac60f922cbe57271",
        "ex_code": "46082",
        "us_id": "46082",
        "upcard_terminal": "95102639",
        "upcard_mer_id": "102951089991011",
        "ex_id": "46082",
        "ex_cost_center_code": "1200846082"
    }
},
{
    "_id": "4432470071329193984",
    "name": "温州龙湾吾悦店",
    "extend_code": {
        "comm_shop_id": "ad0218c0e95f42df8056b5b60f09d22a",
        "ex_code": "46084",
        "alipay_id": "2020122200077000000012946464",
        "us_id": "46084",
        "upcard_terminal": "57701430",
        "upcard_mer_id": "102577089990261",
        "ex_id": "46084",
        "ex_cost_center_code": "1200846084"
    }
},
{
    "_id": "4433591864915951616",
    "name": "温州南站店",
    "extend_code": {
        "comm_shop_id": "9d2269e730bf4581b4e5dcae8033075a",
        "ex_code": "46081",
        "us_id": "46081",
        "upcard_terminal": "57701429",
        "upcard_mer_id": "102577089990260",
        "ex_id": "46081",
        "ex_cost_center_code": "1200846081"
    }
},
{
    "_id": "4433920505130582016",
    "name": "苏州站店",
    "extend_code": {
        "comm_shop_id": "52b0babc6d274451b896f62c7ec83fd4",
        "ex_code": "46089",
        "us_id": "46089",
        "upcard_terminal": "51218019",
        "upcard_mer_id": "102512089993432",
        "ex_id": "46089",
        "ex_cost_center_code": "1200846089"
    }
},
{
    "_id": "4433922116049502208",
    "name": "西安集乐里店",
    "extend_code": {
        "comm_shop_id": "a1cdfd4ef8164a7fb41fe24c18b84bd1",
        "ex_code": "46088",
        "us_id": "46088",
        "upcard_terminal": "02907623",
        "upcard_mer_id": "102290089993521",
        "ex_id": "46088",
        "ex_cost_center_code": "1200846088"
    }
},
{
    "_id": "4433942470004342784",
    "name": "芜湖古城店",
    "extend_code": {
        "comm_shop_id": "34044363a0ad41a2b650a2fc3609d678",
        "ex_code": "46087",
        "alipay_id": "2021102600077000000029272277",
        "us_id": "46087",
        "upcard_terminal": "55301917",
        "upcard_mer_id": "102553089990433",
        "ex_id": "46087",
        "ex_cost_center_code": "1200846087"
    }
},
{
    "_id": "4434268126634377216",
    "name": "广州番禺天河城店",
    "extend_code": {
        "comm_shop_id": "d6cc2cabf7554010815b90025f4acb92",
        "ex_code": "46086",
        "alipay_id": "2020123000077000000013302501",
        "us_id": "46086",
        "upcard_terminal": "02005582",
        "upcard_mer_id": "102200089991278",
        "ex_id": "46086",
        "ex_cost_center_code": "1200046086"
    }
},
{
    "_id": "4436386345595764736",
    "name": "胶州宝龙店",
    "extend_code": {
        "comm_shop_id": "e241196986b549389ddf6fee8de379fa",
        "ex_code": "46096",
        "alipay_id": "2021062200077000000022905442",
        "us_id": "46096",
        "upcard_terminal": "53206622",
        "upcard_mer_id": "102532089991691",
        "ex_id": "46096",
        "ex_cost_center_code": "1200846096"
    }
},
{
    "_id": "4436387402715561984",
    "name": "兰州老街店",
    "extend_code": {
        "comm_shop_id": "20adce2004f844f882687c27f92bdcdf",
        "ex_code": "46077",
        "alipay_id": "2021110500077000000029762930",
        "us_id": "46077",
        "upcard_terminal": "93101719",
        "upcard_mer_id": "102931089990251",
        "ex_id": "46077",
        "ex_cost_center_code": "1200846077"
    }
},
{
    "_id": "4436390072629100544",
    "name": "乐山伊藤洋华堂店",
    "extend_code": {
        "comm_shop_id": "8911a79f07fd447bb14cdf4d6a68c279",
        "ex_code": "46102",
        "alipay_id": "2021010800077000000013701393",
        "us_id": "46102",
        "upcard_terminal": "83390434",
        "upcard_mer_id": "102833089990047",
        "ex_id": "46102",
        "ex_cost_center_code": "1200046102"
    }
},
{
    "_id": "4436391664690462720",
    "name": "苏州悠方店",
    "extend_code": {
        "comm_shop_id": "7dfdefe228e648138efdef0ff34287e3",
        "ex_code": "46100",
        "alipay_id": "2021012900077000000016127858",
        "us_id": "46100",
        "upcard_terminal": "51218038",
        "upcard_mer_id": "102512089993433",
        "ex_id": "46100",
        "ex_cost_center_code": "1200846100"
    }
},
{
    "_id": "4436498380098699264",
    "name": "济南中海环宇城店",
    "extend_code": {
        "comm_shop_id": "11074c99cfe546d8b2e6f9886bed6f85",
        "ex_code": "46052",
        "us_id": "46052",
        "upcard_terminal": "53101866",
        "upcard_mer_id": "102531089990779",
        "ex_id": "46052",
        "ex_cost_center_code": "1200846052"
    }
},
{
    "_id": "4436499811178446849",
    "name": "焦作云台山景区店",
    "extend_code": {
        "comm_shop_id": "ff7befda7ae2460fb96db1110835aa7a",
        "ex_code": "46115",
        "us_id": "46115",
        "upcard_terminal": "39102510",
        "upcard_mer_id": "102391089990195",
        "ex_id": "46115",
        "ex_cost_center_code": "1200846115"
    }
},
{
    "_id": "4438933763541467136",
    "name": "济南莱芜茂业店",
    "extend_code": {
        "comm_shop_id": "7f7850df61fc4d1eb3942f7bfa161051",
        "ex_code": "46097",
        "alipay_id": "2021053100077000000021244575",
        "us_id": "46097",
        "upcard_terminal": "53101865",
        "upcard_mer_id": "102531089990778",
        "ex_id": "46097",
        "ex_cost_center_code": "1200846097"
    }
},
{
    "_id": "4438934844765929472",
    "name": "三亚凤凰国际机场店",
    "extend_code": {
        "comm_shop_id": "446d3d17e2db4551a96dc5b919a2907b",
        "ex_code": "46118",
        "alipay_id": "2021020900077000000016657460",
        "us_id": "46118",
        "upcard_terminal": "89802005",
        "upcard_mer_id": "102898089990652",
        "ex_id": "46118",
        "ex_cost_center_code": "1200846118"
    }
},
{
    "_id": "4439652755734298624",
    "name": "商丘夏邑亿舟城店",
    "extend_code": {
        "comm_shop_id": "d6244cce06e449b48e56ae7568049266",
        "ex_code": "46117",
        "us_id": "46117",
        "upcard_terminal": "37001400",
        "upcard_mer_id": "102370089990046",
        "ex_id": "46117",
        "ex_cost_center_code": "1200846117"
    }
},
{
    "_id": "4439654842731921408",
    "name": "济南融创茂店",
    "extend_code": {
        "comm_shop_id": "09e34caa0fe14df49216f30eed354ea4",
        "ex_code": "46114",
        "alipay_id": "2021062400077000000023016366",
        "us_id": "46114",
        "upcard_terminal": "53101867",
        "upcard_mer_id": "102531089990780",
        "ex_id": "46114",
        "ex_cost_center_code": "1200846114"
    }
},
{
    "_id": "4439657578621894656",
    "name": "郑州新郑新尚天地店",
    "extend_code": {
        "comm_shop_id": "a11a9131c3224a7c8dbf51b26dfac1dd",
        "ex_code": "46135",
        "us_id": "46135",
        "upcard_terminal": "37115166",
        "upcard_mer_id": "102371089993057",
        "ex_id": "46135",
        "ex_cost_center_code": "1200846135"
    }
},
{
    "_id": "4439658471584989184",
    "name": "西安永宁里店",
    "extend_code": {
        "comm_shop_id": "dc99fa8bf90d43459d9b61b91931a938",
        "ex_code": "46140",
        "alipay_id": "2021061000077000000021911193",
        "us_id": "46140",
        "upcard_terminal": "02907667",
        "upcard_mer_id": "102290089993524",
        "ex_id": "46140",
        "ex_cost_center_code": "1200846140"
    }
},
{
    "_id": "4439659372492161024",
    "name": "合肥中环店",
    "extend_code": {
        "comm_shop_id": "54bfe2147ae9405596b22fd6f6a8c2d6",
        "ex_code": "46136",
        "alipay_id": "2021102600077000000029273726",
        "us_id": "46136",
        "upcard_terminal": "55132760",
        "upcard_mer_id": "102551089993981",
        "ex_id": "46136",
        "ex_cost_center_code": "1200846136"
    }
},
{
    "_id": "4441458834847891456",
    "name": "驻马店爱克玖隆茂店",
    "extend_code": {
        "comm_shop_id": "fc319e5fb2d04ec1a22578a270913a3a",
        "ex_code": "46152",
        "alipay_id": "2021012900077000000016127859",
        "us_id": "46152",
        "upcard_terminal": "39601596",
        "upcard_mer_id": "102396089990071",
        "ex_id": "46152",
        "ex_cost_center_code": "1200846152"
    }
},
{
    "_id": "4441460512779173888",
    "name": "重庆光环购物中心店",
    "extend_code": {
        "comm_shop_id": "4b4aa3a7a29c4bc2afdd7c57f49edec4",
        "ex_code": "46151",
        "alipay_id": "2021042700077000000019734196",
        "us_id": "46151",
        "upcard_terminal": "02312459",
        "upcard_mer_id": "102230089992960",
        "ex_id": "46151",
        "ex_cost_center_code": "1200846151"
    }
},
{
    "_id": "4441461591789076480",
    "name": "上海虹桥龙湖天街店",
    "extend_code": {
        "comm_shop_id": "704474d709e9432189afd4ed2dcbf039",
        "ex_code": "46148",
        "alipay_id": "2021062200077000000022902520",
        "us_id": "46148",
        "upcard_terminal": "02113738",
        "upcard_mer_id": "102210089999562",
        "ex_id": "46148",
        "ex_cost_center_code": "1200846148"
    }
},
{
    "_id": "4441467270625001473",
    "name": "海口美兰国际机场店",
    "extend_code": {
        "comm_shop_id": "bf402a491adc42a9bdaf1e4de6eaafa9",
        "ex_code": "46139",
        "us_id": "46139",
        "upcard_terminal": "89802006",
        "upcard_mer_id": "102898089990653",
        "ex_id": "46139",
        "ex_cost_center_code": "1200846139"
    }
},
{
    "_id": "4441471212738084864",
    "name": "昭通彝良世纪金街店",
    "extend_code": {
        "comm_shop_id": "4ee8dc1f68e1476f8f860ccf2b155d48",
        "ex_code": "46137",
        "us_id": "46137",
        "upcard_terminal": "87000274",
        "upcard_mer_id": "102870089990186",
        "ex_id": "46137",
        "ex_cost_center_code": "1200846137"
    }
},
{
    "_id": "4441472500494303232",
    "name": "泰州万象城二期店",
    "extend_code": {
        "comm_shop_id": "7da129c14df8450ea06f01248ffce67f",
        "ex_code": "46141",
        "us_id": "46141",
        "upcard_terminal": "52301064",
        "upcard_mer_id": "102523089990153",
        "ex_id": "46141",
        "ex_cost_center_code": "1200846141"
    }
},
{
    "_id": "4441478441751085056",
    "name": "兰州欣大百货店",
    "extend_code": {
        "comm_shop_id": "753501511a4b4b00a362e4c237cddf0c",
        "ex_code": "46150",
        "alipay_id": "2021110500077000000029762929",
        "us_id": "46150",
        "upcard_terminal": "93101720",
        "upcard_mer_id": "102931089990252",
        "ex_id": "46150",
        "ex_cost_center_code": "1200846150"
    }
},
{
    "_id": "4441505005247201280",
    "name": "铜仁时代商汇店",
    "extend_code": {
        "comm_shop_id": "b8e777a367c04acbb7510fca362f28a2",
        "ex_code": "46112",
        "alipay_id": "2021102600077000000029271050",
        "us_id": "46112",
        "upcard_terminal": "85600036",
        "upcard_mer_id": "102856089990010",
        "ex_id": "46112",
        "ex_cost_center_code": "1200846112"
    }
},
{
    "_id": "4441506408321581056",
    "name": "贵州都匀中寰广场店",
    "extend_code": {
        "comm_shop_id": "ce1752d45d394718bd6a4aa3ac32ee17",
        "ex_code": "46138",
        "us_id": "46138",
        "upcard_terminal": "85410082",
        "upcard_mer_id": "102854089990025",
        "ex_id": "46138",
        "ex_cost_center_code": "1200846138"
    }
},
{
    "_id": "4451246041670418432",
    "name": "扬州砂之船奥特莱斯店",
    "extend_code": {
        "comm_shop_id": "89db22b94d354ccd8d9aefaa406008a9",
        "ex_code": "46154",
        "alipay_id": "2021020900077000000016657461",
        "us_id": "46154",
        "upcard_terminal": "51402161",
        "upcard_mer_id": "102514089990374",
        "ex_id": "46154",
        "ex_cost_center_code": "1200846154"
    }
},
{
    "_id": "4456014294720512000",
    "name": "福州宜家店",
    "extend_code": {
        "comm_shop_id": "ec41fbb5658640e79cd81d326378383e",
        "ex_code": "46155",
        "us_id": "46155",
        "upcard_terminal": "59114246",
        "upcard_mer_id": "102591089990702",
        "ex_id": "46155",
        "ex_cost_center_code": "1200046155"
    }
},
{
    "_id": "4467998408243740672",
    "name": "泰安吾悦店",
    "extend_code": {
        "comm_shop_id": "5e1ca4a9e521471f99870900e9f4986d",
        "ex_code": "46160",
        "us_id": "46160",
        "upcard_terminal": "53800718",
        "upcard_mer_id": "102538089990574",
        "ex_id": "46160",
        "ex_cost_center_code": "1200846160"
    }
},
{
    "_id": "4468006395129659392",
    "name": "南翔印象城店",
    "extend_code": {
        "comm_shop_id": "cae899d009f94eb380e0eeb74dedb1cd",
        "ex_code": "46164",
        "alipay_id": "2021081200077000000025982595",
        "us_id": "46164",
        "upcard_terminal": "02114208",
        "upcard_mer_id": "102210090000001",
        "ex_id": "46164",
        "ex_cost_center_code": "1200046164"
    }
},
{
    "_id": "4472269086652432384",
    "name": "厦门大学访客中心店",
    "extend_code": {
        "comm_shop_id": "ec1c6dbdd9f54c47b0fd36588b2d2ed8",
        "ex_code": "46162",
        "alipay_id": "2021042700077000000019734193",
        "us_id": "46162",
        "upcard_terminal": "59206137",
        "upcard_mer_id": "102592090000001",
        "ex_id": "46162",
        "ex_cost_center_code": "1200846162"
    }
},
{
    "_id": "4472272824813289472",
    "name": "合肥黉街店",
    "extend_code": {
        "comm_shop_id": "40b9483178d64efd83237d11766df92a",
        "ex_code": "46165",
        "alipay_id": "2021053100077000000021247686",
        "us_id": "46165",
        "upcard_terminal": "55132833",
        "upcard_mer_id": "102551090000001",
        "ex_id": "46165",
        "ex_cost_center_code": "1200846165"
    }
},
{
    "_id": "4477056217048481792",
    "name": "成都绿地伊藤468店",
    "extend_code": {
        "comm_shop_id": "a426be686ea945689b67f0169919cb4a",
        "ex_code": "46172",
        "alipay_id": "2021102600077000000029272426",
        "us_id": "46172",
        "upcard_terminal": "02835257",
        "upcard_mer_id": "102280090000015",
        "ex_id": "46172",
        "ex_cost_center_code": "1200046172"
    }
},
{
    "_id": "4477345921962672128",
    "name": "启东凤凰荟购物中心店",
    "extend_code": {
        "comm_shop_id": "2164cac59ebc4e2bade09067069ed306",
        "ex_code": "46183",
        "us_id": "46183",
        "upcard_terminal": "51301654",
        "upcard_mer_id": "102513090000001",
        "ex_id": "46183",
        "ex_cost_center_code": "1200846183"
    }
},
{
    "_id": "4477711045047123968",
    "name": "海南琼海环球春天店",
    "extend_code": {
        "comm_shop_id": "423fd8dbc0b24b7db5a323c95cccf86a",
        "ex_code": "46185",
        "alipay_id": "2021060100077000000021288032",
        "us_id": "46185",
        "upcard_terminal": "89802020",
        "upcard_mer_id": "102898090000002",
        "ex_id": "46185",
        "ex_cost_center_code": "1200846185"
    }
},
{
    "_id": "4479509126600228864",
    "name": "曲靖金都国际店",
    "extend_code": {
        "comm_shop_id": "9d771e01492f44d385ad6ae591687995",
        "ex_code": "46190",
        "us_id": "46190",
        "upcard_terminal": "87400744",
        "upcard_mer_id": "102874090000001",
        "ex_id": "46190",
        "ex_cost_center_code": "1200846190"
    }
},
{
    "_id": "4480596516605558784",
    "name": "盒马",
    "extend_code": {
        "us_id": "46034",
        "ex_id": "46034",
        "ex_cost_center_code": "1200046034",
        "ex_code": "46034"
    }
},
{
    "_id": "4480597822053318656",
    "name": "天猫宅配",
    "extend_code": {
        "us_id": "45929",
        "ex_id": "45929",
        "ex_cost_center_code": "1200045929",
        "ex_code": "45929"
    }
},
{
    "_id": "4480666118190596096",
    "name": "成都天府国际机场店",
    "extend_code": {
        "comm_shop_id": "5f865a7f1aac4d29aa62826923c325e0",
        "ex_code": "46166",
        "us_id": "46166",
        "upcard_terminal": "02835306",
        "upcard_mer_id": "102280090000023",
        "ex_id": "46166",
        "ex_cost_center_code": "1200846166"
    }
},
{
    "_id": "4481777193296363520",
    "name": "上海陆家嘴1885广场店",
    "extend_code": {
        "comm_shop_id": "2f362141dab34a52b17df697f1a6207d",
        "ex_code": "46192",
        "alipay_id": "2021062200077000000022903422",
        "us_id": "46192",
        "upcard_terminal": "02114252",
        "upcard_mer_id": "102210090000012",
        "ex_id": "46192",
        "ex_cost_center_code": "1200046192"
    }
},
{
    "_id": "4481779930155253760",
    "name": "武汉金桥永旺店",
    "extend_code": {
        "comm_shop_id": "9069ae69893f42ee9948041d59446995",
        "ex_code": "46184",
        "us_id": "46184",
        "upcard_terminal": "02732323",
        "upcard_mer_id": "102270090000005",
        "ex_id": "46184",
        "ex_cost_center_code": "1200846184"
    }
},
{
    "_id": "4481784262581846016",
    "name": "宁波万象城店",
    "extend_code": {
        "comm_shop_id": "6db8e587ef4642fb80495a136025088f",
        "ex_code": "46191",
        "alipay_id": "2021053100077000000021245689",
        "us_id": "46191",
        "upcard_terminal": "57405489",
        "upcard_mer_id": "102574090000001",
        "ex_id": "46191",
        "ex_cost_center_code": "1200046191"
    }
},
{
    "_id": "4482494311914569728",
    "name": "济宁吾悦店",
    "extend_code": {
        "comm_shop_id": "f3f560dbbffd4bf8b1e81a5f9de05173",
        "ex_code": "46196",
        "alipay_id": "2021072600077000000025212002",
        "us_id": "46196",
        "upcard_terminal": "53701207",
        "upcard_mer_id": "102537090000001",
        "ex_id": "46196",
        "ex_cost_center_code": "1200846196"
    }
},
{
    "_id": "4484597483533533184",
    "name": "南昌青山湖万象汇店",
    "extend_code": {
        "comm_shop_id": "f3389f20df6b4e8aa6ea862d9e9c3a56",
        "ex_code": "46195",
        "alipay_id": "2021053100077000000021245090",
        "us_id": "46195",
        "upcard_terminal": "79192604",
        "upcard_mer_id": "102791090000002",
        "ex_id": "46195",
        "ex_cost_center_code": "1200846195"
    }
},
{
    "_id": "4485772380838330368",
    "name": "扬中吾悦店",
    "extend_code": {
        "comm_shop_id": "cc8c8e214f1243ea82e659a3f6b54e9a",
        "ex_code": "46208",
        "alipay_id": "2021102600077000000029271049",
        "us_id": "46208",
        "upcard_terminal": "51104702",
        "upcard_mer_id": "102511090000001",
        "ex_id": "46208",
        "ex_cost_center_code": "1200846208"
    }
},
{
    "_id": "4485774483996573696",
    "name": "天水万达店",
    "extend_code": {
        "comm_shop_id": "36c394ab0a9a46659aabc427999e7edc",
        "ex_code": "46204",
        "alipay_id": "2021062200077000000022903853",
        "us_id": "46204",
        "upcard_terminal": "93800016",
        "upcard_mer_id": "102938090000001",
        "ex_id": "46204",
        "ex_cost_center_code": "1200846204"
    }
},
{
    "_id": "4488218227043762176",
    "name": "佛山王府井紫薇港店",
    "extend_code": {
        "comm_shop_id": "f4d73a13c9104eac9f493450cfdc294f",
        "ex_code": "46210",
        "alipay_id": "2021072600077000000025210470",
        "us_id": "46210",
        "upcard_terminal": "75704140",
        "upcard_mer_id": "102757090000008",
        "ex_id": "46210",
        "ex_cost_center_code": "1200846210"
    }
},
{
    "_id": "4488219218455592960",
    "name": "贵阳龙湾万达店",
    "extend_code": {
        "comm_shop_id": "2d319be99e0847e2add6cdc129f30a5c",
        "ex_code": "46211",
        "alipay_id": "2021061100077000000021913566",
        "us_id": "46211",
        "upcard_terminal": "85102973",
        "upcard_mer_id": "102851090000001",
        "ex_id": "46211",
        "ex_cost_center_code": "1200846211"
    }
},
{
    "_id": "4488223581198647296",
    "name": "长沙荟聚店",
    "extend_code": {
        "comm_shop_id": "c64ce2c856c0483e9de4ca8afe86cab5",
        "ex_code": "46217",
        "alipay_id": "2021062800077000000023605505",
        "us_id": "46217",
        "upcard_terminal": "73144432",
        "upcard_mer_id": "102731090000009",
        "ex_id": "46217",
        "ex_cost_center_code": "1200846217"
    }
},
{
    "_id": "4489298264534843392",
    "name": "上海漕河泾印象城店",
    "extend_code": {
        "comm_shop_id": "db7e2899f27f43fd9f7b3f747654f697",
        "ex_code": "46202",
        "alipay_id": "2021053100077000000021247690",
        "us_id": "46202",
        "upcard_terminal": "02114533",
        "upcard_mer_id": "102210090000018",
        "ex_id": "46202",
        "ex_cost_center_code": "1200046202"
    }
},
{
    "_id": "4492544034817998848",
    "name": "上海莘庄维璟印象城店",
    "extend_code": {
        "comm_shop_id": "5ce9676420a74e4eac639bcc6d716283",
        "ex_code": "46219",
        "alipay_id": "2021070700077000000023825635",
        "us_id": "46219",
        "upcard_terminal": "02114537",
        "upcard_mer_id": "102210090000022",
        "ex_id": "46219",
        "ex_cost_center_code": "1200846219"
    }
},
{
    "_id": "4494445383923040256",
    "name": "吴忠万达店",
    "extend_code": {
        "comm_shop_id": "949dc81f282f47ea93ad6343074dc24e",
        "ex_code": "46226",
        "alipay_id": "2021062800077000000023606749",
        "us_id": "46226",
        "upcard_terminal": "95300074",
        "upcard_mer_id": "102953090000001",
        "ex_id": "46226",
        "ex_cost_center_code": "1200846226"
    }
},
{
    "_id": "4494448431881814016",
    "name": "三亚海昌梦幻海洋不夜城店",
    "extend_code": {
        "comm_shop_id": "8946903a29d04ff1afd61878e31e6757",
        "ex_cost_center_code": "1200046229",
        "alipay_id": "2021073000077000000025330817",
        "us_id": "46229",
        "upcard_terminal": "89802021",
        "upcard_mer_id": "102898090000003",
        "ex_id": "46229",
        "ex_code": "46229"
    }
},
{
    "_id": "4495821609061351424",
    "name": "南京桥北万象汇店",
    "extend_code": {
        "comm_shop_id": "b06964d494f042769643c4df7c0106d9",
        "ex_code": "46180",
        "us_id": "46180",
        "upcard_terminal": "02519577",
        "upcard_mer_id": "102250090000002",
        "ex_id": "46180",
        "ex_cost_center_code": "1200846180"
    }
},
{
    "_id": "4495826708391297024",
    "name": "盘州合力生活广场店",
    "extend_code": {
        "comm_shop_id": "94b7831af5de4570ab180c018717b434",
        "ex_code": "46223",
        "alipay_id": "2021071900077000000024884862",
        "us_id": "46223",
        "upcard_terminal": "85800193",
        "upcard_mer_id": "102858090000001",
        "ex_id": "46223",
        "ex_cost_center_code": "1200846223"
    }
},
{
    "_id": "4495879005251076096",
    "name": "西宁中惠万达店",
    "extend_code": {
        "comm_shop_id": "6117deffddbb462fb14886725daaac14",
        "ex_code": "46225",
        "alipay_id": "2021110500077000000029761089",
        "us_id": "46225",
        "upcard_terminal": "97100456",
        "upcard_mer_id": "102971090000001",
        "ex_id": "46225",
        "ex_cost_center_code": "1200846225"
    }
},
{
    "_id": "4495880108218810368",
    "name": "淄博临淄茂业时代广场店",
    "extend_code": {
        "comm_shop_id": "5f3fa11c806047f0b7d9f8e77d8b82ae",
        "ex_code": "46207",
        "alipay_id": "2021071900077000000024887649",
        "us_id": "46207",
        "upcard_terminal": "53311402",
        "upcard_mer_id": "102533090000001",
        "ex_id": "46207",
        "ex_cost_center_code": "1200846207"
    }
},
{
    "_id": "4497287026141331456",
    "name": "宝鸡银泰店",
    "extend_code": {
        "comm_shop_id": "a286de62824e4c758e5f5b878839e6e2",
        "ex_code": "46239",
        "alipay_id": "2021062200077000000022903857",
        "us_id": "46239",
        "upcard_terminal": "91700391",
        "upcard_mer_id": "102917090000001",
        "ex_id": "46239",
        "ex_cost_center_code": "1200846239"
    }
},
{
    "_id": "4497382852150722561",
    "name": "武汉江宸天街店",
    "extend_code": {
        "comm_shop_id": "5a4b100a2c6a49d39984da0d5ff20a27",
        "ex_code": "46230",
        "alipay_id": "2021062200077000000022906327",
        "us_id": "46230",
        "upcard_terminal": "02732425",
        "upcard_mer_id": "102270090000016",
        "ex_id": "46230",
        "ex_cost_center_code": "1200046230"
    }
},
{
    "_id": "4497702865348886528",
    "name": "西安诗经里店",
    "extend_code": {
        "comm_shop_id": "aa6d2a3743fa469193b920cb5aaa500f",
        "ex_code": "46209",
        "us_id": "46209",
        "upcard_terminal": "02907849",
        "upcard_mer_id": "102290090000012",
        "ex_id": "46209",
        "ex_cost_center_code": "1200846209"
    }
},
{
    "_id": "4498408736999636992",
    "name": "深圳龙岗万达店",
    "extend_code": {
        "comm_shop_id": "fa9bdf57b89c4b1cb092219f87aa507b",
        "ex_code": "46238",
        "alipay_id": "2021091000077000000027686676",
        "us_id": "46238",
        "upcard_terminal": "75527364",
        "upcard_mer_id": "102755090000002",
        "ex_id": "46238",
        "ex_cost_center_code": "1200046238"
    }
},
{
    "_id": "4498412415974670336",
    "name": "菏泽佳和城店",
    "extend_code": {
        "comm_shop_id": "395d7b50198b4998b553508aff7c5b0b",
        "ex_code": "46240",
        "us_id": "46240",
        "upcard_terminal": "53000582",
        "upcard_mer_id": "102530090000001",
        "ex_id": "46240",
        "ex_cost_center_code": "1200846240"
    }
},
{
    "_id": "4498414215968620544",
    "name": "揭阳天虹店",
    "extend_code": {
        "comm_shop_id": "7aaf146133524171aa3375cb6b6b112b",
        "ex_code": "46236",
        "alipay_id": "2021072600077000000025208922",
        "us_id": "46236",
        "upcard_terminal": "66300081",
        "upcard_mer_id": "102663090000001",
        "ex_id": "46236",
        "ex_cost_center_code": "1200846236"
    }
},
{
    "_id": "4499794263996923904",
    "name": "眉山仁寿万达店",
    "extend_code": {
        "comm_shop_id": "7b4d8acd8b6644a5af3825ea926f12bc",
        "ex_code": "46237",
        "alipay_id": "2021102100077000000029116580",
        "us_id": "46237",
        "upcard_terminal": "02835871",
        "upcard_mer_id": "102280090000061",
        "ex_id": "46237",
        "ex_cost_center_code": "1200846237"
    }
},
{
    "_id": "4500598271296700416",
    "name": "镇江吾悦店",
    "extend_code": {
        "comm_shop_id": "975b675383b1497faf62e81c68e4c815",
        "ex_code": "46243",
        "alipay_id": "2021081200077000000025710507",
        "us_id": "46243",
        "upcard_terminal": "51104703",
        "upcard_mer_id": "102511090000002",
        "ex_id": "46243",
        "ex_cost_center_code": "1200846243"
    }
},
{
    "_id": "4500599131598782464",
    "name": "成都金牛凯德店",
    "extend_code": {
        "comm_shop_id": "136f6fc8a3f441e9bdd18281362129fc",
        "ex_code": "46246",
        "alipay_id": "2021091800077000000027857303",
        "us_id": "46246",
        "upcard_terminal": "02835911",
        "upcard_mer_id": "102280090000070",
        "ex_id": "46246",
        "ex_cost_center_code": "1200846246"
    }
},
{
    "_id": "4500599916701188096",
    "name": "成都群光广场店",
    "extend_code": {
        "comm_shop_id": "be361b920c9947ec91f6491d128c7649",
        "ex_code": "46245",
        "alipay_id": "2021071900077000000024884863",
        "us_id": "46245",
        "upcard_terminal": "02835910",
        "upcard_mer_id": "102280090000069",
        "ex_id": "46245",
        "ex_cost_center_code": "1200846245"
    }
},
{
    "_id": "4500602802285871104",
    "name": "珠海金湾华发商都店",
    "extend_code": {
        "comm_shop_id": "7b4937d4cc7845f8a663677faeb06033",
        "ex_code": "46244",
        "alipay_id": "2021092200077000000027976696",
        "us_id": "46244",
        "upcard_terminal": "75602328",
        "upcard_mer_id": "102756090000001",
        "ex_id": "46244",
        "ex_cost_center_code": "1200846244"
    }
},
{
    "_id": "4500603791319531520",
    "name": "嘉兴八佰伴华府店",
    "extend_code": {
        "comm_shop_id": "c53efc80d0ef4e20a5c3d7d50ed591a8",
        "ex_code": "46259",
        "alipay_id": "2021073000077000000025330818",
        "us_id": "46259",
        "upcard_terminal": "57304593",
        "upcard_mer_id": "102573090000001",
        "ex_id": "46259",
        "ex_cost_center_code": "1200846259"
    }
},
{
    "_id": "4502705714529075200",
    "name": "西安幸福林带店",
    "extend_code": {
        "comm_shop_id": "2b7b7199bad243c488c1545d49cda769",
        "ex_code": "46260",
        "alipay_id": "2021090200077000000027471584",
        "us_id": "46260",
        "upcard_terminal": "02907853",
        "upcard_mer_id": "102290090000013",
        "ex_id": "46260",
        "ex_cost_center_code": "1200846260"
    }
},
{
    "_id": "4502709321001369600",
    "name": "湖州德清正翔店",
    "extend_code": {
        "comm_shop_id": "d3a4be08ef8d4b3a94e4b733e77aacb7",
        "ex_code": "46258",
        "alipay_id": "2021080200077000000025417359",
        "us_id": "46258",
        "upcard_terminal": "57297336",
        "upcard_mer_id": "102572090000001",
        "ex_id": "46258",
        "ex_cost_center_code": "1200846258"
    }
},
{
    "_id": "4502710944087965696",
    "name": "衡阳酃湖万达店",
    "extend_code": {
        "comm_shop_id": "0f086b02323842f6aa62882926e80306",
        "ex_code": "46256",
        "us_id": "46256",
        "upcard_terminal": "73403778",
        "upcard_mer_id": "102734090000001",
        "ex_id": "46256",
        "ex_cost_center_code": "1200846256"
    }
},
{
    "_id": "4506044308085833728",
    "name": "上海宝山龙湖天街店",
    "extend_code": {
        "comm_shop_id": "17ac7c849add4b46af3031a5ecd3c637",
        "ex_code": "46266",
        "alipay_id": "2021090300077000000027500318",
        "us_id": "46266",
        "upcard_terminal": "02114564",
        "upcard_mer_id": "102210090000040",
        "ex_id": "46266",
        "ex_cost_center_code": "1200846266"
    }
},
{
    "_id": "4506045607728676864",
    "name": "邵阳步步高新天地店",
    "extend_code": {
        "comm_shop_id": "ff0a43ec026448e19274bcf43632821e",
        "ex_code": "46265",
        "us_id": "46265",
        "upcard_terminal": "73903888",
        "upcard_mer_id": "102739090000001",
        "ex_id": "46265",
        "ex_cost_center_code": "1200846265"
    }
},
{
    "_id": "4506047288893800448",
    "name": "衢州吾悦店",
    "extend_code": {
        "comm_shop_id": "efd7fd4c2b8c4683bc4d8c9a054e8ca4",
        "ex_code": "46257",
        "alipay_id": "2021083100077000000027426475",
        "us_id": "46257",
        "upcard_terminal": "57000401",
        "upcard_mer_id": "102570090000001",
        "ex_id": "46257",
        "ex_cost_center_code": "1200846257"
    }
},
{
    "_id": "4508577165077413888",
    "name": "遵义亨特店",
    "extend_code": {
        "comm_shop_id": "636a9218c9334e558872282be6e31698",
        "ex_code": "46262",
        "alipay_id": "2021101100077000000028380358",
        "us_id": "46262",
        "upcard_terminal": "85102978",
        "upcard_mer_id": "102851090000006",
        "ex_id": "46262",
        "ex_cost_center_code": "1200846262"
    }
},
{
    "_id": "4508578562267185152",
    "name": "成都建设路伊藤店",
    "extend_code": {
        "comm_shop_id": "9007d4e9efcf466180b591cdc6644bbc",
        "ex_code": "46274",
        "alipay_id": "2021092800077000000028117155",
        "us_id": "46274",
        "upcard_terminal": "02835954",
        "upcard_mer_id": "102280090000081",
        "ex_id": "46274",
        "ex_cost_center_code": "1200046274"
    }
},
{
    "_id": "4510790398676860928",
    "name": "南宁荟聚店",
    "extend_code": {
        "comm_shop_id": "b8e3c2d1cba3482cac474b44cc2a1b0b",
        "ex_code": "46272",
        "alipay_id": "2021102600077000000029272327",
        "us_id": "46272",
        "upcard_terminal": "77107141",
        "upcard_mer_id": "102771090000001",
        "ex_id": "46272",
        "ex_cost_center_code": "1200046272"
    }
},
{
    "_id": "4510793415975272448",
    "name": "太原泰享里店",
    "extend_code": {
        "comm_shop_id": "46f1986cc1874b258c479d3e44e84820",
        "ex_code": "46275",
        "alipay_id": "2021102600077000000029272261",
        "us_id": "46275",
        "upcard_terminal": "35106273",
        "upcard_mer_id": "102351090000191",
        "ex_id": "46275",
        "ex_cost_center_code": "1200846275"
    }
},
{
    "_id": "4512226906411663360",
    "name": "坊子泰华城店",
    "extend_code": {
        "comm_shop_id": "17514eeaea24444bb95e41d6ce4d6f02",
        "ex_code": "46273\t\t",
        "us_id": "46273",
        "upcard_terminal": "53608445",
        "upcard_mer_id": "102536090000010",
        "ex_id": "46273\t\t",
        "ex_cost_center_code": "1200046273"
    }
},
{
    "_id": "4512231359583911936",
    "name": "太原晋阳里公园店",
    "extend_code": {
        "comm_shop_id": "7cac129e69fe495aa24dbd6e330706be",
        "ex_code": "46277",
        "alipay_id": "2021102600077000000029272262",
        "us_id": "46277",
        "upcard_terminal": "35106301",
        "upcard_mer_id": "102351090000220",
        "ex_id": "46277",
        "ex_cost_center_code": "1200846277"
    }
},
{
    "_id": "4512235829717925888",
    "name": "盱眙苏宁广场店",
    "extend_code": {
        "comm_shop_id": "81901c7880514f299d96c3f368069eac",
        "ex_code": "70627",
        "us_id": "70627",
        "upcard_terminal": "51702563",
        "upcard_mer_id": "102517090000001",
        "ex_id": "70627",
        "ex_cost_center_code": "1200870627"
    }
},
{
    "_id": "4515053764266196992",
    "name": "京东POP店",
    "extend_code": {
        "us_id": "45407",
        "ex_id": "45407",
        "ex_cost_center_code": "1200045407",
        "ex_code": "45407"
    }
},
{
    "_id": "4515384089064275968",
    "name": "贵州遵义国贸店",
    "extend_code": {
        "comm_shop_id": "c9fd150bcaf34801ba314abeed662294",
        "ex_code": "46269",
        "alipay_id": "2015061200077000000000188800",
        "us_id": "46269",
        "upcard_terminal": "85102979",
        "upcard_mer_id": "102851090000007",
        "ex_id": "46269",
        "ex_cost_center_code": "1200046269"
    }
},
{
    "_id": "4516977979630321664",
    "name": "诸城百盛店",
    "extend_code": {
        "comm_shop_id": "72f6284a95bd497fb7cc0d6f29cfa444",
        "ex_code": "70638",
        "us_id": "70638",
        "upcard_terminal": "53608449",
        "upcard_mer_id": "102536090000014",
        "ex_id": "70638",
        "ex_cost_center_code": "1200870638"
    }
},
{
    "_id": "4517241789905666048",
    "name": "徐州新沂吾悦店",
    "extend_code": {
        "comm_shop_id": "efa388fb7a5946ccbe99bdddb450f45b",
        "ex_code": "70635",
        "us_id": "70635",
        "upcard_terminal": "51602013",
        "upcard_mer_id": "102516090000002",
        "ex_id": "70635",
        "ex_cost_center_code": "1200870635"
    }
},
{
    "_id": "4517247785256386560",
    "name": "苏州吴中龙湖天街店",
    "extend_code": {
        "comm_shop_id": "ec0680effccf4c9fb0881568b24f8fd6",
        "ex_code": "46261",
        "alipay_id": "2021100800077000000028308818",
        "us_id": "46261",
        "upcard_terminal": "51218107",
        "upcard_mer_id": "102512090000005",
        "ex_id": "46261",
        "ex_cost_center_code": "1200846261"
    }
},
{
    "_id": "4517266234120634368",
    "name": "揭阳万达店",
    "extend_code": {
        "comm_shop_id": "4af32924d4504b67846b61e89a3e1dce",
        "ex_code": "70634",
        "alipay_id": "2021102100077000000029116579",
        "us_id": "70634",
        "upcard_terminal": "66300082",
        "upcard_mer_id": "102663090000002",
        "ex_id": "70634",
        "ex_cost_center_code": "1200870634"
    }
},
{
    "_id": "4520214932869906432",
    "name": "上海龙茗店",
    "extend_code": {
        "comm_shop_id": "259ea3fe5a77499ea9166be2bbe17db6",
        "ex_code": "70645",
        "alipay_id": "2021102600077000000029272259",
        "us_id": "70645",
        "upcard_terminal": "02114687",
        "upcard_mer_id": "102210090000127",
        "ex_id": "70645",
        "ex_cost_center_code": "1200070645"
    }
},
{
    "_id": "4520875770144391168",
    "name": "永州万达店",
    "extend_code": {
        "comm_shop_id": "09649e26a2d8473097c31fd68ca7cb2f",
        "ex_code": "70648",
        "us_id": "70648",
        "upcard_terminal": "74601090",
        "upcard_mer_id": "102746090000001",
        "ex_id": "70648",
        "ex_cost_center_code": "1200870648"
    }
},
{
    "_id": "4521241753489014784",
    "name": "上海松江印象城店",
    "extend_code": {
        "comm_shop_id": "c2006a9350ed411e806b820bf39226b6",
        "ex_code": "70644",
        "alipay_id": "2021111800077000000030324038",
        "us_id": "70644",
        "upcard_terminal": "02114686",
        "upcard_mer_id": "102210090000126",
        "ex_id": "70644",
        "ex_cost_center_code": "1200070644"
    }
},
{
    "_id": "4521278359335895040",
    "name": "南京建邺吾悦店",
    "extend_code": {
        "comm_shop_id": "2fe63b7e5f0d49a088e9146e601e651c",
        "ex_code": "70649",
        "alipay_id": "2021091400077000000027762915",
        "us_id": "70649",
        "upcard_terminal": "02519981",
        "upcard_mer_id": "102250090000007",
        "ex_id": "70649",
        "ex_cost_center_code": "1200870649"
    }
},
{
    "_id": "4522084409429491712",
    "name": "上海国华广场店",
    "extend_code": {
        "comm_shop_id": "e35e47ed177f41d6b70db96f2f25552e",
        "ex_code": "70654",
        "alipay_id": "2021090700077000000027577645",
        "us_id": "70654",
        "upcard_terminal": "02114707",
        "upcard_mer_id": "102210090000129",
        "ex_id": "70654",
        "ex_cost_center_code": "1200070654"
    }
},
{
    "_id": "4522293316642963456",
    "name": "拉萨柳梧万达店",
    "extend_code": {
        "comm_shop_id": "5524f935f2984ae0b56ec97f8863b7ce",
        "ex_code": "70647",
        "us_id": "70647",
        "upcard_terminal": "89100387",
        "upcard_mer_id": "102891090000001",
        "ex_id": "70647",
        "ex_cost_center_code": "1200870647"
    }
},
{
    "_id": "4522990890479812608",
    "name": "苏州相城大悦春风里店",
    "extend_code": {
        "comm_shop_id": "7eeadf2701db484ebf3a65e946af1dfc",
        "ex_code": "70650",
        "alipay_id": "2021091800077000000027857304",
        "us_id": "70650",
        "upcard_terminal": "51218219",
        "upcard_mer_id": "102512090000006",
        "ex_id": "70650",
        "ex_cost_center_code": "1200870650"
    }
},
{
    "_id": "4523008775239532544",
    "name": "青岛胶东国际机场店",
    "extend_code": {
        "comm_shop_id": "b75985fbfdfb4450b3e28b0a8cb43263",
        "ex_code": "70656",
        "us_id": "70656",
        "upcard_terminal": "53206757",
        "upcard_mer_id": "102532090000059",
        "ex_id": "70656",
        "ex_cost_center_code": "1200070656"
    }
},
{
    "_id": "4523023333148622848",
    "name": "杭州萧山印象城店",
    "extend_code": {
        "comm_shop_id": "9760b2403c0f4eef8d9730998dd30404",
        "ex_code": "46278",
        "alipay_id": "2021101400077000000028489501",
        "us_id": "46278",
        "upcard_terminal": "57115829",
        "upcard_mer_id": "102571090000006",
        "ex_id": "46278",
        "ex_cost_center_code": "1200846278"
    }
},
{
    "_id": "4523025870983561217",
    "name": "南京河西龙湖店",
    "extend_code": {
        "comm_shop_id": "efb2f2d280be49118dd81ed64d03ac9c",
        "ex_code": "70639",
        "alipay_id": "2021091000077000000027690741",
        "us_id": "70639",
        "upcard_terminal": "02519980",
        "upcard_mer_id": "102250090000006",
        "ex_id": "70639",
        "ex_cost_center_code": "1200870639"
    }
},
{
    "_id": "4524930328336302080",
    "name": "荥阳吾悦广场店",
    "extend_code": {
        "comm_shop_id": "223959a613924e04abe21d5d81954ee4",
        "ex_code": "70657",
        "us_id": "70657",
        "upcard_terminal": "37115776",
        "upcard_mer_id": "102371090000023",
        "ex_id": "70657",
        "ex_cost_center_code": "1200870657"
    }
},
{
    "_id": "4524932521827270656",
    "name": "涡阳县绿城青牛广场店",
    "extend_code": {
        "comm_shop_id": "b2fbb4cb24f549d1b97ea95bfbba7ffc",
        "ex_code": "70658",
        "alipay_id": "2021111600077000000030285004",
        "us_id": "70658",
        "upcard_terminal": "55802661",
        "upcard_mer_id": "102558090000002",
        "ex_id": "70658",
        "ex_cost_center_code": "1200870658"
    }
},
{
    "_id": "4528141951385501696",
    "name": "广州花城汇店",
    "extend_code": {
        "comm_shop_id": "0f1b5496f803480ea1f3038e27715b5e",
        "ex_code": "70673",
        "us_id": "70673",
        "upcard_terminal": "02005716",
        "upcard_mer_id": "102200090000001",
        "ex_id": "70673",
        "ex_cost_center_code": "1200070673"
    }
},
{
    "_id": "4528148216052482048",
    "name": "银川建发现代城店",
    "extend_code": {
        "comm_shop_id": "85719f9801384f05aed771886b4f80e1",
        "ex_code": "70671",
        "us_id": "70671",
        "upcard_terminal": "95102690",
        "upcard_mer_id": "102951090000002",
        "ex_id": "70671",
        "ex_cost_center_code": "1200870671"
    }
},
{
    "_id": "4528412149984296960",
    "name": "上海安亭店",
    "extend_code": {
        "comm_shop_id": "04f9437bc55b4960bed9d411bb1968bb",
        "ex_code": "70674",
        "alipay_id": "2021102800077000000029406374",
        "us_id": "70674",
        "upcard_terminal": "02114744",
        "upcard_mer_id": "102210090000133",
        "ex_id": "70674",
        "ex_cost_center_code": "1200070674"
    }
},
{
    "_id": "4528416577919418368",
    "name": "贵阳小河万科店",
    "extend_code": {
        "comm_shop_id": "74d448ac18eb4de4b661de0c3b536027",
        "ex_code": "70676",
        "alipay_id": "2021100800077000000028308817",
        "us_id": "70676",
        "upcard_terminal": "85102983",
        "upcard_mer_id": "102851090000010",
        "ex_id": "70676",
        "ex_cost_center_code": "1200870676"
    }
},
{
    "_id": "4528803627491426304",
    "name": "合肥砂之船奥特莱斯店",
    "extend_code": {
        "comm_shop_id": "657a2fab5e01477da4aea0f54b4e6d99",
        "ex_code": "70667",
        "us_id": "70667",
        "upcard_terminal": "55133219",
        "upcard_mer_id": "102551090000004",
        "ex_id": "70667",
        "ex_cost_center_code": "1200870667"
    }
},
{
    "_id": "4528892160306610176",
    "name": "南京龙湾龙湖店",
    "extend_code": {
        "comm_shop_id": "acf56d06b7f0465ba2988d7617a599ab",
        "ex_code": "70669",
        "us_id": "70669",
        "upcard_terminal": "02523502",
        "upcard_mer_id": "102250090000009",
        "ex_id": "70669",
        "ex_cost_center_code": "1200870669"
    }
},
{
    "_id": "4529177541056954368",
    "name": "长沙大悦城店",
    "extend_code": {
        "comm_shop_id": "00e1e5a182534891af1dfedb7e06c6b8",
        "ex_code": "70672",
        "alipay_id": "2021092800077000000028118888",
        "us_id": "70672",
        "upcard_terminal": "73147576",
        "upcard_mer_id": "102731090000014",
        "ex_id": "70672",
        "ex_cost_center_code": "1200070672"
    }
},
{
    "_id": "4530321313568423936",
    "name": "上海宝杨宝龙店",
    "extend_code": {
        "comm_shop_id": "8ed238b50ace4d2f9ffe09f9f1656cff",
        "ex_code": "70675",
        "alipay_id": "2021111600077000000030257122",
        "us_id": "70675",
        "upcard_terminal": "02114757",
        "upcard_mer_id": "102210090000140",
        "ex_id": "70675",
        "ex_cost_center_code": "1200070675"
    }
},
{
    "_id": "4532479701786066944",
    "name": "普宁COCO City店",
    "extend_code": {
        "comm_shop_id": "d7e758744c994ca8979724b49b381c77",
        "ex_code": "70683",
        "alipay_id": "2021092800077000000028116044",
        "us_id": "70683",
        "upcard_terminal": "66300083",
        "upcard_mer_id": "102663090000003",
        "ex_id": "70683",
        "ex_cost_center_code": "1200870683"
    }
},
{
    "_id": "4533877648373350401",
    "name": "湖州织里吾悦店",
    "extend_code": {
        "comm_shop_id": "5130aed5676b490f8e7f658cfbf76465",
        "ex_code": "46279",
        "us_id": "46279",
        "upcard_terminal": "57297337",
        "upcard_mer_id": "102572090000002",
        "ex_id": "46279",
        "ex_cost_center_code": "1200846279"
    }
},
{
    "_id": "4533880126045192192",
    "name": "宿迁吾悦店",
    "extend_code": {
        "comm_shop_id": "d24c7eeb680d4fa5916379d949ebeeaa",
        "ex_code": "70670",
        "us_id": "70670",
        "upcard_terminal": "52703413",
        "upcard_mer_id": "102527090000001",
        "ex_id": "70670",
        "ex_cost_center_code": "1200870670"
    }
},
{
    "_id": "4534933924821925888",
    "name": "珠海城市阳台店",
    "extend_code": {
        "comm_shop_id": "207074d02c6b4491be35e7dbd5f6ba6b",
        "ex_code": "70684",
        "alipay_id": "2021111800077000000030329962",
        "us_id": "70684",
        "upcard_terminal": "75602378",
        "upcard_mer_id": "102756090000002",
        "ex_id": "70684",
        "ex_cost_center_code": "1200870684"
    }
},
{
    "_id": "4534936828324675584",
    "name": "嘉兴月河古街店",
    "extend_code": {
        "comm_shop_id": "1a7fc631706f40548f350e7de4d45583",
        "ex_code": "70686",
        "us_id": "70686",
        "upcard_terminal": "57304595",
        "upcard_mer_id": "102573090000003",
        "ex_id": "70686",
        "ex_cost_center_code": "1200870686"
    }
},
{
    "_id": "4535785321062793216",
    "name": "忻州古城店",
    "extend_code": {
        "comm_shop_id": "dacd9063d3fc49708d9a4ae9aa715309",
        "ex_code": "70687",
        "us_id": "70687",
        "upcard_terminal": "35000142",
        "upcard_mer_id": "102350090000001",
        "ex_id": "70687",
        "ex_cost_center_code": "1200870687"
    }
},
{
    "_id": "4536146426276675584",
    "name": "昆山万象汇店",
    "extend_code": {
        "comm_shop_id": "d427f757793843e3aab639ab58ca18df",
        "ex_code": "70692",
        "alipay_id": "2021110500077000000029760152",
        "us_id": "70692",
        "upcard_terminal": "51218225",
        "upcard_mer_id": "102512090000007",
        "ex_id": "70692",
        "ex_cost_center_code": "1200870692"
    }
},
{
    "_id": "4536745214791974912",
    "name": "杭州紫荆龙湖店",
    "extend_code": {
        "ex_code": "70694",
        "alipay_id": "2021102600077000000029271051",
        "us_id": "70694",
        "upcard_terminal": "57115834",
        "upcard_mer_id": "102571090000007",
        "ex_id": "70694",
        "ex_cost_center_code": "1200070694"
    }
},
{
    "_id": "4536750507336728576",
    "name": "福州东百店",
    "extend_code": {
        "ex_code": "70690",
        "alipay_id": "2021102600077000000029271068",
        "us_id": "70690",
        "upcard_terminal": "59115519",
        "upcard_mer_id": "102591090000009",
        "ex_id": "70690",
        "ex_cost_center_code": "1200870690"
    }
},
{
    "_id": "4536754436921425920",
    "name": "银川吾悦店",
    "extend_code": {
        "ex_cost_center_code": "1200870691",
        "us_id": "70691",
        "upcard_terminal": "95102693",
        "upcard_mer_id": "102951090000003",
        "ex_id": "70691",
        "ex_code": "70691"
    }
},
{
    "_id": "4538695655142522880",
    "name": "深圳丰盛町蛋糕店",
    "extend_code": {
        "ex_cost_center_code": "1200000003",
        "us_id": "00003",
        "upcard_terminal": "75527757",
        "upcard_mer_id": "102755090000004",
        "ex_id": "00003",
        "ex_code": "00003"
    }
},
{
    "_id": "4540886187553914880",
    "name": "仁怀方圆荟店",
    "extend_code": {
        "ex_code": "70689",
        "alipay_id": "2021111600077000000030253644",
        "us_id": "70689",
        "upcard_terminal": "85102984",
        "upcard_mer_id": "102851090000011",
        "ex_id": "70689",
        "ex_cost_center_code": "1200870689"
    }
},
{
    "_id": "4545582512111419392",
    "name": "商丘帝壹茂店",
    "extend_code": {
        "ex_code": "70716",
        "alipay_id": "2021111600077000000030257126",
        "us_id": "70716",
        "upcard_terminal": "37001404",
        "upcard_mer_id": "102370090000002",
        "ex_id": "70716",
        "ex_cost_center_code": "1200870716"
    }
},
{
    "_id": "4553489936130572288",
    "name": "武汉汉阳万达店",
    "extend_code": {
        "ex_cost_center_code": "1200070735",
        "us_id": "70735",
        "upcard_terminal": "02732837",
        "upcard_mer_id": "102270090000248",
        "ex_id": "70735",
        "ex_code": "70735"
    }
},
{
    "_id": "4553553070937669632",
    "name": "郑州富田新天地店",
    "extend_code": {
        "ex_cost_center_code": "1200870734",
        "us_id": "70734",
        "upcard_terminal": "37116062",
        "upcard_mer_id": "102371090000036",
        "ex_id": "70734",
        "ex_code": "70734"
    }
}]


new_comm_shop_id = {}
new_comm_code = {}
for i in shop_list:
    if i.get('extend_code', {}).get('comm_shop_id'):
        new_comm_shop_id[i.get('name')] = i.get('extend_code', {}).get('comm_shop_id') or ''
    if i.get('extend_code', {}).get('comm_code'):
        new_comm_code[i.get('name')] = i.get('extend_code', {}).get('comm_code') or ''

# print(new_comm_shop_id)
# print(new_comm_code)

# print(new_comm_shop_id.get('3850146064724131841'), new_comm_code.get('3850146064724131841'))

# path = '/Users/hws/Downloads/get_shop_jh_id.xlsx'
# wb = openpyxl.load_workbook(path)

# sh = wb['门店列表']

# rows = sh.max_row
# cols = sh.max_column

# print('==='+str(sh.cell(2, 8).value).replace(' ', '')+'---')
outwb = openpyxl.Workbook()
# outws = outwb.create_sheet('new_sheet')
outws = outwb.create_sheet(index=0)
for i in range(1, len(shop_list)):
    # for j in range(1, cols + 2):
    #     outws.cell(i, j).value = sh.cell(i, j).value if j != 10 else str(sh.cell(i, j).value)
    # 每行多加一列
    if i == 1:
        outws.cell(i, 1).value = '门店名称'
        outws.cell(i, 2).value = '美方ID'
        outws.cell(i, 3).value = '最新交行门店ID'
        outws.cell(i, 4).value = '最新交行积分门店ID'
    else:
        outws.cell(i, 1).value = shop_list[i - 1].get('name')
        outws.cell(i, 2).value = shop_list[i - 1].get('extend_code', {}).get('us_id')
        outws.cell(i, 3).value = shop_list[i - 1].get('extend_code', {}).get('comm_code')
        outws.cell(i, 4).value = shop_list[i - 1].get('extend_code', {}).get('comm_shop_id')

filename2 = '/Users/hws/Downloads/get_shop_jh_id_new.xlsx'
outwb.save(filename2)
print(filename2, '  down!!')
