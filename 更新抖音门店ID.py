import openpyxl
import pymongo
import requests
import json

path = '/Users/hws/Downloads/DQ门店POIID明细表2.xlsx'
wb = openpyxl.load_workbook(path)

sh = wb['Sheet3']

rows = sh.max_row
cols = sh.max_column
print(rows, cols)
update_dic = {}
cf_set = set()
for i in range(2, rows + 1):
	us_id = sh.cell(i, 3).value
	tiktok_id = sh.cell(i, 1).value.replace('\t', '')
	if us_id != '#N/A' and us_id != '美方ID':
		# print("美编ID：%s: 抖音门店ID：%s" % (us_id, tiktok_id))
		if us_id in update_dic:
			cf_set.add(us_id)
		update_dic[us_id] = tiktok_id
print(len(update_dic), rows - 1)
if cf_set:
	for c in cf_set:
		del update_dic[c]
print(cf_set)
print(len(update_dic), rows - 1)
mongo_client_pos = pymongo.MongoClient('127.0.0.1', 27072)

db_name_pos = mongo_client_pos.saas_dq

# print(update_dic)
headers = {
    'authorization': 'Bearer rvCywgEBMSKJeAOYkRD10A',
    'Cookie': 'hex_server_session={}'.format('27b1bbaf-4c41-4037-afe1-4e7f1eb93f0f'),
}

store_id_dict = {}
def get_store_id():
    url = 'http://store.dairyqueen.com.cn' + '/api/v1/store?code=all&include_state=true&include_total=true&relation=all&search_fields=extend_code.ex_code%2Cextend_code.us_id%2Cextend_code.ex_id%2Ccode%2Cname%2Caddress%2Crelation.geo_region.name%2Crelation.branch.name%2Crelation.distribution_region.name%2Crelation.attribute_region.name%2Crelation.formula_region.name%2Crelation.market_region.name%2Crelation.order_region.name&stringified=true&is_task=true&sort=extend_code.ex_code&order=asc&offset=0&limit=1600&state=draft%2Cenabled&status=&include_request=true&is_new=true&include_state=true&_=1625036536949'
    response = requests.get(url, headers=headers)
    data = response.json()
    n = 0
    try:
        dic = data['payload']
        lis = dic['rows']
        total = dic['total']
        for i in lis:
            us_id = i['extend_code']['us_id']
            _id = i['id']
            comm_shop_id = i['extend_code']['comm_shop_id'] if 'comm_shop_id' in i['extend_code'] else '0'
            if us_id in store_id_dict:
                print('请检查{}是否唯一'.format(us_id))
            if not us_id:
                n = n + 1
                print(n, _id)
            store_id_dict[str(us_id)] = str(_id)
            # store_shop_id_dict[us_id] = comm_shop_id
        # print(len(store_id_dict), store_id_dict)
    except IndexError as e:
        print('获取门店id出错', e)
        return
    if total != len(store_id_dict):
        print('门店数量不一致,请检查是否同名美编', 'total:{}'.format(total), 'dict:{}'.format(len(store_id_dict)))
        print('有{}家门店没有美编'.format(n))
        print('----------------------------------')
    return store_id_dict
def change_comm_shop_id(store_id, comm_id):
    store_have_problem_list = []
    put_change_url = 'http://store.dairyqueen.com.cn/api/v1/store/' + str(store_id) + '?stringified=true'
    param_body = {
        "extend_code": {
            "tiktok_shop_id": "{}".format(comm_id)
        }
    }
    put_change_jiaohang_res = requests.put(put_change_url, headers=headers, data=json.dumps(param_body)).json()
    if isinstance(put_change_jiaohang_res, dict):
        if put_change_jiaohang_res['payload']:
            get_store_info_url = 'http://store.dairyqueen.com.cn/api/v1/store/' + str(
                store_id) + '?is_task=true&relation=all&code=all&is_new=true&include_state=true&stringified=true'

            # 生成变更计划
            get_task_id_res = requests.get(get_store_info_url, headers=headers).json()['payload']
            task_id = get_task_id_res['request_id']
            store_name = get_task_id_res['name']
            if task_id:
                post_apply_url = 'http://store.dairyqueen.com.cn/api/v1/store/change/' + str(
                    task_id) + '/to/task?stringified=true'
                post_body = {"name": "{}".format(store_name), "immediate": True, "start_time": ""}
                post_res = requests.post(post_apply_url, headers=headers, data=json.dumps(post_body)).json()
                if isinstance(post_res, dict):
                    if post_res['payload']:
                        # 使变更计划生效
                        get_approve_info_url = 'http://store.dairyqueen.com.cn/api/v1/store/task?stringified=true&include_total=true&record_id=' + str(
                            store_id) + '&order=asc&offset=0&limit=1'
                        get_approve_id_res = requests.get(get_approve_info_url, headers=headers).json()['payload'][
                            'rows']
                        approve_id = get_approve_id_res[0]['id']
                        approve_url = 'http://store.dairyqueen.com.cn/api/v1/store/task/' + str(
                            approve_id) + '/status/APPROVED?stringified=true'
                        approve_res = requests.put(approve_url, headers=headers).json()
                        if isinstance(approve_res, dict):
                            if approve_res['payload']:
                                print('{}修改成功'.format(store_id))
                            else:
                                print('{}修改失败'.format(store_id))
                                store_have_problem_list.append(store_id)
                        else:
                            print('{}修改失败'.format(store_id))
                            store_have_problem_list.append(store_id)
                    else:
                        print('{}修改失败'.format(store_id))
                        store_have_problem_list.append(store_id)
                else:
                    print('{}修改失败'.format(store_id))
                    store_have_problem_list.append(store_id)
        else:
            print('{}未获得payload'.format(store_id))
            store_have_problem_list.append(store_id)
    else:
        print('{}修改失败！'.format(store_id))
        store_have_problem_list.append(store_id)
    return
store_info = get_store_id()
for u, t in update_dic.items():
	print(u, t)
	store_id = store_info.get(str(u))
	change_comm_shop_id(store_id, t)
	print(u, t, '更新成功！！！！！')
	# print(store_id)