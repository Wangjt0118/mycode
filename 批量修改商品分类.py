import requests
import json
import time
import xlrd
from openpyxl import Workbook, load_workbook

#请求头部
authorization = 'Bearer -Nak6qC7MnKdwKF7AYybNQ'
hex_server_session = '525109fb-8292-46bf-91fe-d5c2e0c4afb7'
test_url = 'http://test.dairyqueen.com.cn'
store_url = 'http://store.dairyqueen.com.cn'
store_ppj_url = 'http://store.papajohnshanghai.com'
test_be_url = 'http://hwstest.brutcakecafe.com'
store_be_url = 'http://store.brutcakecafe.com'

headers1 = {
    'authorization': '{}'.format(authorization),
    'Cookie': 'hex_server_session={}'.format(hex_server_session),
}

headers2 = {
        'authorization': '{}'.format(authorization),
        'Content-Type': 'application/json',
        'Cookie': 'hex_server_session={}'.format(hex_server_session),
    }

final_url = store_ppj_url

product_category_dict = {
    # "气泡酒": "4090051089326727169",
    # "红葡萄酒": "4040457617528483841",
    # "白葡萄酒": "4040457445163560961",
    # "桃红葡萄酒": "4090050979716980737",
    # "香槟": "4254888438505357313",
    # "自然酒": "4304491642519732225",
    # "创意特调": "4040458597879934977",
    # "水果特饮": "4129830273809276929",
    # "啤酒": "4297596158832197633",
    # "烈酒": "4297575309534216193",
    # "Brunch": "4225906075867734017",
    # "加料": "4040460792524337153",
    # "鸡尾酒": "4142559571498360833"

    "干货>原材料>干货-食材": "3932463413040226305",
    "干货>包材>干货-包材": "3932463417993699329",
    "干货>市场宣传品>干货-市场宣传品": "3932463420227198977",
    "干货>营运物料>干货-营运物料": "3932463418498736129",
    "干货>清洁用品>干货-清洁用品": "3932463419733700609"
}

def get_product_id(id):
    """
    根据商品编码拿到商品id
    """
    url = final_url + '/api/v1/product?search={}&is_task=true&state=draft%2Cenabled&status=&include_request=true' \
                        '&is_new=true&include_state=true&relation=all'.format(id)
    # print(url)
    response = requests.get(url, headers=headers1)
    # print(response.status_code)
    data = response.json()
    # print(type(data), data)
    product_id = None
    try:
        dic = data['payload']
        for i in dic:
            if i['code'] == str(id):
                product_id = i['id']
                old_category_id = str(i['relation']['product_category'])
    except IndexError as e:
        print('物料编码{}不存在'.format(id))
        return
    # dic = data['payload'][0]
    # product_id = dic['id']
    # print('product_id', product_id)
    # data = json.dumps(data, sort_keys=True, indent=4, separators=(',', ':'))
    # print(type(data), data)
    # print(product_id)
    if not product_id:
        print('物料编码{}不存在'.format(id))
        return
    return product_id, old_category_id


def change_category(id, category_id):
    url = final_url + '/api/v1/product/{}?stringified=true'.format(id)
    s = {"relation": {"product_category": '{}'.format(category_id)}}
    # print(s)
    data = json.dumps(s)
    response = requests.request("PUT", url, headers=headers1, data=data)
    data = response.json()
    return


def get_request_id2(id):
    """
    拿到所有变更计划的 request_id
    """
    url = final_url + '/api/v1/product/task?stringified=true&include_total=true&record_id={}&order=asc&offset=0&limit=1'.format(id)
    # print(url)
    response = requests.get(url, headers=headers1)
    # print(response.status_code)
    data = response.json()
    dic = data['payload']
    lis = dic['rows']
    # print('lis', lis)
    approve_id = lis[0]['id']
    return approve_id


def confirm_changes2(request_id):
    """
    审核变更计划
    """
    url = final_url + '/api/v1/product/task/{}/status/APPROVED?stringified=true'.format(request_id)
    # print(url)
    response = requests.request("PUT", url, headers=headers1)
    data = response.json()
    # print(data)
    # time.sleep(0.1)
    return


def get_request_id(code):
    url = final_url + '/api/v1/product?search={}'.format(code)
    # print(url)
    response = requests.get(url, headers=headers1)
    data = response.json()
    # print(type(data), data)
    request_id = None
    try:
        dic = data['payload']
        for i in dic:
            if i['code'] == str(code):
                request_id = i['request_id']
                name = i['name']
    except IndexError as e:
        print('物料编码{}不存在'.format(id))
        return
    if not request_id:
        print('物料编码{}不存在'.format(id))
        return
    return request_id, name


def accept_action2(name, request_id):
    """
    接受更改操作,用map去拼接正确的名字 eg.DQ全国
    """

    url = final_url + '/api/v1/product/change/{}/to/task?stringified=true'.format(request_id)
    s = {"name": "{}".format(name), "immediate": True, "start_time": ""}
    # print(s)
    data = json.dumps(s, ensure_ascii=False).encode("utf-8")
    # print(data)
    response = requests.post(url, headers=headers2, data=data)
    data2 = response.json()
    # time.sleep(0.1)
    return


def one_trun(code, name_, category, category_name):
    product_id, old_category = get_product_id(code)
    if not product_id:
        print('物料编码{}不存在'.format(id))
        return
    if category == old_category:
        print('商品{}原分类已经为{}不用修改'.format(name_, category_name))
        return
    change_category(product_id, category)
    request_id, name = get_request_id(code)
    accept_action2(name, request_id)
    request_id2 = get_request_id2(product_id)
    confirm_changes2(request_id2)
    print('商品{}更改分类为{}'.format(name, category_name))

    return


def test():
    wb = Workbook()
    ws = wb.active
    file_path = '/Users/hws/Downloads/PPJHEX物料属性分类修改-20221103.xlsx'
    # file_path = '/Users/hws/Downloads/ttttqqq.xlsx'
    # 打开一个已有文件
    wb = load_workbook(file_path)
    sheet_list = wb.sheetnames
    sheet = wb[wb.sheetnames[0]]
    total_list = []
    for r in range(2, sheet.max_row + 1):
        row_list = []  # 每一行建立一个list
        for c in range(1, sheet.max_column + 1):
            v = str(sheet.cell(r, c).value)
            v = v.replace('\n', '')
            row_list.append(v)
        total_list.append(row_list)
    print(len(total_list), total_list)
    return total_list


if __name__ == '__main__':
    lis = test()
    for i in lis:
        product_category = product_category_dict[i[3]]
        # product_category = 4115420549723475969
        one_trun(i[1], i[0], product_category, i[3])
        print('----------------------------------')
        time.sleep(2)

    # 使用说明：打开网页获取此时的authorization; 全局修改url为想要执行的品牌; 转换模版格式
    # 批量执行时, 可能会因执行一直发请求等原因报错, 重新执行即可————这条不确定，建议从执行失败的重新执行
    # 文件格式如下：商品编码｜ 商品名称                           ｜商品分类
    #            9330146 ｜ BTG10004旭金堡副牌                ｜气泡酒
    #         8010000029 ｜ SW10001一束花莫斯卡托阿斯蒂微起泡甜白 ｜红葡萄酒
    # 第三个参数为导入商品的分类,暂时做法从页面上去取，因此product_category_dict里必须有该分类
    # for i in lis:
    #     get_product_id(i[0])





