import requests
import json
import time
import xlrd
from openpyxl import Workbook, load_workbook

#请求头部
authorization = 'Bearer rH47mdvlctEnVE61bBbiD5HhUdAGFm'
hex_server_session = '9ab1418d-f21d-4b53-9db6-aec3509aa04d'
test_url = 'http://test.dairyqueen.com.cn'
store_url = 'http://store.dairyqueen.com.cn'
store_ppj_url = 'http://store.papajohnshanghai.com'
test_be_url = 'http://hwstest.brutcakecafe.com'
store_be_url = 'http://store.brutcakecafe.com'

headers1 = {
    'authorization': '{}'.format(authorization),
    'Cookie': 'hex_server_session={}'.format(hex_server_session),
}

headers2 = {
        'authorization': '{}'.format(authorization),
        'Content-Type': 'application/json',
        'Cookie': 'hex_server_session={}'.format(hex_server_session),
    }


product_category_dict = {
    "蔬菜碗": "4683546227026006017",
    "分享盘": "4683548178824400897",
    "小食": "4683550940505149441",
    "意面": "4683551127923429377",
    "三明治": "4683551245649154049",
    "主菜": "4683551541041401857",
    "蛋糕": "4683551630157778945",
    "咖啡": "4683551793865658369",
    "果汁": "4683551861532364801",
    "茶": "4683551934597140481",
    "水果特饮": "4683552040654311425",
    "鸡尾酒": "4683552203510747137"
}

final_url = store_be_url

def get_product_id(id):
    """
    根据商品编码拿到商品id
    """
    url = final_url + '/api/v1/product?search={}&is_task=true&state=draft%2Cenabled&status=&include_request=true' \
                        '&is_new=true&include_state=true'.format(id)
    # print(url)
    response = requests.get(url, headers=headers1)
    # print(response.status_code)
    data = response.json()
    # print(type(data), data)
    product_id = None
    try:
        dic = data['payload']
        for i in dic:
            if i['code'] == str(id):
                product_id = i['id']
    except IndexError as e:
        print('物料编码{}不存在'.format(id))
        return
    # dic = data['payload'][0]
    # product_id = dic['id']
    # print('product_id', product_id)
    # data = json.dumps(data, sort_keys=True, indent=4, separators=(',', ':'))
    # print(type(data), data)
    # print(product_id)
    if not product_id:
        print('物料编码{}不存在'.format(id))
        return
    return product_id


def create_one_product(code, name, category):
    """
    创建新商品
    """
    url = final_url + '/api/v1/product?stringified=true'
    false = False
    null = None
    s = {"code": "{}".format(code), "name": "{}".format(name), "name_en": "", "main_type": "NORMAL",
         "bom_type": "MADE", "storage_type": "NORMALTP", "category": "{}".format(category), "spec": "", "retail": null,
         "is_cup_measure": false, "extends": {}, "accounting_type": "MA", "brand_id": "3809835199387140099",
         "operation_type": "SELFSPT", "status": "ENABLED", "relation": {"product_category": "{}".format(category),
         "tag": []}, "extend_code": {"spec_1": ""}, "alarm_stock": null, "display_order": null}
    # print(s)
    data = json.dumps(s)
    response = requests.request("POST", url, headers=headers1, data=data)
    data = response.json()
    return


def accept_new_product(product_id):
    url = final_url + '/api/v1/product/{}/state/enable?stringified=true'.format(product_id)
    response = requests.request("PUT", url, headers=headers1)
    data = response.json()
    return


def create_region_attribute(product_id):
    url = final_url + '/api/v1/product/{}/region/attribute?stringified=true'.format(product_id)
    false = False
    true = True
    s = {"attribute": [{"id": "4027811609199616001", "inventory_type": "COUNT", "stocktake_circle": "DWM",
                        "allow_adjust": true, "allow_stocktake": true, "allow_transfer": true}]}
    # 添加属性区域，默认全国属性区域
    data = json.dumps(s)
    response = requests.request("PUT", url, headers=headers2, data=data)
    data1 = response.json()
    return


def get_region_rel_id(product_id):
    url = final_url + '/api/v1/product/{}/region/attribute?stringified=true&include_total=true&is_task=true&order=asc' \
                     '&state=draft%2Cenabled&status=&include_request=true&is_new=true&include_state=true'.format(product_id)
    response = requests.get(url, headers=headers1)
    # print(response.status_code)
    data = response.json()
    dic = data['payload']
    lis = dic['rows']
    rel_id = lis[0]['rel_id']
    return rel_id


def accept_region_attribute(rel_id):
    url = final_url + '/api/v1/product/region/attribute/{}/state/enable?stringified=true'.format(rel_id)
    response = requests.request("PUT", url, headers=headers1)
    data = response.json()
    return


def add_unit(product_id):
    url = final_url + '/api/v1/product/{}/unit?stringified=true'.format(product_id)
    false = False
    true = True
    s = [{"id": "4027815141868937217", "rate": 1, "default": true, "stocktake": true, "sales": true, "bom": true,
          "order": true, "recheck": false}]
    # 默认添加单位：个
    data = json.dumps(s)
    response = requests.request("PUT", url, headers=headers2, data=data)
    data1 = response.json()
    return


def get_request_id(code):
    url = final_url + '/api/v1/product?search={}'.format(code)
    # print(url)
    response = requests.get(url, headers=headers1)
    data = response.json()
    # print(type(data), data)
    request_id = None
    try:
        dic = data['payload']
        for i in dic:
            if i['code'] == str(code):
                request_id = i['request_id']
    except IndexError as e:
        print('物料编码{}不存在'.format(id))
        return
    if not request_id:
        print('物料编码{}不存在'.format(id))
        return
    return request_id


def accept_unit(request_id, name):
    url = final_url + '/api/v1/product/change/{}/to/task?stringified=true'.format(request_id)
    true = True
    s = {"name": "{}".format(name), "immediate": true, "start_time": ""}
    # print(s)
    data = json.dumps(s)
    response = requests.request("POST", url, headers=headers1, data=data)
    data = response.json()
    return


def confirm_unit(product_id):
    url = final_url + '/api/v1/product/task?stringified=true&include_total=true&record_id={}&order=asc'.format(product_id)
    response = requests.get(url, headers=headers1)
    data = response.json()
    dic = data['payload']
    lis = dic['rows']
    # print('lis', lis)
    request_id = lis[0]['id']
    return request_id


def approve_unit(request_id):
    url = final_url + '/api/v1/product/task/{}/status/APPROVED?stringified=true'.format(request_id)
    response = requests.request("PUT", url, headers=headers1)
    data = response.json()
    return


def one_trun(code, name, category):
    create_one_product(code, name, category)
    print('创建商品{}'.format(name))
    product_id = get_product_id(code)
    if not product_id:
        print('物料编码{}不存在'.format(code))
        return
    accept_new_product(product_id)
    print('生效商品{}'.format(name))
    create_region_attribute(product_id)
    rel_id = get_region_rel_id(product_id)
    accept_region_attribute(rel_id)
    print('商品{}添加属性区域'.format(name))
    add_unit(product_id)
    request_id = get_request_id(code)
    accept_unit(request_id, name)
    request_id2 = confirm_unit(product_id)
    approve_unit(request_id2)
    print('商品{}添加单位'.format(name))
    return


def test():
    wb = Workbook()
    ws = wb.active
    file_path = '/Users/hws/Downloads/be合阔主档导入-20221028.xlsx'
    # 打开一个已有文件
    wb = load_workbook(file_path)
    sheet_list = wb.sheetnames
    sheet = wb[wb.sheetnames[0]]
    total_list = []
    for r in range(2, sheet.max_row + 1):
        row_list = []  # 每一行建立一个list
        for c in range(1, sheet.max_column + 1):
            v = str(sheet.cell(r, c).value)
            v = v.replace('\n', '')
            row_list.append(v)
        total_list.append(row_list)
    print(len(total_list), total_list)
    return total_list


if __name__ == '__main__':
    lis = test()
    for i in lis:
        product_category = product_category_dict[i[11]]
        one_trun(i[1], i[2], product_category)
        print('----------------------------------')
        time.sleep(0.1)

    # 使用说明：打开网页获取此时的authorization; 全局修改url为想要执行的品牌; 转换模版格式
    # 批量执行时, 可能会因执行一直发请求等原因报错, 重新执行即可————这条不确定，建议从执行失败的重新执行
    # 文件格式如下：商品编码｜ 商品名称                           ｜商品分类
    #            9330146 ｜ BTG10004旭金堡副牌                ｜气泡酒
    #         8010000029 ｜ SW10001一束花莫斯卡托阿斯蒂微起泡甜白 ｜红葡萄酒
    # 第三个参数为导入商品的分类,暂时做法从页面上去取，因此一次导入的商品分类必须一致
    # 0810：优化成做成map,需要更新的分类放在product_category_dict中
    # for i in lis:
    #     get_product_id(i[0])





