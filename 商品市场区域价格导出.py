import requests
import openpyxl
import pymongo
# from product_ids import product_ids
import time
from random import randint
import multiprocessing
import os

def worker(num, product_ids):

	host = 'http://store.dairyqueen.com.cn'

	region_market_url = '/api/v1/region/market'

	product_url = '/api/v1/product/{}/region/market'

	USER_AGENTS = [
	 "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; AcooBrowser; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
	 "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0; Acoo Browser; SLCC1; .NET CLR 2.0.50727; Media Center PC 5.0; .NET CLR 3.0.04506)",
	 "Mozilla/4.0 (compatible; MSIE 7.0; AOL 9.5; AOLBuild 4337.35; Windows NT 5.1; .NET CLR 1.1.4322; .NET CLR 2.0.50727)",
	 "Mozilla/5.0 (Windows; U; MSIE 9.0; Windows NT 9.0; en-US)",
	 "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Win64; x64; Trident/5.0; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 2.0.50727; Media Center PC 6.0)",
	 "Mozilla/5.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0; WOW64; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 1.0.3705; .NET CLR 1.1.4322)",
	 "Mozilla/4.0 (compatible; MSIE 7.0b; Windows NT 5.2; .NET CLR 1.1.4322; .NET CLR 2.0.50727; InfoPath.2; .NET CLR 3.0.04506.30)",
	 "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN) AppleWebKit/523.15 (KHTML, like Gecko, Safari/419.3) Arora/0.3 (Change: 287 c9dfb30)",
	 "Mozilla/5.0 (X11; U; Linux; en-US) AppleWebKit/527+ (KHTML, like Gecko, Safari/419.3) Arora/0.6",
	 "Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.8.1.2pre) Gecko/20070215 K-Ninja/2.1.1",
	 "Mozilla/5.0 (Windows; U; Windows NT 5.1; zh-CN; rv:1.9) Gecko/20080705 Firefox/3.0 Kapiko/3.0",
	 "Mozilla/5.0 (X11; Linux i686; U;) Gecko/20070322 Kazehakase/0.4.5",
	 "Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.9.0.8) Gecko Fedora/1.9.0.8-1.fc10 Kazehakase/0.5.6",
	 "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11",
	 "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_3) AppleWebKit/535.20 (KHTML, like Gecko) Chrome/19.0.1036.7 Safari/535.20",
	 "Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; fr) Presto/2.9.168 Version/11.52",
	]

	random_agent = USER_AGENTS[randint(0, len(USER_AGENTS)-1)]

	headers = {
		'authorization': '{}'.format('Bearer wJxl5YgdMViyF74xBTWS-w'),  # 该env环境下的authorization
	    'Cookie': 'hex_server_session={}'.format('3ea89622-d1a7-42cc-b686-cef2db6128c9'),  # 该env环境下的Cookie
	}

	region_market_response = requests.get(host + region_market_url, headers=headers)

	region_market_res = region_market_response.json()['payload']
	# print(region_market_res)
	region_market_dic = {market['id']: {
		'name': market['name'],
		'code': market['code']
	} for market in region_market_res}

	# print(region_market_dic)



	mongo_client = pymongo.MongoClient(host='127.0.0.1', port=27072,)
	db_name = mongo_client.saas_dq

	query = {
		'status': 'ENABLED',
		'data_state': 'ENABLED'
	}

	result = db_name.product.find(query, {'_id': 1, 'name': 1, 'code': 1})


	product_dic = {r['_id']: r for r in result}

	return_res = []

	if not product_ids:
		p_ids = product_ids[1000 * (num - 1): 1000 * num]
	else:
		p_ids = product_ids

	for p_id in p_ids:
		print('进程名称: [%s], 商品id: [%s], 当前进度: %s%%' % (multiprocessing.current_process().name, p_id, round((p_ids.index(p_id) + 1) * 100 / len(p_ids), 3) ))
		random_agent = USER_AGENTS[randint(0, len(USER_AGENTS)-1)]
		headers.update({
			'User-Agent':random_agent,
			})
		try:
			product_info = requests.get(host + product_url.format(p_id), headers=headers)
		except Exception as e:
			flg = False
			product_info = None
			print('进程名称: [%s] 商品id: [%s] 失败' % (multiprocessing.current_process().name, p_id), '*' * 10)
			sleep_time = 3
			for i in range(3):
				time.sleep(sleep_time)
				random_agent = USER_AGENTS[randint(0, len(USER_AGENTS)-1)]
				headers.update({
					'User-Agent':random_agent,
					})
				try:
					product_info = requests.get(host + product_url.format(p_id), headers=headers)
					print('进程名称: [%s] 商品id: [%s] 第[%s]次成功' % (multiprocessing.current_process().name, p_id, i + 1), '*' * 10)
					flg = True
				except Exception as e:
					print(e)
				if flg:
					break
				sleep_time = sleep_time * 2
		product_info_res = product_info.json()['payload'] if product_info else []
		for i in product_info_res:
			d = dict()
			d['p_name'] = product_dic[p_id].get('name')
			d['p_code'] = product_dic[p_id].get('code')
			d['market_name'] = region_market_dic[i['id']]['name']
			d['market_code'] = region_market_dic[i['id']]['code']
			d['price'] = i['retail']
			return_res.append(d)
		time.sleep(1)
	print('**' * 20)
	print(return_res)
	print('--' * 20)
	outwb = openpyxl.Workbook()
	outws = outwb.create_sheet(index=0)
	title = ['商品名称', '商品编码', '市场区域名称', '市场区域编码', '价格']
	for i in range(1, 6):
		outws.cell(1, i).value = title[i -1]

	index = 2
	for product in return_res:
		outws.cell(index, 1).value = product['p_name']
		outws.cell(index, 2).value = product['p_code']
		outws.cell(index, 3).value = product['market_name']
		outws.cell(index, 4).value = product['market_code']
		outws.cell(index, 5).value = product['price']
		index += 1

	filename = '/Users/hws/Downloads/product_market_price{}.xlsx'.format(num)
	print(filename + '  down!!')
	outwb.save(filename)


if __name__ == '__main__':
	# print('888')
	# p_list = []
	# for i in range(1, 7):
	# 	p = multiprocessing.Process(target=worker, args=(i, ))
	# 	p_list.append(p)
	# for p in p_list:
	# 	p.start()
	# path = '/Users/hws/Downloads/奶昔.xlsx'
	# wb = openpyxl.load_workbook(path)

	# sh = wb['Sheet1']

	# rows = sh.max_row
	# cols = sh.max_column

	# product_codes = []
	# for i in range(2, rows + 1):
	# 	product_code = str(sh.cell(i, 2).value)
	# 	product_codes.append(product_code)
	# print(product_codes)
	# print(len(product_codes))
	mongo_client_pos = pymongo.MongoClient('127.0.0.1', 27072)

	db_name_pos = mongo_client_pos.saas_dq
	res = db_name_pos.product.find({
		'status': 'ENABLED',
		'data_state': 'ENABLED'
	})
	product_ids = []
	for r in res:
		print(r['_id'])
		product_ids.append(r['_id'])
	worker(1, product_ids)
