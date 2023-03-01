# coding:utf-8
import requests
import json
import time
import xlrd
from openpyxl import Workbook, load_workbook

#请求头部
authorization = 'Bearer LtWcuvA2P5CGMNvWYVIGCg'
hex_server_session = '5cc8acf0-328b-4a17-afa5-b9e46c5f9752'
test_url = 'http://test.dairyqueen.com.cn'
store_url = 'http://store.dairyqueen.com.cn'
store_ppj_url = 'http://store.papajohnshanghai.com'
test_be_url = 'http://hwstest.brutcakecafe.com'
store_be_url = 'http://store.brutcakecafe.com'

headers1 = {
    'authorization': '{}'.format(authorization),
    'Cookie': 'hex_server_session={}'.format(hex_server_session),
}

headers2 = {
        'authorization': '{}'.format(authorization),
        'Content-Type': 'charset=utf8',
        'Cookie': 'hex_server_session={}'.format(hex_server_session),
    }


def get_product_id(id):
    """
    根据商品编码拿到商品id
    """
    url = test_url + '/api/v1/product?search={}&is_task=true&state=draft%2Cenabled&status=&include_request=true' \
                        '&is_new=true&include_state=true'.format(id)
    # print(url)
    response = requests.get(url, headers=headers1)
    # print(response.status_code)
    data = response.json()
    # print(type(data), data)
    product_id = None
    try:
        dic = data['payload']
        for i in dic:
            if i['code'] == str(id):
                product_id = i['id']
    except IndexError as e:
        print('物料编码{}不存在'.format(id))
        return
    # dic = data['payload'][0]
    # product_id = dic['id']
    # print('product_id', product_id)
    # data = json.dumps(data, sort_keys=True, indent=4, separators=(',', ':'))
    # print(type(data), data)
    # print(product_id)
    if not product_id:
        print('物料编码{}不存在'.format(id))
        return
    return product_id


def change_category(id):
    url = test_url + '/api/v1/store/{}?stringified=true'.format(id)
    s = {"relation": {"branch": "3893257008677126145"}}
    # print(s)
    data = json.dumps(s)
    response = requests.request("PUT", url, headers=headers1, data=data)
    data = response.json()
    return


def get_request_id2(product_id):
    """
    拿到所有变更计划的 request_id
    """
    url = test_url + '/api/v1/store/task?stringified=true&include_total=true&record_id={}&order=asc&offset=0&limit=1'.format(product_id)
    # print(url)
    response = requests.get(url, headers=headers1)
    # print(response.status_code)
    data = response.json()
    dic = data['payload']
    lis = dic['rows']
    # print('lis', lis)
    approve_id = lis[0]['id']
    return approve_id


def confirm_changes2(request_id):
    """
    审核变更计划
    """
    url = test_url + '/api/v1/store/task/{}/status/APPROVED?stringified=true'.format(request_id)
    # print(url)
    response = requests.request("PUT", url, headers=headers1)
    data = response.json()
    # print(data)
    # time.sleep(0.1)
    return


def get_request_id(code):
    url = test_url + '/api/v1/store/{}?is_task=true&relation=all&code=all&is_new=true&include_state=true&stringified=true'.format(code)
    # print(url)
    response = requests.get(url, headers=headers1)
    data = response.json()
    # print(type(data), data)
    request_id = None
    try:
        dic = data['payload']
        request_id = dic['request_id']
        name = dic['name']
    except IndexError as e:
        print('门店{}不存在'.format(code))
        return
    if not request_id:
        print('门店{}修改出错'.format(code))
        return
    return request_id, name


def accept_action2(name, request_id):
    """
    接受更改操作,用map去拼接正确的名字 eg.DQ全国
    """

    url = test_url + '/api/v1/store/change/{}/to/task?stringified=true'.format(request_id)
    s = {"name": "{}".format(name), "immediate": True, "start_time": ""}
    # print(s)
    data = json.dumps(s, ensure_ascii=False).encode("utf-8")
    # print(data)
    response = requests.post(url, headers=headers2, data=data)
    data2 = response.json()
    # time.sleep(0.1)
    return


def one_trun(code):
    # product_id = get_product_id(code)
    # if not product_id:
    #     print('物料编码{}不存在'.format(id))
    #     return
    change_category(code)
    # print('生效商品{}'.format(name))

    request_id, name = get_request_id(code)

    accept_action2(name, request_id)
    request_id2 = get_request_id2(code)
    confirm_changes2(request_id2)
    print('门店{}修改区经理'.format(name))
    return


def test():
    wb = Workbook()
    ws = wb.active
    file_path = '/Users/yjq/Desktop/test.xlsx'
    # 打开一个已有文件
    wb = load_workbook(file_path)
    sheet_list = wb.sheetnames
    sheet = wb[wb.sheetnames[0]]
    total_list = []
    for r in range(2, sheet.max_row + 1):
        row_list = []  # 每一行建立一个list
        for c in range(1, sheet.max_column + 1):
            v = str(sheet.cell(r, c).value)
            v = v.replace('\n', '')
            row_list.append(v)
        total_list.append(row_list)
    print(len(total_list), total_list)
    return total_list


if __name__ == '__main__':
    # lis = test()
    # for i in lis:
    #     one_trun(i[0], i[1], '4040457617528483841')
    #     print('----------------------------------')
    #     time.sleep(0.1)
    one_trun(3939604095268151297)

    # 使用说明：打开网页获取此时的authorization; 全局修改url为想要执行的品牌; 转换模版格式
    # 批量执行时, 可能会因执行一直发请求等原因报错, 重新执行即可————这条不确定，建议从执行失败的重新执行
    # 文件格式如下：商品编码｜ 商品名称
    #            9330146 ｜ BTG10004旭金堡副牌
    #         8010000029 ｜ SW10001一束花莫斯卡托阿斯蒂微起泡甜白
    # 第三个参数为导入商品的分类,暂时做法从页面上去取，因此一次导入的商品分类必须一致
    # for i in lis:
    #     get_product_id(i[0])





