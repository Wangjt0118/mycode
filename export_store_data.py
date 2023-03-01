import pymongo
import openpyxl
from dateutil.parser import parse


mongo_client_pos = pymongo.MongoClient('127.0.0.1', 27071)

db_name_pos = mongo_client_pos.saas_dq

start_date = '2021-08-01'
end_date = '2021-08-02'
# 不包括start_date, 包括end_date

start = '{}T16:00:00'.format(start_date)
end = '{}T16:00:00'.format(end_date)
query = {
        "{}".format('sales_date'): {"$gte": parse(start), "$lte": parse(end)},
        # "payload.body.item_count": {"$gte": 3},
        # "payload.body.item_count": 1,
        # "payload.body.items.item_id": {"$in": ["9902369", "9902374", "9902376", "9902377", "9902655", "9902654", "9902653", "9902652", "9902651", "9902650", "9902649", "9902648", "9902647", "9902646", "9902645", "9902644", "TC0001", "TC0003", "TC0005", "TC0006", "TC0007", "TC0008", "TC0009", "TC0010", "TC0011", "TC0014", "TC0015", "TC0016", "TC0017", "TC0018", "TC0019", "TC0025", "TC0026", "TC0027", "TC0028", "TC0029", "TC0030", "TC0032", "TC0033", "TC0034", "TC0035", "TC0036", "TC0037", "TC0038", "TC0039", "TC0040", "TC0041", "TC0042", "TC0043", "TC0044", "TC0045", "TC0046", "TC0047", "TC0048", "TCC0049", "TC0050", "TC0051", "TC0052", "TC0049", "TC0054", "TC0055", "TC0056", "TC0057", "TC0058", "TC0059", "TC0060", "TC0061", "TC0062", "TC0063", "TC0064", "TC0065", "TC0067", "TC0067", "TC0068", "TC0069", "TC0070", "TC0071", "TC0079", "TC0080", "TC0081", "TC0082", "TC0083", "TC0084", "TC0085", "TC0086", "TC0087", "TC0088", "TC0089", "TC0090", "TC0091", "TC0092", "TC0093", "TC0094", "TC0095", "TC0096", "TC0097", "TC0098", "TC0099", "TC0100", "TC0101", "TC0102", "TC0103", "TC0104", "TC0105", "TC0106", "TC0107", "TC0108", "TC0109", "TC0110", "TC0111", "TC0112", "TC0113", "TC0114", "TC0115", "TC0116", "SPU000000", "TC0123", "TC0144", "TC0152", "TC0153", "TC0190", "TC0191", "TC0192", "TC0193", "TC0197", "TC0198", "TC0199", "TC0203", "TC0204", "TC0205", "TC0206", "TC0219", "TC0230", "TC0242", "TC0243", "TC0244", "TC0245", "TC0246", "TC0247", "TC0248", "TC0251", "TC0252", "TC0253", "TC0254", "TC0255", "TC0256", "TC0257", "TC0283", "TC0285", "TC0286", "TC0287", "TC0288", "TC0289", "TC0296", "TC0306", "TC0330", "TC0379", "TC0404", "TC0420", "TC0421", "TC0430", "TC0435", "TC0436", "TC0447", "TC0457", "TC0461", "TC0465", "TC0505", "TC0508", "TC0509", "TC0510", "TC0521", "TC0541", "TC0551", "TC0553", "TC0591", "TC0592", "TC0596", "TC0615", "TC0664", "TC0666", "TC0667", "TC0668", "TC0669", "TC0670", "TC0671", "TC0672", "TC0673", "TC0674", "TC0675", "TC0676", "TC0677", "TC0678", "TC0679", "DP00191", "DP00192", "DP00193", "DP00194", "DP00195", "DP00296", "DP00466", "DP00735", "DP00766", "DP00833", "DP00454", "DP00455", "DP00456", "DP00457", "9902110", "9902111", "9902628", "9902629", "DP00108", "DP00109", "DP00110", "DP00111", "DP00317", "DP00458", "DP00465", "DP00483", "DP00701", "9902375", "DP00512", "DP00521", "DP00251", "DP00297", "DP00298", "DP00299", "9900349", "9900488", "9900492", "9900501", "9900496", "9901859", "9901824", "9901860", "9901861", "9901864", "9901862", "9901878", "9901863", "9901875", "9901876", "9901880", "9901881", "9901879", "9901877", "9901882", "9901887", "9901884", "9901883", "9901886", "9902001", "9902002", "9901888", "9901885", "9902666", "9902667", "DP00048", "DP00049", "DP00141", "DP00142", "DP00291", "DP00334", "DP00692"]},
        # "payload.body.items.item_id": {"$in": ["TC0853"]},
        "payload.body.payments.tender_name": "微营销",
        # "payload.body.item_count": {"$gte": 2}
        "payload.body.hex_discount_list.name": {"$in": ["微营销优惠", "微营销优惠券"]},
        # "payload.body.is_cancel": True,
        # "payload.body.hex_discount_list.name": "复苏买一送一",
        # "store_name": "南通海安万达店"
        # "store_name": "上海宝山安信店",
        "store_type": "DRS",
        # "payload.body.hex_discount_list.code": {"$in": ["201907301", "202007150", "202002051", "202001311", "202001312",
        #                                                 "202001314", "202001315", "202001313", "D201712066"]},
        # "payload.body.items.category.name": "轻食",
        "payload.body.items.category.id": {"$in": ['4356757511066386433', '3895644930529886138', '3895644930529886139', '4427081980707897345', '4420213253303500801']}
    }
mongoret = db_name_pos.pos.find(query, {"store_name": 1, "store_us_id": 1, "payload.body.order_no": 1, "sales_date": 1,
                                    "terminal_open_time": 1, "created": 1,
                                    "payload.body.is_cancel": 1,
                                    "payload.body.order_unique_no": 1, "payload.body.dept_class_name": 1,
                                    "payload.body.pos_name": 1, "payload.body.payments": 1,
                                    "payload.body.items.amount": 1,
                                    "payload.body.items.dis_amount": 1,
                                    "payload.body.items.hex_dis_amount": 1,
                                    "payload.body.items.hex_net_amount": 1,
                                    "payload.body.items.item_id": 1,
                                    "payload.body.items.item_key_id": 1,
                                    "payload.body.items.name": 1,
                                    "payload.body.items.qty": 1,
                                    "payload.body.hex_discount_list": 1,
                                    "payload.body.items.category": 1
                                    })

outwb = openpyxl.Workbook()
outws = outwb.create_sheet(index=0)
title = ['门店编码', '门店id', '销售明细', '品类', '折扣方式']
for i in range(1, len(title) + 1):
	outws.cell(1, i).value = title[i -1]

index = 2
for r in mongoret:
	store_name = r.get('store_name')
	store_id = r.get('store_us_id')

	for p in r['payload']['body']['items']:
		product_name = p.get('name')
		category_name = p.get('category').get('name')

		for z in r['payload']['body']['hex_discount_list']:

			outws.cell(index, 1).value = store_name
			outws.cell(index, 2).value = store_id
			outws.cell(index, 3).value = product_name
			outws.cell(index, 4).value = category_name













