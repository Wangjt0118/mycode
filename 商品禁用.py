import requests
import json
import time
import xlrd
from openpyxl import Workbook, load_workbook

#请求头部
authorization = 'Bearer yL-w9ubKMPu9yH3113mIqg'
hex_server_session = '9f697a29-fc04-4cd4-871b-5ebb814f081f'
test_url = 'http://test.dairyqueen.com.cn'
store_url = 'http://store.dairyqueen.com.cn'
store_ppj_url = 'http://store.papajohnshanghai.com'
test_be_url = 'http://hwstest.brutcakecafe.com'
store_be_url = 'http://store.brutcakecafe.com'


headers1 = {
    'authorization': '{}'.format(authorization),
    'Cookie': 'hex_server_session={}'.format(hex_server_session),
}

headers2 = {
        'authorization': '{}'.format(authorization),
        'Content-Type': 'application/json',
        'Cookie': 'hex_server_session={}'.format(hex_server_session),
    }


final_url = store_ppj_url


def get_product_id(id):
    """
    根据商品编码拿到商品id及单位
    """
    url = final_url + '/api/v1/product?search={}&is_task=true&state=draft%2Cenabled&status=&include_request=true' \
                        '&is_new=true&relation=all&include_state=true'.format(id)
    # print(url)
    response = requests.get(url, headers=headers1)
    # print(response.status_code)
    data = response.json()
    # print(type(data), data)
    product_id = None
    try:
        dic = data['payload']
        for i in dic:
            if i['code'] == str(id):
                product_id = i['id']
                status = i['status']
    except IndexError as e:
        print('物料编码{}不存在'.format(id))
        return 0, 0
    if not product_id:
        print('物料编码{}不存在'.format(id))
        return 0, 0
    return product_id, status


def get_region_order_rel_id(product_id):
    url = final_url + '/api/v1/product/{}/region/order?stringified=true&include_total=true&is_task=true&order=asc' \
                     '&state=draft%2Cenabled&status=&include_request=true&is_new=true&include_state=true'.format(product_id)
    response = requests.get(url, headers=headers1)
    # print(response.status_code)
    data = response.json()
    dic = data['payload']
    lis = dic['rows']
    for i in lis:
        if i['id'] == '4635061286305038336':
            rel_id_new = i['rel_id']
            break
    return rel_id_new


def accept_region_order(rel_id):
    url = final_url + '/api/v1/product/region/order/{}/state/disable?stringified=true'.format(rel_id)
    response = requests.request("PUT", url, headers=headers1)
    data = response.json()
    return


def get_request_id(code):
    url = final_url + '/api/v1/product?search={}'.format(code)
    # print(url)
    response = requests.get(url, headers=headers1)
    data = response.json()
    # print(type(data), data)
    request_id = None
    try:
        dic = data['payload']
        for i in dic:
            if i['code'] == str(code):
                request_id = i['request_id']
                name = i['name']
    except IndexError as e:
        print('物料编码{}不存在'.format(id))
        return
    if not request_id:
        print('物料编码{}不存在'.format(id))
        return
    return request_id, name


def accept_action2(name, request_id):
    """
    接受更改操作,用map去拼接正确的名字 eg.DQ全国
    """

    url = final_url + '/api/v1/product/change/{}/to/task?stringified=true'.format(request_id)
    s = {"name": "{}".format(name), "immediate": True, "start_time": ""}
    # print(s)
    data = json.dumps(s, ensure_ascii=False).encode("utf-8")
    # print(data)
    response = requests.post(url, headers=headers2, data=data)
    data2 = response.json()
    # time.sleep(0.1)
    return


def disable_status(pid):
    url = final_url + '/api/v1/product/{}?stringified=true'.format(pid)
    s = {"status": "DISABLED"}
    # print(s)
    data = json.dumps(s)
    response = requests.request("PUT", url, headers=headers1, data=data)
    data = response.json()
    return


def one_trun(code):
    product_id, status = get_product_id(code)
    if not product_id:
        print('物料编码{}不存在'.format(code))
        return
    if status == 'DISABLED':
        print('商品{}已经是禁用状态'.format(code))
        return
    disable_status(product_id)
    print('商品{}修改为禁用状态'.format(code))
    request_id, name = get_request_id(code)
    accept_action2(name, request_id)
    print('商品{}创建变更计划'.format(code))
    return


def test():
    wb = Workbook()
    ws = wb.active
    file_path = '/Users/yjq/Desktop/test.xlsx'
    # 打开一个已有文件
    wb = load_workbook(file_path)
    sheet_list = wb.sheetnames
    sheet = wb[wb.sheetnames[0]]
    total_list = []
    for r in range(2, sheet.max_row + 1):
        row_list = []  # 每一行建立一个list
        for c in range(1, sheet.max_column + 1):
            v = str(sheet.cell(r, c).value)
            v = v.replace('\n', '')
            row_list.append(v)
        total_list.append(row_list)
    print(len(total_list), total_list)
    return total_list


if __name__ == '__main__':
    lis = test()
    for i in lis:
        one_trun(i[0])
        print('----------------------------------')
        time.sleep(0.1)