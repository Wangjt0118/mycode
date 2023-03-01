import requests
import openpyxl

host = {
	'dq': 'http://store.dairyqueen.com.cn',
	'ppj': 'https://store.papajohnshanghai.com'
}

url = '/api/v1/staff?code=all&include_total=true&include_user=true&relation=all&search_fields=name%2Ccode%2Crelation.store.name&offset={}&limit={}&stringified=true&search=&sort=name&order=desc&state=draft%2Cenabled&status=&include_request=true&is_new=true&include_state=true&_=1650014759358'

def main(env, headers):
	return_res = []
	for i in range(21):
		print(i)
		post_url = '{}{}'.format(host.get(env), url.format(i * 1000, 1000))
		resp = requests.get(post_url, headers=headers)
		# print(resp.json(), post_url, env, headers)
		user_infos = resp.json()['payload']['rows']

		# print(user_infos)
		
		for rs in user_infos:
			d = dict()
			d['code'] = rs.get('code')
			d['name'] = rs.get('name')
			d['login'] = rs.get('user').get('name') if rs.get('user') else '-'
			d['data_state'] = rs.get('data_state')
			d['created'] = rs['created']
			d['updated'] = rs['updated']
			return_res.append(d)



	outwb = openpyxl.Workbook()
	outws = outwb.create_sheet(index=0)
	title = ['员工编号', '名字', '登录名', '状态', '创建时间', '更新时间']

	for i in range(1, 7):
		outws.cell(1, i).value = title[i -1]

	index = 2
	for product in return_res:
		outws.cell(index, 1).value = product['code']
		outws.cell(index, 2).value = product['name']
		outws.cell(index, 3).value = product['login']
		outws.cell(index, 4).value = product['data_state']
		outws.cell(index, 5).value = product['created']
		outws.cell(index, 6).value = product['updated']
		index += 1

	filename = '/Users/hws/Downloads/{}_userinfo.xlsx'.format(env)
	print(filename + '  down!!')
	outwb.save(filename)


if __name__ == '__main__':
	# env = 'ppj'
	# headers = {
	# 	'authorization': '{}'.format('Bearer EOK2CD68OLybgnCqNXgUsw'),  # 该env环境下的authorization
	#     'Cookie': 'hex_server_session={}'.format('f9ae2e68-27c0-47f6-b2c4-7cf9153b615e'),  # 该env环境下的Cookie
	# }
	env = 'dq'
	headers = {
		'authorization': '{}'.format('Bearer vRJt4vZ-POODmZLvTurWmA'),  # 该env环境下的authorization
	    'Cookie': 'hex_server_session={}'.format('91e7a4f8-2b8f-4de8-881a-e6611b5ce77d'),  # 该env环境下的Cookie
	}
	main(env, headers)